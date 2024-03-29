# Django
from ctypes import alignment
from functools import reduce
from math import prod
import operator
from operator import or_
from http.client import HTTPResponse
import re
from tkinter import CENTER, N
import matplotlib.pyplot as plt
import numpy as np
from django import forms
from django.contrib import messages
from django.contrib.auth import views as auth_views
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core import serializers
from django.db.models import Q, F, ExpressionWrapper, IntegerField, Value
from django.forms import DateField, DateTimeField, DurationField
from django.db.models import FloatField, F, Q, Case, When, Value, IntegerField, CharField, ExpressionWrapper, Func
from django.db.models.functions import Least
from django.http import HttpResponseRedirect
from django.http.response import JsonResponse, HttpResponse
from django.core.files.uploadedfile import TemporaryUploadedFile
from django.db.models.aggregates import Min, Max, Count
from django.db.models.functions import Cast, ExtractMonth, ExtractDay, Now, Round, Substr, ExtractYear
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
import pytz
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import CreateView, ListView, TemplateView, DetailView, DeleteView, UpdateView, FormView
from django.views.generic.base import View
from psycopg2 import IntegrityError
from requests import request
from openpyxl.styles import Font
import logging
logging.basicConfig( level=logging.DEBUG )
# Functions
from dashboards.functions import functions, functions_ftp, functions_create, functions_excel
from dashboards.functions.functions import DiffDays, CastDate, mala_entrada, presion_establecida, presupuesto
from aeto import settings
#from dashboards import GoogleDrive

# Forms
from dashboards.forms.forms import EdicionManual, ExcelForm, InspeccionForm, VehiculoForm, ProductoForm, RenovadorForm, DesechoForm, DesechoEditForm, ObservacionForm, ObservacionEditForm, RechazoForm, RechazoEditForm, SucursalForm, TallerForm, UsuarioForm, AplicacionForm, CompaniaForm, UsuarioEditForm, SucursalEditForm, TallerEditForm, AplicacionEditForm, VehiculoEditForm, PresupuestoForm, PresupuestoEditForm

# Models
from django.contrib.auth.models import User, Group
from dashboards.models import Aplicacion, Bitacora_Pro, Inspeccion, InspeccionVehiculo, Llanta, LlantasSeleccionadas, Orden, OrdenDesecho, Producto, Servicio, ServicioLlanta, ServicioVehiculo, Taller, Ubicacion, Vehiculo, Perfil, Bitacora, Compania, Renovador, Desecho, Observacion, Rechazo, User, Observacion, Presupuesto


# Utilities
import math
from multi_form_view import MultiModelFormView
import csv
from datetime import date, datetime, timedelta
from ftplib import FTP as fileTP
import json
import logging
import mimetypes
import openpyxl
from openpyxl.chart import BarChart, Reference
import os
import pandas as pd

import statistics
import xlwt

from dashboards.views.views_rest import CarritoLlantasApi

import dashboards.GoogleDrive as Google

class LoginView(auth_views.LoginView):
    # Vista de Login

    template_name = "login.html"
    redirect_authenticated_user = True

class HomeView(LoginRequiredMixin, TemplateView):
    # Vista de nuevo home

    template_name = "home.html"
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        usuario = self.request.user
        perfil = Perfil.objects.get(user = usuario)
        compania = perfil.compania
        ubicacion = perfil.ubicacion.all()
        aplicacion = perfil.aplicacion.all()
        taller = perfil.taller.all()
        context["compania"] = compania
        context["ubicacion"] = ubicacion
        context["aplicacion"] = aplicacion
        context["taller"] = taller
        return context
    
    def post(self, request, *args, **kwargs):
        #Obtencion de la data
        print(request.POST)
        user = request.user
        perfil = Perfil.objects.get(user = user)
        compania_post = request.POST.getlist('company')[0]
        sucursal_post = request.POST.getlist('sucursal')
        aplicacion_post = request.POST.getlist('aplicacion')
        taller_post = request.POST.getlist('taller')
        #Consulta de la lista del post
        compania = Compania.objects.get(compania = compania_post)
        sucursales = Ubicacion.objects.filter(nombre__in = sucursal_post)
        aplicaciones = Aplicacion.objects.filter(nombre__in = aplicacion_post)
        talleres = Taller.objects.filter(nombre__in = taller_post)
        #Asignmacion del contexto
        perfil.compania = compania
        perfil.ubicacion.clear()
        perfil.aplicacion.clear()
        perfil.taller.clear()
        perfil.ubicacion.set(sucursales)
        perfil.aplicacion.set(aplicaciones)
        perfil.taller.set(talleres)
        perfil.save()
        #Vaciado del carrito
        
        seleccionadas = LlantasSeleccionadas.objects.filter(perfil=perfil)
        if seleccionadas.count() > 0:
            for select in seleccionadas:
                select.llantas.clear()
                select.save()
        else:
            data = []
            inventarios = [
                "Nueva", 
                "Antes de Renovar",
                "Antes de Desechar",
                "Renovada",
                "Con renovador",
                "Desecho final",
                "Servicio",
                "Rodante",
                "Archivado"
                        ]
            for inventario in inventarios:
                data.append(
                    LlantasSeleccionadas(
                        perfil = perfil,
                        inventario = inventario
                    )
                )
            LlantasSeleccionadas.objects.bulk_create(data)
        return redirect('dashboards:home')

class TireDBView(LoginRequiredMixin, TemplateView):
    # Vista de tire-dashboard1

    template_name = "tire-dashboard.html"
    def get_context_data(self, **kwargs):
        
        clase1 = self.request.GET.getlist("clase")
        clase2 = self.request.GET.get("clase2")
        flota1 = self.request.GET.getlist("flota")
        flota2 = self.request.GET.get("flota2")
        pay_boton = self.request.GET.get("boton_intuitivo")

        context = super().get_context_data(**kwargs)
        vehiculos_totales = Vehiculo.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        vehiculo = Vehiculo.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        if clase1:
            vehiculo = vehiculo.filter(functions.reduce(or_, [Q(clase=c.upper()) for c in clase1]))
        if flota1:
            vehiculo = vehiculo.filter(functions.reduce(or_, [Q(ubicacion=Ubicacion.objects.get(nombre=f)) for f in flota1]))

        if flota1:
            flotas_vehiculo = vehiculos_totales.values("ubicacion").distinct()
        else:
            flotas_vehiculo = vehiculo.values("ubicacion").distinct()
        flotas = Ubicacion.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania), id__in=flotas_vehiculo)
        
        if clase1:
            clases = vehiculos_totales.values("clase").distinct()
        else:
            clases = vehiculo.values("clase").distinct()

        bitacora = Bitacora.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania), vehiculo__in=vehiculo)
        llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        inspecciones = Inspeccion.objects.filter(llanta__vehiculo__compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        ultimas_inspecciones = inspecciones.values("llanta")
        hoy = date.today()
        
        periodo_2 = hoy - timedelta(days=self.request.user.perfil.compania.periodo2_inspeccion)
        vehiculos_vistos = vehiculo.filter(ultima_inspeccion__fecha_hora__lte=periodo_2) | vehiculo.filter(ultima_inspeccion__fecha_hora__isnull=True)
        porcentaje_visto = int((vehiculos_vistos.count()/vehiculo.count()) * 100)

        filtro_sospechoso = functions.vehiculo_sospechoso(inspecciones)
        vehiculos_sospechosos = vehiculo.filter(id__in=filtro_sospechoso)
        porcentaje_sospechoso = int((vehiculos_sospechosos.count()/vehiculo.count()) * 100)

        doble_entrada = functions.doble_mala_entrada(bitacora, vehiculo)
        filtro_rojo = functions.vehiculo_rojo(llantas, doble_entrada, vehiculo)
        vehiculos_rojos = vehiculo.filter(id__in=filtro_rojo).exclude(id__in=vehiculos_sospechosos)
        porcentaje_rojo = int((vehiculos_rojos.count()/vehiculo.count()) * 100)

        filtro_amarillo = functions.vehiculo_amarillo(llantas)
        vehiculos_amarillos = vehiculo.filter(id__in=filtro_amarillo).exclude(id__in=vehiculos_rojos).exclude(id__in=vehiculos_sospechosos)
        porcentaje_amarillo = int((vehiculos_amarillos.count()/vehiculo.count()) * 100)


        if pay_boton == "Dualización":
            desdualizacion_encontrada = functions.desdualizacion(llantas, True)
            if desdualizacion_encontrada:
                desdualizacion_encontrada_existente = True
            else:
                desdualizacion_encontrada_existente = False
            desdualizacion_actual = functions.desdualizacion(llantas, True)
            if desdualizacion_actual:
                desdualizacion_actual_existente = True
            else:
                desdualizacion_actual_existente = False
            
            context["pay_boton"] = "dualizacion"
            context["parametro_actual_existente"] = desdualizacion_actual_existente
            context["parametro_encontrado_existente"] = desdualizacion_encontrada_existente
            context["parametro_actual"] = desdualizacion_actual
            context["parametro_encontrado"] = desdualizacion_encontrada
        elif pay_boton == "Presión":
            presion_encontrada = functions.presion_llantas(llantas, True)
            presion_actual = functions.presion_llantas(llantas, True)
            context["pay_boton"] = "presion"
            context["parametro_actual_existente"] = presion_actual.exists()
            context["parametro_encontrado_existente"] = presion_encontrada.exists()
            context["parametro_actual"] = presion_actual
            context["parametro_encontrado"] = presion_encontrada
        else:
            desgaste_irregular_encontrado = functions.desgaste_irregular(llantas, True)
            desgaste_irregular_actual = functions.desgaste_irregular(llantas, True)
            context["pay_boton"] = "desgaste"
            context["parametro_actual_existente"] = bool(desgaste_irregular_actual)
            context["parametro_encontrado_existente"] = bool(desgaste_irregular_encontrado)
            context["parametro_actual"] = desgaste_irregular_actual
            context["parametro_encontrado"] = desgaste_irregular_encontrado

        mala_entrada_ejes = functions.mala_entrada_ejes(llantas, True)

        mes_1 = hoy.strftime("%b")
        mes_2 = functions.mes_anterior(hoy)
        mes_3 = functions.mes_anterior(mes_2)
        mes_4 = functions.mes_anterior(mes_3)

        hoy1 = hoy.strftime("%m")
        hoy2 = mes_2.strftime("%m")
        hoy3 = mes_3.strftime("%m")
        hoy4 = mes_4.strftime("%m")

        vehiculo_mes1 = bitacora.filter(fecha_de_inflado__month=hoy1)
        vehiculo_mes2 = bitacora.filter(fecha_de_inflado__month=hoy2)
        vehiculo_mes3 = bitacora.filter(fecha_de_inflado__month=hoy3)
        vehiculo_mes4 = bitacora.filter(fecha_de_inflado__month=hoy4)
        
        try:
            entrada_correcta_mes_1 = round((functions.contar_entrada_correcta(llantas) / llantas.count()) * 100, 2)
        except:
            entrada_correcta_mes_1 = 0

        vehiculos_vistos_mes_1 = vehiculo.filter(ultima_inspeccion__fecha_hora__month=hoy1)
        vehiculos_vistos_mes_2 = vehiculo.filter(ultima_inspeccion__fecha_hora__month=hoy2)
        vehiculos_vistos_mes_3 = vehiculo.filter(ultima_inspeccion__fecha_hora__month=hoy3)
        vehiculos_vistos_mes_4 = vehiculo.filter(ultima_inspeccion__fecha_hora__month=hoy4)

        vehiculos_rojos_mes_1 = vehiculos_rojos


        print(vehiculos_rojos_mes_1)

        vehiculos_amarillos_mes_1 = vehiculos_amarillos

        vehiculos_sospechosos_mes_1 = vehiculos_sospechosos

        estatus_profundidad = functions.estatus_profundidad(llantas)

        nunca_vistos = functions.nunca_vistos(vehiculo)
        renovables = functions.renovables(llantas, vehiculos_amarillos)
        sin_informacion = functions.sin_informacion(llantas)

        porcentaje_vehiculos_inspeccionados_por_aplicacion = functions.vehiculos_inspeccionados_por_aplicacion(vehiculo)
        porcentaje_vehiculos_inspeccionados_por_clase = functions.vehiculos_inspeccionados_por_clase(vehiculo)

        vehiculos_por_clase_sin_inspeccionar = functions.vehiculos_por_clase_sin_inspeccionar(vehiculo, hoy1, hoy2, hoy3)
        clase_sin_inspeccionar_mes_1 = {}
        clase_sin_inspeccionar_mes_2 = {}
        clase_sin_inspeccionar_mes_3 = {}
        for cls in vehiculos_por_clase_sin_inspeccionar:
            clase_sin_inspeccionar_mes_1[cls["clase"]] = cls["mes1"]
            clase_sin_inspeccionar_mes_2[cls["clase"]] = cls["mes2"]
            clase_sin_inspeccionar_mes_3[cls["clase"]] = cls["mes3"]

        vehiculos_por_aplicacion_sin_inspeccionar = functions.vehiculos_por_aplicacion_sin_inspeccionar(vehiculo, hoy1, hoy2, hoy3)
        aplicacion_sin_inspeccionar_mes_1 = {}
        aplicacion_sin_inspeccionar_mes_2 = {}
        aplicacion_sin_inspeccionar_mes_3 = {}
        for cls in vehiculos_por_aplicacion_sin_inspeccionar:
            aplicacion_sin_inspeccionar_mes_1[cls["aplicacion__nombre"]] = cls["mes1"]
            aplicacion_sin_inspeccionar_mes_2[cls["aplicacion__nombre"]] = cls["mes2"]
            aplicacion_sin_inspeccionar_mes_3[cls["aplicacion__nombre"]] = cls["mes3"]

        context["a"] = False
        context["aplicacion_sin_inspeccionar_mes_1"] = aplicacion_sin_inspeccionar_mes_1
        context["aplicacion_sin_inspeccionar_mes_2"] = aplicacion_sin_inspeccionar_mes_2
        context["aplicacion_sin_inspeccionar_mes_3"] = aplicacion_sin_inspeccionar_mes_3
        context["clase1"] = clase1
        context["clase2"] = clase2
        context["clase_sin_inspeccionar_mes_1"] = clase_sin_inspeccionar_mes_1
        context["clase_sin_inspeccionar_mes_2"] = clase_sin_inspeccionar_mes_2
        context["clase_sin_inspeccionar_mes_3"] = clase_sin_inspeccionar_mes_3
        context["clases_mas_frecuentes_infladas"] = clases
        context["entrada_correcta_mes_1"] = entrada_correcta_mes_1
        context["entrada_correcta_mes_1_cantidad"] = functions.contar_entrada_correcta(llantas)
        context["entrada_correcta_mes_2"] = 0
        context["entrada_correcta_mes_3"] = 0
        context["entrada_correcta_mes_4"] = 0
        context["estatus_profundidad"] = estatus_profundidad
        context["flota1"] = flota1
        context["flota2"] = flota2
        context["flotas"] = flotas
        context["mala_entrada_ejes"] = mala_entrada_ejes
        context["mes_1"] = mes_1
        context["mes_2"] = mes_2.strftime("%b")
        context["mes_3"] = mes_3.strftime("%b")
        context["mes_4"] = mes_4.strftime("%b")
        context["nunca_vistos"] = nunca_vistos
        context["porcentaje_amarillo"] = porcentaje_amarillo
        context["porcentaje_vehiculos_inspeccionados_por_aplicacion"] = porcentaje_vehiculos_inspeccionados_por_aplicacion
        context["porcentaje_vehiculos_inspeccionados_por_clase"] = porcentaje_vehiculos_inspeccionados_por_clase
        context["porcentaje_rojo"] = porcentaje_rojo
        context["porcentaje_sospechoso"] = porcentaje_sospechoso
        context["porcentaje_visto"] = porcentaje_visto
        context["renovables"] = renovables
        context["sin_informacion"] = sin_informacion
        context["vehiculos_amarillos"] = vehiculos_amarillos.count()
        context["vehiculos_amarillos_mes_1"] = vehiculos_amarillos_mes_1.count()
        context["vehiculos_amarillos_mes_2"] = 0
        context["vehiculos_amarillos_mes_3"] = 0
        context["vehiculos_amarillos_mes_4"] = 0
        context["vehiculos_por_aplicacion_sin_inspeccionar"] = vehiculos_por_aplicacion_sin_inspeccionar
        context["vehiculos_por_clase_sin_inspeccionar"] = vehiculos_por_clase_sin_inspeccionar
        context["vehiculos_rojos"] = vehiculos_rojos.count()
        context["vehiculos_rojos_mes_1"] = vehiculos_rojos_mes_1.count()
        context["vehiculos_rojos_mes_2"] = 0
        context["vehiculos_rojos_mes_3"] = 0
        context["vehiculos_rojos_mes_4"] = 0
        context["vehiculos_sospechosos"] = vehiculos_sospechosos.count()
        context["vehiculos_sospechosos_mes_1"] = vehiculos_sospechosos_mes_1.count()
        context["vehiculos_sospechosos_mes_2"] = 0
        context["vehiculos_sospechosos_mes_3"] = 0
        context["vehiculos_sospechosos_mes_4"] = 0
        context["vehiculos_vistos"] = vehiculos_vistos.count()
        context["vehiculos_vistos_mes_1"] = vehiculos_vistos_mes_1.count()
        context["vehiculos_vistos_mes_2"] = vehiculos_vistos_mes_2.count()
        context["vehiculos_vistos_mes_3"] = vehiculos_vistos_mes_3.count()
        context["vehiculos_vistos_mes_4"] = vehiculos_vistos_mes_4.count()
        context["vehiculos_totales"] = vehiculo.count()
        return context



class TireDB2View(LoginRequiredMixin, TemplateView):
    # Vista de tire-dashboard2

    template_name = "tire-dashboard2.html"
    def get_context_data(self, **kwargs):

        clase = self.request.GET.getlist("clase")
        clase2 = self.request.GET.get("clase2")
        flota = self.request.GET.getlist("flota")
        flota2 = self.request.GET.get("flota2")
        aplicacion = self.request.GET.getlist("aplicacion")
        aplicacion2 = self.request.GET.get("aplicacion2")
        compania = Compania.objects.get(compania=self.request.user.perfil.compania)
        
        context = super().get_context_data(**kwargs)
        if clase:
            vehiculo = Vehiculo.objects.filter(functions.reduce(or_, [Q(clase=c.upper()) for c in clase]), compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        elif flota:
            vehiculo = Vehiculo.objects.filter(functions.reduce(or_, [Q(ubicacion=Ubicacion.objects.get(nombre=f)) for f in flota]), compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        elif aplicacion:
            vehiculo = Vehiculo.objects.filter(functions.reduce(or_, [Q(aplicacion=Aplicacion.objects.get(nombre=a)) for a in aplicacion]), compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        else:
            vehiculo = Vehiculo.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))

        llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=self.request.user.perfil.compania), vehiculo__in=vehiculo)
        inspecciones = Inspeccion.objects.filter(llanta__vehiculo__compania=Compania.objects.get(compania=self.request.user.perfil.compania), llanta__vehiculo__in=vehiculo)
        ubicacion = Ubicacion.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))[0]

        reemplazo_actual = functions.reemplazo_actual(llantas)
        # Te elimina los ejes vacíos
        reemplazo_actual_llantas = reemplazo_actual[0]
        reemplazo_actual_ejes = {k: v for k, v in reemplazo_actual[1].items() if v != 0}

        reemplazo_dual = functions.reemplazo_dual(llantas, reemplazo_actual_llantas)
        reemplazo_total = functions.reemplazo_total(reemplazo_actual_ejes, reemplazo_dual)

        """print("llantas", llantas)
        print("reemplazo_actual_llantas", reemplazo_actual_llantas)
        print("reemplazo_actual", reemplazo_actual)
        print("reemplazo_dual", reemplazo_dual)"""

        # Sin regresión
        embudo_vida1 = functions.embudo_vidas(llantas)
        pronostico_de_consumo = {k: v for k, v in embudo_vida1[1].items() if v != 0}
        presupuesto = functions.presupuesto(pronostico_de_consumo, ubicacion)

        # Con regresión
        embudo_vida2 = functions.embudo_vidas_con_regresion(inspecciones, ubicacion, 30)
        embudo_vida3 = functions.embudo_vidas_con_regresion(inspecciones, ubicacion, 60)
        embudo_vida4 = functions.embudo_vidas_con_regresion(inspecciones, ubicacion, 90)
        pronostico_de_consumo2 = {k: v for k, v in embudo_vida2[1].items() if v != 0}
        pronostico_de_consumo3 = {k: v for k, v in embudo_vida3[1].items() if v != 0}
        pronostico_de_consumo4 = {k: v for k, v in embudo_vida4[1].items() if v != 0}
        presupuesto2 = functions.presupuesto(pronostico_de_consumo2, ubicacion)
        presupuesto3 = functions.presupuesto(pronostico_de_consumo3, ubicacion)
        presupuesto4 = functions.presupuesto(pronostico_de_consumo4, ubicacion)

        context["p1"] = 50500
        context["p2"] = 11700
        context["p3"] = 20000
        
        context["p4"] = 24
        context["p5"] = 31
        context["p6"] = 40

        context["p7"] = 4
        context["p8"] = 10
        context["p9"] = 18
        context["p10"] = 25

        context["aplicacion1"] = aplicacion
        context["aplicacion2"] = aplicacion2
        context["aplicaciones"] = Aplicacion.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        context["clase1"] = clase
        context["clase2"] = clase2
        context["clases_mas_frecuentes_infladas"] = functions.clases_mas_frecuentes(vehiculo, self.request.user.perfil.compania)
        context["compania"] = str(compania)
        context["embudo"] = embudo_vida1[1]
        context["flota1"] = flota
        context["flota2"] = flota2
        context["flotas"] = Ubicacion.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        context["presupuesto"] = presupuesto
        context["presupuesto2"] = presupuesto2
        context["presupuesto3"] = presupuesto3
        context["presupuesto4"] = presupuesto4
        context["pronostico_de_consumo"] = pronostico_de_consumo
        context["pronostico_de_consumo2"] = pronostico_de_consumo2
        context["pronostico_de_consumo3"] = pronostico_de_consumo3
        context["pronostico_de_consumo4"] = pronostico_de_consumo4
        context["pronostico_de_consumo_contar"] = len(embudo_vida1[1]) + 1
        context["reemplazo_actual_ejes"] = reemplazo_actual_ejes
        context["reemplazo_dual"] = reemplazo_dual
        context["reemplazo_total"] = reemplazo_total

        return context

class TireDB3View(LoginRequiredMixin, TemplateView):
    # Vista de tire-dashboard3

    template_name = "tire-dashboard3.html"
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        vehiculos_totales = Vehiculo.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        llantas_totales = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=self.request.user.perfil.compania), vehiculo__in=vehiculos_totales)

        flota1 = self.request.GET.getlist("flota")
        flota2 = self.request.GET.get("flota2")
        eje1 = self.request.GET.getlist("eje")
        eje2 = self.request.GET.get("eje2")
        vehiculo1 = self.request.GET.getlist("vehiculo")
        vehiculo2 = self.request.GET.get("vehiculo2")
        aplicacion1 = self.request.GET.getlist("aplicacion")
        aplicacion2 = self.request.GET.getlist("aplicacion2")
        posicion1 = self.request.GET.getlist("posicion")
        posicion2 = self.request.GET.getlist("posicion2")
        clase1 = self.request.GET.getlist("clase")
        clase2 = self.request.GET.getlist("clase2")
        producto1 = self.request.GET.getlist("producto")
        producto2 = self.request.GET.getlist("producto2")

        vehiculos = Vehiculo.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))

        if flota1:
            vehiculos = vehiculos.filter(functions.reduce(or_, [Q(ubicacion=Ubicacion.objects.get(nombre=f)) for f in flota1]))
        if vehiculo1:
            vehiculos = vehiculos.filter(functions.reduce(or_, [Q(numero_economico=v) for v in vehiculo1]))
        if aplicacion1:
            vehiculos = vehiculos.filter(functions.reduce(or_, [Q(aplicacion=Aplicacion.objects.get(nombre=a)) for a in aplicacion1]))
        if clase1:
            vehiculos = vehiculos.filter(functions.reduce(or_, [Q(clase=c.upper()) for c in clase1]))

        llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=self.request.user.perfil.compania), vehiculo__in=vehiculos)

        if eje1:
            llantas = llantas.filter(functions.reduce(or_, [Q(nombre_de_eje=e) for e in eje1]))
        if producto1:
            llantas = llantas.filter(functions.reduce(or_, [Q(producto=Producto.objects.get(producto=p)) for p in producto1]))
        if posicion1:
            llantas = llantas.filter(functions.reduce(or_, [Q(posicion=p) for p in posicion1]))

        inspecciones = Inspeccion.objects.filter(llanta__in=llantas)
        if producto1:
            productos_llanta = llantas_totales.values("producto").distinct()
        else:
            productos_llanta = llantas.values("producto").distinct()
        productos = Producto.objects.filter(id__in=productos_llanta)
        if posicion1:
            posiciones = llantas_totales.values("posicion").distinct()
        else:
            posiciones = llantas.values("posicion").distinct()
        if flota1:
            flotas_vehiculo = vehiculos_totales.values("ubicacion__nombre").distinct()
        else:
            flotas_vehiculo = vehiculos.values("ubicacion__nombre").distinct()
        flotas = Ubicacion.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania), nombre__in=flotas_vehiculo)
        if aplicacion1:
            aplicaciones_vehiculo = vehiculos_totales.values("aplicacion__nombre").distinct()
        else:
            aplicaciones_vehiculo = vehiculos.values("aplicacion__nombre").distinct()
        aplicaciones = Aplicacion.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania), nombre__in=aplicaciones_vehiculo)
        if eje1:
            ejes = llantas_totales.values("nombre_de_eje").distinct()
        else:
            ejes = llantas.values("nombre_de_eje").distinct()
        
        if clase1:
            clases = vehiculos_totales.values("clase").distinct()
        else:
            clases = vehiculos.values("clase").distinct()

        hoy = date.today()
        mes_1 = hoy.strftime("%b")
        mes_2 = functions.mes_anterior(hoy)
        mes_3 = functions.mes_anterior(mes_2)
        mes_4 = functions.mes_anterior(mes_3)
        mes_5 = functions.mes_anterior(mes_4)
        mes_6 = functions.mes_anterior(mes_5)
        mes_7 = functions.mes_anterior(mes_6)
        mes_8 = functions.mes_anterior(mes_7)

        regresion = functions.km_proyectado(inspecciones, True)
        km_proyectado = regresion[0]
        km_x_mm = regresion[1]
        cpk = regresion[2]
        llantas_limpias = regresion[4]
        llantas_analizadas = llantas.filter(numero_economico__in=llantas_limpias)
        print(llantas_analizadas)
        try:
            porcentaje_limpio = (len(llantas_limpias)/llantas.count())*100
        except:
            porcentaje_limpio = 0


        llantas_limpias_nueva = []
        llantas_limpias_1r = []
        llantas_limpias_2r = []
        llantas_limpias_3r = []
        llantas_limpias_4r = []

        for llanta in llantas_limpias:
            if llanta.vida == "Nueva":
                llantas_limpias_nueva.append(llanta)
            elif llanta.vida == "1R":
                llantas_limpias_1r.append(llanta)
            elif llanta.vida == "2R":
                llantas_limpias_2r.append(llanta)
            elif llanta.vida == "3R":
                llantas_limpias_3r.append(llanta)
            elif llanta.vida == "4R":
                llantas_limpias_4r.append(llanta)

        comparativa_de_productos = {}
        cpk_productos = {}
        km_productos = {}
        for producto in productos:
            valores_producto = []
            
            llantas_producto_total = llantas.filter(producto=producto)
            llantas_producto = llantas.filter(producto=producto, numero_economico__in=llantas_limpias)
            desgaste = functions.desgaste_irregular_producto(llantas_producto)
            porcentaje_analizadas = functions.porcentaje(llantas_producto.count(), llantas_producto_total.count())

            inspecciones_producto = Inspeccion.objects.filter(llanta__in=llantas_producto)
            regresion_producto = functions.km_proyectado(inspecciones_producto, False)
            km_proyectado_producto = regresion_producto[0]
            km_x_mm_producto = regresion_producto[1]
            cpk_producto = regresion_producto[2]
            cantidad = llantas_producto.count()
            dibujo = str(producto.dibujo).replace(" ", "_")

            valores_producto.append(km_proyectado_producto)
            valores_producto.append(km_x_mm_producto)
            valores_producto.append(cpk_producto)
            valores_producto.append(cantidad)
            valores_producto.append(desgaste)
            valores_producto.append(porcentaje_analizadas)
            valores_producto.append(dibujo)

            if dibujo != "None":
                if regresion_producto[0] != 0 or regresion_producto[3] != 0:
                    if regresion_producto[3] != []:
                        if  regresion_producto[3].pop() != 0:
                            print(regresion_producto[3])
                            comparativa_de_productos[producto] = valores_producto

                            km_productos[dibujo] = regresion_producto[0]
                            cpk_productos[dibujo] = regresion_producto[3]

        productos_sort = sorted(comparativa_de_productos.items(), key=lambda p:p[1][2])
        comparativa_de_productos = {}
        for c in productos_sort:
            comparativa_de_productos[c[0]] = c[1]
            
        comparativa_de_flotas = {}
        cpk_flotas = {}
        km_flotas = {}
        for flota in flotas:
            valores_flota = []
            
            llantas_flota = llantas.filter(vehiculo__ubicacion=flota, numero_economico__in=llantas_limpias)
            if llantas_flota:
                inspecciones_flota = Inspeccion.objects.filter(llanta__in=llantas_flota)
                regresion_flota = functions.km_proyectado(inspecciones_flota, False)
                km_proyectado_flota = regresion_flota[0]
                km_x_mm_flota = regresion_flota[1]
                cpk_flota = regresion_flota[2]
                cantidad = llantas_flota.count()

                valores_flota.append(km_proyectado_flota)
                valores_flota.append(km_x_mm_flota)
                valores_flota.append(cpk_flota)
                km_flotas[flota] = regresion_flota[0]
                cpk_flotas[flota] = regresion_flota[3]
                valores_flota.append(cantidad)
                
                comparativa_de_flotas[flota] = valores_flota

        comparativa_de_vehiculos = {}
        cpk_vehiculos = []
        for vehiculo in vehiculos:
            valores_vehiculo = []
            
            llantas_vehiculo = llantas.filter(vehiculo=vehiculo, numero_economico__in=llantas_limpias)
            if llantas_vehiculo:
                inspecciones_vehiculo = Inspeccion.objects.filter(llanta__in=llantas_vehiculo)
                regresion_vehiculo = functions.km_proyectado(inspecciones_vehiculo, False)
                km_proyectado_vehiculo = regresion_vehiculo[0]
                km_x_mm_vehiculo = regresion_vehiculo[1]
                cpk_vehiculo = regresion_vehiculo[2]
                cantidad = llantas_vehiculo.count()

                valores_vehiculo.append(km_proyectado_vehiculo)
                valores_vehiculo.append(km_x_mm_vehiculo)
                valores_vehiculo.append(cpk_vehiculo)
                cpk_vehiculos.append(cpk_vehiculo)
                valores_vehiculo.append(cantidad)
                
                comparativa_de_vehiculos[vehiculo.numero_economico] = valores_vehiculo

        comparativa_de_aplicaciones = {}
        cpk_aplicaciones = {}
        km_aplicaciones = {}
        for aplicacion in aplicaciones:
            valores_aplicacion = []

            llantas_aplicacion = llantas.filter(vehiculo__aplicacion =aplicacion, numero_economico__in=llantas_limpias)
            if llantas_aplicacion:

                inspecciones_aplicacion = Inspeccion.objects.filter(llanta__in=llantas_aplicacion)
                regresion_aplicacion = functions.km_proyectado(inspecciones_aplicacion, False)
                km_proyectado_aplicacion = regresion_aplicacion[0]
                km_x_mm_aplicacion = regresion_aplicacion[1]
                cpk_aplicacion = regresion_aplicacion[2]
                cantidad = llantas_aplicacion.count()

                valores_aplicacion.append(km_proyectado_aplicacion)
                valores_aplicacion.append(km_x_mm_aplicacion)
                valores_aplicacion.append(cpk_aplicacion)
                km_aplicaciones[aplicacion] = regresion_aplicacion[0]
                cpk_aplicaciones[aplicacion] = regresion_aplicacion[3]
                valores_aplicacion.append(cantidad)
                
                comparativa_de_aplicaciones[aplicacion] = valores_aplicacion

        comparativa_de_ejes = {}
        for eje in ejes:
            valores_eje = []

            llantas_eje = llantas.filter(nombre_de_eje=eje["nombre_de_eje"], numero_economico__in=llantas_limpias)
            inspecciones_eje = Inspeccion.objects.filter(llanta__in=llantas_eje)
            if inspecciones_eje.exists():
                regresion_eje = functions.km_proyectado(inspecciones_eje, False)
                km_proyectado_eje = regresion_eje[0]
                km_x_mm_eje = regresion_eje[1]
                cpk_eje = regresion_eje[2]
                cantidad = llantas_eje.count()

                valores_eje.append(km_proyectado_eje)
                valores_eje.append(km_x_mm_eje)
                valores_eje.append(cpk_eje)
                valores_eje.append(cantidad)
                
                comparativa_de_ejes[eje["nombre_de_eje"]] = valores_eje

        comparativa_de_posiciones = {}
        for posicion in posiciones:
            valores_posicion = []

            llantas_posicion = llantas.filter(posicion=posicion["posicion"], numero_economico__in=llantas_limpias)
            inspecciones_posicion = Inspeccion.objects.filter(llanta__in=llantas_posicion)
            if inspecciones_posicion.exists():
                regresion_posicion = functions.km_proyectado(inspecciones_posicion, False)
                km_proyectado_posicion = regresion_posicion[0]
                km_x_mm_posicion = regresion_posicion[1]
                cpk_posicion = regresion_posicion[2]
                cantidad = llantas_posicion.count()

                valores_posicion.append(km_proyectado_posicion)
                valores_posicion.append(km_x_mm_posicion)
                valores_posicion.append(cpk_posicion)
                valores_posicion.append(cantidad)
                
                comparativa_de_posiciones[posicion["posicion"]] = valores_posicion

        cpk_vehiculos =  functions.cpk_vehiculo_cantidad(cpk_vehiculos)
        cpk_flotas =  functions.distribucion_cantidad(cpk_flotas)
        cpk_aplicaciones =  functions.distribucion_cantidad(cpk_aplicaciones)
        cpk_productos =  functions.distribucion_cantidad(cpk_productos)

        km_flotas =  functions.distribucion_cantidad(km_flotas)
        km_aplicaciones = functions.distribucion_cantidad(km_aplicaciones)
        km_productos = functions.distribucion_cantidad(km_productos)

        """try:
            tendencia_cpk_mes2 = TendenciaCPK.objects.get(compania=Compania.objects.get(compania=self.request.user.perfil.compania), mes=2)
        except:
            tendencia_cpk_mes2 = 0
        try:
            tendencia_cpk_mes3 = TendenciaCPK.objects.get(compania=Compania.objects.get(compania=self.request.user.perfil.compania), mes=3)
        except:
            tendencia_cpk_mes3 = 0
        try:
            tendencia_cpk_mes4 = TendenciaCPK.objects.get(compania=Compania.objects.get(compania=self.request.user.perfil.compania), mes=4)
        except:
            tendencia_cpk_mes4 = 0
        try:
            tendencia_cpk_mes5 = TendenciaCPK.objects.get(compania=Compania.objects.get(compania=self.request.user.perfil.compania), mes=5)
        except:
            tendencia_cpk_mes5 = 0
        try:
            tendencia_cpk_mes6 = TendenciaCPK.objects.get(compania=Compania.objects.get(compania=self.request.user.perfil.compania), mes=6)
        except:
            tendencia_cpk_mes6 = 0
        try:
            tendencia_cpk_mes7 = TendenciaCPK.objects.get(compania=Compania.objects.get(compania=self.request.user.perfil.compania), mes=7)
        except:
            tendencia_cpk_mes7 = 0
        try:
            tendencia_cpk_mes8 = TendenciaCPK.objects.get(compania=Compania.objects.get(compania=self.request.user.perfil.compania), mes=8)
        except:
            tendencia_cpk_mes8 = 0"""

        print(comparativa_de_productos)
        print(cpk_productos)
        context["aplicacion1"] = aplicacion1
        context["aplicacion2"] = aplicacion2
        context["aplicaciones"] = aplicaciones
        context["clase1"] = clase1
        context["clase2"] = clase2
        context["clases"] = clases
        context["compania"] = str(self.request.user.perfil.compania)
        context["comparativa_de_aplicaciones"] = comparativa_de_aplicaciones
        context["comparativa_de_ejes"] = comparativa_de_ejes
        context["comparativa_de_flotas"] = comparativa_de_flotas
        context["comparativa_de_posiciones"] = comparativa_de_posiciones
        context["comparativa_de_productos"] = comparativa_de_productos
        context["comparativa_de_vehiculos"] = comparativa_de_vehiculos
        context["cpk"] = cpk
        context["cpk_aplicaciones"] = cpk_aplicaciones
        context["cpk_flotas"] = cpk_flotas
        context["cpk_productos"] = cpk_productos
        context["cpk_vehiculos"] = cpk_vehiculos[0]
        context["cpk_vehiculos_cantidad"] = cpk_vehiculos[1]
        context["cpk_vehiculos_cantidad_maxima"] = max(cpk_vehiculos[1])
        context["eje1"] = eje1
        context["eje2"] = eje2
        context["ejes"] = ejes
        context["flota1"] = flota1
        context["flota2"] = flota2
        context["flotas"] = flotas
        context["km_aplicaciones"] = km_aplicaciones
        context["km_flotas"] = km_flotas
        context["km_productos"] = km_productos
        context["km_proyectado"] = km_proyectado
        context["km_x_mm"] = km_x_mm
        context["llantas_analizadas"] = llantas.count()
        context["llantas_limpias"] = len(llantas_limpias)
        context["llantas_limpias_nueva"] = len(llantas_limpias_nueva)
        context["llantas_limpias_1r"] = len(llantas_limpias_1r)
        context["llantas_limpias_2r"] = len(llantas_limpias_2r)
        context["llantas_limpias_3r"] = len(llantas_limpias_3r)
        context["llantas_limpias_4r"] = len(llantas_limpias_4r)
        context["mes_1"] = mes_1
        context["mes_2"] = mes_2.strftime("%b")
        context["mes_3"] = mes_3.strftime("%b")
        context["mes_4"] = mes_4.strftime("%b")
        context["mes_5"] = mes_5.strftime("%b")
        context["mes_6"] = mes_6.strftime("%b")
        context["mes_7"] = mes_7.strftime("%b")
        context["mes_8"] = mes_8.strftime("%b")
        context["porcentaje_limpio"] = porcentaje_limpio
        context["posicion1"] = posicion1
        context["posicion2"] = posicion2
        context["posiciones"] = posiciones
        context["producto1"] = producto1
        context["producto2"] = producto2
        context["productos"] = productos
        """context["tendencia_cpk_mes2"] = tendencia_cpk_mes2
        context["tendencia_cpk_mes3"] = tendencia_cpk_mes3
        context["tendencia_cpk_mes4"] = tendencia_cpk_mes4
        context["tendencia_cpk_mes5"] = tendencia_cpk_mes5
        context["tendencia_cpk_mes6"] = tendencia_cpk_mes6
        context["tendencia_cpk_mes7"] = tendencia_cpk_mes7
        context["tendencia_cpk_mes8"] = tendencia_cpk_mes8"""
        context["vehiculo1"] = vehiculo1
        context["vehiculo2"] = vehiculo2
        context["vehiculos"] = vehiculos
        return context

class hubView(LoginRequiredMixin, TemplateView):
    # Vista de hubView

    template_name = "hub.html"
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        grupo = str(User.objects.get(username=self.request.user).groups.get())

        context["grupo"] = grupo
        return context

class diagramaView(LoginRequiredMixin, TemplateView):
    # Vista de diagramaView
    
    template_name = "diagrama.html"
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        vehiculo = Vehiculo.objects.get(pk = self.kwargs['pk'])
        llantas = Llanta.objects.filter(vehiculo=vehiculo, inventario="Rodante")
        configuracion = vehiculo.configuracion

        cantidad_llantas = functions.cantidad_llantas(configuracion)
        
        context["configuracion"] = configuracion
        
        
        observaciones_vehiculo_act = vehiculo.observaciones.all()
        context['observaciones_vehiculo_act'] = observaciones_vehiculo_act
        
        num_eco_comp = functions.all_num_eco_compania(vehiculo.compania, llantas)
        context['num_eco_comp'] = num_eco_comp
        #Generacion de ejes dinamico
        vehiculo_actual = Vehiculo.objects.get(pk = self.kwargs['pk'])
        llantas_actuales = Llanta.objects.filter(vehiculo = self.kwargs['pk'], inventario="Rodante")
        inspecciones_actuales = Inspeccion.objects.filter(llanta__in=llantas_actuales)
        
        #Obtencion de la data
        num_ejes = vehiculo_actual.configuracion.split('.')
        ejes_no_ordenados = []
        ejes = []
        eje = 1
        color_profundidad = ""
        presiones_establecida=[
            vehiculo_actual.presion_establecida_1,
            vehiculo_actual.presion_establecida_2,
            vehiculo_actual.presion_establecida_3,
            vehiculo_actual.presion_establecida_4,
            vehiculo_actual.presion_establecida_5,
            vehiculo_actual.presion_establecida_6,
            vehiculo_actual.presion_establecida_7,
        ]
        numero = 0
        for num in num_ejes:
            list_temp = []
            for llanta in llantas_actuales:
                objetivo = llanta.vehiculo.compania.objetivo / 100
                presion_act = llanta.presion_actual
                presion_minima = presiones_establecida[numero] - (presiones_establecida[numero] * objetivo)
                presion_maxima = presiones_establecida[numero] + (presiones_establecida[numero] * objetivo)
                try:
                    if presion_act >= presion_minima and presion_act <= presion_maxima:
                        color_presion = 'good'
                    elif presion_act < presion_minima:
                        color_presion = 'bad'
                    elif presion_act > presion_maxima:
                        color_presion = 'yellow'
                except: 
                    color_presion = 'good'
                    
                min_produndidad = functions.min_profundidad(llanta)
                max_produndidad = functions.max_profundidad(llanta)
                
                punto_retiro = functions.punto_de_retiro(llanta)
                color_profundidad = functions.color_profundidad(min_produndidad, punto_retiro)
                mm_dif = llanta.vehiculo.compania.mm_de_diferencia_entre_duales
                context['mm_dif'] = mm_dif
                
                observaciones_act = llanta.observaciones.all()
                
                observaciones = llanta.observaciones.all()
                data = {
                    'presion_minima': presion_minima,
                    'presion_maxima': presion_maxima,
                    'punto_retiro': punto_retiro,
                    'observaciones': observaciones,
                    'observaciones_act': observaciones_act,
                    'min_produndidad': min_produndidad,
                    'max_produndidad': max_produndidad,
                    'desgaste_anterior_llanta': llanta.parametro_desgaste_irregular
                        }
                if llanta.ultima_inspeccion == None:
                    con_inspeccion = False
                else:
                    con_inspeccion = True
                if llanta.eje == eje:
                    color_llanta = functions.color_observaciones_all_one(llanta)
                    eForm = EdicionManual(instance = llanta)
                    list_temp.append([llanta, color_profundidad, eForm, color_presion, min_produndidad, con_inspeccion, data, color_llanta])
            eje += 1
            ejes_no_ordenados.append(list_temp)
            numero += 1
            
        ejes = functions.acomodo_ejes(ejes_no_ordenados)
            
        color = functions.entrada_correcta(vehiculo_actual, None)
        #print(color)
        if color == 'good':
            style = 'good'
        elif color == 'bad':
            style = 'bad'
        else:
            style = 'bad'
        
        cant_ejes = len(ejes)
        
        observaciones_manuales=Observacion.objects.filter(nivel='Llanta', automatico=False)
        observaciones_automaticas=Observacion.objects.filter(nivel='Llanta', automatico=True).exclude(observacion='Ponchado seguro').exclude(observacion='Mala entrada')
        observaciones_vehiculo = Observacion.objects.filter(nivel='Vehiculo')
        
        #print(vehiculo.configuracion)
        #print(ejes)
        #print(f'style: {style}')
        #print(f'llantas_sospechosas: {llantas_sospechosas}')
        #print(f'llantas_rojas: {llantas_rojas}')
        #print(f'llantas_amarillas: {llantas_amarillas}')
        #print(f'llantas_azules: {llantas_azules}')
        mm_de_desgaste_compania = vehiculo_actual.compania.mm_de_desgaste_irregular
        
        context['ejes'] = ejes
        context['style'] = style
        context['cant_ejes'] = cant_ejes
        context['llantas_actuales'] = llantas_actuales
        context['vehiculo_actual'] = vehiculo_actual
        context['observaciones_manuales'] = observaciones_manuales
        context['observaciones_automaticas'] = observaciones_automaticas
        context['observaciones_vehiculo'] = observaciones_vehiculo
        context['mm_de_desgaste_compania'] = mm_de_desgaste_compania
        

        return context
    #DiagramaPost
    def post(self, request, *args, **kwargs):
        print(request.POST)
        #print(self.kwargs['pk'])
        print('---------------------')
        print('---------------------')
        #Vehiclo
        km_vehiculo = request.POST.getlist('km_vehiculo')
        observacion_vehiculo = request.POST.getlist('observation-vehiculo')
        #Llanta
        ids = request.POST.getlist('ids')
        llanta = request.POST.getlist('economico')
        producto = request.POST.getlist('producto')
        vida = request.POST.getlist('vida')
        profundidad_derecha = request.POST.getlist('profundidad_derecha')
        profundidad_central = request.POST.getlist('profundidad_central')
        profundidad_izquierda = request.POST.getlist('profundidad_izquierda')
        presion = request.POST.getlist('presion')
        observaciones = request.POST.getlist('observaciones')
        reemplazar = request.POST.getlist('reemplazar')
        observaciones_list = {}
        for id in ids:
            try:
                cadena = str('manual-observation-') + str(id)
                if len(request.POST.getlist(cadena)) > 0:
                    observaciones_list[id] = request.POST.getlist(cadena)
            except:
                pass
        print(observacion_vehiculo)
        print(observaciones_list)  
        print(ids)
        print(llanta)
        print(producto)
        print(vida)
        print(profundidad_derecha)
        print(profundidad_central)
        print(profundidad_izquierda)
        print(presion)
        print(observaciones)
        print(reemplazar)

        elementos = 0
        diferencias = []
        id_bitacora = None
        cambio_km = False
        print(km_vehiculo[0])
        vehiculo = Vehiculo.objects.get(pk = self.kwargs['pk'])
        vehiculo_ref = Vehiculo.objects.get(pk = self.kwargs['pk'])
        diferencia_presion_duales_list = []
        desdualizacion_list=[]
        print(vehiculo.km)
        km_vehicle = (float(km_vehiculo[0]) if km_vehiculo[0]!='' else None)
        if vehiculo.km != km_vehicle:
            vehiculo.km = km_vehicle
            vehiculo.save()
            cambio_km = True
        for id_actual in ids:
            cambios = []
            ids_post  = ids[elementos]
            economico_post  = llanta[elementos] # <-
            producto_post  = producto[elementos] # <-
            vida_post  = vida[elementos]
            profundidad_derecha_post  = profundidad_derecha[elementos]
            profundidad_central_post  = profundidad_central[elementos]
            profundidad_izquierda_post  = profundidad_izquierda[elementos]
            presion_post  = presion[elementos]
                        
            if id_actual in reemplazar:
                print(f"Remplazando: {id_actual}".center(50, '-'))
                print(economico_post)
                llanta_actual_referencia = Llanta.objects.get(pk = ids_post)
                llanta_actual = Llanta.objects.get(pk = ids_post)
                llanta_nueva = Llanta.objects.get(pk = ids_post)
                llanta_nueva.id = None
                llanta_nueva.ultima_inspeccion = None
                llanta_nueva.km_actual = None
                llanta_nueva.km_montado = None
                llanta_actual.numero_economico = llanta_actual.numero_economico + '- Archivado'
                llanta_nueva.numero_economico = economico_post
                producto_cam = Producto.objects.get(producto = producto_post)
                llanta_nueva.producto = producto_cam
                llanta_nueva.vida = vida_post
                llanta_nueva.profundidad_derecha = (float(profundidad_derecha_post) if profundidad_derecha_post!='' else None)
                llanta_nueva.profundidad_central = (float(profundidad_central_post) if profundidad_central_post!='' else None)
                llanta_nueva.profundidad_izquierda = (float(profundidad_izquierda_post) if profundidad_izquierda_post!='' else None)
                llanta_nueva.presion_actual = (float(presion_post) if presion_post!='' else None)
                llanta_actual.inventario = 'Archivado'
                llanta_actual.save()
                llanta_nueva.save()
                
                
                if llanta_actual.ultima_inspeccion != None:
                    evento_raw = llanta_actual.ultima_inspeccion.evento
                    evento_raw = evento_raw.replace("\'", "\"")
                    evento_act = json.loads(evento_raw)
                    evento_act['llanta_mod'] = llanta_nueva.id
                    llanta_actual.ultima_inspeccion.evento = str(evento_act)
                    llanta_actual.ultima_inspeccion.edicion_manual = True
                    llanta_actual.ultima_inspeccion.save()
                    llanta_actual.save()
                    
            else:
                if id_actual in reemplazar:
                    continue
                print(f"Actualizando: {id_actual}".center(50, '-'))
                llanta_actual_referencia = Llanta.objects.get(pk = ids_post)
                llanta_actual = Llanta.objects.get(pk = ids_post)
                print(llanta_actual)
                
                if llanta_actual.ultima_inspeccion != None:
                    inspeccion_actual = llanta_actual.ultima_inspeccion
                    if llanta_actual.numero_economico != economico_post:
                        llanta_actual.numero_economico = economico_post
                        cambios.append('economico')
                    if str(llanta_actual.producto) != producto_post:
                        producto_cam = Producto.objects.get(producto = producto_post)
                        llanta_actual.producto = producto_cam
                        cambios.append('producto')
                    if llanta_actual.vida != vida_post:
                        
                        inspeccion_actual.vida = vida_post
                        llanta_actual.vida = vida_post
                        cambios.append('vida')
                    if cambio_km == True:
                        inspeccion_actual.km_vehiculo = vehiculo.km
                        cambios.append('cambio_km')
                    if llanta_actual.profundidad_derecha != (float(profundidad_derecha_post) if profundidad_derecha_post!='' else None):
                        inspeccion_actual.profundidad_derecha = (float(profundidad_derecha_post) if profundidad_derecha_post!='' else None)
                        llanta_actual.profundidad_derecha = (float(profundidad_derecha_post) if profundidad_derecha_post!='' else None)
                        cambios.append('profundidad_derecha')
                    if llanta_actual.profundidad_central != (float(profundidad_central_post) if profundidad_central_post!='' else None):
                        inspeccion_actual.profundidad_central = (float(profundidad_central_post) if profundidad_central_post!='' else None)
                        llanta_actual.profundidad_central = (float(profundidad_central_post) if profundidad_central_post!='' else None)
                        cambios.append('profundidad_central')
                    if llanta_actual.profundidad_izquierda != (float(profundidad_izquierda_post) if profundidad_izquierda_post!='' else None):
                        inspeccion_actual.profundidad_izquierda = (float(profundidad_izquierda_post) if profundidad_izquierda_post!='' else None)
                        llanta_actual.profundidad_izquierda = (float(profundidad_izquierda_post) if profundidad_izquierda_post!='' else None)
                        cambios.append('profundidad_izquierda')
                    if  llanta_actual.presion_actual != (float(presion_post) if presion_post!='' else None):
                        inspeccion_actual.presion = (float(presion_post) if presion_post!='' else None)
                        llanta_actual.presion_actual = (float(presion_post) if presion_post!='' else None)
                        cambios.append('presion_actual')
                    inspeccion_actual.save()
                    llanta_actual.save()
                    
                    if vehiculo.km != None:
                        if vehiculo.km != vehiculo_ref.km:
                            cambios.append('0')
                        if llanta_actual.km_montado != None:
                            km_actual_nuevo = functions.actualizar_km_actual(llanta_actual, llanta_actual_referencia, vehiculo, vehiculo_ref)
                            llanta_actual.km_actual = km_actual_nuevo
                            llanta_actual.save()
                        
                        elif llanta_actual.km_montado == None:
                            inspecciones_vehiculos = InspeccionVehiculo.objects.filter(vehiculo = llanta_actual.vehiculo).exclude(km = None)
                            inspecciones = Inspeccion.objects.filter(llanta = llanta_actual, inspeccion_vehiculo__in = inspecciones_vehiculos, vida = llanta_actual.vida)
                            #inspecciones = Inspeccion.objects.filter(llanta = llanta_actual)
                            if len(inspecciones) >= 2:
                                print('Sin km de montado pero con inspecciones suficinetes')
                                print(llanta_actual.id)
                                primer_inspeccion = inspecciones.first()
                                ultima_inspeccion = inspecciones.last()
                                km_teorico = functions.actualizar_km_actual_no_km_montado(primer_inspeccion, ultima_inspeccion)
                                print(km_teorico)
                                if km_teorico != None:
                                    llanta_actual.km_actual = km_teorico
                                    llanta_actual.save()
                            else:
                                print('Sin km de montado y sin inspecciones suficinetes')
                                print(llanta_actual.id)
                    if 'vida' in cambios:
                        functions.cambio_de_vida(llanta_actual_referencia, llanta_actual)
                    print(cambios)
                    if len(cambios) > 0:
                        if inspeccion_actual.evento != None:
                            #print(inspeccion_actual.evento)
                            
                            evento_raw =inspeccion_actual.evento
                            evento_raw = evento_raw.replace("\'", "\"")
                            evento_act = json.loads(evento_raw)
                            if 'economico' in cambios:
                                evento_act['numero_economico_mod'] = economico_post
                                
                            if 'producto' in cambios:
                                evento_act['producto_mod'] = producto_post
                                
                            if 'vida' in cambios:
                                evento_act["vida_mod"] = vida_post
                                
                            if 'cambio_km' in cambios:
                                evento_act['km_mod'] = vehiculo.km 
                                
                            if 'profundidad_derecha' in cambios:
                                evento_act["profundidad_derecha_mod"] = profundidad_derecha_post
                            
                            if 'profundidad_central' in cambios:
                                evento_act["profundidad_central_mod"] = profundidad_central_post
                            
                            if 'profundidad_izquierda' in cambios:
                                evento_act["profundidad_izquierda_mod"] = profundidad_central_post
                            
                            if 'presion_actual' in cambios:
                                evento_act["presion_mod"] = presion_post
                            evento_str = str(evento_act)
                            inspeccion_actual.edicion_manual = True
                            inspeccion_actual.evento = evento_str
                            inspeccion_actual.save()
                        else:
                            #!AQUI
                            evento = str({\
                                "llanta_inicial" : llanta_actual_referencia.id, "llanta_mod" : "",\
                                "producto_inicial" : str(llanta_actual_referencia.producto), "producto_mod" : "",\
                                "vida_inicial" : llanta_actual_referencia.vida, "vida_mod" : "",\
                                "km_inicial" : vehiculo.km, "km_mod" : "",\
                                "presion_inicial" : llanta_actual_referencia.presion_actual, "presion_mod" : "",\
                                "profundidad_izquierda_inicial" : llanta_actual_referencia.profundidad_izquierda, "profundidad_izquierda_mod" : "",\
                                "profundidad_central_inicial" : llanta_actual_referencia.profundidad_central_post, "profundidad_central_mod" : "",\
                                "profundidad_derecha_inicial" : llanta_actual_referencia.profundidad_derecha_post, "profundidad_derecha_mod" : ""\
                            
                            })
                else:
                    print('No paso na') 
            elementos += 1
                     
        elementos = 0
        for id_actual in ids:
            cambios = []
            if id_actual in reemplazar:
                print(f"Remplazando: {id_actual}".center(50, '-'))                
            else:
                print(f"Actualizando: {id_actual}".center(50, '-'))
                
                llanta_referencia = Llanta.objects.get(pk = ids[elementos])
                llanta_actual = Llanta.objects.get(pk = ids[elementos])   
                #observaciones
                llanta_actual.observaciones.clear()
                llanta_actual.ultima_inspeccion.observaciones.clear()
                
                if ids[elementos] in observaciones_list:
                    for i in (observaciones_list[ids[elementos]]):
                        observacion = Observacion.objects.get(observacion = i)
                        llanta_actual.observaciones.add(observacion)
                        llanta_actual.save()
                compania = llanta_actual.vehiculo.compania
                presiones_establecidas = [
                    llanta_actual.vehiculo.presion_establecida_1,
                    llanta_actual.vehiculo.presion_establecida_2,
                    llanta_actual.vehiculo.presion_establecida_3,
                    llanta_actual.vehiculo.presion_establecida_4,
                    llanta_actual.vehiculo.presion_establecida_5,
                    llanta_actual.vehiculo.presion_establecida_6,
                    llanta_actual.vehiculo.presion_establecida_7,
                ]
                establecida = presiones_establecidas[(llanta_actual.eje - 1)]
                presion_minima = establecida - (establecida * (compania.objetivo/100))
                presion_maxima = establecida + (establecida * (compania.objetivo/100))
                if llanta_actual.presion_actual != None:
                    if float(llanta_actual.presion_actual) < presion_minima:
                        baja_presion = Observacion.objects.get(observacion = 'Baja presión')
                        llanta_actual.observaciones.add(baja_presion)
                        print(baja_presion)
                        
                if llanta_actual.presion_actual != None:    
                    if float(llanta_actual.presion_actual) > presion_maxima:
                        alta_presion = Observacion.objects.get(observacion = 'Alta presion')
                        llanta_actual.observaciones.add(alta_presion)
                        print(alta_presion)
                
                if '4' in llanta_actual.tipo_de_eje:
                    print('Duales')
                    eje = str(llanta_actual.eje)
                    posicion = llanta_actual.posicion[1:]
                    #print(eje+posicion)
                    if posicion == 'LO':
                        dual = 'LI'
                    elif posicion == 'LI':
                        dual = 'LO'
                    elif posicion == 'RI':
                        dual = 'RO'
                    elif posicion == 'RO':
                        dual = 'RI'
                    dual_completo = eje + dual
                    llantas_comparacion = Llanta.objects.filter(vehiculo = llanta_actual.vehiculo, inventario="Rodante")
                    for llanta_i in llantas_comparacion:
                        if llanta_i.posicion == dual_completo:
                            dual_llanta = llanta_i
                    if llanta_actual.presion_actual != None and dual_llanta.presion_actual != None:
                        presion_actual = float(llanta_actual.presion_actual) if float(llanta_actual.presion_actual) != 0 else 1
                        presion_dual = dual_llanta.presion_actual if dual_llanta.presion_actual != 0 else 1
                        porcentaje_dif = (float(presion_actual) - presion_dual) / float(presion_actual)
                        #porcentaje_dif = (float(llanta_actual.presion_actual) - dual_llanta.presion_actual) / float(llanta_actual.presion_actual)
                        if llanta_actual in diferencia_presion_duales_list:
                        
                            diferencia_presion_duales = Observacion.objects.get(observacion = 'Diferencia de presión entre los duales')
                            llanta_actual.observaciones.add(diferencia_presion_duales)
                            llanta_actual.ultima_inspeccion.observaciones.add(diferencia_presion_duales)
                            llanta_actual.ultima_inspeccion.save()
                            print(diferencia_presion_duales)
                        else:

                            if porcentaje_dif > 0.1:
                                #Poner a los 2 duales
                                diferencia_presion_duales = Observacion.objects.get(observacion = 'Diferencia de presión entre los duales')
                                llanta_actual.observaciones.add(diferencia_presion_duales)
                                dual_llanta.observaciones.add(diferencia_presion_duales)
                                dual_llanta.save()
                                diferencia_presion_duales_list.append(dual_llanta)
                                print(diferencia_presion_duales)
                            
                    if llanta_actual in desdualizacion_list:
                        desdualización = Observacion.objects.get(observacion = 'Desdualización')
                        llanta_actual.observaciones.add(desdualización)
                        llanta_actual.ultima_inspeccion.observaciones.add(desdualización)
                        llanta_actual.ultima_inspeccion.save()
                        print(desdualización)
                    else:
                        print(functions.min_profundidad(llanta_actual))
                        print(functions.min_profundidad(dual_llanta))
                        print(compania.mm_de_diferencia_entre_duales)
                        if (functions.min_profundidad(llanta_actual) - functions.min_profundidad(dual_llanta)) >= compania.mm_de_diferencia_entre_duales:
                            #Poner a los 2 duales
                            desdualización = Observacion.objects.get(observacion = 'Desdualización')
                            llanta_actual.observaciones.add(desdualización)
                            dual_llanta.observaciones.add(desdualización)
                            dual_llanta.save()
                            desdualizacion_list.append(dual_llanta)
                            print(desdualización)
                    
                if "S" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_direccion
                elif "D" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_traccion
                elif "T" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_arrastre
                elif "C" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_loco
                elif "L" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_retractil
                
                if functions.min_profundidad(llanta_actual) < punto_retiro:
                    baja_profundidad = Observacion.objects.get(observacion = 'Baja profundidad')
                    llanta_actual.observaciones.add(baja_profundidad)
                    print(baja_profundidad)
                    
                if functions.min_profundidad(llanta_actual) == (punto_retiro + 0.6):
                    en_punto_retiro = Observacion.objects.get(observacion = 'En punto de retiro')
                    llanta_actual.observaciones.add(en_punto_retiro)
                    print(en_punto_retiro)
                izquierda = llanta_actual.profundidad_izquierda
                central = llanta_actual.profundidad_central
                derecha = llanta_actual.profundidad_derecha
                functions.desgaste_profundidad(izquierda, central , derecha, llanta_actual)

                
                llanta_actual.save()
                if len(llanta_actual.observaciones.all()) > 0:
                    for observacion in llanta_actual.observaciones.all():
                        llanta_actual.ultima_inspeccion.observaciones.add(observacion)
                    llanta_actual.ultima_inspeccion.save()
                #print(llanta_act.ultima_inspeccion)
                
            elementos += 1
                
        elementos = 0    
        for id_actual in ids: 
            llanta_referencia = Llanta.objects.get(pk = ids[elementos])
            llanta_actual = Llanta.objects.get(pk = ids[elementos]) 
            for obs in llanta_actual.observaciones.all():
                llanta_actual.ultima_inspeccion.observaciones.add(obs)
                llanta_actual.ultima_inspeccion.save()
                llanta_actual.save()
            elementos += 1
        #Diagrma obds
        functions.observaciones_vehiculo(vehiculo)
                        
        return redirect('dashboards:detail', self.kwargs['pk'])
        #return redirect('dashboards:diagrama', self.kwargs['pk'])


class tireDiagramaView(LoginRequiredMixin, TemplateView):
    # Vista de tireDiagramaView

    template_name = "tireDiagrama.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        llanta = Llanta.objects.get(pk=self.kwargs['pk'])
        punto_de_retiro = functions.punto_de_retiro(llanta)
        min_presion = functions.min_presion(llanta)
        max_presion = functions.max_presion(llanta)
        min_profundidad = functions.min_profundidad(llanta)
        color_profundidad = functions.color_profundidad(min_profundidad, punto_de_retiro)
        color_presion = functions.color_presion(llanta.presion_actual, min_presion, max_presion)
        con_inspeccion = False
        if llanta.ultima_inspeccion == None:
            con_inspeccion = True
        productos = Producto.objects.all()
        observaciones_actual = llanta.observaciones.all()
        observaciones_manuales=Observacion.objects.filter(nivel='Llanta', automatico=False)
        observaciones_automaticas=Observacion.objects.filter(nivel='Llanta', automatico=True).exclude(observacion='Ponchado seguro').exclude(observacion='Mala entrada')
        observaciones_todas=Observacion.objects.all()
        llantas_actuales = Llanta.objects.filter(pk=self.kwargs['pk'])
        num_eco_comp = functions.all_num_eco_compania(llanta.compania, llantas_actuales)
        color_observaciones = functions.color_observaciones_all_one(llanta)
        context['llanta'] = llanta
        context['punto_de_retiro'] = punto_de_retiro
        context['min_presion'] = min_presion
        context['max_presion'] = max_presion
        context['min_profundidad'] = min_profundidad
        context['color_presion'] = color_presion
        context['color_profundidad'] = color_profundidad
        context['color_observaciones'] = color_observaciones
        context['con_inspeccion'] = con_inspeccion
        context['productos'] = productos
        context['observaciones_actual'] = observaciones_actual
        context['observaciones_manuales'] = observaciones_manuales
        context['observaciones_automaticas'] = observaciones_automaticas
        context['observaciones_todas'] = observaciones_todas
        context['num_eco_comp'] = num_eco_comp
        
        return context
            
    def post(self, request, *args, **kwargs):
        print('post')
        print(request.POST)
        #print(self.kwargs['pk'])
        llanta = Llanta.objects.get(pk=self.kwargs['pk'])
        llanta_actual_referencia = Llanta.objects.get(pk=self.kwargs['pk'])
        economico = request.POST.getlist('economico')[0]
        producto = request.POST.getlist('producto')[0]
        vida = request.POST.getlist('vida')[0]
        profundidad_izquierda = (float(request.POST.getlist('profundidad_izquierda')[0]) if request.POST.getlist('profundidad_izquierda')[0] != '' else None)
        profundidad_central = (float(request.POST.getlist('profundidad_central')[0]) if request.POST.getlist('profundidad_central')[0] != '' else None)
        profundidad_derecha = (float(request.POST.getlist('profundidad_derecha')[0]) if request.POST.getlist('profundidad_derecha')[0] != '' else None)
        presion = (float(request.POST.getlist('presion')[0]) if request.POST.getlist('presion')[0] != '' else None)
        observaciones_list = request.POST.getlist('manual-observation')
        
        print(economico)
        print(producto)
        print(vida)
        print(profundidad_izquierda)
        print(profundidad_central)
        print(profundidad_derecha)
        print(presion)
        print(observaciones_list)
        inspeccion_actual = llanta.ultima_inspeccion
        cambios = []
        
        if llanta.numero_economico != economico:
            cambios.append('economico')
            llanta.numero_economico = economico
            
        if str(llanta.producto) != producto:
            llanta.producto = Producto.objects.get(producto = producto)
            cambios.append('producto')
            
        if llanta.vida != vida:
            cambios.append('vida')
            llanta.vida = vida
            inspeccion_actual.vida = llanta.vida
            
        if llanta.profundidad_izquierda != profundidad_izquierda:
            cambios.append('profundidad_izquierda')
            llanta.profundidad_izquierda = profundidad_izquierda
            inspeccion_actual.profundidad_izquierda = profundidad_izquierda
            
        if llanta.profundidad_central != profundidad_central:
            cambios.append('profundidad_central')
            llanta.profundidad_central = profundidad_central
            inspeccion_actual.profundidad_central = profundidad_central
            
        if llanta.profundidad_derecha != profundidad_derecha:
            cambios.append('profundidad_derecha')
            llanta.profundidad_derecha = profundidad_derecha
            inspeccion_actual.profundidad_derecha = profundidad_derecha
            
        if llanta.presion_actual != presion:
            cambios.append('presion')
            llanta.presion_actual = presion
            inspeccion_actual.presion = presion
        llanta.save()
        inspeccion_actual.save()
        if 'vida' in cambios:
            functions.cambio_de_vida(llanta_actual_referencia, llanta)
        if len(cambios) > 0:
            if inspeccion_actual.evento != None:
                #print(inspeccion_actual.evento)
                
                evento_raw =inspeccion_actual.evento
                evento_raw = evento_raw.replace("\'", "\"")
                evento_act = json.loads(evento_raw)
                if 'economico' in cambios:
                    print('economico')
                    evento_act['numero_economico_mod'] = str(economico)
                    
                if 'producto' in cambios:
                    print('producto')
                    evento_act['producto_mod'] = str(producto)
                    
                if 'vida' in cambios:
                    print('vida')
                    evento_act["vida_mod"] = vida
                    
                if 'profundidad_derecha' in cambios:
                    print('profundidad_derecha')
                    evento_act["profundidad_derecha_mod"] = str(profundidad_derecha)
                
                if 'profundidad_central' in cambios:
                    print('profundidad_central')
                    evento_act["profundidad_central_mod"] = str(profundidad_central)
                
                if 'profundidad_izquierda' in cambios:
                    print('profundidad_izquierda')
                    evento_act["profundidad_izquierda_mod"] = str(profundidad_central)
                
                if 'presion' in cambios:
                    print('presion_mod')
                    evento_act["presion_mod"] = str(presion)
                    
                evento_str = str(evento_act)
                inspeccion_actual.edicion_manual = True
                inspeccion_actual.evento = evento_str
                inspeccion_actual.save()
                
        #observaciones
        llanta.observaciones.clear()
        llanta.ultima_inspeccion.observaciones.clear()
        
        for i in observaciones_list:
            observacion = Observacion.objects.get(observacion = i)
            llanta.observaciones.add(observacion)
            llanta.ultima_inspeccion.observaciones.add(observacion)
            llanta.save()
            llanta.ultima_inspeccion.save()
            
        compania = llanta.vehiculo.compania
        presiones_establecidas = [
            llanta.vehiculo.presion_establecida_1,
            llanta.vehiculo.presion_establecida_2,
            llanta.vehiculo.presion_establecida_3,
            llanta.vehiculo.presion_establecida_4,
            llanta.vehiculo.presion_establecida_5,
            llanta.vehiculo.presion_establecida_6,
            llanta.vehiculo.presion_establecida_7,
        ]
        establecida = presiones_establecidas[(llanta.eje - 1)]
        presion_minima = establecida - (establecida * (compania.objetivo/100))
        presion_maxima = establecida + (establecida * (compania.objetivo/100))
        if llanta.presion_actual != None:
            if float(llanta.presion_actual) < presion_minima:
                baja_presion = Observacion.objects.get(observacion = 'Baja presión')
                llanta.observaciones.add(baja_presion)
                print(baja_presion)
                
        if llanta.presion_actual != None:    
            if float(llanta.presion_actual) > presion_maxima:
                alta_presion = Observacion.objects.get(observacion = 'Alta presion')
                llanta.observaciones.add(alta_presion)
                print(alta_presion)
        
        if '4' in llanta.tipo_de_eje:
            print('Duales')
            eje = str(llanta.eje)
            posicion = llanta.posicion[1:]
            #print(eje+posicion)
            if posicion == 'LO':
                dual = 'LI'
            elif posicion == 'LI':
                dual = 'LO'
            elif posicion == 'RI':
                dual = 'RO'
            elif posicion == 'RO':
                dual = 'RI'
            dual_completo = eje + dual
            llantas_comparacion = Llanta.objects.filter(vehiculo = llanta.vehiculo, inventario="Rodante")
            for llanta_i in llantas_comparacion:
                if llanta_i.posicion == dual_completo:
                    dual_llanta = llanta_i
            #? Checar pa
            if llanta.presion_actual != None and dual_llanta.presion_actual != None:
                presion_actual = float(llanta.presion_actual) if float(llanta.presion_actual) != 0 else 1
                presion_dual = dual_llanta.presion_actual if dual_llanta.presion_actual != 0 else 1
                
                porcentaje_dif_1 = (float(presion_actual) - presion_dual) / float(presion_actual)
                porcentaje_dif_2 = (presion_dual - float(presion_actual)) / presion_dual
                
                # = (float(llanta.presion_actual) - dual_llanta.presion_actual) / float(llanta.presion_actual)
                # = (dual_llanta.presion_actual - float(llanta.presion_actual)) / dual_llanta.presion_actual
                
            if porcentaje_dif_1 > 0.1 or porcentaje_dif_2 > 0.1:
                #Poner a los 2 duales
                diferencia_presion_duales = Observacion.objects.get(observacion = 'Diferencia de presión entre los duales')
                llanta.observaciones.add(diferencia_presion_duales)
                dual_llanta.observaciones.add(diferencia_presion_duales)
                dual_llanta.save()
                try:
                    dual_llanta.ultima_inspeccion.observaciones.add(diferencia_presion_duales)
                    dual_llanta.ultima_inspeccion.save()
                except:
                    pass
                print(diferencia_presion_duales)
            else:
                diferencia_presion_duales = Observacion.objects.get(observacion = 'Diferencia de presión entre los duales')
                dual_llanta.observaciones.remove(diferencia_presion_duales)
                dual_llanta.save()
                try:
                    dual_llanta.ultima_inspeccion.observaciones.remove(diferencia_presion_duales)
                    dual_llanta.ultima_inspeccion.save()     
                except:
                    pass
            
            if (
                (functions.min_profundidad(llanta) - functions.min_profundidad(dual_llanta)) >= compania.mm_de_diferencia_entre_duales
                or
                (functions.min_profundidad(dual_llanta) - functions.min_profundidad(llanta)) >= compania.mm_de_diferencia_entre_duales
                ):
                #Poner a los 2 duales
                desdualización = Observacion.objects.get(observacion = 'Desdualización')
                llanta.observaciones.add(desdualización)
                dual_llanta.observaciones.add(desdualización)
                dual_llanta.save()
                try:
                    dual_llanta.ultima_inspeccion.observaciones.add(desdualización)
                    dual_llanta.ultima_inspeccion.save()
                except:
                    pass
                print(desdualización)
            else:
                desdualización = Observacion.objects.get(observacion = 'Desdualización')
                dual_llanta.observaciones.remove(desdualización)
                dual_llanta.save()
                try:
                    dual_llanta.ultima_inspeccion.observaciones.remove(desdualización)
                    dual_llanta.ultima_inspeccion.save()
                except:
                    pass
            
        if "S" in llanta.tipo_de_eje:
            punto_retiro = compania.punto_retiro_eje_direccion
        elif "D" in llanta.tipo_de_eje:
            punto_retiro = compania.punto_retiro_eje_traccion
        elif "T" in llanta.tipo_de_eje:
            punto_retiro = compania.punto_retiro_eje_arrastre
        elif "C" in llanta.tipo_de_eje:
            punto_retiro = compania.punto_retiro_eje_loco
        elif "L" in llanta.tipo_de_eje:
            punto_retiro = compania.punto_retiro_eje_retractil
        
        if functions.min_profundidad(llanta) < punto_retiro:
            baja_profundidad = Observacion.objects.get(observacion = 'Baja profundidad')
            llanta.observaciones.add(baja_profundidad)
            print(baja_profundidad)
            
        if functions.min_profundidad(llanta) == (punto_retiro + 0.6):
            en_punto_retiro = Observacion.objects.get(observacion = 'En punto de retiro')
            llanta.observaciones.add(en_punto_retiro)
            print(en_punto_retiro)
        izquierda = llanta.profundidad_izquierda
        central = llanta.profundidad_central
        derecha = llanta.profundidad_derecha
        functions.desgaste_profundidad(izquierda, central , derecha, llanta)

        
        llanta.save()
        try:
            if len(llanta.observaciones.all()) > 0:
                for observacion in llanta.observaciones.all():
                    llanta.ultima_inspeccion.observaciones.add(observacion)
                llanta.ultima_inspeccion.save()
        except:
            pass
        #Tire Diagrama
        functions.observaciones_vehiculo(llanta.vehiculo)
        return redirect('dashboards:tireDetail', self.kwargs['pk'])

class inspeccionLlantaView(LoginRequiredMixin, TemplateView):
    # Vista de inspeccionLlantaView

    template_name = "tire-inspection/inspection-report.html"
    model = Inspeccion
    
    def get_queryset(self):
        """return Bitacora_Edicion.objects.filter(pk = self.kwargs['pk'])"""
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #Data
        inspeccion = Inspeccion.objects.get(pk = self.kwargs['pk'])
        vehiculo = inspeccion.inspeccion_vehiculo.vehiculo
        llanta = inspeccion.llanta
        compania = vehiculo.compania
        presion = inspeccion.presion
        min_profundidad = functions.min_profundidad(inspeccion)
        posicion = inspeccion.posicion
        
        #Obtencion del registro hsitorico
        evento_raw =inspeccion.evento
        evento_raw = evento_raw.replace("\'", "\"")
        evento_act = json.loads(evento_raw)
        producto = Producto.objects.get(producto = evento_act['producto_inicial'])
        
        #Observaciones
        observaciones = inspeccion.observaciones.all()
        print(observaciones)
        #Colores
        min_presion = functions.min_presion(llanta)
        max_presion = functions.max_presion(llanta)
        punto_de_retiro = functions.punto_de_retiro(llanta)
        
        color_presion = functions.color_presion(presion, min_presion, max_presion)
        color_profundidad = functions.color_profundidad(min_profundidad, punto_de_retiro)
        color_observaciones = functions.color_observaciones(observaciones)
        
        #Problemas
        problemas = []
        for observacion in inspeccion.observaciones.all():
            color_obs = functions.color_observaciones_one(observacion)
            if color_obs == 'bad':
                signo = 'icon-cross bad-text'
            elif color_obs == 'yellow':
                signo = 'icon-warning  yellow-text'
            else:
                signo = 'icon-checkmark good-text'
            problemas.append({
                'posicion': inspeccion.posicion, 
                'observacion': observacion, 
                'signo': signo
                })
        print(problemas)
        
        #Paso del contexto
        context['inspeccion'] = inspeccion
        context['llanta'] = llanta
        context['producto'] = producto
        context['presion'] = presion
        context['min_profundidad'] = min_profundidad
        context['posicion'] = posicion
        context['color_presion'] = color_presion
        context['color_profundidad'] = color_profundidad
        context['color_observaciones'] = color_observaciones
        context['observaciones'] = observaciones
        context['problemas'] = problemas
        
        return context

class webservicesView(TemplateView):
    # Vista de webservicesView

    template_name = "webservices/webservices.html"

class ParametroExtractoView(LoginRequiredMixin, TemplateView):
    # Vista de ParametrosExtractoView

    template_name = "parametrosExtracto.html"

class SiteMenuView(LoginRequiredMixin, TemplateView):
    # Vista de SiteMenuView

    template_name = "siteMenu.html"

class catalogoLinksView(LoginRequiredMixin, TemplateView):
    # Vista de catalogoLinksView

    template_name = "catalogoLinks.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        perfiles = Perfil.objects.all()
        context["perfiles"] = perfiles
        context["user"] = False

        return context
    
    def get_success_url(self):
        return reverse_lazy("dashboards:catalogoLinks")

class catalogoLinksEditView(LoginRequiredMixin, DetailView, UpdateView):
    # Vista de catalogoLinksEditView

    template_name = "catalogoLinks.html"
    slug_field = "perfil"
    slug_url_kwarg = "perfil"
    queryset = Perfil.objects.all()
    context_object_name = "perfil"
    model = Perfil
    fields = ["id", 'link_pulpo', 'link_operativo']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        perfiles = Perfil.objects.all()
        context["perfiles"] = perfiles

        return context

    def get_success_url(self):
        return reverse_lazy("dashboards:catalogoLinks")

class catalogoProductoView(LoginRequiredMixin, MultiModelFormView):
    # Vista de catalogoProductosView

    template_name = "catalogoProducto.html"
    form_classes = {"productoForm": ProductoForm}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        if self.request.method =='POST':
            Producto.objects.create(producto=self.request.POST.get("producto"),
                                    compania=self.request.user.perfil.compania,
                                    marca=self.request.POST.get("marca"),
                                    dibujo=self.request.POST.get("dibujo"),
                                    rango=self.request.POST.get("rango"),
                                    dimension=self.request.POST.get("dimension"),
                                    profundidad_inicial=self.request.POST.get("profundidad_inicial"),
                                    aplicacion=self.request.POST.get("aplicacion"),
                                    vida=self.request.POST.get("vida"),
                                    precio=self.request.POST.get("precio"),
                                    km_esperado=self.request.POST.get("km_esperado"),
            )

        productos = Producto.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))[::-1]
        compania = Compania.objects.get(compania=self.request.user.perfil.compania)
        context["compania"] = compania
        context["productos"] = productos

        return context
    
    def get_success_url(self):
        return reverse_lazy("dashboards:catalogoProductos")


class catalogoProductoExcelView(LoginRequiredMixin, TemplateView):
    # Vista de catalogoProductosView

    template_name = "catalogoProductosExcel.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        return context
    
    def post(self, request, *args, **kwargs):
        print(request.FILES)
        user = request.user
        perfil = Perfil.objects.get(user = user)
        compania = perfil.compania
        
        file = self.request.FILES.get("file")
        print(file)
        libro = openpyxl.load_workbook(file)
        sh  = libro.active
        # iterate through excel and display data
        rechazo = []
        productos = []
        for i in range(1, sh.max_row+1):
            
            if i == 1:
                continue
            print("\n")
            print("Row ", i, " data :")
            marca:str
            dibujo:str
            rango:str
            dimension:str
            profundidad_inicial:float
            aplicacion:str
            vida:str
            precio:float
            km_esperado:int

            producto:str
            compania:str
            for j in range(1, sh.max_column+1):
                
                #print(j)
                #print()
                cell_obj = sh.cell(row=i, column=j)
                #if j == 1:
                #    pass
                #elif j == 5:
                #    if str(type(j)) == "<class 'int'>":
                #        print('sdfsdfsdf')
                    
                if j == 1:
                    marca = cell_obj.value
                    
                elif j == 2:
                    dibujo = cell_obj.value
                elif j == 3:
                    rango = cell_obj.value
                    
                elif j == 4:
                    dimension = cell_obj.value
                    
                elif j == 5:
                    profundidad_inicial = cell_obj.value
                        
                elif j == 6:
                    aplicacion = cell_obj.value
                    
                elif j == 7:
    
                    vida = cell_obj.value
                                            
                elif j == 8:
                    precio = cell_obj.value
                    
                elif j == 9:
                    km_esperado = cell_obj.value
                        
            
            
            try:
                float(profundidad_inicial)
            except:
                rechazo.append(i)
                continue   
            
            try:
                float(precio)
            except:
                rechazo.append(i)
                continue
            
            try:
                int(km_esperado)
            except:
                rechazo.append(i)
                continue
            
            
            if vida == 'Nueva' or vida == 'Renovada' or vida == 'Vitacasco':
                pass
            else:
                rechazo.append(i)
                continue
            
            
            
            if (
                aplicacion == 'Dirección' 
                or 
                aplicacion == 'Tracción' 
                or 
                aplicacion == 'Arrastre' 
                or 
                aplicacion == 'Mixta'
                or 
                aplicacion == 'Regional' 
                or 
                aplicacion == 'Urbano'):
                pass
            else:
                rechazo.append(i)
                continue
            
            prodcto = f'{marca} {dibujo} {rango} {dimension}'
            
            coincidencia = Producto.objects.filter(compania = compania, producto = prodcto)
            if coincidencia.count() > 0:
                rechazo.append(i)
                continue
             
            print(f'marca {marca}')
            print(f'dibujo {dibujo}')
            print(f'rango {rango}')
            print(f'dimension {dimension}')
            print(f'profundidad_inicial {profundidad_inicial}')
            print(f'aplicacion {aplicacion}')
            print(f'vida {vida}')
            print(f'precio {precio}')
            print(f'km_esperado {km_esperado}')
            print(f'compania {compania}')
            print(f'prodcto {prodcto}')
            
            
            productos.append(
                Producto(
                marca = marca,
                dibujo = dibujo,
                rango = rango,
                dimension = dimension,
                profundidad_inicial = profundidad_inicial,
                aplicacion = aplicacion,
                vida = vida,
                precio = precio,
                km_esperado = km_esperado,
                compania = compania,
                producto = prodcto
                )
            )
            
            
        Producto.objects.bulk_create(productos)
        print()
        print(rechazo)
        return redirect('dashboards:catalogoProductoExcelView')
    #def get_success_url(self):
    #    return reverse_lazy("dashboards:catalogoProductos")

class catalogoProductoEditView(LoginRequiredMixin, DetailView, UpdateView):
    # Vista de catalogoProductoEditView

    template_name = "catalogoProducto.html"
    slug_field = "producto"
    slug_url_kwarg = "producto"
    queryset = Producto.objects.all()
    context_object_name = "producto"
    model = Producto
    fields = ["id", "producto", 'marca', 'dibujo', 'rango', 'dimension', 'profundidad_inicial', 'aplicacion', 'vida', 'precio', 'km_esperado']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        productos = Producto.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))[::-1]
        context["productos"] = productos

        return context

    def get_success_url(self):
        return reverse_lazy("dashboards:catalogoProductos")

def catalogoProductoDeleteView(request):
    if request.method =="POST":
        producto = Producto.objects.get(id=request.POST.get("id"))
        producto.delete()
        return redirect("dashboards:catalogoProductos")
    return redirect("dashboards:catalogoProductos")


class catalogoRenovadoresView(LoginRequiredMixin, CreateView):
    # Vista de catalogoRenovadoresView

    template_name = "catalogoRenovadores.html"
    form_class = RenovadorForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        compania = Compania.objects.get(compania=self.request.user.perfil.compania)
        companias = self.request.user.perfil.companias.all()
        renovadores = Renovador.objects.filter(compania__in=companias)
        print("companias", companias)
        print("renovadores", renovadores)
        context["renovadores"] = renovadores
        context["compania"] = compania
        context["companias"] = companias

        return context
    
    def get_success_url(self):
        return reverse_lazy("dashboards:catalogoRenovadores")

class catalogoRenovadoresEditView(LoginRequiredMixin, DetailView, UpdateView):
    # Vista de catalogoRenovadoresEditView

    template_name = "catalogoRenovadores.html"
    slug_field = "renovador"
    slug_url_kwarg = "renovador"
    queryset = Renovador.objects.all()
    context_object_name = "renovador"
    model = Renovador
    fields = ["id", "nombre", 'compania', 'ciudad', 'marca']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        compania = Compania.objects.get(compania=self.request.user.perfil.compania)
        companias = self.request.user.perfil.companias.all()
        renovadores = Renovador.objects.filter(compania__in=companias)
        context["renovadores"] = renovadores
        context["compania"] = compania
        context["companias"] = companias

        return context

    def get_success_url(self):
        return reverse_lazy("dashboards:catalogoRenovadores")

def catalogoRenovadoresDeleteView(request):
    if request.method =="POST":
        renovador = Renovador.objects.get(id=request.POST.get("id"))
        renovador.delete()
        return redirect("dashboards:catalogoRenovadores")
    return redirect("dashboards:catalogoRenovadores")

class catalogoDesechosView(LoginRequiredMixin, CreateView):
    # Vista de catalogoDesechosView

    template_name = "catalogoDesechos.html"
    form_class = DesechoForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        desechos = Desecho.objects.all()[::-1]
        llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        context["desechos"] = desechos
        context["llantas"] = llantas

        return context
    
    def get_success_url(self):
        return reverse_lazy("dashboards:catalogoDesechos")

class catalogoDesechosEditView(LoginRequiredMixin, DetailView, UpdateView):
    # Vista de catalogoDesechosEditView

    template_name = "catalogoDesechos.html"
    slug_field = "desecho"
    slug_url_kwarg = "desecho"
    queryset = Desecho.objects.all()
    context_object_name = "desecho"
    form_class = DesechoEditForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        desechos = Desecho.objects.all()[::-1]
        llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        context["desechos"] = desechos
        context["llantas"] = llantas

        return context

    def get_success_url(self):
        return reverse_lazy("dashboards:catalogoDesechos")

def catalogoDesechosDeleteView(request):
    if request.method =="POST":
        desecho = Desecho.objects.get(id=request.POST.get("id"))
        desecho.delete()
        return redirect("dashboards:catalogoDesechos")
    return redirect("dashboards:catalogoDesechos")


class catalogoObservacionesView(LoginRequiredMixin, CreateView):
    # Vista de catalogoObservacionesView

    template_name = "catalogoObservaciones.html"
    form_class = ObservacionForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        observaciones = Observacion.objects.all()[::-1]
        llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        context["observaciones"] = observaciones
        context["llantas"] = llantas

        return context
    
    def get_success_url(self):
        return reverse_lazy("dashboards:catalogoObservaciones")

class catalogoObservacionesEditView(LoginRequiredMixin, DetailView, UpdateView):
    # Vista de catalogoObservacionesEditView

    template_name = "catalogoObservaciones.html"
    slug_field = "observacion"
    slug_url_kwarg = "observacion"
    queryset = Observacion.objects.all()
    context_object_name = "observacion"
    form_class = ObservacionEditForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        observaciones = Observacion.objects.all()[::-1]
        llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        context["observaciones"] = observaciones
        context["llantas"] = llantas

        return context

    def get_success_url(self):
        return reverse_lazy("dashboards:catalogoObservaciones")

def catalogoObservacionesDeleteView(request):
    if request.method =="POST":
        observacion = Observacion.objects.get(id=request.POST.get("id"))
        observacion.delete()
        return redirect("dashboards:catalogoObservaciones")
    return redirect("dashboards:catalogoObservaciones")

class catalogoRechazosView(LoginRequiredMixin, CreateView):
    # Vista de catalogoRechazosView

    template_name = "catalogoRechazo.html"
    form_class = RechazoForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        rechazos = Rechazo.objects.all()[::-1]
        companias = self.request.user.perfil.companias.all()

        context["rechazos"] = rechazos
        context["companias"] = companias

        return context
    
    def get_success_url(self):
        return reverse_lazy("dashboards:catalogoRechazos")

class catalogoRechazosEditView(LoginRequiredMixin, DetailView, UpdateView):
    # Vista de catalogoRechazosEditView

    template_name = "catalogoRechazo.html"
    slug_field = "rechazo"
    slug_url_kwarg = "rechazo"
    queryset = Rechazo.objects.all()
    context_object_name = "rechazo"
    form_class = RechazoEditForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        rechazos = Rechazo.objects.all()[::-1]
        companias = self.request.user.perfil.companias.all()

        context["rechazos"] = rechazos
        context["companias"] = companias

        return context

    def get_success_url(self):
        return reverse_lazy("dashboards:catalogoRechazos")

def catalogoRechazosDeleteView(request):
    if request.method =="POST":
        rechazo = Rechazo.objects.get(id=request.POST.get("id"))
        rechazo.delete()
        return redirect("dashboards:catalogoRechazos")
    return redirect("dashboards:catalogoRechazos")

class companyFormularioView(LoginRequiredMixin, CreateView):
    # Vista de companyFormularioView

    template_name = "formularios/company.html"
    model = Compania
    fields = ["compania", "periodo1_inflado", "periodo2_inflado", "objetivo", "periodo1_inspeccion", "periodo2_inspeccion", "punto_retiro_eje_direccion", "punto_retiro_eje_traccion", "punto_retiro_eje_arrastre", "punto_retiro_eje_loco", "punto_retiro_eje_retractil", "mm_de_desgaste_irregular", "mm_de_diferencia_entre_duales", "mm_parametro_sospechoso", "unidades_presion", "unidades_distancia", "unidades_profundidad", "valor_casco_nuevo", "valor_casco_1r", "valor_casco_2r", "valor_casco_3r", "valor_casco_4r", "valor_casco_5r"]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        companias = self.request.user.perfil.companias.all()
        context["companias"] = companias

        return context

    def get_success_url(self):
        user = self.request.user.perfil
        user.companias.add(Compania.objects.get(compania=self.request.POST.get("compania")))
        user.save()
        return reverse_lazy("dashboards:companyFormulario")

class companyFormularioEditView(LoginRequiredMixin, DetailView, UpdateView):
    # Vista de catalogoCompanyEditView

    template_name = "formularios/company.html"
    slug_field = "compania"
    slug_url_kwarg = "compania"
    queryset = Compania.objects.all()
    context_object_name = "compania"
    model = Compania
    fields = ["id", "compania", "periodo1_inflado", "periodo2_inflado", "objetivo", "periodo1_inspeccion", "periodo2_inspeccion", "punto_retiro_eje_direccion", "punto_retiro_eje_traccion", "punto_retiro_eje_arrastre", "punto_retiro_eje_loco", "punto_retiro_eje_retractil", "mm_de_desgaste_irregular", "mm_de_diferencia_entre_duales", "mm_parametro_sospechoso", "unidades_presion", "unidades_distancia", "unidades_profundidad", "valor_casco_nuevo", "valor_casco_1r", "valor_casco_2r", "valor_casco_3r", "valor_casco_4r", "valor_casco_5r"]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        companias = self.request.user.perfil.companias.all()
        context["companias"] = companias

        return context

    def get_success_url(self):
        return reverse_lazy("dashboards:companyFormulario")

def companyFormularioDeleteView(request):
    if request.method =="POST":
        rechazo = Compania.objects.get(id=request.POST.get("id"))
        rechazo.delete()
        return redirect("dashboards:companyFormulario")
    return redirect("dashboards:companyFormulario")

class sucursalFormularioView(LoginRequiredMixin, CreateView):
    # Vista de sucursalFormularioView

    template_name = "formularios/sucursal.html"
    form_class = SucursalForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sucursales = self.request.user.perfil.ubicaciones.all()
        companias = self.request.user.perfil.companias.all()

        context["companias"] = companias
        context["sucursales"] = sucursales
        return context

    def get_success_url(self):
        user = self.request.user.perfil
        user.ubicaciones.add(Ubicacion.objects.get(nombre=self.request.POST.get("nombre")))
        user.save()
        return reverse_lazy("dashboards:sucursalFormulario")

class sucursalFormularioEditView(LoginRequiredMixin, DetailView, UpdateView):
    # Vista de catalogoSucursalesEditView

    template_name = "formularios/sucursal.html"
    slug_field = "sucursal"
    slug_url_kwarg = "sucursal"
    queryset = Ubicacion.objects.all()
    context_object_name = "sucursal"
    form_class = SucursalEditForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sucursales = self.request.user.perfil.ubicaciones.all()
        companias = self.request.user.perfil.companias.all()
        
        context["companias"] = companias
        context["sucursales"] = sucursales

        return context

    def get_success_url(self):
        return reverse_lazy("dashboards:sucursalFormulario")

def sucursalFormularioDeleteView(request):
    if request.method =="POST":
        sucursal = Ubicacion.objects.get(id=request.POST.get("id"))
        sucursal.delete()
        return redirect("dashboards:sucursalFormulario")
    return redirect("dashboards:sucursalFormulario")

class inpeccionesView(TemplateView):
    # Vista de vehiculosView

    template_name = "webservices/inspeccion.html" 


class sucursalView(TemplateView):
    # Vista de sucursalView

    template_name = "webservices/sucursal.html"   

class aplicacionesView(TemplateView):
    # Vista de aplicacionesView

    template_name = "webservices/ruta.html" 

class vehiculosView(TemplateView):
    # Vista de vehiculosView

    template_name = "webservices/vehiculo.html" 

class llantasView(TemplateView):
    # Vista de vehiculosView

    template_name = "webservices/llanta.html" 

class tallerFormularioView(LoginRequiredMixin, CreateView):
    # Vista de tallerFormularioView

    template_name = "formularios/taller.html"
    form_class = TallerForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        talleres = self.request.user.perfil.talleres.all()
        companias = self.request.user.perfil.companias.all()

        context["companias"] = companias
        context["talleres"] = talleres
        return context

    def get_success_url(self):
        user = self.request.user.perfil
        user.talleres.add(Taller.objects.get(nombre=self.request.POST.get("nombre")))
        user.save()
        return reverse_lazy("dashboards:tallerFormulario")

class tallerFormularioEditView(LoginRequiredMixin, DetailView, UpdateView):
    # Vista de catalogoTalleresEditView

    template_name = "formularios/taller.html"
    slug_field = "taller"
    slug_url_kwarg = "taller"
    queryset = Taller.objects.all()
    context_object_name = "taller"
    form_class = TallerEditForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        talleres = self.request.user.perfil.talleres.all()
        sucursales = self.request.user.perfil.ubicaciones.all()
        companias = self.request.user.perfil.companias.all()

        context["companias"] = companias
        context["sucursales"] = sucursales
        context["talleres"] = talleres

        return context
    
    def post(self, request, pk):
        print(request.POST)
        nombre = request.POST.getlist('nombre')[0]
        compania = request.POST.getlist('compania')[0]
        compania = Compania.objects.get(compania=compania)
        codigo = request.POST.getlist('codigo')[0]
        taller = Taller.objects.get(pk=pk)
        
        taller.nombre = nombre
        taller.compania = compania
        taller.codigo = codigo
        taller.save()
        return redirect("dashboards:tallerFormularioEdit", pk)
    
    #def get_success_url(self):
    #    print("hola")
    #    print(self.request.POST)
    #    return reverse_lazy("dashboards:tallerFormulario")

def tallerFormularioDeleteView(request):
    if request.method =="POST":
        taller = Taller.objects.get(id=request.POST.get("id"))
        taller.delete()
        return redirect("dashboards:tallerFormulario")
    return redirect("dashboards:tallerFormulario")

class usuarioFormularioView(LoginRequiredMixin, CreateView):
    # Vista de usuarioFormularioView

    template_name = "formularios/usuario.html"
    form_class = UsuarioForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        companias = Compania.objects.all()
        ubicaciones = Ubicacion.objects.all()
        aplicaciones = Aplicacion.objects.all()
        talleres = Taller.objects.all()
        usuarios = User.objects.all()
        grupos = Group.objects.all()
        
        context["aplicaciones"] = aplicaciones
        context["companias"] = companias
        context["grupos"] = grupos
        context["sucursales"] = ubicaciones
        context["talleres"] = talleres
        context["usuarios"] = usuarios
        return context

    def form_valid(self, form):
        # Save form data
        print(form.cleaned_data)
        c = {'form': form, }
        user = form.save(commit=False)
        groups = form.cleaned_data['groups']
        groups = Group.objects.get(name=groups)
        compania = self.request.POST.getlist("compania")
        compania = Compania.objects.filter(compania__in=compania)
        ubicacion = self.request.POST.getlist("ubicacion")
        ubicacion = Ubicacion.objects.filter(nombre__in=ubicacion)
        aplicacion = self.request.POST.getlist("aplicacion")
        aplicacion = Aplicacion.objects.filter(nombre__in=aplicacion)
        taller = self.request.POST.getlist("taller")
        taller = Taller.objects.filter(nombre__in=taller)
        password = form.cleaned_data['password']
        repeat_password = form.cleaned_data['repeat_password']
        print(compania)
        print(ubicacion)
        print(aplicacion)
        if password != repeat_password:
            messages.error(self.request, "Passwords do not Match", extra_tags='alert alert-danger')
            return render(self.request, self.template_name, c)
        user.set_password(password)
        user.save()
        
        perfil = Perfil.objects.create(user=user)
        for comp in compania:
            perfil.companias.add(comp)
        for ubi in ubicacion:
            perfil.ubicaciones.add(ubi)
        for apli in aplicacion:
            perfil.aplicaciones.add(apli)
        for tall in taller:
            perfil.talleres.add(tall)
        user.groups.add(groups)
        perfil.save()
        return super(usuarioFormularioView, self).form_valid(form)

    def get_success_url(self):
        return reverse_lazy("dashboards:usuarioFormulario")

class usuarioFormularioEditView(LoginRequiredMixin, DetailView, UpdateView):
    # Vista de usuarioFormularioEditView

    template_name = "formularios/usuario.html"
    slug_field = "usuario"
    slug_url_kwarg = "usuario"
    queryset = User.objects.all()
    context_object_name = "usuario"
    form_class = UsuarioEditForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        usuario = User.objects.get(id=self.kwargs['pk'])
        companias = Compania.objects.all()
        companias_usuario = usuario.perfil.companias.all()
        ubicaciones = Ubicacion.objects.all()
        ubicaciones_usuario = usuario.perfil.ubicaciones.all()
        aplicaciones = Aplicacion.objects.all()
        aplicaciones_usuario = usuario.perfil.aplicaciones.all()
        talleres = Taller.objects.all()
        talleres_usuario = usuario.perfil.talleres.all()
        usuarios = User.objects.all()
        context["aplicaciones"] = aplicaciones
        context["aplicaciones_usuario"] = aplicaciones_usuario
        context["companias"] = companias
        context["companias_usuario"] = companias_usuario
        context["sucursales"] = ubicaciones
        context["sucursales_usuario"] = ubicaciones_usuario
        context["talleres"] = talleres
        context["talleres_usuario"] = talleres_usuario
        context["usuarios"] = usuarios
        return context

    def form_valid(self, form):
        # Save form data
        print(form.cleaned_data)
        c = {'form': form, }
        user = form.save(commit=False)
        groups = form.cleaned_data['groups']
        groups = Group.objects.get(name=groups)
        compania = self.request.POST.getlist("compania")
        compania = Compania.objects.filter(compania__in=compania)
        ubicacion = self.request.POST.getlist("ubicacion")
        ubicacion = Ubicacion.objects.filter(nombre__in=ubicacion)
        aplicacion = self.request.POST.getlist("aplicacion")
        aplicacion = Aplicacion.objects.filter(nombre__in=aplicacion)
        taller = self.request.POST.getlist("taller")
        taller = Taller.objects.filter(nombre__in=taller)
        try:
            password = form.cleaned_data['password']
            repeat_password = form.cleaned_data['repeat_password']
            pass
            if password != repeat_password:
                messages.error(self.request, "Passwords do not Match", extra_tags='alert alert-danger')
                return render(self.request, self.template_name, c)
            user.set_password(password)
        except:
            pass
        user.save()
        
        perfil = Perfil.objects.get(user=user)
        for comp in compania:
            perfil.companias.add(comp)
        for ubi in ubicacion:
            perfil.ubicaciones.add(ubi)
        for apli in aplicacion:
            perfil.aplicaciones.add(apli)
        for tall in taller:
            perfil.talleres.add(tall)
        user.groups.add(groups)
        perfil.save()
        return super(usuarioFormularioEditView, self).form_valid(form)

    def get_success_url(self):
        return reverse_lazy("dashboards:usuarioFormulario")

def usuarioFormularioDeleteView(request):
    if request.method =="POST":
        usuario = User.objects.get(id=request.POST.get("id"))
        usuario.delete()
        return redirect("dashboards:usuarioFormulario")
    return redirect("dashboards:usuarioFormulario")

class nuevoPresupuestoView(LoginRequiredMixin, CreateView):
    # Vista de nuevoPresupuestoView

    template_name = "formularios/presupuesto.html"
    form_class = PresupuestoForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        compania = self.request.user.perfil.compania
        companias = self.request.user.perfil.companias.all()
        sucursales = self.request.user.perfil.ubicaciones.all()
        clases = Vehiculo.opciones_clase
        configuraciones = Vehiculo.opciones_configuracion
        presupuestos = Presupuesto.objects.filter(compania=self.request.user.perfil.compania)

        print("hola1")
        if self.request.method =='POST':
            print("hola")
            mes = self.request.POST.get('mes')
            ano = self.request.POST.get('ano')
            compania = self.request.POST.get('compania')
            presupuesto = self.request.POST.get('presupuesto')
            ubicacion = self.request.POST.get('ubicacion')
            gasto_real = self.request.POST.get('gasto_real')
            km_recorridos = self.request.POST.get('km_recorridos')

            if mes == "Enero":
                mes = 1
            elif mes == "Febrero":
                mes = 2
            elif mes == "Marzo":
                mes = 3
            elif mes == "Abril":
                mes = 4
            elif mes == "Mayo":
                mes = 5
            elif mes == "Junio":
                mes = 6
            elif mes == "Julio":
                mes = 7
            elif mes == "Agosto":
                mes = 8
            elif mes == "Septiembre":
                mes = 9
            elif mes == "Octubre":
                mes = 10
            elif mes == "Noviembre":
                mes = 11
            elif mes == "Diciembre":
                mes = 12

            print("hola2")
            print("ano", ano)
            print("mes", mes)
            print("ano2", type(ano))
            print("mes2", type(mes))

            mes_ano = datetime(int(ano), int(mes), 1)

            Presupuesto.objects.create(mes_ano=mes_ano,
                                        compania=Compania.objects.get(compania=compania),
                                        ubicacion=Ubicacion.objects.get(nombre=ubicacion),
                                        presupuesto=presupuesto,
                                        gasto_real=gasto_real,
                                        km_recorridos=km_recorridos
            )
        print(presupuestos)
        context["clases"] = clases
        context["compania"] = compania
        context["companias"] = companias
        context["configuraciones"] = configuraciones
        context["sucursales"] = sucursales
        context["presupuestos"] = presupuestos

        return context

    def get_success_url(self):
        return reverse_lazy("dashboards:nuevoPresupuesto")

class nuevoPresupuestoEditView(LoginRequiredMixin, DetailView, UpdateView):
    # Vista de catalogonuevoPresupuestotView

    template_name = "formularios/presupuesto.html"
    slug_field = "presupuesto"
    slug_url_kwarg = "presupuesto"
    queryset = Presupuesto.objects.all()
    context_object_name = "presupuesto"
    form_class = PresupuestoEditForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sucursales = self.request.user.perfil.ubicaciones.all()
        companias = self.request.user.perfil.companias.all()
        presupuestos = Presupuesto.objects.filter(compania=self.request.user.perfil.compania)
        
        if self.request.method =='POST':
            print("hola")
            id = self.request.POST.get('id')
            mes = self.request.POST.get('mes')
            ano = self.request.POST.get('ano')
            compania = self.request.POST.get('compania')
            presupuesto = self.request.POST.get('presupuesto')
            ubicacion = self.request.POST.get('ubicacion')
            gasto_real = self.request.POST.get('gasto_real')
            km_recorridos = self.request.POST.get('km_recorridos')

            if mes == "Enero":
                mes = 1
            elif mes == "Febrero":
                mes = 2
            elif mes == "Marzo":
                mes = 3
            elif mes == "Abril":
                mes = 4
            elif mes == "Mayo":
                mes = 5
            elif mes == "Junio":
                mes = 6
            elif mes == "Julio":
                mes = 7
            elif mes == "Agosto":
                mes = 8
            elif mes == "Septiembre":
                mes = 9
            elif mes == "Octubre":
                mes = 10
            elif mes == "Noviembre":
                mes = 11
            elif mes == "Diciembre":
                mes = 12

            print("hola2")
            print("ano", ano)
            print("mes", mes)
            print("ano2", type(ano))
            print("mes2", type(mes))

            mes_ano = datetime(int(ano), int(mes), 1)

            presupuesto_id = Presupuesto.objects.get(id=id)
            presupuesto_id.mes_ano = mes_ano
            presupuesto_id.compania = Compania.objects.get(compania=compania)
            presupuesto_id.ubicacion = Ubicacion.objects.get(nombre=ubicacion)
            presupuesto_id.presupuesto = presupuesto
            presupuesto_id.gasto_real = gasto_real
            presupuesto_id.km_recorridos = km_recorridos
            presupuesto_id.save()

        context["companias"] = companias
        context["sucursales"] = sucursales
        context["presupuestos"] = presupuestos

        return context

    def get_success_url(self):
        return reverse_lazy("dashboards:sucursalFormulario")

def nuevoPresupuestoDeleteView(request):
    if request.method =="POST":
        presupuesto = Presupuesto.objects.get(id=request.POST.get("id"))
        presupuesto.delete()
        return redirect("dashboards:nuevoPresupuesto")
    return redirect("dashboards:nuevoPresupuesto")


class nuevoVehiculoView(LoginRequiredMixin, CreateView):
    # Vista de nuevoVehiculoView

    template_name = "formularios/vehiculo.html"
    form_class = VehiculoForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        aplicaciones = self.request.user.perfil.aplicaciones.all()
        compania = self.request.user.perfil.compania
        sucursales = self.request.user.perfil.ubicaciones.all()
        clases = Vehiculo.opciones_clase
        configuraciones = Vehiculo.opciones_configuracion
        vehiculos = Vehiculo.objects.filter(compania=self.request.user.perfil.compania)

        context["aplicaciones"] = aplicaciones
        context["clases"] = clases
        context["compania"] = compania
        context["configuraciones"] = configuraciones
        context["sucursales"] = sucursales
        context["vehiculos"] = vehiculos

        return context

    def form_valid(self, form):
        # Save form data
        print("hola")
        print(form.cleaned_data)
        c = {'form': form, }
        vehiculo = form.save(commit=False)
        estatus_activo = form.cleaned_data['estatus_activo']
        nuevo = form.cleaned_data['nuevo']
        if estatus_activo == "activo":
            vehiculo.estatus_activo = True
        else:
            vehiculo.estatus_activo = False
        if nuevo == "nuevo":
            vehiculo.nuevo = True
        else:
            vehiculo.nuevo = False
        configuracion = vehiculo.configuracion
        numero_de_llantas = functions.cantidad_llantas(configuracion)
        vehiculo.numero_de_llantas = numero_de_llantas
        vehiculo.save()
        functions.crear_llantas_nuevos_vehiculos(vehiculo)
        
        return super(nuevoVehiculoView, self).form_valid(form)

    def get_success_url(self):
        return reverse_lazy("dashboards:nuevoVehiculo")

class nuevoVehiculoEditView(LoginRequiredMixin, DetailView, UpdateView):
    # Vista de nuevoVehiculoView

    template_name = "formularios/vehiculo.html"
    slug_field = "vehiculo"
    slug_url_kwarg = "vehiculo"
    queryset = Vehiculo.objects.all()
    context_object_name = "vehiculo"
    form_class = VehiculoEditForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.request.method == 'POST':

            vehiculo = Vehiculo.objects.get(id=self.kwargs['pk'])
            numero_economico = self.request.POST.get("numero_economico")
            modelo = self.request.POST.get("modelo")
            marca = self.request.POST.get("marca")
            aplicacion = self.request.POST.get("aplicacion")
            ubicacion = self.request.POST.get("ubicacion")
            clase = self.request.POST.get("clase")
            configuracion = self.request.POST.get("configuracion")
            presion_establecida_1 = self.request.POST.get("presion_establecida_1")
            presion_establecida_2 = self.request.POST.get("presion_establecida_2")
            presion_establecida_3 = self.request.POST.get("presion_establecida_3")
            presion_establecida_4 = self.request.POST.get("presion_establecida_4")
            presion_establecida_5 = self.request.POST.get("presion_establecida_5")
            presion_establecida_6 = self.request.POST.get("presion_establecida_6")
            presion_establecida_7 = self.request.POST.get("presion_establecida_7")
            km_diario_maximo = self.request.POST.get("km_diario_maximo")
            estatus_activo = self.request.POST.get("estatus_activo")
            nuevo = self.request.POST.get("nuevo")
            vehiculo.numero_economico = numero_economico
            vehiculo.modelo = modelo
            vehiculo.marca = marca
            vehiculo.aplicacion = Aplicacion.objects.get(nombre=aplicacion)
            vehiculo.ubicacion = Ubicacion.objects.get(nombre=ubicacion)
            vehiculo.clase = clase
            vehiculo.configuracion = configuracion
            if presion_establecida_1:
                vehiculo.presion_establecida_1 = presion_establecida_1
            if presion_establecida_2:
                vehiculo.presion_establecida_2 = presion_establecida_2
            if presion_establecida_3:
                vehiculo.presion_establecida_3 = presion_establecida_3
            if presion_establecida_4:
                vehiculo.presion_establecida_4 = presion_establecida_4
            if presion_establecida_5:
                vehiculo.presion_establecida_5 = presion_establecida_5
            if presion_establecida_6:
                vehiculo.presion_establecida_6 = presion_establecida_6
            if presion_establecida_7:
                vehiculo.presion_establecida_7 = presion_establecida_7
            vehiculo.km_diario_maximo = km_diario_maximo
            if estatus_activo == "activo":
                vehiculo.estatus_activo = True
            else:
                vehiculo.estatus_activo = False
            if nuevo == "nuevo":
                vehiculo.nuevo = True
            else:
                vehiculo.nuevo = False
            configuracion = vehiculo.configuracion
            numero_de_llantas = functions.cantidad_llantas(configuracion)
            vehiculo.numero_de_llantas = numero_de_llantas
            vehiculo.save()

        aplicaciones = self.request.user.perfil.aplicaciones.all()
        companias = self.request.user.perfil.companias.all()
        sucursales = self.request.user.perfil.ubicaciones.all()
        clases = Vehiculo.opciones_clase
        configuraciones = Vehiculo.opciones_configuracion
        vehiculos = Vehiculo.objects.filter(compania=self.request.user.perfil.compania)

        context["aplicaciones"] = aplicaciones
        context["clases"] = clases
        context["companias"] = companias
        context["configuraciones"] = configuraciones
        context["sucursales"] = sucursales
        context["vehiculos"] = vehiculos
        return context
        
        return super(nuevoVehiculoEditView, self).form_valid(form)

    def get_success_url(self):
        return reverse_lazy("dashboards:nuevoVehiculo")

def nuevoVehiculoDeleteView(request):
    if request.method =="POST":
        vehiculo = Vehiculo.objects.get(id=request.POST.get("id"))
        vehiculo.delete()
        return redirect("dashboards:nuevoVehiculo")
    return redirect("dashboards:nuevoVehiculo")

class aplicacionFormularioView(LoginRequiredMixin, CreateView):
    # Vista de aplicacionFormularioView

    template_name = "formularios/aplicacion.html"
    form_class = AplicacionForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        companias = self.request.user.perfil.companias.all()
        aplicaciones = self.request.user.perfil.aplicaciones.all()
        sucursales = self.request.user.perfil.ubicaciones.all()

        context["aplicaciones"] = aplicaciones
        context["companias"] = companias
        context["sucursales"] = sucursales
        return context

    def get_success_url(self):
        user = self.request.user.perfil
        user.aplicaciones.add(Aplicacion.objects.get(nombre=self.request.POST.get("nombre")))
        user.save()
        return reverse_lazy("dashboards:aplicacionFormulario")

class aplicacionFormularioEditView(LoginRequiredMixin, DetailView, UpdateView):
    # Vista de catalogoTalleresEditView

    template_name = "formularios/aplicacion.html"
    slug_field = "aplicacion"
    slug_url_kwarg = "aplicacion"
    queryset = Aplicacion.objects.all()
    context_object_name = "aplicacion"
    form_class = AplicacionEditForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        companias = self.request.user.perfil.companias.all()
        aplicaciones = self.request.user.perfil.aplicaciones.all()
        sucursales = self.request.user.perfil.ubicaciones.all()

        context["aplicaciones"] = aplicaciones
        context["companias"] = companias
        context["sucursales"] = sucursales

        return context

    def get_success_url(self):
        return reverse_lazy("dashboards:aplicacionFormulario")

def aplicacionFormularioDeleteView(request):
    if request.method =="POST":
        aplicacion = Aplicacion.objects.get(id=request.POST.get("id"))
        aplicacion.delete()
        return redirect("dashboards:aplicacionFormulario")
    return redirect("dashboards:aplicacionFormulario")

class perdidaRendimientoView(LoginRequiredMixin, TemplateView):
# Vista de perdidaRendimientoView

    template_name = "informes/perdida-rendimiento.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        compania = Compania.objects.get(compania=self.request.user.perfil.compania)
        ubicaciones = Ubicacion.objects.filter(compania=compania)
        aplicaciones = Aplicacion.objects.filter(compania=compania)
        
        context["aplicaciones"] = aplicaciones
        context["compania"] = compania
        context["sucursales"] = ubicaciones
        return context

class inspeccionVehiculo(LoginRequiredMixin, TemplateView):
    # Vista de inspeccionVehiculo

    template_name = "inspeccion-vehiculo.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        #Generacion de ejes dinamico
        vehiculo_actual = Vehiculo.objects.get(pk = self.kwargs['pk'])
        llantas_actuales = Llanta.objects.filter(vehiculo = self.kwargs['pk'], inventario="Rodante")
        inspecciones_actuales = Inspeccion.objects.filter(llanta__in=llantas_actuales)
        
        
        num_eco_comp = functions.all_num_eco_compania(vehiculo_actual.compania, llantas_actuales)
        context['num_eco_comp'] = num_eco_comp
        #Obtencion de la data
        num_ejes = vehiculo_actual.configuracion.split('.')
        ejes_no_ordenados = []
        ejes = []
        eje = 1
        color_profundidad = ""
        presiones_establecida=[
            vehiculo_actual.presion_establecida_1,
            vehiculo_actual.presion_establecida_2,
            vehiculo_actual.presion_establecida_3,
            vehiculo_actual.presion_establecida_4,
            vehiculo_actual.presion_establecida_5,
            vehiculo_actual.presion_establecida_6,
            vehiculo_actual.presion_establecida_7,
        ]
        
        numero = 0
        for num in num_ejes:
            list_temp = []
            for llanta in llantas_actuales:
                
                color_profundidad = 'bad'
                # Obtención de los datos de las presiones
                objetivo = llanta.vehiculo.compania.objetivo / 100
                presion_act = llanta.presion_actual
                presion_minima = presiones_establecida[numero] - (presiones_establecida[numero] * objetivo)
                presion_maxima = presiones_establecida[numero] + (presiones_establecida[numero] * objetivo)

                try:
                    if presion_act >= presion_minima and presion_act <= presion_maxima:
                        color_presion = 'good'
                    elif presion_act < presion_minima:
                        color_presion = 'bad'
                    elif presion_act > presion_maxima:
                        color_presion = 'yellow'
                except:
                    color_presion = 'good'    
                    
                inspecciones_llanta = Inspeccion.objects.filter(llanta = llanta)
                total_inspecciones = len(inspecciones_llanta)
                
                min_produndidad = functions.min_profundidad(llanta)
                max_produndidad = functions.max_profundidad(llanta)
                
                punto_retiro = functions.punto_de_retiro(llanta)
                
                minimos = functions.min_profundidad_template(llanta)
                mm_dif = llanta.vehiculo.compania.mm_de_diferencia_entre_duales
                context['mm_dif'] = mm_dif
                try:
                    #km_max = functions.km_max(vehiculo_actual)
                    km_max = ''
                    
                except:
                    km_max = ''
                inspeccion = True
                if llanta.ultima_inspeccion == None:
                    inspeccion = False
                context['km_max'] = km_max
                data = {
                    'presion_minima': presion_minima,
                    'presion_maxima': presion_maxima,
                    'punto_retiro': punto_retiro,
                    'min_izquierda': minimos['min_izquierda'],
                    'min_central': minimos['min_central'],
                    'min_derecha': minimos['min_derecha'],
                    'min_produndidad': min_produndidad,
                    'max_produndidad': max_produndidad,
                    'desgaste_anterior_llanta': llanta.parametro_desgaste_irregular,
                    'inspeccion': inspeccion
                        }
                                  
                if llanta.eje == eje:
                    color_llanta = functions.color_observaciones_all_one(llanta)
                    eForm = InspeccionForm(instance = llanta.ultima_inspeccion) 
                    list_temp.append([llanta, color_profundidad, eForm, color_presion, total_inspecciones, min_produndidad, data, color_llanta])
            eje += 1
            ejes_no_ordenados.append(list_temp)
            numero += 1
            
        ejes = functions.acomodo_ejes(ejes_no_ordenados)
            
        color = functions.entrada_correcta(vehiculo_actual, None)
        #print(color)
        if color == 'good':
            style = 'good'
        elif color == 'bad':
            style = 'bad'
        else:
            style = 'bad'
        
        cant_ejes = len(ejes)
        
        #obvervaciones manuales
        observaciones_manuales=Observacion.objects.filter(nivel='Llanta', automatico=False)
        observaciones_automaticas=Observacion.objects.filter(nivel='Llanta', automatico=True).exclude(observacion='Ponchado seguro').exclude(observacion='Mala entrada')
        observaciones_vehiculo = Observacion.objects.filter(nivel='Vehiculo')
        #print(vehiculo.configuracion)
        #print(ejes)
        #print(f'style: {style}')
        #print(f'llantas_sospechosas: {llantas_sospechosas}')
        #print(f'llantas_rojas: {llantas_rojas}')
        #print(f'llantas_amarillas: {llantas_amarillas}')
        #print(f'llantas_azules: {llantas_azules}')
        
        mm_de_desgaste_compania = vehiculo_actual.compania.mm_de_desgaste_irregular
        
        context['ejes'] = ejes
        context['style'] = style
        context['cant_ejes'] = cant_ejes
        context['llantas_actuales'] = llantas_actuales
        #Generacion de productos
        productos = Producto.objects.filter(compania=vehiculo_actual.compania)
        context['productos'] = productos
        context['vehiculo_actual'] = vehiculo_actual
        context['observaciones_manuales'] = observaciones_manuales
        context['observaciones_automaticas'] = observaciones_automaticas
        context['observaciones_vehiculo'] = observaciones_vehiculo
        context['mm_de_desgaste_compania'] = mm_de_desgaste_compania
        print(date.today())
        return context
    
    def post(self, request, *args, **kwargs):
        print(request.POST)
        usuario_actual = Perfil.objects.get(user = self.request.user)
        
        #print(self.kwargs['pk'])
        print('---------------------')
        print('---------------------')
        #Vehiculo
        id_vehiculo = request.POST.getlist('vehiculo_id')
        kilometraje = request.POST.getlist('kilometraje')
        observacion_vehiculo = request.POST.getlist('observation-vehiculo')
        print(f'id_vehiculo: {id_vehiculo}')
        print(f'kilometraje: {kilometraje}')
        print(f'observacion_vehiculo: {observacion_vehiculo}')
        #Llantas
        ids = request.POST.getlist('ids')
        llanta = request.POST.getlist('llanta')
        producto = request.POST.getlist('producto')
        vida = request.POST.getlist('vida')
        profundidad_derecha = request.POST.getlist('profundidad_derecha')
        profundidad_central = request.POST.getlist('profundidad_central')
        profundidad_izquierda = request.POST.getlist('profundidad_izquierda')
        presion = request.POST.getlist('presion')
        observaciones = request.POST.getlist('observaciones')
        reemplazar = request.POST.getlist('reemplazar')
        observaciones_list = {}
        for id in ids:
            try:
                cadena = str('manual-observation-') + str(id)
                if len(request.POST.getlist(cadena)) > 0:
                    observaciones_list[id] = request.POST.getlist(cadena)
            except:
                pass
        #print(f'ids: {ids}')
        #print(f'llanta: {llanta}')
        #print(f'producto: {producto}')
        #print(f'vida: {vida}')
        #print(f'profundidad_derecha: {profundidad_derecha}')
        #print(f'profundidad_central: {profundidad_central}')
        #print(f'profundidad_izquierda: {profundidad_izquierda}')
        #print(f'presion: {presion}')
        #print(f'observaciones: {observaciones}')
        #print(f'reemplazar: {reemplazar}')
        cambios = []
        diferencia_presion_duales_list = []
        desdualizacion_list=[]
        elementos = 0
        
        #Cambio al vehiculo
        vehiculo_referencia = Vehiculo.objects.get(pk = self.kwargs['pk'])
        km_ref = (float(vehiculo_referencia.km) if vehiculo_referencia.km!=None else -99999999)
        vehiculo = Vehiculo.objects.get(pk = self.kwargs['pk'])
        
            #Creacion de la bitacora del vehicul
        vehiculo_inspeccion = InspeccionVehiculo.objects.create(
            usuario = usuario_actual,
            vehiculo = vehiculo,
            fecha = date.today()
        )
        
            #Km del vehiculo
        if kilometraje[0] != vehiculo_referencia.km:
            if kilometraje[0] == "":
                vehiculo.km = None
                vehiculo_inspeccion.km = None
            else:
                vehiculo.km = kilometraje[0]
                vehiculo_inspeccion.km = kilometraje[0]
            #Observaciones del vehiculo
        vehiculo.observaciones.clear()
        if len(observacion_vehiculo) > 0:
            for observacion in observacion_vehiculo:
                obs = Observacion.objects.get(observacion=observacion)
                vehiculo.observaciones.add(obs)
                vehiculo_inspeccion.observaciones.add(obs)
        vehiculo.save()
        vehiculo_inspeccion.save()
        
        #Cambio de llantas
        for id_actual in ids:
            cambios = []
            if id_actual in reemplazar:
                print(f"Remplazando: {id_actual}".center(50, '-'))                
            else:
                print(f"Actualizando: {id_actual}".center(50, '-'))
                
                llanta_referencia = Llanta.objects.get(pk = ids[elementos])
                llanta_actual = Llanta.objects.get(pk = ids[elementos])
                #Actualzacion de km llanta
                if vehiculo.km != None:
                    if float(kilometraje[0]) != km_ref:
                        cambios.append('0')
                        if llanta_actual.km_montado != None:
                            km_actual_nuevo = functions.actualizar_km_actual(llanta_actual, llanta_referencia, vehiculo, vehiculo_referencia)
                            llanta_actual.km_actual = km_actual_nuevo
                            llanta_actual.save()
                        
                #Creando Nueva Inspeccion
                economico = llanta[elementos]
                producto_post = Producto.objects.get(producto=producto[elementos])
                vida_post = vida[elementos]
                profundidad_izquierda_post = (float(profundidad_izquierda[elementos]) if profundidad_izquierda[elementos]!='' else None)
                profundidad_central_post = (float(profundidad_central[elementos]) if profundidad_central[elementos]!='' else None)
                profundidad_derecha_post = (float(profundidad_derecha[elementos]) if profundidad_derecha[elementos]!='' else None)
                presion_post = (presion[elementos] if presion[elementos] != '' else llanta_actual.presion_actual)
                if vida_post != llanta_referencia.vida:
                    print('Cambio de vida')
                    cambios.append('1')
                if str(profundidad_izquierda_post) != str(llanta_referencia.profundidad_izquierda):
                    cambios.append('2')
                if str(profundidad_central_post) != str(llanta_referencia.profundidad_central):
                    cambios.append('3')
                if str(profundidad_derecha_post) != str(llanta_referencia.profundidad_derecha):
                    cambios.append('4')
                if str(presion_post) != str(llanta_referencia.presion_actual):
                    cambios.append('5')
                
                if llanta_actual.km_actual == None:
                    km_act = 0
                else:
                    km_act = llanta_actual.km_actual
                cambios.append('default')    
                if len(cambios) > 0:
                    
                    inspeccion_nueva = Inspeccion.objects.create(
                        tipo_de_evento = 'Inspección',
                        inspeccion_vehiculo = vehiculo_inspeccion,
                        llanta = llanta_actual,
                        posicion = llanta_actual.posicion,
                        tipo_de_eje = llanta_actual.tipo_de_eje,
                        eje = llanta_actual.eje,
                        usuario = usuario_actual,
                        vehiculo = vehiculo,
                        fecha_hora = date.today(),
                        vida = vida_post,
                        km_vehiculo = vehiculo.km,
                        presion = presion_post,
                        presion_establecida = functions.presion_establecida(llanta_actual),
                        profundidad_izquierda = profundidad_izquierda_post,
                        profundidad_central = profundidad_central_post,
                        profundidad_derecha = profundidad_derecha_post,
                        evento = str({\
                        "llanta_inicial" : str(llanta_actual.id), "llanta_mod" : "",\
                        "numero_economico": economico, "numero_economico_mod": "",
                        "producto_inicial" : str(producto_post), "producto_mod" : "",\
                        "vida_inicial" : vida_post, "vida_mod" : "",\
                        "km_inicial" : str(vehiculo.km), "km_mod" : "",\
                        "presion_inicial" : str(presion_post), "presion_mod" : "",\
                        "profundidad_izquierda_inicial" : str(profundidad_izquierda_post), "profundidad_izquierda_mod" : "",\
                        "profundidad_central_inicial" : str(profundidad_central_post), "profundidad_central_mod" : "",\
                        "profundidad_derecha_inicial" : str(profundidad_derecha_post), "profundidad_derecha_mod" : ""\
                        })
                    )
                    inspeccion_nueva.save()
                    
                    if llanta_actual.primera_inspeccion == None and inspeccion_nueva.vida == llanta_actual.vida:
                        llanta_actual.primera_inspeccion = inspeccion_nueva
                        
                    llanta_actual.ultima_inspeccion = inspeccion_nueva
                    llanta_actual.save()
                    #Actualizacion de los datos de la llante
                    llanta_actual.numero_economico = economico
                    llanta_actual.producto = producto_post
                    llanta_actual.vida = vida_post
                    llanta_actual.presion_actual = presion_post
                    llanta_actual.profundidad_izquierda = profundidad_izquierda_post
                    llanta_actual.profundidad_central = profundidad_central_post
                    llanta_actual.profundidad_derecha = profundidad_derecha_post
                    llanta_actual.save() 
                    if 'vida' in cambios:
                        functions.cambio_de_vida(llanta_referencia, llanta_actual)
                    inspecciones_vehiculos = InspeccionVehiculo.objects.filter(vehiculo = llanta_actual.vehiculo).exclude(km = None)
                    if vehiculo.km != None:
                        if kilometraje[0] != vehiculo_referencia.km:
                            if llanta_actual.km_montado == None:
                                inspecciones = Inspeccion.objects.filter(llanta = llanta_actual, inspeccion_vehiculo__in = inspecciones_vehiculos, vida = llanta_actual.vida)
                                if len(inspecciones) >= 2:
                                    print(llanta_actual.vida)
                                    print(inspecciones)
                                    print('Sin km de montado pero con inspecciones suficinetes')
                                    print(llanta_actual.id)
                                    primer_inspeccion = inspecciones.first()
                                    ultima_inspeccion = inspecciones.last()
                                    km_teorico = functions.actualizar_km_actual_no_km_montado(primer_inspeccion, ultima_inspeccion)
                                    print(km_teorico)
                                    if km_teorico != None:
                                        llanta_actual.km_actual = km_teorico
                                        llanta_actual.save()
                                else:
                                    print('Sin km de montado y sin inspecciones suficinetes')
                                    print(llanta_actual.id)
                    
                    print('Si paso to')
                else:
                    print('No paso na')
            elementos += 1
            
        elementos = 0
            
        for id_actual in ids:
            cambios = []
            if id_actual in reemplazar:
                print(f"Remplazando: {id_actual}".center(50, '-'))                
            else:
                print(f"Actualizando: {id_actual}".center(50, '-'))
                
                llanta_referencia = Llanta.objects.get(pk = ids[elementos])
                llanta_actual = Llanta.objects.get(pk = ids[elementos])   
                #observaciones
                llanta_actual.observaciones.clear()
                
                if ids[elementos] in observaciones_list:
                    for i in (observaciones_list[ids[elementos]]):
                        observacion = Observacion.objects.get(observacion = i)
                        llanta_actual.observaciones.add(observacion)
                        llanta_actual.save()
                compania = llanta_actual.vehiculo.compania
                presiones_establecidas = [
                    llanta_actual.vehiculo.presion_establecida_1,
                    llanta_actual.vehiculo.presion_establecida_2,
                    llanta_actual.vehiculo.presion_establecida_3,
                    llanta_actual.vehiculo.presion_establecida_4,
                    llanta_actual.vehiculo.presion_establecida_5,
                    llanta_actual.vehiculo.presion_establecida_6,
                    llanta_actual.vehiculo.presion_establecida_7,
                ]
                establecida = presiones_establecidas[(llanta_actual.eje - 1)]
                presion_minima = establecida - (establecida * (compania.objetivo/100))
                presion_maxima = establecida + (establecida * (compania.objetivo/100))
                if llanta_actual.presion_actual != None:
                    if float(llanta_actual.presion_actual) < presion_minima:
                        baja_presion = Observacion.objects.get(observacion = 'Baja presión')
                        llanta_actual.observaciones.add(baja_presion)
                        print(baja_presion)
                        
                if llanta_actual.presion_actual != None:    
                    if float(llanta_actual.presion_actual) > presion_maxima:
                        alta_presion = Observacion.objects.get(observacion = 'Alta presion')
                        llanta_actual.observaciones.add(alta_presion)
                        print(alta_presion)
                
                if '4' in llanta_actual.tipo_de_eje:
                    print('Duales')
                    eje = str(llanta_actual.eje)
                    posicion = llanta_actual.posicion[1:]
                    #print(eje+posicion)
                    if posicion == 'LO':
                        dual = 'LI'
                    elif posicion == 'LI':
                        dual = 'LO'
                    elif posicion == 'RI':
                        dual = 'RO'
                    elif posicion == 'RO':
                        dual = 'RI'
                    dual_completo = eje + dual
                    llantas_comparacion = Llanta.objects.filter(vehiculo = llanta_actual.vehiculo, inventario="Rodante")
                    for llanta_i in llantas_comparacion:
                        if llanta_i.posicion == dual_completo:
                            dual_llanta = llanta_i
                    if llanta_actual.presion_actual != None and dual_llanta.presion_actual != None:
                        presion_actual = float(llanta_actual.presion_actual) if float(llanta_actual.presion_actual) != 0 else 1
                        presion_dual = dual_llanta.presion_actual if dual_llanta.presion_actual != 0 else 1
                        porcentaje_dif = (float(presion_actual) - presion_dual) / float(presion_actual)
                        print(f'presion_actual {presion_actual}')
                        print(f'presion_dual {presion_dual}')
                        print(f'porcentaje_dif {porcentaje_dif}')
                        #porcentaje_dif = (float(llanta_actual.presion_actual) - dual_llanta.presion_actual) / float(llanta_actual.presion_actual)
                        if llanta_actual in diferencia_presion_duales_list:
                        
                            diferencia_presion_duales = Observacion.objects.get(observacion = 'Diferencia de presión entre los duales')
                            llanta_actual.observaciones.add(diferencia_presion_duales)
                            llanta_actual.ultima_inspeccion.observaciones.add(diferencia_presion_duales)
                            llanta_actual.ultima_inspeccion.save()
                            print(diferencia_presion_duales)
                        else:

                            if porcentaje_dif > 0.1:
                                #Poner a los 2 duales
                                diferencia_presion_duales = Observacion.objects.get(observacion = 'Diferencia de presión entre los duales')
                                llanta_actual.observaciones.add(diferencia_presion_duales)
                                dual_llanta.observaciones.add(diferencia_presion_duales)
                                dual_llanta.save()
                                diferencia_presion_duales_list.append(dual_llanta)
                                print(diferencia_presion_duales)
                            
                    if llanta_actual in desdualizacion_list:
                        desdualización = Observacion.objects.get(observacion = 'Desdualización')
                        llanta_actual.observaciones.add(desdualización)
                        llanta_actual.ultima_inspeccion.observaciones.add(desdualización)
                        llanta_actual.ultima_inspeccion.save()
                        print(desdualización)
                    else:
                        print(functions.min_profundidad(llanta_actual))
                        print(functions.min_profundidad(dual_llanta))
                        print(compania.mm_de_diferencia_entre_duales)
                        if (functions.min_profundidad(llanta_actual) - functions.min_profundidad(dual_llanta)) >= compania.mm_de_diferencia_entre_duales:
                            #Poner a los 2 duales
                            desdualización = Observacion.objects.get(observacion = 'Desdualización')
                            llanta_actual.observaciones.add(desdualización)
                            dual_llanta.observaciones.add(desdualización)
                            dual_llanta.save()
                            desdualizacion_list.append(dual_llanta)
                            print(desdualización)
                    
                if "S" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_direccion
                elif "D" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_traccion
                elif "T" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_arrastre
                elif "C" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_loco
                elif "L" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_retractil
                
                if functions.min_profundidad(llanta_actual) <= punto_retiro:
                    baja_profundidad = Observacion.objects.get(observacion = 'Baja profundidad')
                    llanta_actual.observaciones.add(baja_profundidad)
                    print(baja_profundidad)
                    
                if functions.min_profundidad(llanta_actual) == (punto_retiro + 0.6):
                    en_punto_retiro = Observacion.objects.get(observacion = 'En punto de retiro')
                    llanta_actual.observaciones.add(en_punto_retiro)
                    print(en_punto_retiro)
                izquierda = llanta_actual.profundidad_izquierda
                central = llanta_actual.profundidad_central
                derecha = llanta_actual.profundidad_derecha
                functions.desgaste_profundidad(izquierda, central , derecha, llanta_actual)

                
                llanta_actual.save()
                if len(llanta_actual.observaciones.all()) > 0:
                    for observacion in llanta_actual.observaciones.all():
                        llanta_actual.ultima_inspeccion.observaciones.add(observacion)
                    llanta_actual.ultima_inspeccion.save()
                #print(llanta_act.ultima_inspeccion)
                
            elementos += 1
                
        elementos = 0    
        for id_actual in ids: 
            llanta_referencia = Llanta.objects.get(pk = ids[elementos])
            llanta_actual = Llanta.objects.get(pk = ids[elementos]) 
            for obs in llanta_actual.observaciones.all():
                llanta_actual.ultima_inspeccion.observaciones.add(obs)
                llanta_actual.ultima_inspeccion.save()
                llanta_actual.save()
            elementos += 1
            #Inspeccion      
        functions.observaciones_vehiculo(vehiculo)
        Vehiculo.objects.filter(pk=vehiculo.id).update(fecha_ultima_inspeccion = date.today())
        return redirect('dashboards:detail', self.kwargs['pk'])
        #return redirect('dashboards:inspeccionVehiculo', self.kwargs['pk'])

class CuatroUmbralesView(LoginRequiredMixin, TemplateView):
    # Vista de CatalogoDesechosView

    template_name = "cuatroUmbrales.html"

class vehiculosformularioView(LoginRequiredMixin, TemplateView):
    # Vista de formularioView

    template_name = "formularioVehiculos.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        vehiculos = Vehiculo.objects.filter(compania=self.request.user.perfil.compania)

        context["vehiculos"] = vehiculos
        return context

    def get_success_url(self):
        return reverse_lazy("dashboards:webservices")


class formularioView(LoginRequiredMixin, TemplateView):
    # Vista de formularioView

    template_name = "formulario.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        compania = Compania.objects.get(compania=self.request.user.perfil.compania)
        vehiculo = Vehiculo.objects.get(id=self.kwargs['pk'])
        llantas = Llanta.objects.filter(vehiculo=vehiculo)
        usuario = Compania.objects.get(compania=self.request.user.perfil)

        context["compania"] = compania
        context["llantas"] = llantas
        context["usuario"] = usuario
        context["vehiculo"] = vehiculo
        return context

    def post(self, request, *args, **kwargs):
        print(request.POST)
        #print(self.kwargs['pk'])
        print('---------------------')
        print('---------------------')
        #Vehiculo
        id_vehiculo = request.POST.getlist('vehiculo_id')
        kilometraje = request.POST.getlist('kilometraje')
        observacion_vehiculo = request.POST.getlist('observation-vehiculo')
        print(f'id_vehiculo: {id_vehiculo}')
        print(f'kilometraje: {kilometraje}')
        print(f'observacion_vehiculo: {observacion_vehiculo}')
        #Llantas
        ids = request.POST.getlist('ids')
        llanta = request.POST.getlist('llanta')
        producto = request.POST.getlist('producto')
        vida = request.POST.getlist('vida')
        profundidad_derecha = request.POST.getlist('profundidad_derecha')
        profundidad_central = request.POST.getlist('profundidad_central')
        profundidad_izquierda = request.POST.getlist('profundidad_izquierda')
        presion = request.POST.getlist('presion')
        observaciones = request.POST.getlist('observaciones')
        reemplazar = request.POST.getlist('reemplazar')
        observaciones_list = {}
        for id in ids:
            try:
                cadena = str('manual-observation-') + str(id)
                if len(request.POST.getlist(cadena)) > 0:
                    observaciones_list[id] = request.POST.getlist(cadena)
            except:
                pass
        #print(f'ids: {ids}')
        #print(f'llanta: {llanta}')
        #print(f'producto: {producto}')
        #print(f'vida: {vida}')
        #print(f'profundidad_derecha: {profundidad_derecha}')
        #print(f'profundidad_central: {profundidad_central}')
        #print(f'profundidad_izquierda: {profundidad_izquierda}')
        #print(f'presion: {presion}')
        #print(f'observaciones: {observaciones}')
        #print(f'reemplazar: {reemplazar}')
        cambios = []
        diferencia_presion_duales_list = []
        desdualizacion_list=[]
        elementos = 0
        
        #Cambio al vehiculo
        vehiculo_referencia = Vehiculo.objects.get(pk = self.kwargs['pk'])
        km_ref = (float(vehiculo_referencia.km) if vehiculo_referencia.km!=None else -99999999)
        vehiculo = Vehiculo.objects.get(pk = self.kwargs['pk'])
        
            #Creacion de la bitacora del vehicul
        vehiculo_inspeccion = InspeccionVehiculo.objects.create(
            vehiculo = vehiculo,
            fecha = date.today()
        )
        
            #Km del vehiculo
        if kilometraje[0] != vehiculo_referencia.km:
            if kilometraje[0] == "":
                vehiculo.km = None
                vehiculo_inspeccion.km = None
            else:
                vehiculo.km = kilometraje[0]
                vehiculo_inspeccion.km = kilometraje[0]
            #Observaciones del vehiculo
        vehiculo.observaciones.clear()
        if len(observacion_vehiculo) > 0:
            for observacion in observacion_vehiculo:
                obs = Observacion.objects.get(observacion=observacion)
                vehiculo.observaciones.add(obs)
                vehiculo_inspeccion.observaciones.add(obs)
        vehiculo.save()
        vehiculo_inspeccion.save()
        
        #Cambio de llantas
        for id_actual in ids:
            cambios = []
            if id_actual in reemplazar:
                print(f"Remplazando: {id_actual}".center(50, '-'))                
            else:
                print(f"Actualizando: {id_actual}".center(50, '-'))
                
                llanta_referencia = Llanta.objects.get(pk = ids[elementos])
                llanta_actual = Llanta.objects.get(pk = ids[elementos])
                #Actualzacion de km llanta
                if vehiculo.km != None:
                    if float(kilometraje[0]) != km_ref:
                        cambios.append('0')
                        if llanta_actual.km_montado != None:
                            km_actual_nuevo = functions.actualizar_km_actual(llanta_actual, llanta_referencia, vehiculo, vehiculo_referencia)
                            llanta_actual.km_actual = km_actual_nuevo
                            llanta_actual.save()
                        
                #Creando Nueva Inspeccion
                economico = llanta[elementos]
                producto_post = Producto.objects.get(producto=producto[elementos])
                vida_post = vida[elementos]
                profundidad_izquierda_post = (float(profundidad_izquierda[elementos]) if profundidad_izquierda[elementos]!='' else None)
                profundidad_central_post = (float(profundidad_central[elementos]) if profundidad_central[elementos]!='' else None)
                profundidad_derecha_post = (float(profundidad_derecha[elementos]) if profundidad_derecha[elementos]!='' else None)
                presion_post = (presion[elementos] if presion[elementos] != '' else llanta_actual.presion_actual)
                if vida_post != llanta_referencia.vida:
                    functions.cambio_de_vida(llanta_referencia)
                    print('Cambio de vida')
                    cambios.append('1')
                if str(profundidad_izquierda_post) != str(llanta_referencia.profundidad_izquierda):
                    cambios.append('2')
                if str(profundidad_central_post) != str(llanta_referencia.profundidad_central):
                    cambios.append('3')
                if str(profundidad_derecha_post) != str(llanta_referencia.profundidad_derecha):
                    cambios.append('4')
                if str(presion_post) != str(llanta_referencia.presion_actual):
                    cambios.append('5')
                
                if llanta_actual.km_actual == None:
                    km_act = 0
                else:
                    km_act = llanta_actual.km_actual
                cambios.append('default')    
                if len(cambios) > 0:
                    
                    usuario_actual = Perfil.objects.get(user = self.request.user)
                    inspeccion_nueva = Inspeccion.objects.create(
                        tipo_de_evento = 'Inspección',
                        inspeccion_vehiculo = vehiculo_inspeccion,
                        llanta = llanta_actual,
                        posicion = llanta_actual.posicion,
                        tipo_de_eje = llanta_actual.tipo_de_eje,
                        eje = llanta_actual.eje,
                        usuario = usuario_actual,
                        vehiculo = vehiculo,
                        fecha_hora = date.today(),
                        vida = vida_post,
                        km_vehiculo = vehiculo.km,
                        presion = presion_post,
                        presion_establecida = functions.presion_establecida(llanta_actual),
                        profundidad_izquierda = profundidad_izquierda_post,
                        profundidad_central = profundidad_central_post,
                        profundidad_derecha = profundidad_derecha_post,
                        evento = str({\
                        "llanta_inicial" : str(llanta_actual.id), "llanta_mod" : "",\
                        "numero_economico": economico, "numero_economico_mod": "",
                        "producto_inicial" : str(producto_post), "producto_mod" : "",\
                        "vida_inicial" : vida_post, "vida_mod" : "",\
                        "km_inicial" : str(vehiculo.km), "km_mod" : "",\
                        "presion_inicial" : str(presion_post), "presion_mod" : "",\
                        "profundidad_izquierda_inicial" : str(profundidad_izquierda_post), "profundidad_izquierda_mod" : "",\
                        "profundidad_central_inicial" : str(profundidad_central_post), "profundidad_central_mod" : "",\
                        "profundidad_derecha_inicial" : str(profundidad_derecha_post), "profundidad_derecha_mod" : ""\
                        })
                    )
                    inspeccion_nueva.save()
                    llanta_actual.ultima_inspeccion = inspeccion_nueva
                    llanta_actual.save()
                    #Actualizacion de los datos de la llante
                    llanta_actual.numero_economico = economico
                    llanta_actual.producto = producto_post
                    llanta_actual.vida = vida_post
                    llanta_actual.presion_actual = presion_post
                    llanta_actual.profundidad_izquierda = profundidad_izquierda_post
                    llanta_actual.profundidad_central = profundidad_central_post
                    llanta_actual.profundidad_derecha = profundidad_derecha_post
                    llanta_actual.save()
                    
                    inspecciones_vehiculos = InspeccionVehiculo.objects.filter(vehiculo = llanta_actual.vehiculo).exclude(km = None)
                    if vehiculo.km != None:
                        if kilometraje[0] != vehiculo_referencia.km:
                            if llanta_actual.km_montado == None:
                                inspecciones = Inspeccion.objects.filter(llanta = llanta_actual, inspeccion_vehiculo__in = inspecciones_vehiculos, vida = llanta_actual.vida)
                                #inspecciones = Inspeccion.objects.filter(llanta = llanta_actual)
                                if len(inspecciones) >= 2:
                                    print('Sin km de montado pero con inspecciones suficinetes')
                                    print(llanta_actual.id)
                                    primer_inspeccion = inspecciones.first()
                                    ultima_inspeccion = inspecciones.last()
                                    km_teorico = functions.actualizar_km_actual_no_km_montado(primer_inspeccion, ultima_inspeccion)
                                    print(km_teorico)
                                    if km_teorico != None:
                                        llanta_actual.km_actual = km_teorico
                                        llanta_actual.save()
                                else:
                                    print('Sin km de montado y sin inspecciones suficinetes')
                                    print(llanta_actual.id)
                    
                    print('Si paso to')
                else:
                    print('No paso na')
            elementos += 1
            
        elementos = 0
            
        for id_actual in ids:
            cambios = []
            if id_actual in reemplazar:
                print(f"Remplazando: {id_actual}".center(50, '-'))                
            else:
                print(f"Actualizando: {id_actual}".center(50, '-'))
                
                llanta_referencia = Llanta.objects.get(pk = ids[elementos])
                llanta_actual = Llanta.objects.get(pk = ids[elementos])   
                #observaciones
                llanta_actual.observaciones.clear()
                
                if ids[elementos] in observaciones_list:
                    for i in (observaciones_list[ids[elementos]]):
                        observacion = Observacion.objects.get(observacion = i)
                        llanta_actual.observaciones.add(observacion)
                        llanta_actual.save()
                compania = llanta_actual.vehiculo.compania
                presiones_establecidas = [
                    llanta_actual.vehiculo.presion_establecida_1,
                    llanta_actual.vehiculo.presion_establecida_2,
                    llanta_actual.vehiculo.presion_establecida_3,
                    llanta_actual.vehiculo.presion_establecida_4,
                    llanta_actual.vehiculo.presion_establecida_5,
                    llanta_actual.vehiculo.presion_establecida_6,
                    llanta_actual.vehiculo.presion_establecida_7,
                ]
                establecida = presiones_establecidas[(llanta_actual.eje - 1)]
                presion_minima = establecida - (establecida * (compania.objetivo/100))
                presion_maxima = establecida + (establecida * (compania.objetivo/100))
                if llanta_actual.presion_actual != None:
                    if float(llanta_actual.presion_actual) < presion_minima:
                        baja_presion = Observacion.objects.get(observacion = 'Baja presión')
                        llanta_actual.observaciones.add(baja_presion)
                        print(baja_presion)
                        
                if llanta_actual.presion_actual != None:    
                    if float(llanta_actual.presion_actual) > presion_maxima:
                        alta_presion = Observacion.objects.get(observacion = 'Alta presion')
                        llanta_actual.observaciones.add(alta_presion)
                        print(alta_presion)
                
                if '4' in llanta_actual.tipo_de_eje:
                    print('Duales')
                    eje = str(llanta_actual.eje)
                    posicion = llanta_actual.posicion[1:]
                    #print(eje+posicion)
                    if posicion == 'LO':
                        dual = 'LI'
                    elif posicion == 'LI':
                        dual = 'LO'
                    elif posicion == 'RI':
                        dual = 'RO'
                    elif posicion == 'RO':
                        dual = 'RI'
                    dual_completo = eje + dual
                    llantas_comparacion = Llanta.objects.filter(vehiculo = llanta_actual.vehiculo, inventario="Rodante")
                    for llanta_i in llantas_comparacion:
                        if llanta_i.posicion == dual_completo:
                            dual_llanta = llanta_i
                    if llanta_actual.presion_actual != None and dual_llanta.presion_actual != None:
                        presion_actual = float(llanta_actual.presion_actual) if float(llanta_actual.presion_actual) != 0 else 1
                        presion_dual = dual_llanta.presion_actual if dual_llanta.presion_actual != 0 else 1
                        porcentaje_dif = (float(presion_actual) - presion_dual) / float(presion_actual)
                        #porcentaje_dif = (float(llanta_actual.presion_actual) - dual_llanta.presion_actual) / float(llanta_actual.presion_actual)
                        if llanta_actual in diferencia_presion_duales_list:
                        
                            diferencia_presion_duales = Observacion.objects.get(observacion = 'Diferencia de presión entre los duales')
                            llanta_actual.observaciones.add(diferencia_presion_duales)
                            llanta_actual.ultima_inspeccion.observaciones.add(diferencia_presion_duales)
                            llanta_actual.ultima_inspeccion.save()
                            print(diferencia_presion_duales)
                        else:

                            if porcentaje_dif > 0.1:
                                #Poner a los 2 duales
                                diferencia_presion_duales = Observacion.objects.get(observacion = 'Diferencia de presión entre los duales')
                                llanta_actual.observaciones.add(diferencia_presion_duales)
                                dual_llanta.observaciones.add(diferencia_presion_duales)
                                dual_llanta.save()
                                diferencia_presion_duales_list.append(dual_llanta)
                                print(diferencia_presion_duales)
                            
                    if llanta_actual in desdualizacion_list:
                        desdualización = Observacion.objects.get(observacion = 'Desdualización')
                        llanta_actual.observaciones.add(desdualización)
                        llanta_actual.ultima_inspeccion.observaciones.add(desdualización)
                        llanta_actual.ultima_inspeccion.save()
                        print(desdualización)
                    else:
                        print(functions.min_profundidad(llanta_actual))
                        print(functions.min_profundidad(dual_llanta))
                        print(compania.mm_de_diferencia_entre_duales)
                        if (functions.min_profundidad(llanta_actual) - functions.min_profundidad(dual_llanta)) >= compania.mm_de_diferencia_entre_duales:
                            #Poner a los 2 duales
                            desdualización = Observacion.objects.get(observacion = 'Desdualización')
                            llanta_actual.observaciones.add(desdualización)
                            dual_llanta.observaciones.add(desdualización)
                            dual_llanta.save()
                            desdualizacion_list.append(dual_llanta)
                            print(desdualización)
                    
                if "S" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_direccion
                elif "D" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_traccion
                elif "T" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_arrastre
                elif "C" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_loco
                elif "L" in llanta_actual.tipo_de_eje:
                    punto_retiro = compania.punto_retiro_eje_retractil
                
                if functions.min_profundidad(llanta_actual) < punto_retiro:
                    baja_profundidad = Observacion.objects.get(observacion = 'Baja profundidad')
                    llanta_actual.observaciones.add(baja_profundidad)
                    print(baja_profundidad)
                    
                if functions.min_profundidad(llanta_actual) == (punto_retiro + 0.6):
                    en_punto_retiro = Observacion.objects.get(observacion = 'En punto de retiro')
                    llanta_actual.observaciones.add(en_punto_retiro)
                    print(en_punto_retiro)
                izquierda = llanta_actual.profundidad_izquierda
                central = llanta_actual.profundidad_central
                derecha = llanta_actual.profundidad_derecha
                functions.desgaste_profundidad(izquierda, central , derecha, llanta_actual)

                
                llanta_actual.save()
                if len(llanta_actual.observaciones.all()) > 0:
                    for observacion in llanta_actual.observaciones.all():
                        llanta_actual.ultima_inspeccion.observaciones.add(observacion)
                    llanta_actual.ultima_inspeccion.save()
                #print(llanta_act.ultima_inspeccion)
                
            elementos += 1
                
        elementos = 0    
        for id_actual in ids: 
            llanta_referencia = Llanta.objects.get(pk = ids[elementos])
            llanta_actual = Llanta.objects.get(pk = ids[elementos]) 
            for obs in llanta_actual.observaciones.all():
                llanta_actual.ultima_inspeccion.observaciones.add(obs)
                llanta_actual.ultima_inspeccion.save()
                llanta_actual.save()
            elementos += 1
                    
            
            
        #CREACION DE BITACORA
        """print(diferencias)
        if len(diferencias) > 0:
            if id_bitacora == None:
                ""bitacora_cambios = Bitacora_Edicion.objects.create(
                    vehiculo = llanta_ref.vehiculo,
                    tipo = 'inspeccion'
                    )
                bitacora_cambios.save()
                id_bitacora = bitacora_cambios.id
                
            else:
                try:
                    bitacora_cambios = Bitacora_Edicion.objects.get(pk = id_bitacora)
                except:
                    pass""
            
            for diferencia in diferencias:
                registro = Registro_Bitacora.objects.create(
                    bitacora = bitacora_cambios,
                    evento = diferencia
                )
                registro.save()"""
        return redirect('dashboards:detail', self.kwargs['pk'])
        #return redirect('dashboards:inspeccionVehiculo', self.kwargs['pk'])


class SerialVehiculoView(LoginRequiredMixin, TemplateView):
    # Vista de CatalogoDesechosView

    template_name = "serialVehiculo.html"

class TireEyeView(LoginRequiredMixin, TemplateView):
    # Vista de TireEyeViewView

    template_name = "tireEyeView.html"

class reporteVehiculoView(LoginRequiredMixin, TemplateView):
    # Vista de reporteVehiculoView

    template_name = "reporteVehiculo.html"   
    context_object_name = "bitacora"

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)
        if self.kwargs['tipo'] == 'pulpo':
            tipo_bit = 'pulpo'
            bitacora = Bitacora.objects.get(pk = self.kwargs['pk'])
        elif self.kwargs['tipo'] == 'pulpopro':
            bitacora = Bitacora_Pro.objects.get(pk = self.kwargs['pk'])
            tipo_bit = 'pulpopro'

        vehiculo = bitacora.vehiculo
        bitacoraa = Bitacora.objects.filter(vehiculo=Vehiculo.objects.get(numero_economico=vehiculo.numero_economico, compania=Compania.objects.get(compania=self.request.user.perfil.compania)), compania=Compania.objects.get(compania=self.request.user.perfil.compania)).order_by("id")
        bitacoraa_pro = Bitacora_Pro.objects.filter(vehiculo=Vehiculo.objects.get(numero_economico=vehiculo.numero_economico, compania=Compania.objects.get(compania=self.request.user.perfil.compania)), compania=Compania.objects.get(compania=self.request.user.perfil.compania)).order_by("id")

        bitacoraas = bitacoraa.annotate(presion_de_entrada_1=F("presion_de_entrada"), presion_de_salida_1=F("presion_de_salida"), presion_de_entrada_2=Value(None, output_field=IntegerField()), presion_de_salida_2=Value(None, output_field=IntegerField()), presion_de_entrada_3=Value(None, output_field=IntegerField()), presion_de_salida_3=Value(None, output_field=IntegerField()), presion_de_entrada_4=Value(None, output_field=IntegerField()), presion_de_salida_4=Value(None, output_field=IntegerField()), presion_de_entrada_5=Value(None, output_field=IntegerField()), presion_de_salida_5=Value(None, output_field=IntegerField()), presion_de_entrada_6=Value(None, output_field=IntegerField()), presion_de_salida_6=Value(None, output_field=IntegerField()), presion_de_entrada_7=Value(None, output_field=IntegerField()), presion_de_salida_7=Value(None, output_field=IntegerField()), presion_de_entrada_8=Value(None, output_field=IntegerField()), presion_de_salida_8=Value(None, output_field=IntegerField()), presion_de_entrada_9=Value(None, output_field=IntegerField()), presion_de_salida_9=Value(None, output_field=IntegerField()), presion_de_entrada_10=Value(None, output_field=IntegerField()), presion_de_salida_10=Value(None, output_field=IntegerField()), presion_de_entrada_11=Value(None, output_field=IntegerField()), presion_de_salida_11=Value(None, output_field=IntegerField()), presion_de_entrada_12=Value(None, output_field=IntegerField()), presion_de_salida_12=Value(None, output_field=IntegerField())).order_by("id")
        bitacoraas = list(bitacoraas.values("id", "vehiculo__id", "vehiculo__configuracion", "compania__id", "fecha_de_inflado", "tiempo_de_inflado", "presion_de_entrada_1", "presion_de_salida_1", "presion_de_entrada_2", "presion_de_salida_2", "presion_de_entrada_3", "presion_de_salida_3", "presion_de_entrada_4", "presion_de_salida_4", "presion_de_entrada_5", "presion_de_salida_5", "presion_de_entrada_6", "presion_de_salida_6", "presion_de_entrada_7", "presion_de_salida_7", "presion_de_entrada_8", "presion_de_salida_8", "presion_de_entrada_9", "presion_de_salida_9", "presion_de_entrada_10", "presion_de_salida_10", "presion_de_entrada_11", "presion_de_salida_11", "presion_de_entrada_12", "presion_de_salida_12", "vehiculo__presion_establecida_1", "vehiculo__presion_establecida_2", "vehiculo__presion_establecida_3", "vehiculo__presion_establecida_4", "vehiculo__presion_establecida_5", "vehiculo__presion_establecida_6", "vehiculo__presion_establecida_7"))

        bitacoraas_pro = list(bitacoraa_pro.values("id", "vehiculo__id", "vehiculo__configuracion", "compania__id", "fecha_de_inflado", "tiempo_de_inflado", "presion_de_entrada_1", "presion_de_salida_1", "presion_de_entrada_2", "presion_de_salida_2", "presion_de_entrada_3", "presion_de_salida_3", "presion_de_entrada_4", "presion_de_salida_4", "presion_de_entrada_5", "presion_de_salida_5", "presion_de_entrada_6", "presion_de_salida_6", "presion_de_entrada_7", "presion_de_salida_7", "presion_de_entrada_8", "presion_de_salida_8", "presion_de_entrada_9", "presion_de_salida_9", "presion_de_entrada_10", "presion_de_salida_10", "presion_de_entrada_11", "presion_de_salida_11", "presion_de_entrada_12", "presion_de_salida_12", "vehiculo__presion_establecida_1", "vehiculo__presion_establecida_2", "vehiculo__presion_establecida_3", "vehiculo__presion_establecida_4", "vehiculo__presion_establecida_5", "vehiculo__presion_establecida_6", "vehiculo__presion_establecida_7"))
       
        bitacoraas.extend(bitacoraas_pro)

        bitacoras = sorted(bitacoraas, key=lambda x:x["fecha_de_inflado"], reverse=False)

        entradas_correctas = functions.entrada_correcta_ambas(bitacoras)

        hoy = date.today()
        user = User.objects.get(username=self.request.user)
        llantas = Llanta.objects.filter(vehiculo = vehiculo, inventario="Rodante")
        
        objetivo = vehiculo.compania.objetivo
        color1 = functions.entrada_correcta(bitacora, None)
        color2 = functions.salida_correcta(bitacora)
        if color1 == "good":
            signo1 = "icon-checkmark"
        else:
            signo1 = "icon-cross"
        if color2 == "good":
            signo2 = "icon-checkmark"
        else:
            signo2 = "icon-cross"
          
        #Generacion de ejes dinamico
        llantas_actuales = Llanta.objects.filter(vehiculo = vehiculo, inventario="Rodante")
        inspecciones_actuales = Inspeccion.objects.filter(llanta__in=llantas_actuales)

        #Obtencion de la data
        num_ejes = vehiculo.configuracion.split('.')
        ejes_no_ordenados = []
        ejes = []
        eje = 1
        color_profundidad = ""
        if self.kwargs['tipo'] == 'pulpo':
            for num in num_ejes:
                list_temp = []
                for llanta in llantas_actuales:
                    if llanta.tipo_de_eje != 'SP1':
                        if llanta.eje == eje:
                            list_temp.append(
                                [llanta, 
                                 bitacora.presion_de_entrada, 
                                 bitacora.presion_de_salida
                                 ])
                eje += 1
                ejes_no_ordenados.append(list_temp)
        elif self.kwargs['tipo'] == 'pulpopro':
            for num in num_ejes:
                list_temp = []
                for llanta in llantas_actuales:
                
                    if llanta.eje == eje:
                        list_temp.append(
                            [llanta])
                eje += 1
                ejes_no_ordenados.append(list_temp)
        
        for eje in ejes_no_ordenados:
            if len(eje) == 2:
                lista_temp = ['', '']
                for llanta_act in eje:
                    if 'LI' in llanta_act[0].posicion:
                        #print(llanta_act)
                        lista_temp[0] = llanta_act
                        
                    elif 'RI' in llanta_act[0].posicion:
                        lista_temp[1] = llanta_act
                ejes.append(lista_temp)
                print(' 0---0')
            
            elif len(eje) == 4:
                lista_temp = ['', '', '', '']
                for llanta_act in eje:
                    if 'LO' in llanta_act[0].posicion:
                        lista_temp[0] = llanta_act
                    elif 'LI' in llanta_act[0].posicion:
                        lista_temp[1] = llanta_act
                    elif 'RI' in llanta_act[0].posicion:
                        lista_temp[2] = llanta_act
                    elif 'RO' in llanta_act[0].posicion:
                        lista_temp[3] = llanta_act
                ejes.append(lista_temp)
                print('00---00')
            else:
                pass
        #if self.kwargs['tipo'] == 'pulpopro':
        #    for eje in ejes:
        #        print(eje)
        #print(ejes)
        if self.kwargs['tipo'] == 'pulpopro':
            presiones_entrada = [
                bitacora.presion_de_entrada_1,
                bitacora.presion_de_entrada_2,
                bitacora.presion_de_entrada_3,
                bitacora.presion_de_entrada_4,
                bitacora.presion_de_entrada_5,
                bitacora.presion_de_entrada_6,
                bitacora.presion_de_entrada_7,
                bitacora.presion_de_entrada_8,
                bitacora.presion_de_entrada_9,
                bitacora.presion_de_entrada_10,
                bitacora.presion_de_entrada_11,
                bitacora.presion_de_entrada_12,
            ]
            presiones_salida = [
                bitacora.presion_de_salida_1,
                bitacora.presion_de_salida_2,
                bitacora.presion_de_salida_3,
                bitacora.presion_de_salida_4,
                bitacora.presion_de_salida_5,
                bitacora.presion_de_salida_6,
                bitacora.presion_de_salida_7,
                bitacora.presion_de_salida_8,
                bitacora.presion_de_salida_9,
                bitacora.presion_de_salida_10,
                bitacora.presion_de_salida_11,
                bitacora.presion_de_salida_12,
            ]

            numero = 0
            for eje in ejes:
                for ej in eje:
                    ej.append(presiones_entrada[numero])
                    ej.append(presiones_salida[numero])
                    numero += 1
                    
        presiones_establecida=[
            vehiculo.presion_establecida_1,
            vehiculo.presion_establecida_2,
            vehiculo.presion_establecida_3,
            vehiculo.presion_establecida_4,
            vehiculo.presion_establecida_5,
            vehiculo.presion_establecida_6,
            vehiculo.presion_establecida_7,
        ]
        numero = 0
        #print(vehiculo.id)
        
        observacion_doble = Observacion.objects.get(observacion="Doble mala entrada")
        for eje in ejes:
            for ej in eje:
                
                presion_entrada_act = ej[1]
                presion_salida_act  = ej[2]
                #Cheking precion_ entrada
                precion_establecida = presiones_establecida[numero]
                presion_minima = precion_establecida - (precion_establecida * (objetivo/100))
                presion_maxima = precion_establecida + (precion_establecida * (objetivo/100))
                if presion_entrada_act >= presion_minima and presion_entrada_act <= presion_maxima:
                    color_entrada = 'good'
                    icono_entrada = "icon-checkmark"
                else:
                    color_entrada = 'bad'
                    icono_entrada = "icon-cross"

                ej.append(color_entrada) # 3
                ej.append(icono_entrada) # 4
                #Cheking precion_ salida
                if presion_salida_act >= presion_minima and presion_salida_act <= presion_maxima:
                    color_salida = 'good'
                    icono_salida = "icon-checkmark"
                else:
                    color_salida = 'bad'
                    icono_salida = "icon-cross"

                if entradas_correctas[bitacora.id] == "Doble":
                    doble_entrada = True
                else:
                    doble_entrada = False

                if doble_entrada and icono_entrada == "icon-cross":
                    icono_entrada2 = "icon-cross"
                else:
                    icono_entrada2 = None

                if doble_entrada and icono_salida == "icon-cross":
                    icono_salida2 = "icon-cross"
                else:
                    icono_salida2 = None

                print("icono_entrada", icono_entrada)
                print("icono_salida2", icono_salida2)

                ej.append(color_salida) # 5
                ej.append(icono_salida) # 6
                ej.append(icono_entrada2) # 7
                ej.append(icono_salida2) # 8
            numero += 1
            #print(precion_establecida)    
        if self.kwargs['tipo'] == 'pulpo':
            context["pulpo"] = 'Pulpo'
        elif self.kwargs['tipo'] == 'pulpopro':
            context["pulpo"] = 'Pulpo Pro'
        context["color1"] = color1
        context["color2"] = color2
        context["hoy"] = hoy
        context["user"] = user
        context["signo1"] = signo1
        context["signo2"] = signo2
        context['ejes'] = ejes
        context['bitacora'] = bitacora
        #Prueba
        #functions.send_mail(bitacora, tipo_bit)
        return context   
 
class reporteLlantaView(LoginRequiredMixin, TemplateView):
    # Vista de reporteLlantaView
          
    template_name = "reporteLlanta.html"
    context_object_name = "bitacora"
    def get(self, *args, **kwargs):
        if self.kwargs['pulpo'] == 'pulpo':
            bitacora = Bitacora.objects.get(pk=self.kwargs['pk'])
        elif self.kwargs['pulpo'] == 'pulpopro':
            bitacora = Bitacora_Pro.objects.get(pk=self.kwargs['pk'])
        
        if functions.check_bitacora_actual(bitacora):
            return redirect('dashboards:reporteVehiculo', bitacora.id, self.kwargs['pulpo'])
        
        self.bitacora = bitacora

        return super().get(*args, **kwargs)
        
    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)
        bitacora = self.bitacora
        #bitacora = Bitacora.objects.get(id=self.kwargs['pk'])
        llanta = Llanta.objects.get(id=self.kwargs['llanta'])
        vehiculo = Vehiculo.objects.get(pk=llanta.vehiculo.id)
        hoy = date.today()
        user = User.objects.get(username=self.request.user)
        presiones_establecida=[
            vehiculo.presion_establecida_1,
            vehiculo.presion_establecida_2,
            vehiculo.presion_establecida_3,
            vehiculo.presion_establecida_4,
            vehiculo.presion_establecida_5,
            vehiculo.presion_establecida_6,
            vehiculo.presion_establecida_7,
        ]
        if self.kwargs['pulpo'] == 'pulpo':
            entrada = bitacora.presion_de_entrada
            salida = bitacora.presion_de_salida
            objetivo = (vehiculo.compania.objetivo / 100)
            
            
            color_entrada = functions.check_color_presion_bitacora_pulpo(llanta, bitacora, 'entrada')
            signo_entrada = functions.check_icon_pulpo(color_entrada)
            
            color_salida = functions.check_color_presion_bitacora_pulpo(llanta, bitacora, 'salida')
            signo_salida = functions.check_icon_pulpo(color_salida)
            
                     
        elif self.kwargs['pulpo'] == 'pulpopro':
            datos_entrada = functions.datos_entrada_o_salida_pulpopro(bitacora, llanta, 'entrada')
            datos_salida = functions.datos_entrada_o_salida_pulpopro(bitacora, llanta, 'salida')
            print(datos_entrada)
            print(datos_salida)
            entrada = datos_entrada['presion']
            salida = datos_salida['presion']
            objetivo = (vehiculo.compania.objetivo / 100)
            
            
            color_entrada = functions.check_color_presion_bitacora_pulpopro(datos_entrada['presion'], datos_entrada['presion_establecida'], objetivo)
            signo_entrada = functions.check_icon_pulpo(color_entrada)
            
            color_salida = functions.check_color_presion_bitacora_pulpopro(datos_salida['presion'], datos_salida['presion_establecida'], objetivo)
            signo_salida = functions.check_icon_pulpo(color_salida)
            

        context["entrada"] = entrada
        context["salida"] = salida
        context["color_entrada"] = color_entrada
        context["color_salida"] = color_salida
        context["signo_entrada"] = signo_entrada
        context["signo_salida"] = signo_salida
        context["hoy"] = hoy
        context["llanta"] = llanta
        context["user"] = user
        return context

class configuracionVehiculoView(LoginRequiredMixin, TemplateView):
    # Vista de configuracionVehiculoView

    template_name = "configuracionVehiculo.html" 
    
    def get_context_data(self, **kwargs):
        
        context = super().get_context_data(**kwargs)
        #Declaracion de variables
        
        #Generacion de ejes dinamico
        vehiculo_actual = Vehiculo.objects.get(pk = self.kwargs['pk'])
        llantas_actuales = Llanta.objects.filter(vehiculo = self.kwargs['pk'])
        inspecciones_actuales = Inspeccion.objects.filter(llanta__in=llantas_actuales)
        
        #Obtencion de la lista de las llantas
        
        filtro_sospechoso = functions.vehiculo_sospechoso_llanta(inspecciones_actuales)
        llantas_sospechosas = llantas_actuales.filter(numero_economico__in=filtro_sospechoso)

        filtro_rojo = functions.vehiculo_rojo_llanta(llantas_actuales)
        llantas_rojas = llantas_actuales.filter(numero_economico__in=filtro_rojo).exclude(id__in=llantas_sospechosas)
        
        filtro_amarillo = functions.vehiculo_amarillo_llanta(llantas_actuales)
        llantas_amarillas = llantas_actuales.filter(numero_economico__in=filtro_amarillo).exclude(id__in=llantas_sospechosas).exclude(id__in=llantas_rojas)
        
        llantas_azules = llantas_actuales.exclude(id__in=llantas_sospechosas).exclude(id__in=llantas_rojas).exclude(id__in=llantas_amarillas)
        
        #Obtencion de la data
        num_ejes = vehiculo_actual.configuracion.split('.')
        ejes_no_ordenados = []
        ejes = []
        eje = 1
        color_profundidad = ""
        presiones_establecida=[
            vehiculo_actual.presion_establecida_1,
            vehiculo_actual.presion_establecida_2,
            vehiculo_actual.presion_establecida_3,
            vehiculo_actual.presion_establecida_4,
            vehiculo_actual.presion_establecida_5,
            vehiculo_actual.presion_establecida_6,
            vehiculo_actual.presion_establecida_7,
        ]
        numero = 0
        for num in num_ejes:
            list_temp = []
            for llanta in llantas_actuales:
                if llanta in llantas_sospechosas:
                    color_profundidad = 'purple'
                elif llanta in llantas_rojas:
                    color_profundidad = 'bad'
                elif llanta in llantas_amarillas:
                    color_profundidad = 'yellow'
                elif llanta in llantas_azules:
                    color_profundidad = 'good'
                
                objetivo = llanta.vehiculo.compania.objetivo / 100
                presion_act = llanta.presion_actual
                presion_minima = presiones_establecida[numero] - (100 * objetivo)
                presion_maxima = presiones_establecida[numero] + (100 * objetivo)
                #print(f'{objetivo}'.center(50, "-"))
                #print(f'{presion_minima}'.center(50, "-"))
                #print(f'{presion_maxima}'.center(50, "-"))
                #print(f'{presion_act}'.center(50, "-"))
                #print(presion_act > presion_minima)
                #print(presion_act < presion_maxima)
                #print('***********************************')
                
                if presion_act >= presion_minima and presion_act <= presion_maxima:
                    color_presion = 'good'
                elif presion_act < presion_minima:
                    color_presion = 'bad'
                elif presion_act > presion_maxima:
                    color_presion = 'yellow'
                
                if llanta.eje == eje:
                    list_temp.append([llanta, color_profundidad, color_presion])
            eje += 1
            ejes_no_ordenados.append(list_temp)
            numero += 1
        for eje in ejes_no_ordenados:
            if len(eje) == 2:
                lista_temp = ['', '']
                for llanta_act in eje:
                    if 'LI' in llanta_act[0].posicion:
                        lista_temp[0] = llanta_act
                        
                    elif 'RI' in llanta_act[0].posicion:
                        lista_temp[1] = llanta_act
                ejes.append(lista_temp)
                print(' 0---0')
            
            else:
                lista_temp = ['', '', '', '']
                for llanta_act in eje:
                    if 'LO' in llanta_act[0].posicion:
                        lista_temp[0] = llanta_act
                    elif 'LI' in llanta_act[0].posicion:
                        lista_temp[1] = llanta_act
                    elif 'RI' in llanta_act[0].posicion:
                        lista_temp[2] = llanta_act
                    elif 'RO' in llanta_act[0].posicion:
                        lista_temp[3] = llanta_act
                ejes.append(lista_temp)
                print('00---00')
            
            
        color = functions.entrada_correcta(vehiculo_actual, None)
        #print(color)
        if color == 'good':
            style = 'good'
        elif color == 'bad':
            style = 'bad'
        else:
            style = 'bad'
        
        cant_ejes = len(ejes)
        
        
        #print(vehiculo.configuracion)
        #print(ejes)
        #print(f'style: {style}')
        #print(f'llantas_sospechosas: {llantas_sospechosas}')
        #print(f'llantas_rojas: {llantas_rojas}')
        #print(f'llantas_amarillas: {llantas_amarillas}')
        #print(f'llantas_azules: {llantas_azules}')
        context['ejes'] = ejes
        context['style'] = style
        context['cant_ejes'] = cant_ejes
        
        
        return context

class configuracionLlantaView(LoginRequiredMixin, TemplateView):
    # Vista de configuracionLlantaView

    template_name = "configuracionLlanta.html" 

class carritoStockView(LoginRequiredMixin, TemplateView):
    # Vista de carritoStockView

    template_name = "carrito-stock.html" 
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #Obtencion datos del usuario
        user = self.request.user
        perfil = Perfil.objects.get(user = user)
        #Obtencion del inventario
        inventario = self.request.GET.get('inventario').replace('_', ' ')
        print(inventario)
        #Lista de colores
        azul = []
        amarillo = []
        rojo = []
        morado = []
        #Obtencion del carrito y las llantas
        carrito = LlantasSeleccionadas.objects.get(perfil=perfil, inventario=inventario)
        llantas = carrito.llantas.all()
        llantas_dict = []
        for llanta in llantas:
            status = functions.color_observaciones(llanta.observaciones.all())
            if status == 'good':
                azul.append(1)
            elif status == 'yellow':
                amarillo.append(1)
            elif status == 'bad':
                rojo.append(1)
            elif status == 'purple':
                morado.append(1)
            producto = llanta.producto.producto if llanta.producto != None else None
            llantas_dict.append( {
                'id': llanta.id,
                'eco': llanta.numero_economico,
                'producto': producto,
                'fecha': f'{llanta.fecha_de_entrada_inventario.day}/{llanta.fecha_de_entrada_inventario.month}/{llanta.fecha_de_entrada_inventario.year}',
                'status': status
            })
        opciones = functions.opciones_redireccion(inventario)
        #Paso del contexto
        context['inventario'] = inventario
        context['llantas_dict'] = llantas_dict
        context['opciones'] = opciones
        context['azul'] = len(azul)
        context['rojo'] = len(rojo)
        context['amarillo'] = len(amarillo)
        context['morado'] = len(morado)
        context['total'] = len(azul) + len(rojo) + len(amarillo) + len(morado)
        return context
    
    
    def post(self, request, *args, **kwargs):
        print(request.POST)
        #Obtencion datos del usuario
        user = self.request.user
        perfil = Perfil.objects.get(user = user)
        #Obtencion del inventario
        inventario = request.POST.getlist('inventario')[0].replace('_', ' ')
        
        if 'vaciar' in request.POST:
    
            functions.eliminarCarrito(inventario, perfil)
            url = f"%s?inventario={inventario}" % reverse('dashboards:carritoStock')
            return redirect(url)
        
        #Se verifica si se quiere eliminar un elento del carrito
        if 'eliminar' in request.POST:
            id = int(request.POST.getlist('eliminar')[0])
            inventario = request.POST.getlist('inventario')[0].replace('_', ' ')
            functions.eliminarElementoCarrito(id, inventario, perfil)
            url = f"%s?inventario={inventario}" % reverse('dashboards:carritoStock')
            return redirect(url)
            
            
       
        #En caso que no se elimina nade continua
        print('-------------')
        print(inventario)
        destino = request.POST.getlist('destino')[0]
        carrito = LlantasSeleccionadas.objects.get(perfil=perfil, inventario=inventario).llantas.all()
        ids = []
        for llanta in carrito:
            ids.append(llanta.id)

        ids = str(ids).replace(" ", "")
        if destino == 'Taller Destino':
            url = f"%s?ids={ids}&inventario={inventario}" % reverse('dashboards:tallerDestino')
            
        elif destino == 'Panel de Renovado':
            url = f"%s?ids={ids}" % reverse('dashboards:procesoRenovado')
            
        elif destino == 'Enviar con renovador':
            url = f"%s?ids={ids}" % reverse('dashboards:ordenSalidaRen')
        
        elif destino == 'Stock Destino':
            url = f"%s?ids={ids}&inventario={inventario}" % reverse('dashboards:stockDestino')
        
        elif destino == 'Panel de Desecho':
            url = f"%s?ids={ids}&inventario={inventario}" % reverse('dashboards:procesoDesecho')
            
            
        return redirect(url)
class almacenView(LoginRequiredMixin, TemplateView, View):
    # Vista de almacenView

    template_name = "almacen.html"

    """def form_valid(self, request):
        print("hola")
        get_llanta = self.request.GET.get("llanta")
        if get_llanta:
            llanta = Llanta.objects.get(numero_economico=get_llanta).id
            print(get_llanta)
            print(llanta)
            return reverse('dashboards:detail', kwargs={"pk": llanta})"""

    def post(self, request):
        print("hola")
        get_llanta = self.request.POST.get("llanta")
        print(get_llanta)
        if get_llanta:
            try:
                llanta = Llanta.objects.get(numero_economico=get_llanta).id
                print(llanta)
                return HttpResponseRedirect(reverse('dashboards:tireDetail', kwargs={"pk": llanta}))
            except:
                return HttpResponseRedirect(reverse_lazy('dashboards:almacen'))


    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)
        llantas = Llanta.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))

        llantas_nuevas = llantas.filter(inventario="Nueva")
        llantas_antes_de_renovar = llantas.filter(inventario="Antes de Renovar")
        llantas_antes_de_desechar = llantas.filter(inventario="Antes de Desechar")
        llantas_renovadas = llantas.filter(inventario="Renovada")
        llantas_con_renovador = llantas.filter(inventario="Con renovador")
        llantas_desecho_final = llantas.filter(inventario="Desecho final")
        llantas_servicio = llantas.filter(inventario="Servicio")
        llantas_rodante = llantas.filter(inventario="Rodante")

        context["llantas_nuevas"] = llantas_nuevas.count()
        context["llantas_antes_de_renovar"] = llantas_antes_de_renovar.count()
        context["llantas_antes_de_desechar"] = llantas_antes_de_desechar.count()
        context["llantas_renovadas"] = llantas_renovadas.count()
        context["llantas_con_renovador"] = llantas_con_renovador.count()
        context["llantas_desecho_final"] = llantas_desecho_final.count()
        context["llantas_servicio"] = llantas_servicio.count()
        context["llantas_rodantes"] = llantas_rodante.count()
        return context

class historialTransacciones(LoginRequiredMixin, TemplateView):
    # Vista de nuevo home

    template_name = "historial-transacciones.html"

class antesDesecharView(LoginRequiredMixin, TemplateView):
    # Vista de antesDesecharView

    template_name = "antesDesechar.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        usuario = self.request.user
        perfil = Perfil.objects.get(user = usuario)
        compania = perfil.compania
        renovadores = Renovador.objects.filter(compania=compania)
        context['renovadores'] = renovadores
        return context
    
class antesRenovarView(LoginRequiredMixin, TemplateView):
    # Vista de antesRenovarView

    template_name = "antesRenovar.html"

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)
        
        return context
    
    def post(self, request, *args, **kwargs):
        print(request.POST)
        ids = request.POST.getlist('id')
        form = request.POST.getlist('form')[0]
        print(ids)
        ids = str(ids).replace("'", "")
        ids = str(ids).replace(" ", "")
        if form == 'antes-renovar':
            url = f"%s?ids={ids}" % reverse('dashboards:ordenSalidaRen')
        if form == 'tallerDestino':
            url = f"%s?ids={ids}" % reverse('dashboards:tallerDestino')
        return redirect(url)
    
class conRenovadorView(LoginRequiredMixin, TemplateView):
    # Vista de conRenovadorView

    template_name = "conRenovador.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        usuario = self.request.user
        perfil = Perfil.objects.get(user = usuario)
        compania = perfil.compania
        renovadores = Renovador.objects.filter(compania=compania)
        context['renovadores'] = renovadores
        return context
class desechoFinalView(LoginRequiredMixin, TemplateView):
    # Vista de desechoFinalView

    template_name = "desechoFinal.html"

class vistaordenesView(LoginRequiredMixin, TemplateView):
    # Vista de vistaordenesView

    template_name = "vistaordenes.html"

class nuevaView(LoginRequiredMixin, TemplateView):
    # Vista de nuevaView

    template_name = "nueva.html"

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)
        llantas_nuevas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=self.request.user.perfil.compania), vida="Nueva")

        context["llantas_nuevas"] = llantas_nuevas

        return context

class formordennuevaView(LoginRequiredMixin, TemplateView):
    # Vista de formordennuevaView

    template_name = "formordennueva.html"
    
    def get_context_data(self, **kwargs):
        
        context = super().get_context_data(**kwargs)

        usuario = self.request.user
        perfil = Perfil.objects.get(user = usuario)
        compania = perfil.compania
        ordenes = Orden.objects.filter(compania = compania, status = 'PreOrden')

        context["ordenes"] = ordenes

        return context
class renovadaView(LoginRequiredMixin, TemplateView):
    # Vista de renovadaView

    template_name = "renovada.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        usuario = self.request.user
        perfil = Perfil.objects.get(user = usuario)
        compania = perfil.compania
        renovadores = Renovador.objects.filter(compania=compania)
        context['renovadores'] = renovadores
        return context
    
class servicioView(LoginRequiredMixin, TemplateView):
    # Vista de servicioView

    template_name = "servicio.html"

class rodanteView(LoginRequiredMixin, TemplateView):
    # Vista de rodanteView

    template_name = "rodante.html"

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)
        llantas_nuevas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=self.request.user.perfil.compania), vida="Nueva")

        context["llantas_nuevas"] = llantas_nuevas

        return context

class procesoDesechoView(LoginRequiredMixin, TemplateView):
    # Vista de procesoDesechoView

    template_name = "procesoDesecho.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        usuario = self.request.user
        perfil = Perfil.objects.get(user = usuario)
        context['perfil'] = perfil.id
        return context
    
    def post(self, request, *args, **kwargs):
        id = request.POST.getlist('id')[0]
        
        ids = request.GET.get('ids')
        ids = functions.int_list_element(ids.replace('[', '').replace(']', '').split(','))
        
        ids.pop( ids.index( int(id) ) )
        
        usuario = self.request.user
        perfil = Perfil.objects.get(user = usuario)
        compania = perfil.compania
        print(request.POST)
        print(request.FILES)
        condicion = request.POST.getlist('condicion')[0]
        zona = request.POST.getlist('zona')[0]
        razon = request.POST.getlist('razon')[0]
        profundidad = request.POST.getlist('profundidad')[0]
        
        try:
            image = request.FILES.getlist('image')[0]
            print()
            ruta_archivo = TemporaryUploadedFile.temporary_file_path(image)
            id_folder = '1SwbkExxc9pF25QFQUPJXr_pu9heE2LcG'
            nombre = ''
            print()
            a = Google.subir_archivo(ruta_archivo,id_folder, nombre)
            print(a)
        except:
            image = ''
            a = ''
            
        print(condicion)
        print(zona)
        print(razon)
        print(profundidad)
        print(id)
        print(image)
        hoy = date.today()
        llanta = Llanta.objects.get(pk = id)
        desecho = Desecho.objects.get(condicion = condicion, zona_de_llanta = zona, razon = razon)
        orden = OrdenDesecho.objects.create(
            usuario = perfil,
            llanta = llanta,
            compania = compania,
            fecha = hoy,
            min_profundidad = profundidad,
            desecho = desecho,
            imagen = a,
        )
        folio = str('OD' + str(perfil) + str(hoy.year)+ str(hoy.month)+ str(hoy.day) + str(orden.id))
        orden.folio = folio
        orden.perfil = perfil
        orden.save()
        min_prof = functions.min_profundidad(llanta)
        if llanta.profundidad_derecha == min_prof:
            llanta.profundidad_derecha = float(profundidad)
        if llanta.profundidad_central == min_prof:
            llanta.profundidad_central = float(profundidad)
        if llanta.profundidad_izquierda == min_prof:
            llanta.profundidad_izquierda = float(profundidad)
        llanta.inventario = 'Desecho final'
        llanta.save()
        #Borrar llanta del carrito
        LlantasSeleccionadas.objects.get(perfil=perfil, inventario = 'Antes de Desechar').llantas.remove(llanta)
        if len(ids) == 0:
            return redirect('dashboards:almacen')
        else:
            current_id = str(ids).replace(' ', '')
            url = f"%s?ids={current_id}&inventario=Antes de Desechar" % reverse("dashboards:procesoDesecho")
            return redirect(url)
class ordenSalidaRenView(LoginRequiredMixin, TemplateView):
    # Vista de ordenSalidaRenView
    template_name = "ordenSalidaRen.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ids = self.request.GET.get('ids')
        temp = ids
        #Formateo de la data
        ids = str(ids).replace("[", "")
        ids = str(ids).replace("]", "")
        ids = str(ids).split(',')
        ids = functions.int_list_element(ids)
        llantas = Llanta.objects.filter(pk__in = ids)
        
        fecha = date.today()
        usuario = self.request.user
        perfil = Perfil.objects.get(user = usuario)
        renovadores = Renovador.objects.filter(compania = perfil.compania)
        
        print(llantas)
        print(renovadores)
        print(ids)
        context['llantas'] = llantas
        context['fecha'] = fecha
        context['perfil'] = perfil
        context['renovadores'] = renovadores
        context['temp'] = temp
        context['ids'] = ids
        
        return context
    
    def post(self, request, *args, **kwargs):
        print(request.POST)
        usuario = request.user
        perfil = Perfil.objects.get(user = usuario)
        compania = perfil.compania
        
        renovador = request.POST.getlist('destino')
        ids = request.POST.getlist('ids')
        ids = str(ids[0]).replace("[", "")
        ids = str(ids).replace("]", "")
        ids = str(ids).split(',')
        ids = functions.int_list_element(ids)
        llantas = Llanta.objects.filter(pk__in = ids)
        productos = list(llantas.values('producto__producto').distinct())
        productos = functions.get_product_list(productos)
        hoy = date.today()
        renovador = Renovador.objects.get(pk = int(renovador[0]))
        print('--------------------------')
        print(renovador)
        print(productos)
        print('--------------------------')
        
        data = []
        for producto in productos:
            list_temp = []
            for llanta in llantas:
                if llanta.producto.producto == producto:
                    list_temp.append(llanta.id)
            data.append(
                {
                    'producto': producto,
                    'ids': list_temp
                }
            )

        orden = Orden.objects.create(
            status = 'AntesRenovar',
            renovador = renovador,
            compania = compania,
            fecha = date.today()
        )

        folio = str('OR' + str(perfil) + str(hoy.year)+ str(hoy.month)+ str(hoy.day) + str(orden.id))
        datos = str({
            "folio": folio, 
            "usuario": str(perfil),
            "fecha": str(datetime.today()), 
            "renovador": str(renovador), 
            "productos": data
            })
        orden.folio = folio
        orden.datos = datos
        orden.save()
        for llanta in llantas:
            llanta.inventario = 'Con renovador'
            llanta.fecha_de_entrada_inventario = date.today()
            llanta.renovador = renovador
            llanta.save()
        print(datos)
        print(ids)
        print(llantas)
        
        #Vaciado del carrito
        carrito = LlantasSeleccionadas.objects.get(perfil=perfil, inventario='Antes de Renovar')
        carrito.llantas.clear()
        return redirect('dashboards:conRenovador')
    
def ordenSalidaRenDef(request):   
    return redirect('dashboards:conRenovador')

class ordenEntradaView(LoginRequiredMixin, TemplateView):
    # Vista de ordenEntradaView

    template_name = "ordenEntrada.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        id = self.request.GET.get('id', None)
        if id != None:
            orden = Orden.objects.get(pk=int(id))
            print('Muestra')
            print(orden.datos)
            evento_raw = orden.datos
            evento_raw = evento_raw.replace("\'", "\"")
            datos = json.loads(evento_raw)
            context['folio'] = orden.folio
            context['renovador'] = orden.renovador
            context['hoy'] = orden.fecha
            context['status'] = datos['status']
            context['taller'] = datos['taller']
            context['total'] = datos['total']
            try:
                context['producto'] = datos['producto']
            except:
                context['razon'] = datos['razon']
            context['llantas_list'] = datos['llantas']
            context['solovista'] = True
            return context
        status = self.request.GET.get('status')
        producto = self.request.GET.get('producto', None)
        razon = self.request.GET.get('razon', None)
        taller = self.request.GET.get('taller')
        ids = self.request.GET.get('ids')
        ids = ids.split(',')
        ids = functions.int_list_element(ids)
        total = len(ids)
        hoy = date.today()
        print(status)
        print(producto)
        print(taller)
        print(ids)
        print(total)
        
        llantas = Llanta.objects.filter(id__in = ids)
        llantas_list = list(llantas.annotate(
            dias_fuera = DiffDays(CastDate(Now())-CastDate(F('fecha_de_entrada_inventario')), output_field=IntegerField()) + 1
        ).values('numero_economico', 'producto__producto', 'vida', 'dias_fuera', 'taller__nombre'))
        print(llantas_list)
        renovador = llantas[0].renovador
        
        context["renovador"] = renovador
        context["ids"] = ids
        context["hoy"] = hoy
        context["llantas_list"] = llantas_list
        if producto != None:
            context["producto"] = producto
        else:
            context['razon'] = razon
        context["status"] = status
        context["taller"] = taller
        context["total"] = total
        context["user"] = self.request.user
        return context
    def post(self, request, *args, **kwargs):
        print(request.POST)
        usuario = request.user
        perfil = Perfil.objects.get(user = usuario)
        compania = perfil.compania
        
        ids = request.POST.getlist('ids')[0]
        ids = functions.int_list_element(ids.replace('[', '').replace(']', '').split(','))
        try:
            producto = request.POST.getlist('producto')[0]
        except:
            razon = request.POST.getlist('razon')[0]
        taller = request.POST.getlist('taller')[0]
        status = request.POST.getlist('status')[0]
        #Llantas
        llantas = Llanta.objects.filter(id__in = ids)
        llantas_list = llantas.annotate(
            dias_fuera = DiffDays(CastDate(Now())-CastDate(F('fecha_de_entrada_inventario')), output_field=IntegerField()) + 1
        ).values('numero_economico', 'producto__producto', 'vida', 'dias_fuera', 'taller__nombre')
        hoy = date.today()
        print(llantas_list)
        renovador = llantas[0].renovador
        #! Guarda la orden
        data = []
        
        for llanta in llantas_list:
            data.append(
                {
                    'numero_economico': llanta['numero_economico'],
                    'producto__producto': llanta['producto__producto'],
                    'vida': llanta['vida'],
                    'dias_fuera': llanta['dias_fuera'],
                    'taller__nombre': llanta['taller__nombre'],
                }
            )

        orden = Orden.objects.create(
            status = 'ConRenovador',
            renovador = renovador,
            compania = compania,
            fecha = date.today()
        )

        folio = str('CR' + str(perfil) + str(hoy.year)+ str(hoy.month)+ str(hoy.day) + str(orden.id))
        if status == 'aprobado':
            datos = str({
                "folio": folio, 
                "usuario": str(perfil),
                "fecha": str(datetime.today()),
                "total": str(len(ids)),
                "status": status,
                "taller": taller,
                "producto": producto,
                "llantas": data
                })
        else:
            datos = str({
                "folio": folio, 
                "usuario": str(perfil),
                "fecha": str(datetime.today()),
                "total": str(len(ids)),
                "status": status,
                "taller": taller,
                "razon": razon,
                "llantas": data
                })
        orden.folio = folio
        orden.datos = datos
        orden.save()
        #!Cambia los datos de la llanta (Vida Incluida)
        taller = Taller.objects.get(nombre=taller)
        if status == 'aprobado':
            producto = Producto.objects.get(producto=producto)
        for llanta in llantas:
            llanta.taller = taller
            if status == 'aprobado':
                functions.cambio_de_vida(llanta, llanta)
                llanta.vida = functions.actualizar_vida(llanta)
                llanta.producto = producto
                llanta.profundidad_izquierda = producto.profundidad_inicial
                llanta.profundidad_central = producto.profundidad_inicial
                llanta.profundidad_derecha = producto.profundidad_inicial
                llanta.km_actual = 0
                llanta.primera_inspeccion = None
                functions.quitar_todo_profundidad(llanta)
            if status == 'aprobado':
                llanta.inventario = 'Renovada'
            else:
                llanta.inventario = 'Antes de Desechar'
            
        Llanta.objects.bulk_update(llantas, 
                                   ['taller', 
                                    'producto', 
                                    'vida', 
                                    'inventario', 
                                    'profundidad_izquierda', 
                                    'profundidad_central', 
                                    'profundidad_derecha',
                                    'km_actual',
                                    'primera_inspeccion'
                                    ])
        #! Elimina del carrito solo los ids actuales
        carrito = LlantasSeleccionadas.objects.get(perfil = perfil, inventario = 'Con renovador')
        for id in ids:
            carrito.llantas.remove(id)
                
        #!Si hay ids regresar al panel de renovado
        #!Si no hay ids, regresar al panel de almacenes
        current_id = []
        for llanta in carrito.llantas.all():
            print(llanta.id)
            current_id.append(llanta.id)
        current_id = str(current_id).replace(' ', '').replace("\'", "")
        if carrito.llantas.all().count() > 0:
            url = f"%s?ids={current_id}" % reverse("dashboards:procesoRenovado")
            return redirect(url)
        else:
            return redirect("dashboards:almacen")
        

class ordenEntradaTallerView(LoginRequiredMixin, TemplateView):
    # Vista de ordenEntradaTallerView

    template_name = "orden-entrada-taller.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        orden = Orden.objects.get(pk=self.kwargs['pk'])
        print(orden)
        orden_raw = orden.datos
        orden_raw = orden_raw.replace("\'", "\"")
        datos = json.loads(orden_raw)
        print('----')
        print(datos)
        fecha = datos['fecha']
        usuario = datos['usuario']
        taller_destino = datos['taller_destino']
        total_llantas = datos['total_llantas']
        data = datos['data']
        
        context['folio'] = orden.folio
        context['fecha'] = fecha
        context['usuario'] = usuario
        context['taller_destino'] = taller_destino
        context['total_llantas'] = total_llantas
        context['data'] = data
        return context
    
class ordenEntradaStockView(LoginRequiredMixin, TemplateView):
    # Vista de ordenEntradaTallerView

    template_name = "orden-entrada-stock.html"
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        orden = Orden.objects.get(pk=self.kwargs['pk'])
        print(orden)
        orden_raw = orden.datos
        orden_raw = orden_raw.replace("\'", "\"")
        datos = json.loads(orden_raw)
        print('----')
        print(datos)
        fecha = datos['fecha']
        usuario = datos['usuario']
        stock_destino = datos['inventario_destino']
        total_llantas = datos['total_llantas']
        data = datos['data']
        
        context['folio'] = orden.folio
        context['fecha'] = fecha
        context['usuario'] = usuario
        context['stock_destino'] = stock_destino
        context['total_llantas'] = total_llantas
        context['data'] = data
        return context

class reporteDesechoView(LoginRequiredMixin, TemplateView):
    # Vista de reporteDesechoView

    template_name = "reporte-desecho.html"
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #Obtener id
        fecha = self.request.GET.get('fecha')
        ids = str(self.request.GET.get('ids')).replace('[', '').replace(']', '').split(',')
        ordenes = OrdenDesecho.objects.filter(id__in = ids)
        print(ordenes)
        context['fecha'] = fecha
        context['total'] = ordenes.count()
        context['ordenes'] = ordenes
        
        return context

class imagenDesechoView(TemplateView):
    # Vista de reporteDesechoView
    template_name = "imagenes/imagen-desecho.html"
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        desecho = OrdenDesecho.objects.get(id = self.kwargs['id'])
        context['desecho'] = desecho
        return context
    
    
class historialDesechoView(LoginRequiredMixin, TemplateView):
    # Vista de historialDesechoView

    template_name = "historial-desecho.html"
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        usuario = self.request.user
        perfil = Perfil.objects.get(user = usuario)
        compania = perfil.compania
        ordenes = OrdenDesecho.objects.select_related().filter(compania = compania).order_by('id').values('id', 'fecha')
        fechas = OrdenDesecho.objects.select_related().filter(compania = compania).values('fecha').distinct('fecha')
        ordenes_list =  []
        
        for fecha in fechas:
            temp = []
            for orden in ordenes:
                if fecha['fecha'] == orden['fecha']:
                    temp.append(orden['id'])
            ordenes_list.append({
                'fecha': fecha['fecha'],
                'ids': temp
            })
        
        for orden in ordenes_list:
            print(orden)
            print(orden['fecha'])
            print(orden['ids'])
        context['ordenes_list'] = ordenes_list
        return context
class ordenLlantaView(LoginRequiredMixin, TemplateView):
    # Vista de ordenLlantaView

    template_name = "ordenLlanta.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        usuario = self.request.user
        perfil = Perfil.objects.get(user = usuario)
        
        orden = Orden.objects.get(pk=self.kwargs['pk'])
        fecha = orden.fecha
        datos = json.loads(orden.datos)
        data_list = datos['productos'][0]
        folio = orden.folio
        taller = orden.taller
        status = orden.status
        
        productos = data_list['producto']
        cantidad = data_list['cantidad']
        num_ecos = data_list['num_economicos']
    
        print(productos)
        print(cantidad)
        print(num_ecos)
        cont = 0
        main_cont = 0
        templeta_data = []
        for num in cantidad:
            produc = productos[main_cont]
            nums_ecos = []
            for i in range(0, int(num)):
                #print(i)
                #print(produc)
                #print(num_ecos[cont])
                nums_ecos.append(num_ecos[cont])
                cont += 1
            templeta_data.append(
            {
                'producto': produc,
                'nums_ecos': nums_ecos
            }
            )
            main_cont += 1
        
        context["perfil"] = perfil
        context["fecha"] = fecha
        context["folio"] = folio
        context["id"] = orden.id
        context["taller"] = taller
        context["total"] = cont
        context["status"] = status
        context["templeta_data"] = templeta_data
        

        return context


def ordenLlantaDef(request, id):
    usuario = request.user
    perfil = Perfil.objects.get(user = usuario)
    compania = perfil.compania
    
    orden = Orden.objects.get(pk=id)
    fecha = orden.fecha
    datos = json.loads(orden.datos)
    data_list = datos['productos'][0]
    folio = orden.folio
    
    productos = data_list['producto']
    cantidad = data_list['cantidad']
    num_ecos = data_list['num_economicos']

    print(productos)
    print(cantidad)
    print(num_ecos)
    cont = 0
    main_cont = 0
    templeta_data = []
    for num in cantidad:
        produc = productos[main_cont]
        nums_ecos = []
        for i in range(0, int(num)):
            #print(i)
            #print(produc)
            #print(num_ecos[cont])
            nums_ecos.append(num_ecos[cont])
            cont += 1
        templeta_data.append(
        {
            'producto': produc,
            'nums_ecos': nums_ecos
        }
        )
        main_cont += 1
    for data in templeta_data:
        producto_act = Producto.objects.get(producto = data['producto'])
        for num in data['nums_ecos']:
            print(f'{producto_act}: {num}')
            Llanta.objects.create(
                numero_economico = num,
                compania = compania,
                taller = orden.taller,
                vida = 'Nueva',
                producto = producto_act,
                fecha_de_entrada_inventario = date.today(),
                profundidad_izquierda = producto_act.profundidad_inicial,
                profundidad_central = producto_act.profundidad_inicial,
                profundidad_derecha = producto_act.profundidad_inicial,
                inventario = 'Nueva'
            )
    orden.status = 'Orden'
    orden.save()
    return redirect('dashboards:nueva')

def ordenLlantaDeleteDef(request, id):
    orden = get_object_or_404(Orden, pk=id)
    orden.delete()
    return redirect('dashboards:vistaordenes')

class tallerDestinoView(LoginRequiredMixin, TemplateView):
# Vista de tallerDestinoView

    template_name = "tallerDestino.html"
    
    def get_context_data(self, **kwargs):
        
        context = super().get_context_data(**kwargs)
        #Ontencion de la data
        fecha = date.today()
        usuario = self.request.user
        perfil = Perfil.objects.get(user = usuario)
        talleres_context = perfil.taller.all()
        talleres_select = Taller.objects.filter(compania = perfil.compania).exclude(id__in = talleres_context)
        inventario = self.request.GET.get('inventario', [])
        
        #Formateo de los id
        ids = self.request.GET.get('ids', [])
        ids = str(ids).replace("[", "")
        ids = str(ids).replace("]", "")
        ids_list = ids
        ids = str(ids).split(',')
        ids = functions.int_list_element(ids)
        
        #Obtencion de las llantas
        llantas = Llanta.objects.filter(pk__in = ids)
        #Obtencion de los talleres
        talleres = llantas.values('taller__nombre').distinct()
        talleres_list = []
        for taller in talleres:
            talleres_list.append(taller['taller__nombre'])
        
        #Generacion de la data del template
        llantas_list = []
        for taller in talleres_list:
            temp = []
            for llanta in llantas:
                if llanta.taller.nombre == taller:
                    temp.append(llanta.numero_economico)
            llantas_list.append(
                {
                    'taller': taller,
                    'ecos': temp
                }
            )
        
        #Envio del contexto
        context['fecha'] = fecha
        context['inventario'] = inventario
        context['ids_list'] = ids_list
        context['talleres_select'] = talleres_select
        context['perfil'] = perfil
        context['llantas_list'] = llantas_list
        return context
    
    def post(self, request, *args, **kwargs):
        print(request.POST)
        #Obtencion de la data
        hoy = date.today()
        user = request.user
        perfil = Perfil.objects.get(user = user)
        inventario = request.POST.getlist('inventario')[0]
        taller = request.POST.getlist('taller')[0]
        taller_destino = Taller.objects.get(nombre=taller)
        ids = request.POST.getlist('ids')[0]
        ids = ids.split(',')
        ids = functions.int_list_element(ids)
        
        #Obtencion de las llantas
        llantas = Llanta.objects.filter(id__in = ids)
        
        #Generacion de la lista de los datos de las llantas
        llanta_data = []
        for llanta in llantas:
            llanta_data.append(
                {
                    'num_eco': llanta.numero_economico,
                    'producto': llanta.producto.producto if llanta.producto != None else None ,
                    'taller_origen': llanta.taller.nombre
                }
            )
        
        #Generacion del campo datos
        data = str({
            'fecha': str(date.today()),
            'usuario': str(user),
            'taller_destino': taller,
            'total_llantas': len(ids),
            'data': llanta_data
        })
        
        #Creacion de la orden
        orden = Orden.objects.create(
            status = 'CambioTaller',
            datos = data,
            compania = perfil.compania,
            fecha = hoy
        )
        
        folio = str('CT' + str(perfil) + str(hoy.year)+ str(hoy.month)+ str(hoy.day) + str(orden.id))
        orden.folio = folio
        orden.save()
        
        #Cambio de taller
        for llanta in llantas:
            llanta.taller = taller_destino
            llanta.fecha_de_entrada_inventario = date.today()
        Llanta.objects.bulk_update(llantas, ['taller', 'fecha_de_entrada_inventario'])
        
        #Vaciado del carrito
        carrito = LlantasSeleccionadas.objects.get(perfil=perfil, inventario=inventario)
        carrito.llantas.clear()
        carrito.save()
        
        return redirect('dashboards:almacen')
        
        
class stockDestinoView(LoginRequiredMixin, TemplateView):
# Vista de stockDestinoView

    template_name = "stockDestino.html"

    def get_context_data(self, **kwargs):
        
        context = super().get_context_data(**kwargs)
        #Obtencion de la data
        stock = self.request.GET.get('inventario', [])
        ids = self.request.GET.get('ids', [])
        
        #Estructuracion de los id
        ids = str(ids).replace("[", "")
        ids = str(ids).replace("]", "")
        ids_list = ids
        ids = str(ids).split(',')
        ids = functions.int_list_element(ids)
        
        #Llantas
        llantas = Llanta.objects.filter(id__in = ids)
        #Obtecion de la lista de destinos
        inventarios = [
            "Antes de Renovar", 
            "Antes de Desechar", 
            "Servicio",
        ]
        
        #Paso del contexto
        context['stock'] = stock
        context['ids_list'] = ids_list
        context['llantas'] = llantas
        context['inventarios'] = inventarios
        return context
    def post(self, request):
        print(request.POST)
        #Obtencion de la data
        hoy = date.today()
        user = request.user
        perfil = Perfil.objects.get(user = user)
        inventario = request.POST.getlist('inventario')[0]
        inventario_destino = request.POST.getlist('inventario_destino')[0]
        ids = request.POST.getlist('ids')[0]
        ids = ids.split(',')
        ids = functions.int_list_element(ids)
        
        #Obtencion de las llantas
        llantas = Llanta.objects.filter(id__in = ids)
        
        #Generacion de la lista de los datos de las llantas
        llanta_data = []
        for llanta in llantas:
            llanta_data.append(
                {
                    'num_eco': llanta.numero_economico,
                    'producto': llanta.producto.producto if llanta.producto != None else None,
                    'stock_origen': llanta.inventario
                }
            )
        
        #Generacion del campo datos
        data = str({
            'fecha': str(date.today()),
            'usuario': str(user),
            'inventario_destino': inventario_destino,
            'total_llantas': len(ids),
            'data': llanta_data
        })
        
        #Creacion de la orden
        orden = Orden.objects.create(
            status = 'CambioStock',
            datos = data,
            compania = perfil.compania,
            fecha = hoy
        )
        
        folio = str('CS' + str(perfil) + str(hoy.year)+ str(hoy.month)+ str(hoy.day) + str(orden.id))
        orden.folio = folio
        orden.save()
        
        #Cambio de taller
        for llanta in llantas:
            llanta.inventario = inventario_destino
            llanta.fecha_de_entrada_inventario = date.today()
        Llanta.objects.bulk_update(llantas, ['inventario', 'fecha_de_entrada_inventario'])
        
        #Vaciado del carrito
        carrito = LlantasSeleccionadas.objects.get(perfil=perfil, inventario=inventario)
        carrito.llantas.clear()
        carrito.save()
        
        return redirect('dashboards:almacen')

class procesoRenovadoView(LoginRequiredMixin, TemplateView):
# Vista de procesoRenovadoView

    template_name = "procesoRenovado.html"
    
    def get_context_data(self, **kwargs):
        
        context = super().get_context_data(**kwargs)
        ids = self.request.GET.get('ids', [])
        ids = str(ids).replace("[", "")
        ids = str(ids).replace("]", "")
        ids_list = ids
        ids = str(ids).split(',')
        print(ids)
        ids = functions.int_list_element(ids)
        print(ids)
        context["ids"] = ids
        return context

    def post(self, request, *args, **kwargs):
        print(request.POST)
        status = self.request.POST.get("status")
        producto = self.request.POST.get("producto", None)
        razon = self.request.POST.get("razon", None)
        taller = str(self.request.POST.get("taller"))
        ids = self.request.POST.getlist("tire-id")
        ids = (str(ids).replace(' ', '')).replace("\'", '').replace('[', '').replace(']', '')
        print(status)
        print(producto)
        print(taller)
        print(ids)
        #folio = str('CS' + str(perfil) + str(hoy.year)+ str(hoy.month)+ str(hoy.day) + str(orden.id))
        if producto != None:
            url = f"%s?status={status}&producto={producto}&taller={taller}&ids={ids}" % reverse("dashboards:ordenEntrada")
        else:
            url = f"%s?status={status}&razon={razon}&taller={taller}&ids={ids}" % reverse("dashboards:ordenEntrada")
        return redirect(url)

class calendarView(LoginRequiredMixin, TemplateView):
# Vista de calendarView

    template_name = "calendar/calendar.html"

class serviciosAbiertosView(LoginRequiredMixin, TemplateView):
# Vista de serviciosAbiertosView

    template_name = "serviciosAbiertos.html"

class resumenView(LoginRequiredMixin, TemplateView):
# Vista de resumenView

    template_name = "resumen.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        perfil = Perfil.objects.get(user=user)
        compania = perfil.compania
        vehiculos = Vehiculo.objects.select_related().filter(compania = compania).exclude(configuracion=None).exclude(estatus_activo=False).values('id').order_by('id')
        total_vehiculos = vehiculos.count()
        current_vehiculos = functions.list_vehicle_id(vehiculos)
        llantas = Llanta.objects.filter(vehiculo__id__in = current_vehiculos, inventario = 'Rodante')

        alta_presion = [5]
        baja_presion = [4]
        baja_profundidad = [7]
        en_punto_retiro = [8]
        desgaste_irregular = [10, 11, 12, 13]
        dualizacion = [9]
        
        libre = 'C'
        direccion = 'S'
        traccion = 'D'
        retractil = 'L'
        arrastre = 'T'
        #? Query's
        clauses_direccion = (Q(tipo_de_eje__icontains=p) for p in [direccion])
        query_direccion = reduce(operator.or_, clauses_direccion)
        
        clauses_libre = (Q(tipo_de_eje__icontains=p) for p in [libre])
        query_libre = reduce(operator.or_, clauses_libre)
        
        clauses_retractil = (Q(tipo_de_eje__icontains=p) for p in [retractil])
        query_retractil = reduce(operator.or_, clauses_retractil)
        
        clauses_traccion = (Q(tipo_de_eje__icontains=p) for p in [traccion])
        query_traccion = reduce(operator.or_, clauses_traccion)
        
        clauses_arrastre = (Q(tipo_de_eje__icontains=p) for p in [arrastre])
        query_arrastre = reduce(operator.or_, clauses_arrastre)
        
        #? Problemas Query's
        query_baja_presion = ({'observaciones__id__in': baja_presion})
        query_alta_presion = ({'observaciones__id__in': alta_presion})
        query_baja_profundidad = ({'observaciones__id__in': baja_profundidad})
        query_desgaste_irregular = ({'observaciones__id__in': desgaste_irregular})
        query_dualizacion = ({'observaciones__id__in': dualizacion})
        
        #? Listas
        
        direccion_baja_presion = llantas.filter(query_direccion, **query_baja_presion).count()
        direccion_alta_presion = llantas.filter(query_direccion, **query_alta_presion).count()
        direccion_baja_profundidad = llantas.filter(query_direccion, **query_baja_profundidad).count()
        direccion_desgaste_irregular = llantas.filter(query_direccion, **query_desgaste_irregular).count()
        
        
        libre_baja_presion = llantas.filter(query_libre, **query_baja_presion).count()
        libre_alta_presion = llantas.filter(query_libre, **query_alta_presion).count()
        libre_baja_profundidad = llantas.filter(query_libre, **query_baja_profundidad).count()
        libre_desgaste_irregular = llantas.filter(query_libre, **query_desgaste_irregular).count()
        
        retractil_baja_presion = llantas.filter(query_retractil, **query_baja_presion).count()
        retractil_alta_presion = llantas.filter(query_retractil, **query_alta_presion).count()
        retractil_baja_profundidad = llantas.filter(query_retractil, **query_baja_profundidad).count()
        retractil_desgaste_irregular = llantas.filter(query_retractil, **query_desgaste_irregular).count()
        
        traccion_baja_presion = llantas.filter(query_traccion, **query_baja_presion).count()
        traccion_alta_presion = llantas.filter(query_traccion, **query_alta_presion).count()
        traccion_baja_profundidad = llantas.filter(query_traccion, **query_baja_profundidad).count()
        traccion_desgaste_irregular = llantas.filter(query_traccion, **query_desgaste_irregular).count()
        
        arrastre_baja_presion = llantas.filter(query_arrastre, **query_baja_presion).count()
        arrastre_alta_presion = llantas.filter(query_arrastre, **query_alta_presion).count()
        arrastre_baja_profundidad = llantas.filter(query_arrastre, **query_baja_profundidad).count()
        arrastre_desgaste_irregular = llantas.filter(query_arrastre, **query_desgaste_irregular).count()
        
        
        #Sumas
        
        total_alta_presion = (
                                direccion_baja_presion +
                                libre_baja_presion +
                                retractil_baja_presion +
                                traccion_baja_presion +
                                arrastre_baja_presion
                                )
        
        total_baja_presion = (
                                direccion_alta_presion +
                                libre_alta_presion +
                                retractil_alta_presion +
                                traccion_alta_presion +
                                arrastre_alta_presion
                                )
        total_baja_profundidad = (
                                    direccion_baja_profundidad +
                                    libre_baja_profundidad +
                                    retractil_baja_profundidad +
                                    traccion_baja_profundidad +
                                    arrastre_baja_profundidad
                                    )
        total_desgaste_irregular = (
                                    direccion_desgaste_irregular +
                                    libre_desgaste_irregular +
                                    retractil_desgaste_irregular +
                                    traccion_desgaste_irregular +
                                    arrastre_desgaste_irregular
                                    )
        duales_impares = round(llantas.filter(**query_dualizacion).count() / 2)
        hoy = datetime.today()
        
        vehiculos_sin_alineacion = vehiculos.annotate(
            dias_sin_alinear=DiffDays(
                hoy-F('fecha_ultima_alineacion'), output_field=IntegerField()
                )).exclude(
                    dias_alinear=0
                    ).exclude(
                        dias_alinear=None
                        ).exclude(
                            dias_sin_alinear__lte=F('dias_alinear')
                            ).count()

        vehiculos_sin_inspeccion = vehiculos.annotate(
            dias_sin_inspeccion=DiffDays(
                hoy-F('fecha_ultima_inspeccion'), output_field=IntegerField()
                )).exclude(
                    compania__periodo2_inspeccion=0
                    ).exclude(
                        compania__periodo2_inspeccion=None
                        ).exclude(
                            dias_sin_inspeccion__lte=F('compania__periodo2_inspeccion')
                            ).values('numero_economico').count()
        
        vehiculos_sin_inflado = vehiculos.annotate(
            dias_sin_inflado=DiffDays(
                hoy-F('fecha_de_inflado'), output_field=IntegerField()
                )).exclude(
                    compania__periodo2_inflado=0
                    ).exclude(
                        compania__periodo2_inflado=None
                        ).exclude(
                            dias_sin_inflado__lte=F('compania__periodo2_inflado')
                            ).count()
        
        #? Contexto
        
        #* Eje-Direccion
        context['direccion_baja_presion'] = direccion_baja_presion
        context['direccion_alta_presion'] = direccion_alta_presion
        context['direccion_baja_profundidad'] = direccion_baja_profundidad
        context['direccion_desgaste_irregular'] = direccion_desgaste_irregular
        
        #* Eje-Libre
        context['libre_baja_presion'] = libre_baja_presion
        context['libre_alta_presion'] = libre_alta_presion
        context['libre_baja_profundidad'] = libre_baja_profundidad
        context['libre_desgaste_irregular'] = libre_desgaste_irregular
        
        #* Eje-Retractil
        context['retractil_baja_presion'] = retractil_baja_presion
        context['retractil_alta_presion'] = retractil_alta_presion
        context['retractil_baja_profundidad'] = retractil_baja_profundidad
        context['retractil_desgaste_irregular'] = retractil_desgaste_irregular
        
        #* Eje-Traccion
        context['traccion_baja_presion'] = traccion_baja_presion
        context['traccion_alta_presion'] = traccion_alta_presion
        context['traccion_baja_profundidad'] = traccion_baja_profundidad
        context['traccion_desgaste_irregular'] = traccion_desgaste_irregular
        
        #* Eje-Arrastre
        context['arrastre_baja_presion'] = arrastre_baja_presion
        context['arrastre_alta_presion'] = arrastre_alta_presion
        context['arrastre_baja_profundidad'] = arrastre_baja_profundidad
        context['arrastre_desgaste_irregular'] = arrastre_desgaste_irregular
        
        #* Totales
        context['total_alta_presion'] = total_alta_presion
        context['total_baja_presion'] = total_baja_presion
        context['total_baja_profundidad'] = total_baja_profundidad
        context['total_desgaste_irregular'] = total_desgaste_irregular
        
        #* Datos individuales
        context['vehiculos_sin_alineacion'] = vehiculos_sin_alineacion
        context['vehiculos_sin_inspeccion'] = vehiculos_sin_inspeccion
        context['vehiculos_sin_inflado'] = vehiculos_sin_inflado
        context['duales_impares'] = duales_impares
        
        #* Total de vehiculos
        context['total_vehiculos'] = total_vehiculos
        return context
    
class planTallerView(LoginRequiredMixin, TemplateView):
# Vista de planTallerView
 
    template_name = "planTaller.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #? Se obtiene el vehiculo
        vehiculo = Vehiculo.objects.filter(pk=self.kwargs['pk'])
        
        
        km = vehiculo[0].km
        km_max = functions.km_max(vehiculo[0])
        km_min = functions.km_min(vehiculo[0])
        #km_max = ''
        #km_min = ''
        print('-------')
        print(km_max)
        print(km_min)
        print('-------')
        
        
        #? Se convirte el id en una lista(Por convenvion de las funciones)
        ids_vehiculo = functions.list_vehicle_id(vehiculo.values('id'))
        #? Se obtiene las llantas 
        llantas = Llanta.objects.filter(
            vehiculo__id__in = ids_vehiculo,
            inventario = 'Rodante'
            ).annotate(
                min_profundidad=Least("profundidad_izquierda", "profundidad_central", "profundidad_derecha")
                ) 
        #? Se acomodan las llantas
        vehiculos_llantas_acomodadas = functions.acomodo_de_llantas_por_vehiculo(llantas, ids_vehiculo)
        acomodo_ejes_vehicle = functions.acomodo_ejes_vehicle(vehiculos_llantas_acomodadas)
        acomodo_posicion_ejes_vehicle = functions.acomodo_pocisiones_vehicle(acomodo_ejes_vehicle)
        vehiculo_acomodado = acomodo_posicion_ejes_vehicle
        
        if vehiculo_acomodado[0]['sp1'] == True:
            llanta_sp1 = Llanta.objects.get(vehiculo = vehiculo[0], inventario = 'Rodante', tipo_de_eje = 'SP1')
            context['llanta_sp1'] = llanta_sp1
        talleres = Taller.objects.filter(compania = vehiculo[0].compania)
        
        usuarios = Perfil.objects.filter(companias__in = [vehiculo[0].compania])
        
        
        #? Checar si existe un taller abierto
        servicios = ServicioVehiculo.objects.filter(vehiculo = vehiculo[0], estado = 'abierto')
        
        
        llantas_no_validas = llantas.filter(producto=None)
        num_llantas_no_validas = llantas_no_validas.filter(producto=None).count()

        
        context['vehiculo_acomodado'] = vehiculo_acomodado
        context['talleres'] = talleres
        context['vehiculo'] = vehiculo
        context['fecha'] = datetime.now()
        context['servicios'] = servicios.count()
        try:
            context['idServicio'] = servicios.first().id
        except:
            context['idServicio'] = None
        context['usuarios'] = usuarios
        context['km'] = km
        context['km_max'] = km_max
        context['km_min'] = km_min
        context['llantas_no_validas'] = llantas_no_validas
        context['num_llantas_no_validas'] = num_llantas_no_validas
        print(datetime.now())
        return context
    
    def post(self, request, pk):
        print(request.POST.getlist('formulario')[0])
        #? Usuarios
        usuario = self.request.user
        perfil = Perfil.objects.get(user = usuario)
        compania_act = perfil.compania
        if request.POST.getlist('formulario')[0] == 'agendar':
            servicio_vehiculo = functions.servicio_vehiculo_preguardado(pk, request)
            functions.archivar_taller(request.POST, pk, servicio_vehiculo)
            return redirect('dashboards:calendario')
        print(request.POST)
        #? Obtencion de los datos de la hoja de servicio
        hoja = request.POST['hoja']
        hoja = json.loads(hoja)
        
        #? Acciones del vehiculo
        vehiculo = Vehiculo.objects.get(pk = pk)
        vehiculos_lista = [vehiculo]
        
        try:
            acciones_vehiculo = json.loads(request.POST.getlist('vehiculo')[0])
        except:
            acciones_vehiculo = []
        if 'alinearVehiculo' in acciones_vehiculo:
            vehiculo.fecha_ultima_alineacion = date.today()
        #Actualizar km del vehiculo
        km_imposible = True if 'no_km' in hoja else False
        print(f'km_imposible = {km_imposible}')
        if km_imposible == False:
            km_vehiculo = int(hoja['km_montado']) if hoja['km_montado'] != '' and hoja['km_montado'] != None else None
            vehiculo.km = km_vehiculo
        
        
        inflarVehiculo = True if 'inflarVehiculo' in acciones_vehiculo else False
        Vehiculo.objects.bulk_update([vehiculo], ['fecha_ultima_alineacion', 'km'])
        
        #? Acciones de llanta
        
        try:
            dataPOST = json.loads(request.POST.getlist('data')[0])
        except:
            dataPOST = []
        if len(dataPOST) == 0 and inflarVehiculo == True:
            llantas = Llanta.objects.filter(vehiculo = vehiculo, inventario = 'Rodante')
            for llanta in llantas:
                dataPOST.append(
                    {
                    "tipoServicio":"sr",
                    "srvcarretera":"",
                    "inflar":"on",
                    "balancear":"",
                    "reparar":"",
                    "valvula":"",
                    "costado":"",
                    "rotar":"no",
                    "otroVehiculo":"",
                    "llantaOrigen":"",
                    "kmMontadoOtro":"",
                    "stock":"",
                    "nuevaLlanta":"",
                    "razon":"",
                    "almacen_desmontaje":"",
                    "taller_desmontaje":"",
                    "llantaId":llanta.id,
                    "numero_economico":"",
                    "posicion":"",
                    "id_servicio":""}
                )
                
        if len(dataPOST) != 0 and inflarVehiculo == True:
            ids_in_data = functions.ids_in_data(dataPOST)
            llantas = Llanta.objects.filter(vehiculo = vehiculo, inventario = 'Rodante').exclude(id__in = ids_in_data)
           
            for data in dataPOST:
                if data['tipoServicio'] == 'sr':
                    data['inflar'] = 'on'
            for llanta in llantas:
                dataPOST.append(
                    {
                    "tipoServicio":"sr",
                    "srvcarretera":"",
                    "inflar":"on",
                    "balancear":"",
                    "reparar":"",
                    "valvula":"",
                    "costado":"",
                    "rotar":"no",
                    "otroVehiculo":"",
                    "llantaOrigen":"",
                    "kmMontadoOtro":"",
                    "stock":"",
                    "nuevaLlanta":"",
                    "razon":"",
                    "almacen_desmontaje":"",
                    "taller_desmontaje":"",
                    "llantaId":llanta.id,
                    "numero_economico":"",
                    "posicion":"",
                    "id_servicio":""}
                )
                   
        #raise Exception("user text")
        llantas_desmontadas = []
        llantas_rotadas = []
        print('---------------')
        for data in dataPOST:
            if data['tipoServicio'] == 'desmontaje':
                print('Desmontaje')
                #? Se llama la llanta actual
                llanta = Llanta.objects.get(pk=data['llantaId'])
                #? Presion establecida
                presion_establecida = functions.presion_establecida(llanta)
                vehiculo = llanta.vehiculo
                ubicacion = llanta.ubicacion
                aplicacion = llanta.aplicacion
                tipo_eje = llanta.tipo_de_eje
                nombre_de_eje = llanta.nombre_de_eje
                eje = llanta.eje
                posicion = llanta.posicion
                if km_imposible == False:
                    km_montado = int(hoja['km_montado']) if hoja['km_montado'] != '' and hoja['km_montado'] != None else None                    
                else:
                    km_montado = None
                    
                #? Se llama la llanta a montar
                nuevaLlanta = data['nuevaLlanta']
                llanta_nueva = Llanta.objects.get(numero_economico = nuevaLlanta, compania = perfil.compania)
                
                #? Se obtienen los datos del desmontaje
                razon = data['razon']
                almacen_desmontaje = data['almacen_desmontaje']
                taller_desmontaje = Taller.objects.get(id = int(data['taller_desmontaje']))
                
                #? Se desmonta la llanta
                llanta.inventario = almacen_desmontaje
                llanta.taller = taller_desmontaje
                
                #?Montaje de la nueva llanta
                llanta_nueva.vehiculo = vehiculo
                llanta_nueva.posicion = posicion
                llanta_nueva.ubicacion = ubicacion
                llanta_nueva.aplicacion = aplicacion
                llanta_nueva.tipo_de_eje = tipo_eje
                llanta_nueva.eje = eje
                llanta_nueva.nombre_de_eje = nombre_de_eje
                llanta_nueva.inventario = 'Rodante'
                llanta_nueva.km_montado = km_montado
                llanta_nueva.presion_actual = presion_establecida
                
                #? Se actualizan las llantas respectivamente
                Llanta.objects.bulk_update([llanta_nueva], [
                    'vehiculo',
                    'posicion',
                    'ubicacion',
                    'aplicacion',
                    'tipo_de_eje',
                    'eje',
                    'inventario',
                    'km_montado',
                    'presion_actual',
                    'nombre_de_eje'
                    ])
                
                Llanta.objects.bulk_update([llanta], [
                    'taller', 'inventario'
                    ])
                llantas_desmontadas.append(llanta)
                functions.quitar_llanta_de_carritos(llanta_nueva, perfil)
                functions.check_dualizacion(llanta_nueva)
                functions.check_dif_presion_duales(llanta_nueva)
                
                functions.quitar_todo_observaciones(llanta)
                print('---------------')
                                
        for data in dataPOST:
            if data['tipoServicio'] == 'sr':
                print('Servicio')
                #?Se llama la llanta
                llanta = Llanta.objects.get(pk=data['llantaId'])
                #? Se verifican que servicios se realizaron
                srvcarretera = True if data['srvcarretera'] == 'on' else False
                inflar = True if data['inflar'] == 'on' else False
                balancear = True if data['balancear'] == 'on' else False
                reparar = True if data['reparar'] == 'on' else False
                valvula = True if data['valvula'] == 'on' else False
                costado = True if data['costado'] == 'on' else False
                
                rotar = data['rotar'] if data['rotar'] != 'no' else False
                
                print(f'srvcarretera:{srvcarretera}')
                print(f'inflar:{inflar}')
                print(f'balancear:{balancear}')
                print(f'reparar:{reparar}')
                print(f'valvula:{valvula}')
                print(f'costado:{costado}')
                
                print(f'rotar:{rotar}')
                
                #? Servicios
                if srvcarretera:
                    producto, created = Producto.objects.get_or_create(producto ='CAMBIO EN CARRETERA NULL NULL NULL', defaults={
                        'producto': 'CAMBIO EN CARRETERA NULL NULL NULL',
                        'compania': compania_act,
                        'marca': 'NULL',
                        'dibujo': 'NULL',
                        'rango': 'NULL',
                        'dimension': 'NULL',
                        'profundidad_inicial': 25,
                        'aplicacion': 'Dirección',
                        'vida': 'Nueva',
                        'precio': 0,
                        'km_esperado': 0,
                    }
                    )

                    llanta_nueva = Llanta.objects.create(
                        numero_economico = f'{vehiculo.numero_economico}{llanta.eje}{llanta.posicion}',
                        compania = llanta.compania,
                        vehiculo = llanta.vehiculo,
                        ubicacion = llanta.ubicacion,
                        aplicacion = llanta.aplicacion,
                        vida = 'Nueva',
                        tipo_de_eje = llanta.tipo_de_eje,
                        eje = llanta.eje,
                        posicion = llanta.posicion,
                        nombre_de_eje = llanta.nombre_de_eje,
                        presion_actual = functions.presion_establecida(llanta),
                        profundidad_izquierda = llanta.profundidad_izquierda,
                        profundidad_central = llanta.profundidad_central,
                        profundidad_derecha = llanta.profundidad_derecha,
                        km_montado = None,
                        producto = producto,
                        inventario = llanta.inventario,
                        fecha_de_entrada_inventario = timezone.now(),
                    )
                    llanta_nueva.numero_economico = llanta_nueva.numero_economico + str(llanta_nueva.id)
                    llanta.inventario = 'Archivado'
                    llanta.numero_economico = llanta.numero_economico + 'Archivado'
                    Llanta.objects.bulk_update([llanta, llanta_nueva], ['numero_economico', 'inventario'])
                    llanta = None
                    llanta = llanta_nueva
                    
                if inflar:
                    print('Se inflo')
                    presion_establecida = functions.presion_establecida(llanta)
                    llanta.presion_actual = presion_establecida
                    presion = Observacion.objects.get(observacion = 'Baja presión')
                    alta_presion = Observacion.objects.get(observacion = 'Alta presion')
                    mala_entrada = Observacion.objects.get(observacion = 'Mala entrada')
                    doble_mala_entrada = Observacion.objects.get(observacion = 'Doble mala entrada')
                    llanta.observaciones.remove(presion)
                    llanta.observaciones.remove(alta_presion)
                    llanta.observaciones.remove(mala_entrada)
                    llanta.observaciones.remove(doble_mala_entrada)
                if balancear:
                    print('Se balanceo')
                    llanta.fecha_de_balanceado = date.today()
                if reparar:
                    print('Se reparo')
                    obj_penetrado = Observacion.objects.get(observacion = 'Objeto penetrado')
                    llanta.observaciones.remove(obj_penetrado)
                if valvula:
                    print('Se reparo valvula')
                    valvula_fuga = Observacion.objects.get(observacion = 'Válvula fugando')
                    valvula_inaccesible = Observacion.objects.get(observacion = 'Válvula inaccesible')
                    llanta.observaciones.remove(valvula_fuga)
                    llanta.observaciones.remove(valvula_inaccesible)
                if costado:
                    print('Se reparo costado')
                    costado = Observacion.objects.get(observacion = 'Ruptura en costado')
                    llanta.observaciones.remove(costado)
                
                Llanta.objects.bulk_update([llanta], [
                    'presion_actual', 
                    'fecha_de_balanceado'
                    ])
                
                #? Rotar en el mismo vehiculo
                if rotar == 'mismo':
                    print('Rotar en el mismo vehiculo')
                    id_llanta_rotar = int(data['origenLlanta'])
                    llanta_rotar = Llanta.objects.get(id = id_llanta_rotar)
                    
                    if llanta_rotar in llantas_rotadas or llanta_rotar in llantas_desmontadas:
                        continue
                    
                    #? Datos de la llanta a rotar
                    posicion_llanta_rotar = llanta_rotar.posicion
                    eje_llanta_rotar = llanta_rotar.eje
                    tipo_de_eje_llanta_rotar = llanta_rotar.tipo_de_eje
                    
                    #? Datos de la llanta montada
                    posicion_llanta_montada = llanta.posicion
                    eje_llanta_montada = llanta.eje
                    tipo_de_eje_llanta_montada = llanta.tipo_de_eje
                    
                    #? llanta montada
                    llanta.posicion = posicion_llanta_rotar
                    llanta.eje = eje_llanta_rotar
                    llanta.tipo_de_eje = tipo_de_eje_llanta_rotar
                    #? Llanta a rotar
                    llanta_rotar.posicion = posicion_llanta_montada
                    llanta_rotar.eje = eje_llanta_montada
                    llanta_rotar.tipo_de_eje = tipo_de_eje_llanta_montada
                    
                    #? Quitar observaciones                  
                    functions.quitar_desgaste(llanta, llanta_rotar)
                    
                    
                    #? Corregir presiones
                    presion_establecida_montada= functions.presion_establecida(llanta)
                    presion_establecida_rotar = functions.presion_establecida(llanta_rotar)
                    print(f'presion_establecida_montada {presion_establecida_montada}')
                    print(f'presion_establecida_rotar {presion_establecida_rotar}')
                    llanta.presion_actual = presion_establecida_montada
                    llanta_rotar.presion_actual = presion_establecida_rotar
                    
                    
                    Llanta.objects.bulk_update([llanta, llanta_rotar], [
                        'posicion', 
                        'tipo_de_eje', 
                        'eje', 
                        'presion_actual'
                        ])
                    llantas_rotadas.append(llanta)
                    llantas_rotadas.append(llanta_rotar)
                    
                    functions.quitar_todo_de_presion(llanta)
                    functions.quitar_todo_de_presion(llanta_rotar)
                    
                    functions.check_dualizacion(llanta_rotar)
                    functions.check_dif_presion_duales(llanta_rotar)
                    
                #? Rotar en diferente vehiculo
                if rotar == 'otro':
                    print('Rotar en diferentes vehiculo')
                    id_llanta_rotar = int(data['llantaOrigen'])
                    llanta_rotar = Llanta.objects.get(id = id_llanta_rotar)
                    #km_imposible_otro = True if 'no_km' in 'no_km' in data else False
                    km_imposible_otro = True if 'no_km' in data else False
                    
                    if km_imposible_otro == False:
                        km_montado_otro = int(data['kmMontadoOtro']) if data['kmMontadoOtro'] != '' and data['kmMontadoOtro'] != None else None
                    else:
                        km_montado_otro = None

                    if km_imposible == False:
                        km_montado = int(hoja['km_montado']) if hoja['km_montado'] != '' and hoja['km_montado'] != None else None
                    else:
                        km_montado = None
                        
                    
                    if llanta_rotar in llantas_rotadas or llanta_rotar in llantas_desmontadas:
                        continue
                    print('************************')
                    print(km_montado_otro)
                    print(km_montado)
                    print('************************')
                    #? Datos de la llanta a rotar
                    vehiculo_llanta_rotar = llanta_rotar.vehiculo
                    posicion_llanta_rotar = llanta_rotar.posicion
                    eje_llanta_rotar = llanta_rotar.eje
                    tipo_de_eje_llanta_rotar = llanta_rotar.tipo_de_eje
                    #? Actualizar km del vehiculo
                    
                    try:
                        if km_montado_otro > vehiculo_llanta_rotar.km and km_montado_otro != None:
                            vehiculo_llanta_rotar.km = km_montado_otro
                            vehiculo_llanta_rotar.save()
                    except:
                        pass
                    print(vehiculo_llanta_rotar)
                    print(vehiculo_llanta_rotar.km)
                    print('************************')
                    
                    #? Datos de la llanta montada
                    vehiculo_llanta_montada = llanta.vehiculo
                    posicion_llanta_montada = llanta.posicion
                    eje_llanta_montada = llanta.eje
                    tipo_de_eje_llanta_montada = llanta.tipo_de_eje
                    
                    #? Checar presiones objetivos
                    
                    #? llanta montada
                    llanta.vehiculo = vehiculo_llanta_rotar
                    llanta.posicion = posicion_llanta_rotar
                    llanta.eje = eje_llanta_rotar
                    llanta.tipo_de_eje = tipo_de_eje_llanta_rotar
                    llanta.km_montado = km_montado_otro
                    #? Llanta a rotar
                    llanta_rotar.vehiculo = vehiculo_llanta_montada
                    llanta_rotar.posicion = posicion_llanta_montada
                    llanta_rotar.eje = eje_llanta_montada
                    llanta_rotar.tipo_de_eje = tipo_de_eje_llanta_montada
                    llanta_rotar.km_montado = km_montado
                    
                    print(f'vehiculo_llanta_rotar {vehiculo_llanta_rotar}')
                    print(f'posicion_llanta_rotar {posicion_llanta_rotar}')
                    print(f'eje_llanta_rotar {eje_llanta_rotar}')
                    print(f'tipo_de_eje_llanta_rotar {tipo_de_eje_llanta_rotar}')
                    print()
                    print(f'vehiculo_llanta_montada {vehiculo_llanta_montada}')
                    print(f'posicion_llanta_montada {posicion_llanta_montada}')
                    print(f'eje_llanta_montada {eje_llanta_montada}')
                    print(f'tipo_de_eje_llanta_montada {tipo_de_eje_llanta_montada}')
                    #? Quitar observaciones                  
                    functions.quitar_desgaste(llanta, llanta_rotar)
                    
                    #? Corregir presiones
                    presion_establecida_montada= functions.presion_establecida(llanta)
                    presion_establecida_rotar = functions.presion_establecida(llanta_rotar)
                    print(f'presion_establecida_montada {presion_establecida_montada}')
                    print(f'presion_establecida_rotar {presion_establecida_rotar}')
                    llanta.presion_actual = presion_establecida_montada
                    llanta_rotar.presion_actual = presion_establecida_rotar
                    
                    Llanta.objects.bulk_update([llanta, llanta_rotar], [
                        'vehiculo',
                        'posicion', 
                        'tipo_de_eje', 
                        'eje', 
                        'km_montado',
                        'presion_actual'
                        ])
                    llantas_rotadas.append(llanta)
                    llantas_rotadas.append(llanta_rotar)
                    if vehiculo_llanta_montada not in vehiculos_lista:
                        vehiculos_lista.append(vehiculo_llanta_montada)
                    
                    functions.quitar_todo_de_presion(llanta)
                    functions.quitar_todo_de_presion(llanta_rotar)
                    
                    functions.check_dualizacion(llanta_rotar)
                    functions.check_dif_presion_duales(llanta_rotar)
                    
                    functions.check_dualizacion(llanta)
                    functions.check_dif_presion_duales(llanta)
                    
                print('---------------')
        
        #? Rectificar observaciones deñ vehiculo
        functions.rectificar_observaciones_vehiculo(vehiculos_lista)
        #? Guardado de la bitacora
        functions.bitacora_servicios(pk, request, acciones_vehiculo, dataPOST, llantas_desmontadas)
        
        
        
        return redirect('dashboards:calendario')
    
class reporteTallerView(LoginRequiredMixin, TemplateView):
    # Vista del reporteTallerView
    template_name = "reporte-taller.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #? LLamada de los 
        servicio = ServicioVehiculo.objects.get(pk = kwargs['pk'])
        servicios_llanta = ServicioLlanta.objects.filter(serviciovehiculo__id = servicio.id)
        vehiculo_acomodado = functions.eje_a_list(servicio.configuracion)
        cant_ejes = functions.cant_ejes(vehiculo_acomodado)
        
        problemas = functions.lista_problemas_taller(servicios_llanta, servicio)
        context['vehiculo_acomodado'] = vehiculo_acomodado
        context['cant_ejes'] = cant_ejes
        context['problemas'] = problemas
        context['vehiculo'] = servicio.vehiculo
        context['folio'] = servicio.folio
        context['fecha'] = servicio.fecha_inicio
        context['usuario'] = servicio.usuario
        return context
    
class vehicleListView(LoginRequiredMixin, TemplateView):
# Vista de vehicleListView

    template_name = "vehicleList.html"
    
    def get(self, *args, **kwargs):
        #Obtencion de la informacion del usuario
        user = self.request.user
        perfil = Perfil.objects.get(user = user)
        compania = perfil.compania
        
        #Obtencion de los filtros
        filtro = self.request.GET.get('filtro', None)
        filtro_query = ({'observaciones_llanta__id__in': filtro.split(',')} if filtro != '' and  filtro != None else {})
        exclude = self.request.GET.get('exclude', None)
        exclude_query = ({'observaciones_llanta__id__in': exclude.split(',')} if exclude != '' and  exclude != None else {})
        ejes = self.request.GET.get('ejes', None)
        search = self.request.GET.get('search', None)
        search_query = ({'numero_economico__icontains': search} if search != '' and  search != None else {})
        dateFormatter = "%Y-%m-%d"
        
        #? Infaldo
        inflado_inicio = self.request.GET.get('inflado_inicio', None)
        inflado_inicio_query = ({'fecha_de_inflado__gte': datetime.strptime(inflado_inicio, dateFormatter)} if inflado_inicio != '' and  inflado_inicio != None else {})

        inflado_final = self.request.GET.get('inflado_final', None)
        inflado_final_query = ({'fecha_de_inflado__lte': datetime.strptime(inflado_final, dateFormatter)} if inflado_final != '' and  inflado_final != None else {})

        #? Inspeccion
        inspeccion_inicio = self.request.GET.get('inspeccion_inicio', None)
        inspeccion_inicio_query = ({'fecha_ultima_inspeccion__gte': datetime.strptime(inspeccion_inicio, dateFormatter)} if inspeccion_inicio != '' and  inspeccion_inicio != None else {})

        inspeccion_final = self.request.GET.get('inspeccion_final', None)
        inspeccion_final_query = ({'fecha_ultima_inspeccion__lte': datetime.strptime(inspeccion_final, dateFormatter)} if inspeccion_final != '' and  inspeccion_final != None else {})

        #? Alineacion
        alineacion_inicio = self.request.GET.get('alineacion_inicio', None)
        alineacion_inicio_query = ({'fecha_ultima_alineacion__gte': datetime.strptime(alineacion_inicio, dateFormatter)} if alineacion_inicio != '' and  alineacion_inicio != None else {})

        alineacion_final = self.request.GET.get('alineacion_final', None)
        alineacion_final_query = ({'fecha_ultima_alineacion__lte': datetime.strptime(alineacion_final, dateFormatter)} if alineacion_final != '' and  alineacion_final != None else {})

        #Se obtienen los vehiculos de la compañia
        if ejes == None or ejes == '':
            vehiculos = Vehiculo.objects.select_related().filter(
                compania = compania,
                **filtro_query,
                **inflado_inicio_query,
                **inflado_final_query,
                **inspeccion_inicio_query,
                **inspeccion_final_query,
                **alineacion_inicio_query,
                **alineacion_final_query
                ).filter(
                    **search_query).exclude(
                        **exclude_query
                        ).exclude(configuracion=None).exclude(estatus_activo=False).values('id').order_by('id')
                    
        else:
            vehiculos = Vehiculo.objects.select_related().filter(
                compania = compania,
                #! **filtro_query,
                #! **search_query
                ).exclude(
                    **exclude_query
                    ).exclude(configuracion=None).exclude(estatus_activo=False).values('id').order_by('id')
                
            clauses = (Q(configuracion__icontains=p) for p in ejes.split(','))
            query = reduce(operator.or_, clauses)
            vehiculos = vehiculos.filter(query)
            
            clauses2 = (Q(tipo_de_eje__icontains=p) for p in ejes.split(','))
            query2 = reduce(operator.or_, clauses2)
            vehiculos_validos = functions.list_vehicles_valid_filter(vehiculos, filtro, query2)
            vehiculos = vehiculos.filter(id__in = vehiculos_validos,
                                        **inflado_inicio_query,
                                        **inflado_final_query,
                                        **inspeccion_inicio_query,
                                        **inspeccion_final_query,
                                        **alineacion_inicio_query,
                                        **alineacion_final_query)
        vehiculos = vehiculos.filter(**search_query).distinct()
        #vehiculos = functions.ordenar_por_status(vehiculos)
        ids_vehiculo = functions.list_vehicle_id(vehiculos)
        
        #Obtencion de los parametros iniciales de la paginación
        size = (int(self.request.GET['size']) if 'size' in self.request.GET else 12)
        page = (int(self.request.GET['page']) if 'page' in self.request.GET else 1)
        #Se ejecuta la paguinacion para poder trabajar solo en el fragmento que se mostrara
        datos = vehiculos.count()  
        pages = (math.ceil(datos/size))
        limit = page * size
        offset = limit - size
        
        
        print(f'datos: {datos}')
        print(f'pages: {pages}')
        print(f'page: {page}')
        print(f'limit: {limit}')
        print(f'offset: {offset}')
        
        
        #Se verifica que el page se encuentre dentro de un rango valido
        #if page < 1:
        #    url = f"%s?page=1" % reverse('dashboards:vehicleList')
        #    return redirect(url)
        #if page > pages:
        #    url = f"%s?page={pages}" % reverse('dashboards:vehicleList')
        #    return redirect(url)
        
        #pagination = functions.pagination(page, pages)
        url_complemento = functions.pagination_url(filtro, 
                                                   exclude, 
                                                   ejes, 
                                                   search, 
                                                   inflado_inicio,
                                                   inflado_final,
                                                   inspeccion_inicio,
                                                   inspeccion_final,
                                                   alineacion_inicio,
                                                   alineacion_final)
        
        prev = functions.pagination_prev(page, pages, url_complemento)
        next = functions.pagination_next(page, pages, url_complemento)

        self.limit = limit
        self.offset = offset
        self.prev = prev
        self.next = next
        self.ids_vehiculo = ids_vehiculo
        self.ejes = ejes
        self.exclude = exclude
        self.filtro = filtro
        self.search = search
        self.inflado_inicio =  inflado_inicio
        self.inflado_final =  inflado_final
        self.inspeccion_inicio =  inspeccion_inicio
        self.inspeccion_final =  inspeccion_final
        self.alineacion_inicio =  alineacion_inicio
        self.alineacion_final =  alineacion_final
        
        print(f'prev: {prev}')
        print(f'next: {next}')
        
        return super().get(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #Se realiza la obtencion de la rebajada para obtener las llantas
        current_vehiculos = self.ids_vehiculo[self.offset:self.limit]
        llantas = Llanta.objects.filter(vehiculo__id__in = current_vehiculos, inventario = 'Rodante')
        
        #Se acomodan las llantas por vehiculo.
        vehiculos_llantas_acomodadas = functions.acomodo_de_llantas_por_vehiculo(llantas, current_vehiculos)
        acomodo_ejes_vehicle = functions.acomodo_ejes_vehicle(vehiculos_llantas_acomodadas)
        acomodo_posicion_ejes_vehicle = functions.acomodo_pocisiones_vehicle(acomodo_ejes_vehicle)
        vehiculos = acomodo_posicion_ejes_vehicle
        
        grupo = str(User.objects.get(username=self.request.user).groups.get())

        
        """print(llantas.count())
        print(len(current_vehiculos))"""

        context['grupo'] = grupo
        context['ejes'] = self.ejes
        context['exclude'] = self.exclude
        context['filtro'] = self.filtro
        context['prev'] = self.prev
        context['next'] = self.next
        context['llantas_acomodadas'] = vehiculos
        context['search'] = self.search
        context['inflado_inicio'] = self.inflado_inicio
        context['inflado_final'] = self.inflado_final
        context['inspeccion_inicio'] = self.inspeccion_inicio
        context['inspeccion_final'] = self.inspeccion_final
        context['alineacion_inicio'] = self.alineacion_inicio
        context['alineacion_final'] = self.alineacion_final
        return context

    def post(self, request):
        print('---------------------------------')
        print(request.POST)
        print('---------------------------------')
        
        if 'buscar' in request.POST:
            lista_observaciones = functions.lista_de_id_observaciones_get(request.GET)
            lista_observaciones_exclude = functions.lista_de_id_observaciones_exclude_get(request.GET)
            lista_ejes = functions.lista_de_ejes_get(request.GET)
            
            inflado_inicio = functions.inflado_inicio_get(request.GET)
            inflado_final = functions.inflado_final_get(request.GET)

            inspeccion_inicio = functions.inspeccion_inicio_get(request.GET)
            inspeccion_final = functions.inspeccion_final_get(request.GET)

            alineacion_inicio = functions.alineacion_inicio_get(request.GET)
            alineacion_final = functions.alineacion_final_get(request.GET)
            
            vehiculo_search = request.POST['search']
            url = f'%s?filtro={lista_observaciones}&exclude={lista_observaciones_exclude}&ejes={lista_ejes}&inflado_inicio={inflado_inicio}&inflado_final={inflado_final}&inspeccion_inicio={inspeccion_inicio}&inspeccion_final={inspeccion_final}&alineacion_inicio={alineacion_inicio}&alineacion_final={alineacion_final}&search={vehiculo_search}' % reverse('dashboards:vehicleList')
            return redirect(url)
        lista_observaciones = functions.lista_de_id_observaciones(request.POST)
        lista_observaciones_exclude = functions.lista_de_id_observaciones_exclude(request.POST)
        lista_ejes = functions.lista_de_ejes(request.POST)
        
        inflado_inicio = functions.inflado_inicio(request.POST)
        inflado_final = functions.inflado_final(request.POST)
        
        inspeccion_inicio = functions.inspeccion_inicio(request.POST)
        inspeccion_final = functions.inspeccion_final(request.POST)
        
        alineacion_inicio = functions.alineacion_inicio(request.POST)
        alineacion_final = functions.alineacion_final(request.POST)

        #vehiculos = Vehiculo.objects.filter(observaciones_llanta__id__in = lista_observaciones)
        url = f'%s?filtro={lista_observaciones}&exclude={lista_observaciones_exclude}&ejes={lista_ejes}&inflado_inicio={inflado_inicio}&inflado_final={inflado_final}&inspeccion_inicio={inspeccion_inicio}&inspeccion_final={inspeccion_final}&alineacion_inicio={alineacion_inicio}&alineacion_final={alineacion_final}' % reverse('dashboards:vehicleList')
        return redirect(url)
    
    
class dashboardOperativoView(LoginRequiredMixin, TemplateView):
# Vista de dashboardOperativoView

    template_name = "dashboardOperativo.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        print(user)
        perfil = Perfil.objects.get(user=user)
        link_operativo = perfil.link_operativo
        context['link_operativo'] = link_operativo
        return context
    

class LogoutView(LoginRequiredMixin, auth_views.LogoutView):
    # Vista de Logout
    pass

class VehiculoAPI(View):
    #? Api pulpo
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request):
        jd = json.loads(request.body)
        vehiculo = Vehiculo.objects.get(numero_economico=jd['numero_economico'], compania=Compania.objects.get(compania=jd['compania']))
        vehiculo.fecha_de_inflado=datetime.now()
        vehiculo.tiempo_de_inflado=jd['tiempo_de_inflado']
        vehiculo.presion_de_entrada=jd['presion_de_entrada']
        vehiculo.presion_de_salida=jd['presion_de_salida']
        vehiculo.save()
        bi = Bitacora.objects.create(vehiculo=vehiculo,
                compania=Compania.objects.get(compania=jd['compania']),
                fecha_de_inflado=datetime.now(),
                tiempo_de_inflado=jd['tiempo_de_inflado'],
                presion_de_entrada=jd['presion_de_entrada'],
                presion_de_salida=jd['presion_de_salida']
                )
        bi.save()
        llantas = Llanta.objects.filter(vehiculo=vehiculo, inventario = 'Rodante')
        condicional = True
        numero = 1
        condiciones_ = []
        for llanta in llantas:
            if llanta.tipo_de_eje != "SP1":

                llanta.fecha_de_inflado=datetime.now()
                llanta.presion_de_entrada=jd['presion_de_entrada']
                llanta.presion_de_salida=jd['presion_de_salida']
                llanta.presion_actual=jd['presion_de_salida']
                min_presion = functions.min_presion(llanta)
                max_presion = functions.max_presion(llanta)
                
                presion_ent = float(jd['presion_de_entrada'])
                
                nueva_condicional = functions.check_presion_pulpo(llanta, min_presion, max_presion, condicional)
                condicional = nueva_condicional
                condiciones_.append(condicional)
                functions.check_dif_presion_duales(llanta)
                llanta.save()
                if numero == 1:
                    bi.llanta_1 = llanta
                    bi.presion_establecida_1 = int(functions.presion_establecida(llanta))
                elif numero == 2:
                    bi.llanta_2 = llanta
                    bi.presion_establecida_2 = int(functions.presion_establecida(llanta))
                elif numero == 3:
                    bi.llanta_3 = llanta
                    bi.presion_establecida_3 = int(functions.presion_establecida(llanta))
                elif numero == 4:
                    bi.llanta_4 = llanta
                    bi.presion_establecida_4 = int(functions.presion_establecida(llanta))
                elif numero == 5:
                    bi.llanta_5 = llanta
                    bi.presion_establecida_5 = int(functions.presion_establecida(llanta))
                elif numero == 6:
                    bi.llanta_6 = llanta
                    bi.presion_establecida_6 = int(functions.presion_establecida(llanta))
                elif numero == 7:
                    bi.llanta_7 = llanta
                    bi.presion_establecida_7 = int(functions.presion_establecida(llanta))
                elif numero == 8:
                    bi.llanta_8 = llanta
                    bi.presion_establecida_8 = int(functions.presion_establecida(llanta))
                elif numero == 9:
                    bi.llanta_9 = llanta
                    bi.presion_establecida_9 = int(functions.presion_establecida(llanta))
                elif numero == 10:
                    bi.llanta_10 = llanta
                    bi.presion_establecida_10 = int(functions.presion_establecida(llanta))
                elif numero == 11:
                    bi.llanta_11 = llanta
                    bi.presion_establecida_11 = int(functions.presion_establecida(llanta))
                elif numero == 12:
                    bi.llanta_12 = llanta
                    bi.presion_establecida_12 = int(functions.presion_establecida(llanta))
                
                bi.llantas.add(llanta)
                numero += 1
                
        mala_entrada = Observacion.objects.get(observacion='Mala entrada')
        doble_mala_entrada = Observacion.objects.get(observacion='Doble mala entrada')
        
        if "Doble" in condiciones_:
            bi.estado = 'Doble'
            bi.vehiculo.observaciones_llanta.add(doble_mala_entrada)
        elif "Mala" in condiciones_:
            bi.estado = 'Mala'
            bi.vehiculo.observaciones_llanta.add(mala_entrada)
        else:
            bi.estado = 'Buena'
            bi.vehiculo.observaciones_llanta.remove(mala_entrada)
            bi.vehiculo.observaciones_llanta.remove(doble_mala_entrada)

        bi.save()
        
        functions.asignar_obs_pulpo(bi.vehiculo)
        functions.rectificar_observaciones_vehiculo(bi.vehiculo)
        
        functions.send_mail(bi, 'pulpo')
        return JsonResponse(jd)


class PulpoProAPI(View):
    #? Api pulpopro
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request):
        jd = json.loads(request.body)
        
        print(f'Pulpo Pro: {jd} ')
        
        vehiculo = Vehiculo.objects.get(numero_economico=jd['numero_economico'], compania=Compania.objects.get(compania=jd['compania']))
        
        vehiculo.fecha_de_inflado=timezone.now()
        llantas = Llanta.objects.filter(vehiculo=vehiculo, inventario='Rodante').exclude(tipo_de_eje = 'SP1')
        presiones_de_entrada = eval(jd['presiones_de_entrada'])
        presiones_de_salida = eval(jd['presiones_de_salida'])
        numero_de_llantas = vehiculo.numero_de_llantas

        num_ejes = vehiculo.configuracion.split('.')
        if num_ejes[-1] == "SP1":
            numero_de_llantas = numero_de_llantas - 1

        if numero_de_llantas == len(presiones_de_entrada):
            num_ejes = vehiculo.configuracion.split('.')
            ejes_no_ordenados = []
            ejes = []
            eje = 1
            for num in num_ejes:
                if num != "SP1":
                    list_temp = []
                    for llanta in llantas:
                        if llanta.eje == eje:
                            list_temp.append([llanta])
                    eje += 1
                    ejes_no_ordenados.append(list_temp)
            
            for eje in ejes_no_ordenados:
                if len(eje) == 2:
                    lista_temp = ['', '']
                    for llanta_act in eje:
                        if 'LI' in llanta_act[0].posicion:
                            lista_temp[0] = llanta_act
                            
                        elif 'RI' in llanta_act[0].posicion:
                            lista_temp[1] = llanta_act
                    ejes.append(lista_temp)
                    #print(' 0---0')
                
                elif len(eje) == 4:
                    lista_temp = ['', '', '', '']
                    for llanta_act in eje:
                        if 'LO' in llanta_act[0].posicion:
                            lista_temp[0] = llanta_act
                        elif 'LI' in llanta_act[0].posicion:
                            lista_temp[1] = llanta_act
                        elif 'RI' in llanta_act[0].posicion:
                            lista_temp[2] = llanta_act
                        elif 'RO' in llanta_act[0].posicion:
                            lista_temp[3] = llanta_act
                    ejes.append(lista_temp)
                    #print('00---00')
                else:
                    pass
            loop_llantas = 0
            for eje in ejes:
                for llanta in eje:
                    llanta[0].presion_de_entrada = presiones_de_entrada[loop_llantas]
                    llanta[0].presion_de_salida = presiones_de_salida[loop_llantas]
                    llanta[0].presion_actual = presiones_de_salida[loop_llantas]
                    llanta[0].fecha_de_inflado=timezone.now()
                    llanta[0].save()
                    loop_llantas += 1
                    
            bi = Bitacora_Pro.objects.create(vehiculo=vehiculo,
                                    compania=Compania.objects.get(compania=jd['compania']),
                                    fecha_de_inflado=timezone.now(),
                                    tiempo_de_inflado=jd['tiempo_de_inflado'],
            )
            #print("jd['presiones_de_entrada']", jd['presiones_de_entrada'])

            #print("presiones_de_entrada", presiones_de_entrada)
            if numero_de_llantas == len(presiones_de_entrada):

                if len(presiones_de_entrada) >= 1:
                    bi.presion_de_entrada_1 = presiones_de_entrada[0]
                    bi.presion_de_salida_1 = presiones_de_salida[0]
                    if len(presiones_de_entrada) >= 2:
                        bi.presion_de_entrada_2 = presiones_de_entrada[1]
                        bi.presion_de_salida_2 = presiones_de_salida[1]
                        if len(presiones_de_entrada) >= 3:
                            bi.presion_de_entrada_3 = presiones_de_entrada[2]
                            bi.presion_de_salida_3 = presiones_de_salida[2]
                            if len(presiones_de_entrada) >= 4:
                                bi.presion_de_entrada_4 = presiones_de_entrada[3]
                                bi.presion_de_salida_4 = presiones_de_salida[3]
                                if len(presiones_de_entrada) >= 5:
                                    bi.presion_de_entrada_5 = presiones_de_entrada[4]
                                    bi.presion_de_salida_5 = presiones_de_salida[4]
                                    if len(presiones_de_entrada) >= 6:
                                        bi.presion_de_entrada_6 = presiones_de_entrada[5]
                                        bi.presion_de_salida_6 = presiones_de_salida[5]
                                        if len(presiones_de_entrada) >= 7:
                                            bi.presion_de_entrada_7 = presiones_de_entrada[6]
                                            bi.presion_de_salida_7 = presiones_de_salida[6]
                                            if len(presiones_de_entrada) >= 8:
                                                bi.presion_de_entrada_8 = presiones_de_entrada[7]
                                                bi.presion_de_salida_8 = presiones_de_salida[7]
                                                if len(presiones_de_entrada) >= 9:
                                                    bi.presion_de_entrada_9 = presiones_de_entrada[8]
                                                    bi.presion_de_salida_9 = presiones_de_salida[8]
                                                    if len(presiones_de_entrada) >= 10:
                                                        bi.presion_de_entrada_10 = presiones_de_entrada[9]
                                                        bi.presion_de_salida_10 = presiones_de_salida[9]
                                                        if len(presiones_de_entrada) >= 11:
                                                            bi.presion_de_entrada_11 = presiones_de_entrada[10]
                                                            bi.presion_de_salida_11 = presiones_de_salida[10]
                                                            if len(presiones_de_entrada) >= 12:
                                                                bi.presion_de_entrada_12 = presiones_de_entrada[11]
                                                                bi.presion_de_salida_12 = presiones_de_salida[11]
                bi.save()
        
                vehiculo.ultima_bitacora_pro= bi
                vehiculo.save()
                condicional = True
                numero = 1
                condiciones_ = []
                #print()
                #print(ejes)
                #print()
                for eje in ejes:
                    for ej in eje:
                        for llant in ej:
                            llanta = llant
                            #print(llanta)
                            min_presion = functions.min_presion(llanta)
                            max_presion = functions.max_presion(llanta)
                            
                            nueva_condicional = functions.check_presion_pulpo(llanta, min_presion, max_presion, condicional)
                            condicional = nueva_condicional
                            condiciones_.append(condicional)
                            functions.check_dif_presion_duales(llanta)
                            if numero == 1:
                                bi.llanta_1 = llanta
                                bi.presion_establecida_1 = functions.presion_establecida(llanta)
                                bi.posicion_1  = llanta.posicion
                                bi.estado_llanta_1 = functions.estado_llanta(llanta)
                            elif numero == 2:
                                bi.llanta_2 = llanta
                                bi.presion_establecida_2 = functions.presion_establecida(llanta)
                                bi.posicion_2  = llanta.posicion
                                bi.estado_llanta_2 = functions.estado_llanta(llanta)
                            elif numero == 3:
                                bi.llanta_3 = llanta
                                bi.presion_establecida_3 = functions.presion_establecida(llanta)
                                bi.posicion_3  = llanta.posicion
                                bi.estado_llanta_3 = functions.estado_llanta(llanta)
                            elif numero == 4:
                                bi.llanta_4 = llanta
                                bi.presion_establecida_4 = functions.presion_establecida(llanta)
                                bi.posicion_4  = llanta.posicion
                                bi.estado_llanta_4 = functions.estado_llanta(llanta)
                            elif numero == 5:
                                bi.llanta_5 = llanta
                                bi.presion_establecida_5 = functions.presion_establecida(llanta)
                                bi.posicion_5  = llanta.posicion
                                bi.estado_llanta_5 = functions.estado_llanta(llanta)
                            elif numero == 6:
                                bi.llanta_6 = llanta
                                bi.presion_establecida_6 = functions.presion_establecida(llanta)
                                bi.posicion_6  = llanta.posicion
                                bi.estado_llanta_6 = functions.estado_llanta(llanta)
                            elif numero == 7:
                                bi.llanta_7 = llanta
                                bi.presion_establecida_7 = functions.presion_establecida(llanta)
                                bi.posicion_7  = llanta.posicion
                                bi.estado_llanta_7 = functions.estado_llanta(llanta)
                            elif numero == 8:
                                bi.llanta_8 = llanta
                                bi.presion_establecida_8 = functions.presion_establecida(llanta)
                                bi.posicion_8  = llanta.posicion
                                bi.estado_llanta_8 = functions.estado_llanta(llanta)
                            elif numero == 9:
                                bi.llanta_9 = llanta
                                bi.presion_establecida_9 = functions.presion_establecida(llanta)
                                bi.posicion_9  = llanta.posicion
                                bi.estado_llanta_9 = functions.estado_llanta(llanta)
                            elif numero == 10:
                                bi.llanta_10 = llanta
                                bi.presion_establecida_10 = functions.presion_establecida(llanta)
                                bi.posicion_10 = llanta.posicion
                                bi.estado_llanta_10 = functions.estado_llanta(llanta)
                            elif numero == 11:
                                bi.llanta_11 = llanta
                                bi.presion_establecida_11 = functions.presion_establecida(llanta)
                                bi.posicion_11 = llanta.posicion
                                bi.estado_llanta_11 = functions.estado_llanta(llanta)
                            elif numero == 12:
                                bi.llanta_12 = llanta
                                bi.presion_establecida_12 = functions.presion_establecida(llanta)
                                bi.posicion_12 = llanta.posicion
                                bi.estado_llanta_12 = functions.estado_llanta(llanta)
                                
                            bi.llantas.add(llanta)
                                
                            numero += 1
                    
                mala_entrada = Observacion.objects.get(observacion='Mala entrada')
                doble_mala_entrada = Observacion.objects.get(observacion='Doble mala entrada')
                
                if "Doble" in condiciones_:
                    bi.estado = 'Doble'
                    bi.vehiculo.observaciones_llanta.add(doble_mala_entrada)
                elif "Mala" in condiciones_:
                    bi.estado = 'Mala'
                    bi.vehiculo.observaciones_llanta.add(mala_entrada)
                else:
                    bi.estado = 'Buena'
                    bi.vehiculo.observaciones_llanta.remove(mala_entrada)
                    bi.vehiculo.observaciones_llanta.remove(doble_mala_entrada)

                bi.aplicacion = vehiculo.aplicacion
                bi.ubicacion = vehiculo.ubicacion
                bi.save()
                
                functions.check_presion__entre_duales_pulpo(vehiculo)
                
                #Correccion de obs:
                functions.asignar_obs_pulpo(bi.vehiculo)
                functions.rectificar_observaciones_vehiculo(bi.vehiculo)
                
                functions.send_mail(bi, 'pulpopro')
                return JsonResponse(jd)

class TireEyeAPI(View):

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request):
        jd = json.loads(request.body)
        llanta = Llanta.objects.get(numero_economico=jd['numero_economico'], compania=Compania.objects.get(compania=jd['compania']))
        Inspeccion.objects.create(llanta=llanta,
                                fecha_hora=date.today(),
                                km=jd['km'],
                                min_profundidad=jd['min_profundidad'],
                                max_profundidad=jd['max_profundidad']
        )
        return JsonResponse(jd)

class PulpoView(LoginRequiredMixin, ListView):
    # Vista del dashboard pulpo
    template_name = "pulpo.html"
    model = Vehiculo
    
    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)
        vehiculo = Vehiculo.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        bitacora = Bitacora.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        bitacora_pro = Bitacora_Pro.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        hoy = date.today()

        #functions_create.crear_de_bitacora_el_vehiculo()
        #functions_create.crear_inspecciones(self.request.user.perfil)
        #functions_excel.excel_productos()
        #functions_excel.excel_observaciones()
        #functions_create.borrar_ultima_inspeccion_vehiculo()
        #functions_create.asignar_ultima_fecha_de_inflado()
        #functions_ftp.ftp_diario()
        ultimo_mes = hoy - timedelta(days=31)

        #functions_create.tirecheck_llanta()
        #functions_create.borrar_ultima_inspeccion_vehiculo()
        #functions_create.crear_llantas()
        #functions_excel.excel_llantas(User.objects.get(username="equipo-logistico"))
        #functions_excel.excel_inspecciones()

        mes_1 = hoy.strftime("%b")
        mes_2 = functions.mes_anterior(hoy)
        mes_3 = functions.mes_anterior(mes_2)
        mes_4 = functions.mes_anterior(mes_3)
        mes_5 = functions.mes_anterior(mes_4)

        hoy1 = hoy.strftime("%m")
        hoy2 = mes_2.strftime("%m")
        hoy3 = mes_3.strftime("%m")
        hoy4 = mes_4.strftime("%m")
        hoy5 = mes_5.strftime("%m")

        vehiculo_fecha = vehiculo.filter(fecha_de_inflado__range=[ultimo_mes, hoy])
        vehiculo_fecha_pro = vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__range=[ultimo_mes, hoy])
        vehiculo_fecha_total = vehiculo.filter(fecha_de_inflado__range=[ultimo_mes, hoy]) | vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__range=[ultimo_mes, hoy])
        vehiculo_fecha_barras_1 = vehiculo.filter(fecha_de_inflado__month=hoy1) | vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__month=hoy1)
        vehiculo_fecha_barras_2 = vehiculo.filter(fecha_de_inflado__month=hoy2) | vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__month=hoy2)
        vehiculo_fecha_barras_3 = vehiculo.filter(fecha_de_inflado__month=hoy3) | vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__month=hoy3)
        vehiculo_fecha_barras_4 = vehiculo.filter(fecha_de_inflado__month=hoy4) | vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__month=hoy4)
        vehiculo_fecha_barras_5 = vehiculo.filter(fecha_de_inflado__month=hoy5) | vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__month=hoy5)
        
        vehiculo_mes1 = bitacora.filter(fecha_de_inflado__month=hoy1)
        vehiculo_mes2 = bitacora.filter(fecha_de_inflado__month=hoy2)
        vehiculo_mes3 = bitacora.filter(fecha_de_inflado__month=hoy3)
        vehiculo_mes4 = bitacora.filter(fecha_de_inflado__month=hoy4)

        vehiculo_pro_mes1 = bitacora_pro.filter(fecha_de_inflado__month=hoy1)
        vehiculo_pro_mes2 = bitacora_pro.filter(fecha_de_inflado__month=hoy2)
        vehiculo_pro_mes3 = bitacora_pro.filter(fecha_de_inflado__month=hoy3)
        vehiculo_pro_mes4 = bitacora_pro.filter(fecha_de_inflado__month=hoy4)

        entrada_correcta_contar = functions.contar_entrada_correcta(vehiculo_fecha)
        entrada_correcta_contar += functions.contar_entrada_correcta_pro(vehiculo_fecha_pro)
        mala_entrada_contar_mes1 = functions.contar_mala_entrada(vehiculo_mes1)
        mala_entrada_contar_mes2 = functions.contar_mala_entrada(vehiculo_mes2)
        mala_entrada_contar_mes3 = functions.contar_mala_entrada(vehiculo_mes3)
        mala_entrada_contar_mes4 = functions.contar_mala_entrada(vehiculo_mes4)

        mala_entrada_contar_mes1 += functions.contar_mala_entrada_pro(vehiculo_pro_mes1)
        mala_entrada_contar_mes2 += functions.contar_mala_entrada_pro(vehiculo_pro_mes2)
        mala_entrada_contar_mes3 += functions.contar_mala_entrada_pro(vehiculo_pro_mes3)
        mala_entrada_contar_mes4 += functions.contar_mala_entrada_pro(vehiculo_pro_mes4)

        entrada_correcta_contar_barras_mes1 = functions.contar_entrada_correcta(vehiculo_fecha_barras_1)
        entrada_correcta_contar_barras_mes2 = functions.contar_entrada_correcta(vehiculo_fecha_barras_2)
        entrada_correcta_contar_barras_mes3 = functions.contar_entrada_correcta(vehiculo_fecha_barras_3)
        entrada_correcta_contar_barras_mes4 = functions.contar_entrada_correcta(vehiculo_fecha_barras_4)
        entrada_correcta_contar_barras_mes5 = functions.contar_entrada_correcta(vehiculo_fecha_barras_5)

        doble_entrada = functions.doble_entrada(bitacora, bitacora_pro) 
        doble_mala_entrada = functions.doble_mala_entrada(bitacora, vehiculo)

        doble_mala_entrada_pro = functions.doble_mala_entrada_pro(bitacora_pro, vehiculo)

        vehiculo_periodo = vehiculo.filter(fecha_de_inflado__lte=ultimo_mes).filter(ultima_bitacora_pro=None) | vehiculo.filter(fecha_de_inflado=None).filter(ultima_bitacora_pro__fecha_de_inflado__lte=ultimo_mes) | vehiculo.filter(fecha_de_inflado=None).filter(ultima_bitacora_pro=None) | vehiculo.filter(fecha_de_inflado__lte=ultimo_mes).filter(ultima_bitacora_pro__fecha_de_inflado__lte=ultimo_mes)
        vehiculo_periodo_status = {}
        mala_entrada_periodo = functions.mala_entrada(vehiculo_periodo) | functions.mala_entrada_pro(vehiculo_periodo)
        for v in vehiculo_periodo:
            try:
                if v in doble_mala_entrada or v in doble_mala_entrada_pro:
                    vehiculo_periodo_status[v] = "Doble Entrada"
                elif v in mala_entrada_periodo:
                    vehiculo_periodo_status[v] = "Mala Entrada"
                else:
                    vehiculo_periodo_status[v] = "Entrada Correctas"
            except:
                vehiculo_periodo_status[v] = "Entrada Correctas"

        vehiculo_malos_status = {}
        mala_entrada = functions.mala_entrada(vehiculo) | functions.mala_entrada_pro(vehiculo)
        for v in vehiculo:
            try:
                if v in doble_mala_entrada or v in doble_mala_entrada_pro:
                    vehiculo_malos_status[v] = "Doble Entrada"
                elif v in mala_entrada:
                    vehiculo_malos_status[v] = "Mala Entrada"
            except:
                pass

        my_profile = Perfil.objects.get(user=self.request.user)

        radar_min = functions.radar_min(vehiculo_fecha_total, self.request.user.perfil.compania)
        radar_max = functions.radar_max(vehiculo_fecha_total, self.request.user.perfil.compania)
        radar_min_resta = functions.radar_min_resta(radar_min, radar_max)

        """functions.crear_1(numero_economico="P5",
        compania="pruebacal",
        ubicacion="DOS",
        aplicacion="FORANEO",
        clase="CAMIONETA",
        configuracion="S2.D2", 
        tiempo_de_inflado=2.5,
        presion_de_entrada=100,
        presion_de_salida=100,
        presion_establecida=100)"""

        #functions.crear_3(Vehiculo.objects.get(numero_economico="P5"))

        user = self.request.user
        perfil = Perfil.objects.get(user = user)
        link_pulpo = perfil.link_pulpo
        
        context["aplicaciones"] = Aplicacion.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        context["aplicaciones_mas_frecuentes_infladas"] = functions.aplicaciones_mas_frecuentes(vehiculo_fecha_total, vehiculo, self.request.user.perfil.compania)
        context["bitacoras"] = bitacora
        context["bitacoras_pro"] = bitacora_pro
        context["boton_intuitivo"] = "Vehículos Vencidos"
        context["cantidad_inflado"] = vehiculo_fecha_total.count()
        context["cantidad_inflado_1"] = vehiculo_fecha_barras_1.count()
        context["cantidad_inflado_2"] = vehiculo_fecha_barras_2.count()
        context["cantidad_inflado_3"] = vehiculo_fecha_barras_3.count()
        context["cantidad_inflado_4"] = vehiculo_fecha_barras_4.count()
        context["cantidad_inflado_5"] = vehiculo_fecha_barras_5.count()
        context["cantidad_entrada"] = entrada_correcta_contar
        context["cantidad_entrada_barras_mes1"] = entrada_correcta_contar_barras_mes1
        context["cantidad_entrada_barras_mes2"] = entrada_correcta_contar_barras_mes2
        context["cantidad_entrada_barras_mes3"] = entrada_correcta_contar_barras_mes3
        context["cantidad_entrada_barras_mes4"] = entrada_correcta_contar_barras_mes4
        context["cantidad_entrada_barras_mes5"] = entrada_correcta_contar_barras_mes5
        context["cantidad_entrada_mes1"] = mala_entrada_contar_mes1
        context["cantidad_entrada_mes2"] = mala_entrada_contar_mes2
        context["cantidad_entrada_mes3"] = mala_entrada_contar_mes3
        context["cantidad_entrada_mes4"] = mala_entrada_contar_mes4
        context["cantidad_total"] = vehiculo.count()
        context["clases_compania"] = functions.clases_mas_frecuentes(vehiculo, self.request.user.perfil.compania)
        context["clases_mas_frecuentes_infladas"] = functions.clases_mas_frecuentes(vehiculo_fecha_total, self.request.user.perfil.compania)
        context["compania"] = self.request.user.perfil.compania
        context["doble_entrada"] = doble_entrada
        context["flotas"] = Ubicacion.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        context["hoy"] = hoy
        context["mes_1"] = mes_1
        context["mes_2"] = mes_2.strftime("%b")
        context["mes_3"] = mes_3.strftime("%b")
        context["mes_4"] = mes_4.strftime("%b")
        context["mes_5"] = mes_5.strftime("%b")
        context["porcentaje_inflado"] = functions.porcentaje(vehiculo_fecha_total.count(), vehiculo.count())
        context["porcentaje_entrada_correcta"] = functions.porcentaje(entrada_correcta_contar, vehiculo_fecha_total.count())
        context["radar_min"] = radar_min_resta
        context["radar_max"] = radar_max
        context["rango_1"] = my_profile.compania.periodo1_inflado
        context["rango_2"] = my_profile.compania.periodo2_inflado
        context["rango_3"] = my_profile.compania.periodo1_inflado + 1
        context["rango_4"] = my_profile.compania.periodo2_inflado + 1
        context["tiempo_promedio"] = functions.inflado_promedio(vehiculo_fecha_total)
        context["vehiculos"] = vehiculo_fecha_total
        context["vehiculos_malos"] = vehiculo_malos_status
        context["vehiculos_periodo"] = vehiculo_periodo_status
        context["vehiculos_todos"] = vehiculo
        context["link_pulpo"] = link_pulpo
        return context


def buscar(request):
    # Busca por fecha, localidad o clase
    
    my_profile = Perfil.objects.get(user=request.user)
    vehiculo = Vehiculo.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania))
    vehiculos_totales = Vehiculo.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania))
    bitacora = Bitacora.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania))
    bitacora_pro = Bitacora_Pro.objects.filter(compania=Compania.objects.get(compania= request.user.perfil.compania))
    hoy = date.today()
    ultimo_mes = hoy - timedelta(days=31)
    clase1 = request.GET.getlist("clase")
    flota1 = request.GET.getlist("flota")
    fecha1 = request.GET.get("fechaInicio")
    fecha2 = request.GET.get("fechaFin")

    # Buscar por fecha

    fecha_con_formato1 = None
    fecha_con_formato2 = None
    clase = None
    flota = None
    if clase1:
        clase = clase1
        vehiculo = vehiculo.filter(functions.reduce(or_, [Q(clase=c.upper()) for c in clase1]))
    if flota1:
        flota = flota1
        vehiculo = vehiculo.filter(functions.reduce(or_, [Q(ubicacion=Ubicacion.objects.get(nombre=f)) for f in flota1]))

    if flota1:
        flotas_vehiculo = vehiculos_totales.values("ubicacion").distinct()
    else:
        flotas_vehiculo = vehiculo.values("ubicacion").distinct()

    flotas = Ubicacion.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania), id__in=flotas_vehiculo)
    
    if clase1:
        clases = vehiculos_totales.values_list("clase", flat=True).distinct()
    else:
        clases = vehiculo.values_list("clase", flat=True).distinct()
    
    if fecha1 and fecha2:
        fecha1 = functions.convertir_rango(fecha1)
        fecha2 = functions.convertir_rango(fecha2)
        primera_fecha = datetime.strptime(fecha1, "%Y/%m/%d").date()
        segunda_fecha = datetime.strptime(fecha2, "%Y/%m/%d").date()
        # Convertir formato de fecha
        fecha_con_formato1 = functions.convertir_fecha(fecha1)
        fecha_con_formato2 = functions.convertir_fecha(fecha2)

        vehiculo_fecha = vehiculo.filter(fecha_de_inflado__range=[primera_fecha, segunda_fecha])
        vehiculo_fecha_pro = vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__range=[primera_fecha, segunda_fecha])
    else:
        vehiculo_fecha = vehiculo.filter(fecha_de_inflado__range=[ultimo_mes, hoy])
        vehiculo_fecha_pro = vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__range=[ultimo_mes, hoy])
        vehiculo_fecha_total = vehiculo.filter(fecha_de_inflado__range=[ultimo_mes, hoy]) | vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__range=[ultimo_mes, hoy])
    
    bitacora = bitacora.filter(vehiculo__in=vehiculo)
    bitacora_pro = bitacora_pro.filter(vehiculo__in=vehiculo)


    mes_1 = hoy.strftime("%b")
    mes_2 = functions.mes_anterior(hoy)
    mes_3 = functions.mes_anterior(mes_2)
    mes_4 = functions.mes_anterior(mes_3)
    mes_5 = functions.mes_anterior(mes_4)

    hoy1 = hoy.strftime("%m")
    hoy2 = mes_2.strftime("%m")
    hoy3 = mes_3.strftime("%m")
    hoy4 = mes_4.strftime("%m")
    hoy5 = mes_5.strftime("%m")
    
    vehiculo_fecha_barras_1 = vehiculo.filter(fecha_de_inflado__month=hoy1) | vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__month=hoy1)
    vehiculo_fecha_barras_2 = vehiculo.filter(fecha_de_inflado__month=hoy2) | vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__month=hoy2)
    vehiculo_fecha_barras_3 = vehiculo.filter(fecha_de_inflado__month=hoy3) | vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__month=hoy3)
    vehiculo_fecha_barras_4 = vehiculo.filter(fecha_de_inflado__month=hoy4) | vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__month=hoy4)
    vehiculo_fecha_barras_5 = vehiculo.filter(fecha_de_inflado__month=hoy5) | vehiculo.filter(ultima_bitacora_pro__fecha_de_inflado__month=hoy5)

    vehiculo_mes1 = bitacora.filter(fecha_de_inflado__month=hoy1)
    vehiculo_mes2 = bitacora.filter(fecha_de_inflado__month=hoy2)
    vehiculo_mes3 = bitacora.filter(fecha_de_inflado__month=hoy3)
    vehiculo_mes4 = bitacora.filter(fecha_de_inflado__month=hoy4)

    vehiculo_pro_mes1 = bitacora_pro.filter(fecha_de_inflado__month=hoy1)
    vehiculo_pro_mes2 = bitacora_pro.filter(fecha_de_inflado__month=hoy2)
    vehiculo_pro_mes3 = bitacora_pro.filter(fecha_de_inflado__month=hoy3)
    vehiculo_pro_mes4 = bitacora_pro.filter(fecha_de_inflado__month=hoy4)


    entrada_correcta_contar = functions.contar_entrada_correcta(vehiculo_fecha)
    entrada_correcta_contar += functions.contar_entrada_correcta_pro(vehiculo_fecha_pro)
    mala_entrada_contar_mes1 = functions.contar_mala_entrada(vehiculo_mes1)
    mala_entrada_contar_mes2 = functions.contar_mala_entrada(vehiculo_mes2)
    mala_entrada_contar_mes3 = functions.contar_mala_entrada(vehiculo_mes3)
    mala_entrada_contar_mes4 = functions.contar_mala_entrada(vehiculo_mes4)

    mala_entrada_contar_mes1 += functions.contar_mala_entrada_pro(vehiculo_pro_mes1)
    mala_entrada_contar_mes2 += functions.contar_mala_entrada_pro(vehiculo_pro_mes2)
    mala_entrada_contar_mes3 += functions.contar_mala_entrada_pro(vehiculo_pro_mes3)
    mala_entrada_contar_mes4 += functions.contar_mala_entrada_pro(vehiculo_pro_mes4)


    entrada_correcta_contar_barras_mes1 = functions.contar_entrada_correcta(vehiculo_fecha_barras_1)
    entrada_correcta_contar_barras_mes2 = functions.contar_entrada_correcta(vehiculo_fecha_barras_2)
    entrada_correcta_contar_barras_mes3 = functions.contar_entrada_correcta(vehiculo_fecha_barras_3)
    entrada_correcta_contar_barras_mes4 = functions.contar_entrada_correcta(vehiculo_fecha_barras_4)
    entrada_correcta_contar_barras_mes5 = functions.contar_entrada_correcta(vehiculo_fecha_barras_5)

    doble_entrada = functions.doble_entrada(bitacora, bitacora_pro)
    doble_mala_entrada = functions.doble_mala_entrada(bitacora, vehiculo)

    doble_mala_entrada_pro = functions.doble_mala_entrada_pro(bitacora_pro, vehiculo)

    vehiculo_periodo = vehiculo.filter(fecha_de_inflado__lte=ultimo_mes).filter(ultima_bitacora_pro=None) | vehiculo.filter(fecha_de_inflado=None).filter(ultima_bitacora_pro__fecha_de_inflado__lte=ultimo_mes) | vehiculo.filter(fecha_de_inflado=None).filter(ultima_bitacora_pro=None) | vehiculo.filter(fecha_de_inflado__lte=ultimo_mes).filter(ultima_bitacora_pro__fecha_de_inflado__lte=ultimo_mes)
    vehiculo_periodo_status = {}
    mala_entrada_periodo = functions.mala_entrada(vehiculo_periodo) | functions.mala_entrada_pro(vehiculo_periodo)
    for v in vehiculo_periodo:
        try:
            if v in doble_mala_entrada or v in doble_mala_entrada_pro:
                vehiculo_periodo_status[v] = "Doble Entrada"
            elif v in mala_entrada_periodo:
                vehiculo_periodo_status[v] = "Mala Entrada"
            else:
                vehiculo_periodo_status[v] = "Entrada Correctas"
        except:
            vehiculo_periodo_status[v] = "Entrada Correctas"

    vehiculo_malos_status = {}
    mala_entrada = functions.mala_entrada(vehiculo) | functions.mala_entrada_pro(vehiculo)
    print(doble_mala_entrada_pro)
    for v in vehiculo:
        try:
            if v in doble_mala_entrada or v in doble_mala_entrada_pro:
                vehiculo_malos_status[v] = "Doble Entrada"
            elif v in mala_entrada:
                vehiculo_malos_status[v] = "Mala Entrada"
        except:
            pass

    my_profile = Perfil.objects.get(user=request.user)

    radar_min = functions.radar_min(vehiculo_fecha_total, request.user.perfil.compania)
    radar_max = functions.radar_max(vehiculo_fecha_total, request.user.perfil.compania)
    radar_min_resta = functions.radar_min_resta(radar_min, radar_max)

    return render(request, "pulpo.html", {
                                        "aplicaciones": Aplicacion.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania)),
                                        "aplicaciones_mas_frecuentes_infladas": functions.aplicaciones_mas_frecuentes(vehiculo_fecha_total, vehiculo, request.user.perfil.compania),
                                        "bitacoras": bitacora,
                                        "bitacoras_pro": bitacora_pro,
                                        "boton_intuitivo": "Vehículos Vencidos",
                                        "cantidad_entrada": entrada_correcta_contar,
                                        "cantidad_entrada_barras_mes1": entrada_correcta_contar_barras_mes1,
                                        "cantidad_entrada_barras_mes2": entrada_correcta_contar_barras_mes2,
                                        "cantidad_entrada_barras_mes3": entrada_correcta_contar_barras_mes3,
                                        "cantidad_entrada_barras_mes4": entrada_correcta_contar_barras_mes4,
                                        "cantidad_entrada_barras_mes5": entrada_correcta_contar_barras_mes5,
                                        "cantidad_entrada_mes1": mala_entrada_contar_mes1,
                                        "cantidad_entrada_mes2": mala_entrada_contar_mes2,
                                        "cantidad_entrada_mes3": mala_entrada_contar_mes3,
                                        "cantidad_entrada_mes4": mala_entrada_contar_mes4,
                                        "cantidad_inflado": vehiculo_fecha_total.count(),
                                        "cantidad_inflado_1": vehiculo_fecha_barras_1.count(),
                                        "cantidad_inflado_2": vehiculo_fecha_barras_2.count(),
                                        "cantidad_inflado_3": vehiculo_fecha_barras_3.count(),
                                        "cantidad_inflado_4": vehiculo_fecha_barras_4.count(),
                                        "cantidad_inflado_5": vehiculo_fecha_barras_5.count(),
                                        "cantidad_total": vehiculo.count(),
                                        "clase": clase,
                                        "clase1": clase1,
                                        "clases_compania": clases,
                                        "clases_mas_frecuentes_infladas": functions.clases_mas_frecuentes(vehiculo_fecha_total, request.user.perfil.compania),
                                        "compania": request.user.perfil.compania,
                                        "doble_entrada": doble_entrada,
                                        "fecha1":fecha1,
                                        "fecha2":fecha2,
                                        "fecha_con_formato1":fecha_con_formato1,
                                        "fecha_con_formato2":fecha_con_formato2,
                                        "flota": flota,
                                        "flota1": flota1,
                                        "flotas": flotas,
                                        "hoy": hoy,
                                        "mes_1": mes_1,
                                        "mes_2": mes_2.strftime("%b"),
                                        "mes_3": mes_3.strftime("%b"),
                                        "mes_4": mes_4.strftime("%b"),
                                        "mes_5": mes_5.strftime("%b"),
                                        "porcentaje_inflado": functions.porcentaje(vehiculo_fecha_total.count(), vehiculo.count()),
                                        "porcentaje_entrada_correcta": functions.porcentaje(entrada_correcta_contar, vehiculo_fecha_total.count()),
                                        "radar_min": radar_min_resta,
                                        "radar_max": radar_max,
                                        "rango_1": my_profile.compania.periodo1_inflado,
                                        "rango_2": my_profile.compania.periodo2_inflado,
                                        "rango_3": my_profile.compania.periodo1_inflado + 1,
                                        "rango_4": my_profile.compania.periodo2_inflado + 1,
                                        "tiempo_promedio": functions.inflado_promedio(vehiculo_fecha_total),
                                        "vehiculos": vehiculo_fecha_total,
                                        "vehiculos_malos": vehiculo_malos_status,
                                        "vehiculos_periodo": vehiculo_periodo_status,
                                        "vehiculos_todos": vehiculo
                                    })

class ConfigView(LoginRequiredMixin, MultiModelFormView):
    # Vista del dashboard configuración
    template_name = "config.html"
    form_classes = {"companiaform": CompaniaForm,
                    "usuarioform": UsuarioEditForm,
                    "excelform": ExcelForm}
    success_url = reverse_lazy('dashboards:config')

    """def forms_valid(self, form):
        # Save form data
        user = form.save(commit=False)
        compania_form = forms['companiaform'].save(commit=False)
        groups = form.cleaned_data['groups']
        groups = Group.objects.get(name=groups)
        idioma = form.cleaned_data['idioma']
        user.save()
        
        perfil = Perfil.objects.get(user=user)
        perfil.idioma = idioma
        perfil.save()
        user.groups.add(groups)
        
        return super(ConfigView, self).forms_valid(form)"""

    def get_context_data(self, **kwargs):

        context = super(ConfigView, self).get_context_data(**kwargs)
        if_vehiculo = None
        user = User.objects.get(username=self.request.user)
        groups_names = Group.objects.all()
        compania = Compania.objects.get(compania=self.request.user.perfil.compania)
        
        separador = os.path.sep
        dir_actual = os.path.dirname(os.path.abspath(__file__))
        dir = separador.join(dir_actual.split(separador)[:-2])
        print(dir)
        print('hola')
        if self.request.method=='POST' and 'periodo1_inflado' in self.request.POST:
            print(request.POST)
            compania.periodo1_inflado = self.request.POST.get("periodo1_inflado")
            compania.periodo2_inflado = self.request.POST.get("periodo2_inflado")
            compania.objetivo = self.request.POST.get("objetivo")
            compania.periodo1_inspeccion = self.request.POST.get("periodo1_inspeccion")
            compania.periodo2_inspeccion = self.request.POST.get("periodo2_inspeccion")
            compania.punto_retiro_eje_direccion = self.request.POST.get("punto_retiro_eje_direccion")
            compania.punto_retiro_eje_traccion = self.request.POST.get("punto_retiro_eje_traccion")
            compania.punto_retiro_eje_arrastre = self.request.POST.get("punto_retiro_eje_arrastre")
            compania.mm_de_desgaste_irregular = self.request.POST.get("mm_de_desgaste_irregular")
            compania.mm_de_diferencia_entre_duales = self.request.POST.get("mm_de_diferencia_entre_duales")
            compania.save()
        elif self.request.method=='POST' and 'email' in self.request.POST:
            user.email = self.request.POST.get("email")
            user.username = self.request.POST.get("username")
            user.idioma = self.request.POST.get("idioma")
            groups = Group.objects.get(name=self.request.POST.get("groups"))
            user.groups.clear()
            user.groups.add(groups)
            user.save()
        elif self.request.method=='POST' and self.request.FILES.get("file"):
            file = self.request.FILES.get("file")
            archivo = dir + r"\files.xlsx"
            fp = open(archivo,'wb')
            for chunk in file.chunks():
                fp.write(chunk)
  
            wb_obj = openpyxl.load_workbook(file)
            sheet_obj = wb_obj.active
            
            if_vehiculo = "Nunca"
            llantas_validadas = []
            cantidad_llantas_vehiculo = 0
            llantas_por_vehiculo = []
            for i in range(sheet_obj.max_row):
                try:
                    numero_economico = sheet_obj.cell(row=i + 2, column=1).value
                    vehiculo = Vehiculo.objects.get(numero_economico=numero_economico, compania=Compania.objects.get(compania=compania))
                    cantidad_llantas_vehiculo += 1
                    llantas_por_vehiculo.append(numero_economico)
                except:
                    pass             

            for i in range(sheet_obj.max_row):
                numero_economico = sheet_obj.cell(row=i + 2, column=1).value
                try:
                    try:
                        vehiculo = Vehiculo.objects.get(numero_economico=numero_economico, compania=Compania.objects.get(compania=compania))
                    except:
                        modelo = sheet_obj.cell(row=i + 2, column=2).value
                        marca = sheet_obj.cell(row=i + 2, column=3).value
                        ubicacion = sheet_obj.cell(row=i + 2, column=4).value
                        aplicacion = sheet_obj.cell(row=i + 2, column=5).value
                        clase = sheet_obj.cell(row=i + 2, column=6).value
                        configuracion = sheet_obj.cell(row=i + 2, column=7).value
                        presion_establecida_1 = sheet_obj.cell(row=i + 2, column=8).value
                        presion_establecida_2 = sheet_obj.cell(row=i + 2, column=9).value
                        presion_establecida_3 = sheet_obj.cell(row=i + 2, column=10).value
                        presion_establecida_4 = sheet_obj.cell(row=i + 2, column=11).value
                        presion_establecida_5 = sheet_obj.cell(row=i + 2, column=12).value
                        presion_establecida_6 = sheet_obj.cell(row=i + 2, column=13).value
                        presion_establecida_7 = sheet_obj.cell(row=i + 2, column=14).value
                        km_diario_maximo = sheet_obj.cell(row=i + 2, column=15).value
                        estatus_activo = sheet_obj.cell(row=i + 2, column=16).value
                        nuevo = sheet_obj.cell(row=i + 2, column=17).value

                        if estatus_activo.lower() == "activo":
                            estatus_activo = True
                        else:
                            estatus_activo = False

                        try:
                            if nuevo.lower() == "nuevo":
                                nuevo = True
                            else:
                                nuevo = False
                        except:
                            nuevo = False

                        try:
                            presion_establecida_1 = int(presion_establecida_1)
                        except:
                            presion_establecida_1 = None

                        try:
                            presion_establecida_2 = int(presion_establecida_2)
                        except:
                            presion_establecida_2 = None

                        try:
                            presion_establecida_3 = int(presion_establecida_3)
                        except:
                            presion_establecida_3 = None

                        try:
                            presion_establecida_4 = int(presion_establecida_4)
                        except:
                            presion_establecida_4 = None

                        try:
                            presion_establecida_5 = int(presion_establecida_5)
                        except:
                            presion_establecida_5 = None

                        try:
                            presion_establecida_6 = int(presion_establecida_6)
                        except:
                            presion_establecida_6 = None

                        try:
                            presion_establecida_7 = int(presion_establecida_7)
                        except:
                            presion_establecida_7 = None


                        numero_de_llantas = functions.cantidad_llantas(configuracion)
                        
                        if self.request.user.groups.filter(name__in=['Aeto Manager']):
                            ubicacion = Ubicacion.objects.get(nombre=ubicacion)

                            aplicacion = Aplicacion.objects.get(nombre=aplicacion)

                            compania = ubicacion.compania
                            
                        else:   
                            ubicacion = Ubicacion.objects.get(nombre=ubicacion, compania=compania)

                            aplicacion = Aplicacion.objects.get(nombre=aplicacion, compania=compania)
                            

                        vehiculo_nuevo = Vehiculo.objects.create(numero_economico=numero_economico,
                                            modelo=modelo,
                                            marca=marca,
                                            compania=compania,
                                            ubicacion=ubicacion,
                                            aplicacion=aplicacion,
                                            numero_de_llantas=numero_de_llantas,
                                            clase=clase.upper(),
                                            configuracion=configuracion,
                                            presion_establecida_1=presion_establecida_1,
                                            presion_establecida_2=presion_establecida_2,
                                            presion_establecida_3=presion_establecida_3,
                                            presion_establecida_4=presion_establecida_4,
                                            presion_establecida_5=presion_establecida_5,
                                            presion_establecida_6=presion_establecida_6,
                                            presion_establecida_7=presion_establecida_7,
                                            km_diario_maximo=int(km_diario_maximo),
                                            estatus_activo=estatus_activo,
                                            nuevo=nuevo
                                            )
                        if if_vehiculo == "Algunos":
                            if_vehiculo = "Algunos"
                        else:
                            if_vehiculo = "Todos"
                        #! No eliminar, esta funcion crea las llantas
                        functions.crear_llantas_nuevos_vehiculos(vehiculo_nuevo)
                except:
                    if if_vehiculo == "Todos" and numero_economico != None:
                        if_vehiculo = "Algunos"
                    pass
            fp.close()
            wb_obj.close()
            os.remove(os.path.abspath(archivo))
        elif self.request.method=='POST' and self.request.FILES.get("file2"):
            try:
                file = self.request.FILES.get("file2")
                archivo = dir + r"\files.xlsx"
                fp = open(archivo,'wb')
                for chunk in file.chunks():
                    fp.write(chunk)
    
                print("archivo", archivo)
                wb_obj = openpyxl.load_workbook(file)
                sheet_obj = wb_obj.active
                print(sheet_obj.max_row)
                print()
                for i in range(sheet_obj.max_row):
                    try:
                        numero_economico = str(sheet_obj.cell(row=i + 2, column=1).value)
                        numero_economico_nuevo = str(sheet_obj.cell(row=i + 2, column=2).value)
                        vehiculo = str(sheet_obj.cell(row=i + 2, column=3).value)
                        ubicacion = str(sheet_obj.cell(row=i + 2, column=4).value)
                        aplicacion = str(sheet_obj.cell(row=i + 2, column=5).value)
                        vida = str(sheet_obj.cell(row=i + 2, column=6).value)
                        posicion = str(sheet_obj.cell(row=i + 2, column=7).value)
                        presion_actual = str(sheet_obj.cell(row=i + 2, column=8).value)
                        profundidad_izquierda = str(sheet_obj.cell(row=i + 2, column=9).value)
                        profundidad_central = str(sheet_obj.cell(row=i + 2, column=10).value)
                        profundidad_derecha = str(sheet_obj.cell(row=i + 2, column=11).value)
                        km_actual = str(sheet_obj.cell(row=i + 2, column=12).value)
                        producto = str(sheet_obj.cell(row=i + 2, column=13).value)
                        inventario = str(sheet_obj.cell(row=i + 2, column=14).value)
                        km_montado = str(sheet_obj.cell(row=i + 2, column=15).value)                            

                        vehiculo = Vehiculo.objects.get(numero_economico=vehiculo, compania=compania)
                        ubicacion = Ubicacion.objects.get(nombre=ubicacion, compania=compania)
                        aplicacion = Aplicacion.objects.get(nombre=aplicacion, compania=compania)
                        producto = Producto.objects.get(producto=producto, compania=compania)
                        inventario = "Rodante"

                        try:
                            presion_actual = int(presion_actual)
                        except:
                            presion_actual = None

                        try:
                            profundidad_izquierda = float(profundidad_izquierda)
                            if producto.profundidad_inicial:
                                if 0 < profundidad_izquierda <= producto.profundidad_inicial:
                                    pass
                                else:
                                    profundidad_izquierda = producto.profundidad_inicial
                        except:
                            profundidad_izquierda = producto.profundidad_inicial

                        try:
                            profundidad_central = float(profundidad_central)
                            if producto.profundidad_inicial:
                                if 0 < profundidad_central <= producto.profundidad_inicial:
                                    pass
                                else:
                                    profundidad_central = producto.profundidad_inicial
                        except:
                            profundidad_central = producto.profundidad_inicial

                        try:
                            profundidad_derecha = float(profundidad_derecha)
                            if producto.profundidad_inicial:
                                if 0 < profundidad_derecha <= producto.profundidad_inicial:
                                    pass
                                else:
                                    profundidad_derecha = producto.profundidad_inicial
                        except:
                            profundidad_derecha = producto.profundidad_inicial

                        try:
                            km_actual = int(km_actual)
                        except:
                            km_actual = None

                        try:
                            km_montado = int(km_montado)
                            if km_montado > vehiculo.km:
                                km_montado = int(vehiculo.km)
                        except:
                            km_montado = None


                        llanta_id = Llanta.objects.get(numero_economico=numero_economico, compania=Compania.objects.get(compania=compania))
                        if numero_economico_nuevo:
                            
                            if numero_economico_nuevo != "None":
                                llanta_id.numero_economico=numero_economico_nuevo
                                llanta_id.patito = False
                                
                        llanta_id.ubicacion=ubicacion
                        llanta_id.aplicacion=aplicacion
                        llanta_id.vida=vida
                        llanta_id.presion_actual=presion_actual
                        llanta_id.profundidad_izquierda=profundidad_izquierda
                        llanta_id.profundidad_central=profundidad_central
                        llanta_id.profundidad_derecha=profundidad_derecha
                        llanta_id.km_actual=km_actual
                        llanta_id.producto=producto
                        llanta_id.inventario=inventario
                        llanta_id.fecha_de_entrada_inventario=date.today()
                        llanta_id.km_montado=km_montado

                        llanta_id.save()
                    except:
                        pass
            except Exception as ex:
                print(ex)
                    
        context["if_vehiculo"] = if_vehiculo
        context["user"] = user
        context["groups_names"] = groups_names
        return context



class SearchView(LoginRequiredMixin, ListView):
    # Vista del dashboard buscar_vehiculosf
    template_name = "buscar_vehiculos.html"
    model = Vehiculo
    ordering = ("-fecha_de_creacion")

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)
        vehiculos = Vehiculo.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania), tirecheck=False)
        bitacora = Bitacora.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        bitacora_pro = Bitacora_Pro.objects.filter(compania=Compania.objects.get(compania=self.request.user.perfil.compania))
        llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=self.request.user.perfil.compania), tirecheck=False)
        inspecciones = Inspeccion.objects.filter(llanta__vehiculo__compania=Compania.objects.get(compania=self.request.user.perfil.compania), llanta__tirecheck=False)
        
        filtro_sospechoso = functions.vehiculo_sospechoso(inspecciones)
        vehiculos_sospechosos = vehiculos.filter(id__in=filtro_sospechoso)

        doble_mala_entrada = functions.doble_mala_entrada(bitacora, vehiculos)
        filtro_rojo = functions.vehiculo_rojo(llantas, doble_mala_entrada, vehiculos)
        doble_mala_entrada_pro = functions.doble_mala_entrada_pro(bitacora_pro, vehiculos)
        filtro_rojo_pro = functions.vehiculo_rojo(llantas, doble_mala_entrada_pro, vehiculos)
        vehiculos_rojos = vehiculos.filter(id__in=filtro_rojo).exclude(id__in=vehiculos_sospechosos) | vehiculos.filter(id__in=filtro_rojo_pro).exclude(id__in=vehiculos_sospechosos)

        filtro_amarillo = functions.vehiculo_amarillo(llantas)
        vehiculos_amarillos = vehiculos.filter(id__in=filtro_amarillo).exclude(id__in=vehiculos_rojos).exclude(id__in=vehiculos_sospechosos)

        vehiculos_verdes = vehiculos.exclude(id__in=vehiculos_rojos).exclude(id__in=vehiculos_sospechosos).exclude(id__in=vehiculos_amarillos)

        context["vehiculos_amarillos"] = vehiculos_amarillos
        context["vehiculos_rojos"] = vehiculos_rojos
        context["vehiculos_sospechosos"] = vehiculos_sospechosos
        context["vehiculos_verdes"] = vehiculos_verdes
        context["cantidad_amarillos"] = vehiculos_amarillos.count()
        context["cantidad_rojos"] = vehiculos_rojos.count()
        context["cantidad_sospechosos"] = vehiculos_sospechosos.count()
        context["cantidad_total"] = vehiculos.count()
        context["cantidad_verdes"] = vehiculos_verdes.count()

        return context

def search(request):
    num = request.GET.get("numero_economico")
    fecha = request.GET.get("fecha1")
    print(fecha)
    if num:
        vehiculos = Vehiculo.objects.filter(numero_economico__icontains=num, compania=Compania.objects.get(compania=request.user.perfil.compania), tirecheck=False)
        bitacora = Bitacora.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania))
        bitacora_pro = Bitacora_Pro.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania))
        llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=request.user.perfil.compania), tirecheck=False)
        inspecciones = Inspeccion.objects.filter(llanta__vehiculo__compania=Compania.objects.get(compania=request.user.perfil.compania), llanta__tirecheck=False)
        
        filtro_sospechoso = functions.vehiculo_sospechoso(inspecciones)
        vehiculos_sospechosos = vehiculos.filter(id__in=filtro_sospechoso)

        doble_mala_entrada = functions.doble_mala_entrada(bitacora, vehiculos)
        filtro_rojo = functions.vehiculo_rojo(llantas, doble_mala_entrada, vehiculos)
        doble_mala_entrada_pro = functions.doble_mala_entrada_pro(bitacora_pro, vehiculos)
        filtro_rojo_pro = functions.vehiculo_rojo(llantas, doble_mala_entrada_pro, vehiculos)
        vehiculos_rojos = vehiculos.filter(id__in=filtro_rojo).exclude(id__in=vehiculos_sospechosos) | vehiculos.filter(id__in=filtro_rojo_pro).exclude(id__in=vehiculos_sospechosos)

        filtro_amarillo = functions.vehiculo_amarillo(llantas)
        vehiculos_amarillos = vehiculos.filter(id__in=filtro_amarillo).exclude(id__in=vehiculos_rojos).exclude(id__in=vehiculos_sospechosos)

        vehiculos_verdes = vehiculos.exclude(id__in=vehiculos_rojos).exclude(id__in=vehiculos_sospechosos).exclude(id__in=vehiculos_amarillos)
        return render(request, "buscar_vehiculos.html", {
                                                "vehiculos_amarillos": vehiculos_amarillos,
                                                "vehiculos_rojos": vehiculos_rojos,
                                                "vehiculos_sospechosos": vehiculos_sospechosos,
                                                "vehiculos_verdes": vehiculos_verdes,
                                                "cantidad_amarillos": vehiculos_amarillos.count(),
                                                "cantidad_rojos": vehiculos_rojos.count(),
                                                "cantidad_sospechosos": vehiculos_sospechosos.count(),
                                                "cantidad_total": vehiculos.count(),
                                                "cantidad_verdes": vehiculos_verdes.count()
        })
    elif fecha and fecha != "Seleccionar Fecha":
        dividir_fecha = functions.convertir_rango2(fecha)
        primera_fecha = datetime.strptime(dividir_fecha[0], "%m/%d/%Y").date()
        segunda_fecha = datetime.strptime(dividir_fecha[1], "%m/%d/%Y").date()

        vehiculos = Vehiculo.objects.filter(fecha_de_inflado__range=[primera_fecha, segunda_fecha], compania=Compania.objects.get(compania=request.user.perfil.compania), tirecheck=False) | Vehiculo.objects.filter(ultima_bitacora_pro__fecha_de_inflado__range=[primera_fecha, segunda_fecha], compania=Compania.objects.get(compania=request.user.perfil.compania), tirecheck=False)
        bitacora = Bitacora.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania))
        bitacora_pro = Bitacora_Pro.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania))
        llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=request.user.perfil.compania), tirecheck=False)
        inspecciones = Inspeccion.objects.filter(llanta__vehiculo__compania=Compania.objects.get(compania=request.user.perfil.compania), llanta__tirecheck=False)
        
        filtro_sospechoso = functions.vehiculo_sospechoso(inspecciones)
        vehiculos_sospechosos = vehiculos.filter(id__in=filtro_sospechoso)

        doble_mala_entrada = functions.doble_mala_entrada(bitacora, vehiculos)
        filtro_rojo = functions.vehiculo_rojo(llantas, doble_mala_entrada, vehiculos)
        doble_mala_entrada_pro = functions.doble_mala_entrada_pro(bitacora_pro, vehiculos)
        filtro_rojo_pro = functions.vehiculo_rojo(llantas, doble_mala_entrada_pro, vehiculos)
        vehiculos_rojos = vehiculos.filter(id__in=filtro_rojo).exclude(id__in=vehiculos_sospechosos) | vehiculos.filter(id__in=filtro_rojo_pro).exclude(id__in=vehiculos_sospechosos)


        filtro_amarillo = functions.vehiculo_amarillo(llantas)
        vehiculos_amarillos = vehiculos.filter(id__in=filtro_amarillo).exclude(id__in=vehiculos_rojos).exclude(id__in=vehiculos_sospechosos)

        vehiculos_verdes = vehiculos.exclude(id__in=vehiculos_rojos).exclude(id__in=vehiculos_sospechosos).exclude(id__in=vehiculos_amarillos)

        return render(request, "buscar_vehiculos.html", {
                                                "vehiculos_amarillos": vehiculos_amarillos,
                                                "vehiculos_rojos": vehiculos_rojos,
                                                "vehiculos_sospechosos": vehiculos_sospechosos,
                                                "vehiculos_verdes": vehiculos_verdes,
                                                "cantidad_amarillos": vehiculos_amarillos.count(),
                                                "cantidad_rojos": vehiculos_rojos.count(),
                                                "cantidad_sospechosos": vehiculos_sospechosos.count(),
                                                "cantidad_total": vehiculos.count(),
                                                "cantidad_verdes": vehiculos_verdes.count(),
        })
    else:

        vehiculos = Vehiculo.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania), tirecheck=False)
        bitacora = Bitacora.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania))
        bitacora_pro = Bitacora_Pro.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania))
        llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=request.user.perfil.compania), tirecheck=False)
        inspecciones = Inspeccion.objects.filter(llanta__vehiculo__compania=Compania.objects.get(compania=request.user.perfil.compania), llanta__tirecheck=False)
        
        filtro_sospechoso = functions.vehiculo_sospechoso(inspecciones)
        vehiculos_sospechosos = vehiculos.filter(id__in=filtro_sospechoso)

        doble_mala_entrada = functions.doble_mala_entrada(bitacora, vehiculos)
        filtro_rojo = functions.vehiculo_rojo(llantas, doble_mala_entrada, vehiculos)
        doble_mala_entrada_pro = functions.doble_mala_entrada_pro(bitacora_pro, vehiculos)
        filtro_rojo_pro = functions.vehiculo_rojo(llantas, doble_mala_entrada_pro, vehiculos)
        vehiculos_rojos = vehiculos.filter(id__in=filtro_rojo).exclude(id__in=vehiculos_sospechosos) | vehiculos.filter(id__in=filtro_rojo_pro).exclude(id__in=vehiculos_sospechosos)

        filtro_amarillo = functions.vehiculo_amarillo(llantas)
        vehiculos_amarillos = vehiculos.filter(id__in=filtro_amarillo).exclude(id__in=vehiculos_rojos).exclude(id__in=vehiculos_sospechosos)

        vehiculos_verdes = vehiculos.exclude(id__in=vehiculos_rojos).exclude(id__in=vehiculos_sospechosos).exclude(id__in=vehiculos_amarillos)

        return render(request, "buscar_vehiculos.html", {
                                                "vehiculos_amarillos": vehiculos_amarillos,
                                                "vehiculos_rojos": vehiculos_rojos,
                                                "vehiculos_sospechosos": vehiculos_sospechosos,
                                                "vehiculos_verdes": vehiculos_verdes,
                                                "cantidad_amarillos": vehiculos_amarillos.count(),
                                                "cantidad_rojos": vehiculos_rojos.count(),
                                                "cantidad_sospechosos": vehiculos_sospechosos.count(),
                                                "cantidad_total": vehiculos.count(),
                                                "cantidad_verdes": vehiculos_verdes.count(),
        })

class tireDetailView(LoginRequiredMixin, DetailView):
    # Vista de tireDetailView
    
    template_name = "tireDetail.html"
    slug_field = "llanta"
    slug_url_kwarg = "llanta"
    queryset = Llanta.objects.all()
    context_object_name = "llanta"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        llanta = self.get_object()
        inspecciones_llanta = Inspeccion.objects.filter(llanta=llanta)
        vehiculo = llanta.vehiculo
        if llanta.inventario == "Rodante":
            llantas = Llanta.objects.filter(vehiculo=vehiculo)
            context["vehiculo__boolean"] = True
        else:
            llantas = Llanta.objects.filter(numero_economico=self.get_object())
            context["vehiculo__boolean"] = False

        inspecciones = Inspeccion.objects.filter(llanta__in=llantas)

        if llanta.inventario == "Rodante":
            try:
                bitacora = Bitacora.objects.filter(
                    #vehiculo=Vehiculo.objects.get(numero_economico=vehiculo.numero_economico), 
                    compania=Compania.objects.get(compania=self.request.user.perfil.compania),
                    llantas__in = [llanta]                    
                    )
            except:
                bitacora = None
            try:
                bitacora_normal = Bitacora.objects.filter(
                    #vehiculo=Vehiculo.objects.get(numero_economico=vehiculo.numero_economico), 
                    compania=Compania.objects.get(compania=self.request.user.perfil.compania),
                    llantas__in = [llanta]   
                    ).order_by("id")
                bitacora_pro = Bitacora_Pro.objects.filter(
                    #vehiculo=Vehiculo.objects.get(numero_economico=vehiculo.numero_economico), 
                    compania=Compania.objects.get(compania=self.request.user.perfil.compania),
                    llantas__in = [llanta]   
                    ).order_by("id")
            except:
                bitacora_normal = None 
                bitacora_pro = None

            bitacora = Bitacora.objects.filter(vehiculo=Vehiculo.objects.get(numero_economico=vehiculo.numero_economico, compania=Compania.objects.get(compania=self.request.user.perfil.compania)), compania=Compania.objects.get(compania=self.request.user.perfil.compania)).order_by("id")
            bitacora_pro = Bitacora_Pro.objects.filter(vehiculo=Vehiculo.objects.get(numero_economico=vehiculo.numero_economico, compania=Compania.objects.get(compania=self.request.user.perfil.compania)), compania=Compania.objects.get(compania=self.request.user.perfil.compania)).order_by("id")

            bitacoras = bitacora.annotate(presion_de_entrada_1=F("presion_de_entrada"), presion_de_salida_1=F("presion_de_salida"), presion_de_entrada_2=Value(None, output_field=IntegerField()), presion_de_salida_2=Value(None, output_field=IntegerField()), presion_de_entrada_3=Value(None, output_field=IntegerField()), presion_de_salida_3=Value(None, output_field=IntegerField()), presion_de_entrada_4=Value(None, output_field=IntegerField()), presion_de_salida_4=Value(None, output_field=IntegerField()), presion_de_entrada_5=Value(None, output_field=IntegerField()), presion_de_salida_5=Value(None, output_field=IntegerField()), presion_de_entrada_6=Value(None, output_field=IntegerField()), presion_de_salida_6=Value(None, output_field=IntegerField()), presion_de_entrada_7=Value(None, output_field=IntegerField()), presion_de_salida_7=Value(None, output_field=IntegerField()), presion_de_entrada_8=Value(None, output_field=IntegerField()), presion_de_salida_8=Value(None, output_field=IntegerField()), presion_de_entrada_9=Value(None, output_field=IntegerField()), presion_de_salida_9=Value(None, output_field=IntegerField()), presion_de_entrada_10=Value(None, output_field=IntegerField()), presion_de_salida_10=Value(None, output_field=IntegerField()), presion_de_entrada_11=Value(None, output_field=IntegerField()), presion_de_salida_11=Value(None, output_field=IntegerField()), presion_de_entrada_12=Value(None, output_field=IntegerField()), presion_de_salida_12=Value(None, output_field=IntegerField())).order_by("id")
            bitacoras = list(bitacoras.values("id", "vehiculo__id", "vehiculo__configuracion", "compania__id", "fecha_de_inflado", "tiempo_de_inflado", "presion_de_entrada_1", "presion_de_salida_1", "presion_de_entrada_2", "presion_de_salida_2", "presion_de_entrada_3", "presion_de_salida_3", "presion_de_entrada_4", "presion_de_salida_4", "presion_de_entrada_5", "presion_de_salida_5", "presion_de_entrada_6", "presion_de_salida_6", "presion_de_entrada_7", "presion_de_salida_7", "presion_de_entrada_8", "presion_de_salida_8", "presion_de_entrada_9", "presion_de_salida_9", "presion_de_entrada_10", "presion_de_salida_10", "presion_de_entrada_11", "presion_de_salida_11", "presion_de_entrada_12", "presion_de_salida_12", "vehiculo__presion_establecida_1", "vehiculo__presion_establecida_2", "vehiculo__presion_establecida_3", "vehiculo__presion_establecida_4", "vehiculo__presion_establecida_5", "vehiculo__presion_establecida_6", "vehiculo__presion_establecida_7"))

            bitacoras_pro = list(bitacora_pro.values("id", "vehiculo__id", "vehiculo__configuracion", "compania__id", "fecha_de_inflado", "tiempo_de_inflado", "presion_de_entrada_1", "presion_de_salida_1", "presion_de_entrada_2", "presion_de_salida_2", "presion_de_entrada_3", "presion_de_salida_3", "presion_de_entrada_4", "presion_de_salida_4", "presion_de_entrada_5", "presion_de_salida_5", "presion_de_entrada_6", "presion_de_salida_6", "presion_de_entrada_7", "presion_de_salida_7", "presion_de_entrada_8", "presion_de_salida_8", "presion_de_entrada_9", "presion_de_salida_9", "presion_de_entrada_10", "presion_de_salida_10", "presion_de_entrada_11", "presion_de_salida_11", "presion_de_entrada_12", "presion_de_salida_12", "vehiculo__presion_establecida_1", "vehiculo__presion_establecida_2", "vehiculo__presion_establecida_3", "vehiculo__presion_establecida_4", "vehiculo__presion_establecida_5", "vehiculo__presion_establecida_6", "vehiculo__presion_establecida_7"))
        
            bitacoras.extend(bitacoras_pro)

            bitacoras = sorted(bitacoras, key=lambda x:x["fecha_de_inflado"], reverse=False)

            entradas_correctas = functions.entrada_correcta_ambas(bitacoras)

            print(f'bitacora_normal: {bitacora_normal}')
            print(f'bitacora_pro: {bitacora_pro}')
            eventos = []
            for bit in bitacora_normal:
                eventos.append([bit, 'pulpo'])
            for bit in bitacora_pro:
                eventos.append([bit, 'pulpopro'])
            print(f'Eventos: {eventos}')
        
        hoy = date.today()
        mes_1 = hoy.strftime("%b")
        mes_2 = functions.mes_anterior(hoy)
        mes_3 = functions.mes_anterior(mes_2)
        mes_4 = functions.mes_anterior(mes_3)
        mes_5 = functions.mes_anterior(mes_4)
        mes_6 = functions.mes_anterior(mes_5)
        mes_7 = functions.mes_anterior(mes_6)
        mes_8 = functions.mes_anterior(mes_7)

        hoy1 = hoy.strftime("%m")
        hoy2 = mes_2.strftime("%m")
        hoy3 = mes_3.strftime("%m")
        hoy4 = mes_4.strftime("%m")
        hoy5 = mes_5.strftime("%m")
        hoy6 = mes_6.strftime("%m")
        hoy7 = mes_7.strftime("%m")
        hoy8 = mes_8.strftime("%m")


        if llanta.inventario == "Rodante":
            color = functions.entrada_correcta_actual(vehiculo)
            vehiculo_mes1 = bitacora.filter(fecha_de_inflado__month=hoy1)
            vehiculo_mes2 = bitacora.filter(fecha_de_inflado__month=hoy2)
            vehiculo_mes3 = bitacora.filter(fecha_de_inflado__month=hoy3)
            vehiculo_mes4 = bitacora.filter(fecha_de_inflado__month=hoy4)
            vehiculo_mes5 = bitacora.filter(fecha_de_inflado__month=hoy5)
            vehiculo_mes6 = bitacora.filter(fecha_de_inflado__month=hoy6)
            vehiculo_mes7 = bitacora.filter(fecha_de_inflado__month=hoy7)
            vehiculo_mes8 = bitacora.filter(fecha_de_inflado__month=hoy8)

            vehiculo_pro_mes1 = bitacora_pro.filter(fecha_de_inflado__month=hoy1)
            vehiculo_pro_mes2 = bitacora_pro.filter(fecha_de_inflado__month=hoy2)
            vehiculo_pro_mes3 = bitacora_pro.filter(fecha_de_inflado__month=hoy3)
            vehiculo_pro_mes4 = bitacora_pro.filter(fecha_de_inflado__month=hoy4)
            vehiculo_pro_mes5 = bitacora_pro.filter(fecha_de_inflado__month=hoy5)
            vehiculo_pro_mes6 = bitacora_pro.filter(fecha_de_inflado__month=hoy6)
            vehiculo_pro_mes7 = bitacora_pro.filter(fecha_de_inflado__month=hoy7)
            vehiculo_pro_mes8 = bitacora_pro.filter(fecha_de_inflado__month=hoy8)

            mala_entrada_contar_mes1 = functions.contar_mala_entrada(vehiculo_mes1)
            mala_entrada_contar_mes2 = functions.contar_mala_entrada(vehiculo_mes2)
            mala_entrada_contar_mes3 = functions.contar_mala_entrada(vehiculo_mes3)
            mala_entrada_contar_mes4 = functions.contar_mala_entrada(vehiculo_mes4)
            mala_entrada_contar_mes5 = functions.contar_mala_entrada(vehiculo_mes5)
            mala_entrada_contar_mes6 = functions.contar_mala_entrada(vehiculo_mes6)
            mala_entrada_contar_mes7 = functions.contar_mala_entrada(vehiculo_mes7)
            mala_entrada_contar_mes8 = functions.contar_mala_entrada(vehiculo_mes8)

            mala_entrada_contar_mes1 += functions.contar_mala_entrada_pro(vehiculo_pro_mes1)
            mala_entrada_contar_mes2 += functions.contar_mala_entrada_pro(vehiculo_pro_mes2)
            mala_entrada_contar_mes3 += functions.contar_mala_entrada_pro(vehiculo_pro_mes3)
            mala_entrada_contar_mes4 += functions.contar_mala_entrada_pro(vehiculo_pro_mes4)
            mala_entrada_contar_mes5 += functions.contar_mala_entrada_pro(vehiculo_pro_mes5)
            mala_entrada_contar_mes6 += functions.contar_mala_entrada_pro(vehiculo_pro_mes6)
            mala_entrada_contar_mes7 += functions.contar_mala_entrada_pro(vehiculo_pro_mes7)
            mala_entrada_contar_mes8 += functions.contar_mala_entrada_pro(vehiculo_pro_mes8)

            doble_entrada = functions.doble_entrada(bitacora, bitacora_pro)
            doble_mala_entrada = functions.doble_mala_entrada(bitacora, vehiculo)

            doble_mala_entrada_pro = functions.doble_mala_entrada_pro(bitacora_pro, vehiculo)

        filtro_sospechoso = functions.vehiculo_sospechoso_llanta(inspecciones)
        llantas_sospechosas = llantas.filter(numero_economico__in=filtro_sospechoso)
        print('----------')
                
        filtro_rojo = functions.vehiculo_rojo_llanta(llantas)
        llantas_rojas = llantas.filter(numero_economico__in=filtro_rojo).exclude(id__in=llantas_sospechosas)

        filtro_amarillo = functions.vehiculo_amarillo_llanta(llantas)
        llantas_amarillas = llantas.filter(numero_economico__in=filtro_amarillo).exclude(id__in=llantas_sospechosas).exclude(id__in=llantas_rojas)

        llantas_azules = llantas.exclude(id__in=llantas_sospechosas).exclude(id__in=llantas_rojas).exclude(id__in=llantas_amarillas)
        
        if llanta in llantas_sospechosas:
            color_profundidad = 'purple'
        elif llanta in llantas_rojas:
            color_profundidad = 'bad'
        elif llanta in llantas_amarillas:
            color_profundidad = 'yellow'
        elif llanta in llantas_azules:
            color_profundidad = 'good'

        message = None
        message_pro = {}
        if llanta.inventario == "Rodante":
            try:
                if doble_mala_entrada:
                    message = "Doble mala entrada"
                    color = "bad"
                elif doble_mala_entrada_pro:
                    for entrada in doble_mala_entrada_pro:
                        message_pro[entrada[-1]] = "Doble mala entrada"
                    color = "bad"
                elif color == "bad":
                    message = "Baja presión"
            except:
                pass

        problemas_abiertos = []
        if message:
            problemas_abiertos.append([llanta, message])
        if color_profundidad == "bad":
            problemas_abiertos.append([llanta, "Profundidad abajo del punto de retiro"])
        if color_profundidad == "yellow":
            problemas_abiertos.append([llanta, "Profundidad en el punto de retiro"])

        regresion_llanta = functions.km_proyectado_llanta(inspecciones_llanta)
        try:
            km_proyectado = round(regresion_llanta[0])
            km_x_mm = round(regresion_llanta[1])
            cpk = round(regresion_llanta[2], 3)
        except:
            km_proyectado = 0
            km_x_mm = 0
            cpk = 0
        try:
            comportamiento_de_desgaste = functions.comportamiento_de_desgaste(inspecciones_llanta)
        except:
            pass
        desgaste_mensual = functions.desgaste_mensual(inspecciones_llanta)

        if llanta.inventario == "Rodante":
            context["bitacoras"] = bitacora
            context["cantidad_doble_entrada_mes1"] = doble_entrada[2]["mes1"] + doble_entrada[3]["mes1"]
            context["cantidad_doble_entrada_mes2"] = doble_entrada[2]["mes2"] + doble_entrada[3]["mes2"]
            context["cantidad_doble_entrada_mes3"] = doble_entrada[2]["mes3"] + doble_entrada[3]["mes3"]
            context["cantidad_doble_entrada_mes4"] = doble_entrada[2]["mes4"] + doble_entrada[3]["mes4"]
            context["cantidad_doble_entrada_mes5"] = doble_entrada[2]["mes5"] + doble_entrada[3]["mes5"]
            context["cantidad_doble_entrada_mes6"] = doble_entrada[2]["mes6"] + doble_entrada[3]["mes6"]
            context["cantidad_doble_entrada_mes7"] = doble_entrada[2]["mes7"] + doble_entrada[3]["mes7"]
            context["cantidad_doble_entrada_mes8"] = doble_entrada[2]["mes8"] + doble_entrada[3]["mes8"]
            context["cantidad_entrada_mes1"] = mala_entrada_contar_mes1
            context["cantidad_entrada_mes2"] = mala_entrada_contar_mes2
            context["cantidad_entrada_mes3"] = mala_entrada_contar_mes3
            context["cantidad_entrada_mes4"] = mala_entrada_contar_mes4
            context["cantidad_entrada_mes5"] = mala_entrada_contar_mes5
            context["cantidad_entrada_mes6"] = mala_entrada_contar_mes6
            context["cantidad_entrada_mes7"] = mala_entrada_contar_mes7
            context["cantidad_entrada_mes8"] = mala_entrada_contar_mes8
            context["color"] = color
        context["color_profundidad"] = color_profundidad
        context["cpk"] = cpk
        try:
            context["desgastes"] = comportamiento_de_desgaste
        except:
            pass
        context["desgaste_mensual"] = desgaste_mensual
        context["hoy"] = hoy
        context["inspecciones"] = inspecciones_llanta
        context["km_proyectado"] = km_proyectado
        context["km_x_mm"] = km_x_mm
        context["mes_1"] = mes_1
        context["mes_2"] = mes_2.strftime("%b")
        context["mes_3"] = mes_3.strftime("%b")
        context["mes_4"] = mes_4.strftime("%b")
        context["mes_5"] = mes_5.strftime("%b")
        context["mes_6"] = mes_6.strftime("%b")
        context["mes_7"] = mes_7.strftime("%b")
        context["mes_8"] = mes_8.strftime("%b")
        context["message"] = message
        context["problemas_abiertos"] = problemas_abiertos
        if llanta.inventario == "Rodante":
            context["vehiculo"] = vehiculo
            context["entradas"] = entradas_correctas
        if llanta.inventario == "Rodante":
            context["vehiculo_mes1"] = vehiculo_mes1.count() + vehiculo_pro_mes1.count()
            context["vehiculo_mes2"] = vehiculo_mes2.count() + vehiculo_pro_mes2.count()
            context["vehiculo_mes3"] = vehiculo_mes3.count() + vehiculo_pro_mes3.count()
            context["vehiculo_mes4"] = vehiculo_mes4.count() + vehiculo_pro_mes4.count()
            context["vehiculo_mes5"] = vehiculo_mes5.count() + vehiculo_pro_mes5.count()
            context["vehiculo_mes6"] = vehiculo_mes6.count() + vehiculo_pro_mes6.count()
            context["vehiculo_mes7"] = vehiculo_mes7.count() + vehiculo_pro_mes7.count()
            context["vehiculo_mes8"] = vehiculo_mes8.count() + vehiculo_pro_mes8.count()
        
        llanta_actual = llanta
        
        
        if llanta_actual.inventario == 'Rodante':
            color_presion = functions.check_color_presion(llanta_actual)      
        else:
            color_presion = 'good'
        
        color_medio = functions.color_observaciones_all_one(llanta_actual)
        color_profundidad = functions.check_color_profundidad(llanta_actual)   
        
        
        
        context['color_presion'] = color_presion
        context['color_medio'] = color_medio
        context['color_profundidad'] = color_profundidad
        context['llanta'] = llanta
        context['min_profundidad'] = functions.min_profundidad(llanta)
              
        context['llanta_info'] = 0
        context['numero_eje_check'] = 0
        context['numero_llanta_check'] = 0
        
        #Problemas abiertos Observaciones
        observaciones = llanta_actual.observaciones.all()
        context['observaciones'] = observaciones

        if llanta.inventario != "Rodante":
            presion_act = llanta.presion_actual
            min_produndidad = llanta.producto.profundidad_inicial
            color_prof = "good"
            llanta_info = [llanta, color_prof, "good", min_produndidad, None, None, None, None]
            context["llanta_info"] = llanta_info

        problemas = []
        ultima_inspeccion = llanta_actual.ultima_inspeccion

        for observacion in llanta_actual.observaciones.all():
            color_obs = functions.color_observaciones_one(observacion)
            if color_obs == 'bad':
                signo = 'icon-cross bad-text'
            elif color_obs == 'yellow':
                signo = 'icon-warning  yellow-text'
            else:
                signo = 'icon-checkmark good-text'
            problemas.append({
                'posicion': llanta_actual.posicion, 
                'observacion': observacion, 
                'signo': signo
                })
        
        context['problemas'] = problemas
        #Bitacorasllanta
        if llanta.inventario == "Rodante":
            inpecciones_llanta_act = Inspeccion.objects.filter(llanta = llanta_actual)
            eventos = []
            inspecciones_list = []
            servicios = ServicioLlanta.objects.filter(llanta = llanta_actual).order_by('-id')
            for bit in bitacora:
                signo = ''
                eventos.append([bit.fecha_de_inflado, bit, 'pulpo', signo])
            for bit in bitacora_pro:
                signo = ''
                eventos.append([bit.fecha_de_inflado, bit, 'pulpopro', signo])
                
            for inspeccion in inpecciones_llanta_act:
                color_insp = ''
                if color_insp == 'bad':
                    signo = 'icon-cross bad-text'
                elif color_insp == 'yellow':
                    signo = 'icon-warning  yellow-text'
                else:
                    signo = 'icon-checkmark good-text'
                print(inspeccion)
                if functions.exist_edicion_manual_one(inspeccion):
                    inspecciones_list.append([inspeccion.fecha_hora, inspeccion, 'edicion', signo])
                inspecciones_list.append([inspeccion.fecha_hora, inspeccion, 'inspeccion', signo])
            inspecciones_list = sorted(inspecciones_list, key=lambda x:x[1].id, reverse=True)
            for ins in inspecciones_list:
                eventos.append([ins[0], ins[1], ins[2], ins[3]])
            for servicio in servicios:
                eventos.append([datetime.combine(servicio.serviciovehiculo.fecha_inicio, datetime.min.time(), timezone.utc)  , servicio, 'servicio', 'icon-checkmark good-text'])
            eventos = sorted(eventos, key=lambda x:x[0], reverse=True)
            print('----------------0')
            for evento in eventos:
                print(evento)
            context['eventos'] = eventos
            
        if llanta.inventario == 'Desecho final':
            desechos = OrdenDesecho.objects.filter(llanta = llanta)
            context['desechos'] = desechos
        return context



class DetailView(LoginRequiredMixin, DetailView):
    # Vista del dashboard detail
    template_name = "detail.html"
    slug_field = "vehiculo"
    slug_url_kwarg = "vehiculo"
    queryset = Vehiculo.objects.all()
    context_object_name = "vehiculo"

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)
        vehiculo = Vehiculo.objects.get(pk=self.kwargs['pk'])
        llantas = Llanta.objects.filter(vehiculo=vehiculo, tirecheck=False, inventario = 'Rodante')
        llantas_ordenadas = functions.llantas_ordenadas(llantas)
        inspecciones = Inspeccion.objects.filter(llanta__in=llantas)
        inspecciones_vehiculo = InspeccionVehiculo.objects.filter(vehiculo=vehiculo)
        ultima_inspeccion_img = []
        for llanta in llantas:
            if llanta.ultima_inspeccion != None:
                if llanta.ultima_inspeccion.imagen :
                    print(llanta.ultima_inspeccion.imagen)
                    ultima_inspeccion_img.append(llanta)
        
        context['ultima_inspeccion_img'] = ultima_inspeccion_img
        
        bitacora = Bitacora.objects.filter(vehiculo=Vehiculo.objects.get(numero_economico=vehiculo.numero_economico, compania=Compania.objects.get(compania=self.request.user.perfil.compania)), compania=Compania.objects.get(compania=self.request.user.perfil.compania)).order_by("id")
        bitacora_pro = Bitacora_Pro.objects.filter(vehiculo=Vehiculo.objects.get(numero_economico=vehiculo.numero_economico, compania=Compania.objects.get(compania=self.request.user.perfil.compania)), compania=Compania.objects.get(compania=self.request.user.perfil.compania)).order_by("id")

        bitacoras = bitacora.annotate(presion_de_entrada_1=F("presion_de_entrada"), presion_de_salida_1=F("presion_de_salida"), presion_de_entrada_2=Value(None, output_field=IntegerField()), presion_de_salida_2=Value(None, output_field=IntegerField()), presion_de_entrada_3=Value(None, output_field=IntegerField()), presion_de_salida_3=Value(None, output_field=IntegerField()), presion_de_entrada_4=Value(None, output_field=IntegerField()), presion_de_salida_4=Value(None, output_field=IntegerField()), presion_de_entrada_5=Value(None, output_field=IntegerField()), presion_de_salida_5=Value(None, output_field=IntegerField()), presion_de_entrada_6=Value(None, output_field=IntegerField()), presion_de_salida_6=Value(None, output_field=IntegerField()), presion_de_entrada_7=Value(None, output_field=IntegerField()), presion_de_salida_7=Value(None, output_field=IntegerField()), presion_de_entrada_8=Value(None, output_field=IntegerField()), presion_de_salida_8=Value(None, output_field=IntegerField()), presion_de_entrada_9=Value(None, output_field=IntegerField()), presion_de_salida_9=Value(None, output_field=IntegerField()), presion_de_entrada_10=Value(None, output_field=IntegerField()), presion_de_salida_10=Value(None, output_field=IntegerField()), presion_de_entrada_11=Value(None, output_field=IntegerField()), presion_de_salida_11=Value(None, output_field=IntegerField()), presion_de_entrada_12=Value(None, output_field=IntegerField()), presion_de_salida_12=Value(None, output_field=IntegerField())).order_by("id")
        bitacoras = list(bitacoras.values("id", "vehiculo__id", "vehiculo__configuracion", "compania__id", "fecha_de_inflado", "tiempo_de_inflado", "presion_de_entrada_1", "presion_de_salida_1", "presion_de_entrada_2", "presion_de_salida_2", "presion_de_entrada_3", "presion_de_salida_3", "presion_de_entrada_4", "presion_de_salida_4", "presion_de_entrada_5", "presion_de_salida_5", "presion_de_entrada_6", "presion_de_salida_6", "presion_de_entrada_7", "presion_de_salida_7", "presion_de_entrada_8", "presion_de_salida_8", "presion_de_entrada_9", "presion_de_salida_9", "presion_de_entrada_10", "presion_de_salida_10", "presion_de_entrada_11", "presion_de_salida_11", "presion_de_entrada_12", "presion_de_salida_12", "vehiculo__presion_establecida_1", "vehiculo__presion_establecida_2", "vehiculo__presion_establecida_3", "vehiculo__presion_establecida_4", "vehiculo__presion_establecida_5", "vehiculo__presion_establecida_6", "vehiculo__presion_establecida_7"))

        bitacoras_pro = list(bitacora_pro.values("id", "vehiculo__id", "vehiculo__configuracion", "compania__id", "fecha_de_inflado", "tiempo_de_inflado", "presion_de_entrada_1", "presion_de_salida_1", "presion_de_entrada_2", "presion_de_salida_2", "presion_de_entrada_3", "presion_de_salida_3", "presion_de_entrada_4", "presion_de_salida_4", "presion_de_entrada_5", "presion_de_salida_5", "presion_de_entrada_6", "presion_de_salida_6", "presion_de_entrada_7", "presion_de_salida_7", "presion_de_entrada_8", "presion_de_salida_8", "presion_de_entrada_9", "presion_de_salida_9", "presion_de_entrada_10", "presion_de_salida_10", "presion_de_entrada_11", "presion_de_salida_11", "presion_de_entrada_12", "presion_de_salida_12", "vehiculo__presion_establecida_1", "vehiculo__presion_establecida_2", "vehiculo__presion_establecida_3", "vehiculo__presion_establecida_4", "vehiculo__presion_establecida_5", "vehiculo__presion_establecida_6", "vehiculo__presion_establecida_7"))
       
        bitacoras.extend(bitacoras_pro)

        bitacoras = sorted(bitacoras, key=lambda x:x["fecha_de_inflado"], reverse=False)

        entradas_correctas = functions.entrada_correcta_ambas(bitacoras)
        fecha = functions.convertir_fecha(str(vehiculo.fecha_de_inflado))
        servicios = ServicioVehiculo.objects.filter(vehiculo = vehiculo).order_by('-id').exclude(estado='abierto')
        
        if vehiculo.fecha_de_inflado:
            inflado = 1
        else:
            inflado = 0

        eventos = []
        inspecciones_list = []
        for bit in bitacora:
            eventos.append([bit.fecha_de_inflado, bit, 'pulpo'])
        for bit in bitacora_pro:
            eventos.append([bit.fecha_de_inflado, bit, 'pulpopro'])

        for inspeccion in inspecciones_vehiculo:
            color_insp = functions.color_observaciones_all(inspeccion)
            if color_insp == 'bad':
                signo = 'icon-cross bad-text'
            elif color_insp == 'yellow':
                signo = 'icon-warning  yellow-text'
            else:
                signo = 'icon-checkmark good-text'
            if functions.exist_edicion_manual(inspeccion):
                inspecciones_list.append([inspeccion.fecha, inspeccion, 'edicion', signo])
            inspecciones_list.append([inspeccion.fecha, inspeccion, 'inspeccion', signo])
        inspecciones_list = sorted(inspecciones_list, key=lambda x:x[1].id, reverse=True)
        for ins in inspecciones_list:
            eventos.append([ins[0], ins[1], ins[2], ins[3]])
        for servicio in servicios:
            eventos.append([datetime.combine(servicio.fecha_inicio, datetime.min.time(), timezone.utc), servicio, 'servicio', 'icon-checkmark good-text'])
        eventos = sorted(eventos, key=lambda x:x[0], reverse=True)
        context["eventos"] = eventos
        hoy = date.today()

        mes_1 = hoy.strftime("%b")
        mes_2 = functions.mes_anterior(hoy)
        mes_3 = functions.mes_anterior(mes_2)
        mes_4 = functions.mes_anterior(mes_3)
        mes_5 = functions.mes_anterior(mes_4)
        mes_6 = functions.mes_anterior(mes_5)
        mes_7 = functions.mes_anterior(mes_6)
        mes_8 = functions.mes_anterior(mes_7)

        hoy1 = hoy.strftime("%m")
        hoy2 = mes_2.strftime("%m")
        hoy3 = mes_3.strftime("%m")
        hoy4 = mes_4.strftime("%m")
        hoy5 = mes_5.strftime("%m")
        hoy6 = mes_6.strftime("%m")
        hoy7 = mes_7.strftime("%m")
        hoy8 = mes_8.strftime("%m")

        color = functions.entrada_correcta_actual(vehiculo)
        message = None
        message_pro = {}

        if bitacora or bitacora_pro:
            vehiculo_mes1 = bitacora.filter(fecha_de_inflado__month=hoy1)
            vehiculo_mes2 = bitacora.filter(fecha_de_inflado__month=hoy2)
            vehiculo_mes3 = bitacora.filter(fecha_de_inflado__month=hoy3)
            vehiculo_mes4 = bitacora.filter(fecha_de_inflado__month=hoy4)
            vehiculo_mes5 = bitacora.filter(fecha_de_inflado__month=hoy5)
            vehiculo_mes6 = bitacora.filter(fecha_de_inflado__month=hoy6)
            vehiculo_mes7 = bitacora.filter(fecha_de_inflado__month=hoy7)
            vehiculo_mes8 = bitacora.filter(fecha_de_inflado__month=hoy8)

            vehiculo_pro_mes1 = bitacora_pro.filter(fecha_de_inflado__month=hoy1)
            vehiculo_pro_mes2 = bitacora_pro.filter(fecha_de_inflado__month=hoy2)
            vehiculo_pro_mes3 = bitacora_pro.filter(fecha_de_inflado__month=hoy3)
            vehiculo_pro_mes4 = bitacora_pro.filter(fecha_de_inflado__month=hoy4)
            vehiculo_pro_mes5 = bitacora_pro.filter(fecha_de_inflado__month=hoy5)
            vehiculo_pro_mes6 = bitacora_pro.filter(fecha_de_inflado__month=hoy6)
            vehiculo_pro_mes7 = bitacora_pro.filter(fecha_de_inflado__month=hoy7)
            vehiculo_pro_mes8 = bitacora_pro.filter(fecha_de_inflado__month=hoy8)

            mala_entrada_contar_mes1 = functions.contar_mala_entrada(vehiculo_mes1)
            mala_entrada_contar_mes2 = functions.contar_mala_entrada(vehiculo_mes2)
            mala_entrada_contar_mes3 = functions.contar_mala_entrada(vehiculo_mes3)
            mala_entrada_contar_mes4 = functions.contar_mala_entrada(vehiculo_mes4)
            mala_entrada_contar_mes5 = functions.contar_mala_entrada(vehiculo_mes5)
            mala_entrada_contar_mes6 = functions.contar_mala_entrada(vehiculo_mes6)
            mala_entrada_contar_mes7 = functions.contar_mala_entrada(vehiculo_mes7)
            mala_entrada_contar_mes8 = functions.contar_mala_entrada(vehiculo_mes8)

            mala_entrada_contar_mes1 += functions.contar_mala_entrada_pro(vehiculo_pro_mes1)
            mala_entrada_contar_mes2 += functions.contar_mala_entrada_pro(vehiculo_pro_mes2)
            mala_entrada_contar_mes3 += functions.contar_mala_entrada_pro(vehiculo_pro_mes3)
            mala_entrada_contar_mes4 += functions.contar_mala_entrada_pro(vehiculo_pro_mes4)
            mala_entrada_contar_mes5 += functions.contar_mala_entrada_pro(vehiculo_pro_mes5)
            mala_entrada_contar_mes6 += functions.contar_mala_entrada_pro(vehiculo_pro_mes6)
            mala_entrada_contar_mes7 += functions.contar_mala_entrada_pro(vehiculo_pro_mes7)
            mala_entrada_contar_mes8 += functions.contar_mala_entrada_pro(vehiculo_pro_mes8)

            doble_entrada = functions.doble_entrada(bitacora, bitacora_pro)
            doble_mala_entrada = functions.doble_mala_entrada2(bitacora, vehiculo)

            doble_mala_entrada_pro = functions.doble_mala_entrada_pro(bitacora_pro, vehiculo)
            try:
                if doble_mala_entrada:
                    message = "Doble mala entrada"
                    color = "bad"
                elif doble_mala_entrada_pro:
                    for entrada in doble_mala_entrada_pro:
                        message_pro[entrada[-1]] = "Doble mala entrada"
                    color = "bad"
            except:
                pass

            

            configuracion = vehiculo.configuracion
            cantidad_llantas = functions.cantidad_llantas(configuracion)

            posiciones = llantas.values("posicion").distinct()
            ejes = llantas.values("nombre_de_eje").distinct()

            comparativa_de_posiciones = {}
            for posicion in posiciones:
                valores_posicion = []

                llantas_posicion = llantas.filter(posicion=posicion["posicion"])
                inspecciones_posicion = Inspeccion.objects.filter(llanta__in=llantas_posicion)
                if inspecciones_posicion.exists():
                    regresion_posicion = functions.km_proyectado(inspecciones_posicion, False)
                    km_proyectado = regresion_posicion[0]
                    profundidad = regresion_posicion[6]

                    valores_posicion.append(km_proyectado)
                    valores_posicion.append(profundidad)

                    comparativa_de_posiciones[posicion["posicion"]] = valores_posicion

            #print("comparativa_de_posiciones", comparativa_de_posiciones)
            comparativa_de_ejes = {}
            for eje in ejes:
                valores_eje = []

                llantas_eje = llantas.filter(nombre_de_eje=eje["nombre_de_eje"])
                inspecciones_eje = Inspeccion.objects.filter(llanta__in=llantas_eje)
                if inspecciones_eje.exists():
                    regresion_eje = functions.km_proyectado(inspecciones_eje, False)
                    km_x_mm_eje = regresion_eje[1]

                    valores_eje.append(km_x_mm_eje)

                    comparativa_de_ejes[eje["nombre_de_eje"]] = valores_eje

            regresion_vehiculo = functions.km_proyectado(inspecciones, False)
            cpk_vehiculo = regresion_vehiculo[3]
            cpk_vehiculo = round(sum(cpk_vehiculo), 3)

            reemplazo_actual = functions.reemplazo_actual2(llantas)
            reemplazo_actual_ejes = {k: v for k, v in reemplazo_actual.items() if v != 0}

            grupo = str(User.objects.get(username=self.request.user).groups.get())

            context["bitacoras"] = bitacora
            context["cantidad_doble_entrada_mes1"] = doble_entrada[2]["mes1"] + doble_entrada[3]["mes1"]
            context["cantidad_doble_entrada_mes2"] = doble_entrada[2]["mes2"] + doble_entrada[3]["mes2"]
            context["cantidad_doble_entrada_mes3"] = doble_entrada[2]["mes3"] + doble_entrada[3]["mes3"]
            context["cantidad_doble_entrada_mes4"] = doble_entrada[2]["mes4"] + doble_entrada[3]["mes4"]
            context["cantidad_doble_entrada_mes5"] = doble_entrada[2]["mes5"] + doble_entrada[3]["mes5"]
            context["cantidad_doble_entrada_mes6"] = doble_entrada[2]["mes6"] + doble_entrada[3]["mes6"]
            context["cantidad_doble_entrada_mes7"] = doble_entrada[2]["mes7"] + doble_entrada[3]["mes7"]
            context["cantidad_doble_entrada_mes8"] = doble_entrada[2]["mes8"] + doble_entrada[3]["mes8"]
            context["cantidad_entrada_mes1"] = mala_entrada_contar_mes1
            context["cantidad_entrada_mes2"] = mala_entrada_contar_mes2
            context["cantidad_entrada_mes3"] = mala_entrada_contar_mes3
            context["cantidad_entrada_mes4"] = mala_entrada_contar_mes4
            context["cantidad_entrada_mes5"] = mala_entrada_contar_mes5
            context["cantidad_entrada_mes6"] = mala_entrada_contar_mes6
            context["cantidad_entrada_mes7"] = mala_entrada_contar_mes7
            context["cantidad_entrada_mes8"] = mala_entrada_contar_mes8
            context["cantidad_inflado"] = inflado
            context["cantidad_llantas"] = cantidad_llantas
            context["color"] = color
            context["comparativa_de_ejes"] = comparativa_de_ejes
            context["comparativa_de_posiciones"] = comparativa_de_posiciones
            context["configuracion"] = configuracion
            context["cpk_vehiculo"] = cpk_vehiculo
            context["doble_entrada"] = doble_entrada
            context["entradas"] = entradas_correctas
            context["fecha"] = fecha
            context["grupo"] = grupo
            context["hoy"] = hoy
            context["mes_1"] = mes_1
            context["mes_2"] = mes_2.strftime("%b")
            context["mes_3"] = mes_3.strftime("%b")
            context["mes_4"] = mes_4.strftime("%b")
            context["mes_5"] = mes_5.strftime("%b")
            context["mes_6"] = mes_6.strftime("%b")
            context["mes_7"] = mes_7.strftime("%b")
            context["mes_8"] = mes_8.strftime("%b")
            context["message"] = message
            context["reemplazo_actual_ejes"] = reemplazo_actual_ejes
            context["tiempo_promedio"] = functions.inflado_promedio(vehiculo)
            context["vehiculo_mes1"] = vehiculo_mes1.count() + vehiculo_pro_mes1.count()
            context["vehiculo_mes2"] = vehiculo_mes2.count() + vehiculo_pro_mes2.count()
            context["vehiculo_mes3"] = vehiculo_mes3.count() + vehiculo_pro_mes3.count()
            context["vehiculo_mes4"] = vehiculo_mes4.count() + vehiculo_pro_mes4.count()
            context["vehiculo_mes5"] = vehiculo_mes5.count() + vehiculo_pro_mes5.count()
            context["vehiculo_mes6"] = vehiculo_mes6.count() + vehiculo_pro_mes6.count()
            context["vehiculo_mes7"] = vehiculo_mes7.count() + vehiculo_pro_mes7.count()
            context["vehiculo_mes8"] = vehiculo_mes8.count() + vehiculo_pro_mes8.count()

        #Generacion de ejes dinamico
        vehiculo_actual = Vehiculo.objects.get(pk = self.kwargs['pk'])

        llantas_actuales = Llanta.objects.filter(vehiculo = vehiculo_actual, tirecheck=False, inventario="Rodante")
        inspecciones_actuales = Inspeccion.objects.filter(llanta__in=llantas_actuales)

        

        #Obtencion de la data
        num_ejes = vehiculo_actual.configuracion.split('.')
        ejes_no_ordenados = []
        ejes = []
        eje = 1
        color_profundidad = ""
        presiones_establecida=[
            vehiculo_actual.presion_establecida_1,
            vehiculo_actual.presion_establecida_2,
            vehiculo_actual.presion_establecida_3,
            vehiculo_actual.presion_establecida_4,
            vehiculo_actual.presion_establecida_5,
            vehiculo_actual.presion_establecida_6,
            vehiculo_actual.presion_establecida_7,
        ]
        numero = 0
        for num in num_ejes:
            list_temp = []
            for llanta in llantas_actuales:

                objetivo = llanta.vehiculo.compania.objetivo / 100
                presion_act = llanta.presion_actual
                presion_minima = presiones_establecida[numero] - (presiones_establecida[numero] * objetivo)
                presion_maxima = presiones_establecida[numero] + (presiones_establecida[numero] * objetivo)
                min_produndidad = functions.min_profundidad(llanta)
                punto_retiro = functions.punto_de_retiro(llanta)
                color_profundidad = functions.color_profundidad(min_produndidad, punto_retiro)

                problema = None
                try:
                    if presion_act < presion_minima:
                        color_presion = 'bad'
                    elif presion_act > presion_maxima:
                        color_presion = 'bad'
                    else:
                        color_presion = 'good'
                except:
                    color_presion = 'good'

                if message or problema:
                    color_llanta = "bad"
                else:
                    color_llanta = "good"
                color_llanta = functions.color_observaciones_all_one(llanta)
                try:
                    if presion_act >= presion_minima and presion_act <= presion_maxima:
                        color_presion = 'good'
                    elif presion_act < presion_minima:
                        color_presion = 'bad'
                    elif presion_act > presion_maxima:
                        color_presion = 'yellow'
                except:
                    color_presion = 'good'
                if llanta.eje == eje:
                    #print(presion_act > presion_minima and presion_act < presion_maxima)
                    list_temp.append([llanta, color_profundidad, color_presion, min_produndidad, color_llanta, problema, int(presion_minima), int(presion_maxima)])
            eje += 1
            ejes_no_ordenados.append(list_temp)
            numero += 1
        
        ejes = functions.acomodo_ejes(ejes_no_ordenados)
        color = functions.entrada_correcta_actual(vehiculo_actual)
        
        #print(color)
        if bitacora:
            if doble_mala_entrada:
                color == 'bad'
                style = 'bad'
            elif color == 'good':
                style = 'good'
            elif color == 'bad':
                style = 'bad'
            else:
                style = 'bad'
        else:
            style = 'good'

        cant_ejes = len(ejes)

        if len(llantas_actuales) == 0:
            sin_llantas = True
        else:
            sin_llantas = False

        problemas_abiertos = []

        try:
            num_llanta = 0
            for eje in ejes:
                #print(message)
                if len(eje) == 2:
                    num_llanta += 1
                    if num_llanta in message_pro:
                        problemas_abiertos.append([eje[0][0], message_pro[num_llanta]])
                        eje[0][4] = "bad"
                    elif message:
                        problemas_abiertos.append([eje[0][0], message])
                    if eje[0][5] or eje[0][5] == 0:
                        problemas_abiertos.append([eje[0][0], eje[0][5]])
                    num_llanta += 1
                    if num_llanta in message_pro:
                        problemas_abiertos.append([eje[1][0], message_pro[num_llanta]])
                        eje[1][4] = "bad"
                    elif message:
                        problemas_abiertos.append([eje[1][0], message])
                    if eje[1][5] or eje[1][5] == 0:
                        problemas_abiertos.append([eje[1][0], eje[1][5]])
                else:
                    num_llanta += 1
                    if num_llanta in message_pro:
                        problemas_abiertos.append([eje[0][0], message_pro[num_llanta]])
                        eje[0][4] = "bad"
                    elif message:
                        problemas_abiertos.append([eje[0][0], message])
                    if eje[0][5] or eje[0][5] == 0:
                        problemas_abiertos.append([eje[0][0], eje[0][5]])
                    num_llanta += 1
                    if num_llanta in message_pro:
                        problemas_abiertos.append([eje[1][0], message_pro[num_llanta]])
                        eje[1][4] = "bad"
                    elif message:
                        problemas_abiertos.append([eje[1][0], message])
                    if eje[1][5] or eje[1][5] == 0:
                        problemas_abiertos.append([eje[1][0], eje[1][5]])
                    num_llanta += 1
                    if num_llanta in message_pro:
                        problemas_abiertos.append([eje[2][0], message_pro[num_llanta]])
                        eje[2][4] = "bad"
                    elif message:
                        problemas_abiertos.append([eje[2][0], message])
                    if eje[2][5] or eje[2][5] == 0:
                        problemas_abiertos.append([eje[2][0], eje[2][5]])
                    num_llanta += 1
                    if num_llanta in message_pro:
                        problemas_abiertos.append([eje[3][0], message_pro[num_llanta]])
                        eje[3][4] = "bad"
                    elif message:
                        problemas_abiertos.append([eje[3][0], message])
                    if eje[3][5] or eje[3][5] == 0:
                        problemas_abiertos.append([eje[3][0], eje[3][5]])
        except:
            pass

        #print(problemas_abiertos)
        #print(len(llantas_actuales))
        #print(sin_llantas)
        #print(problemas_abiertos)
        #print(vehiculo.configuracion)
        #print(ejes)
        #print(f'style: {style}')
        #print(f'llantas_sospechosas: {llantas_sospechosas}')
        #print(f'llantas_rojas: {llantas_rojas}')
        #print(f'llantas_amarillas: {llantas_amarillas}')
        #print(f'llantas_azules: {llantas_azules}')
        context['ejes'] = ejes
        context['style'] = style
        context['cant_ejes'] = cant_ejes
        context['problemas_abiertos'] = problemas_abiertos
        context['sin_llantas'] = sin_llantas
        #try:
        #    context['presion_maxima'] = int(float(presion_maxima))
        #    context['presion_minima'] = int(float(presion_minima))
        #except:
        #    pass
        #Generacion de las bitacoras
        """bitacora_edicion = Bitacora_Edicion.objects.filter(vehiculo = vehiculo_actual)
        context['bitacora_edicion'] = bitacora_edicion"""

        #Generacion de la dimencion
        dimension = 'Desconocido'
        for llanta in llantas_actuales:
            try:
                if llanta.producto.dimension != "" and llanta.producto.dimension != None:
                    dimension = llanta.producto.dimension
                    break
            except:
                pass
        #print(dimension)
        context['dimension'] = dimension


        #observaciones:

        problemas = []
        ultima_inspeccion_vehiculo = InspeccionVehiculo.objects.filter(vehiculo=vehiculo).last()
        current_isnpec = False
        if ultima_inspeccion_vehiculo != None:
            ultimas_inspecciones = Inspeccion.objects.filter(inspeccion_vehiculo = ultima_inspeccion_vehiculo).exclude(
                llanta__inventario = 'Archivado'
            )
            current_isnpec = True
        if current_isnpec:
            for observacion in vehiculo.observaciones.all():
                color_obs = functions.color_observaciones_one(observacion)
                if color_obs == 'bad':
                    signo = 'icon-cross bad-text'
                elif color_obs == 'yellow':
                    signo = 'icon-warning  yellow-text'
                else:
                    signo = 'icon-checkmark good-text'
                problemas.append(['VH', observacion, signo])
            #for inspeccion in ultimas_inspecciones:
            #    for observacion in inspeccion.observaciones.all():
            #        color_obs = functions.color_observaciones_one(observacion)
            #        if color_obs == 'bad':
            #            signo = 'icon-cross bad-text'
            #        elif color_obs == 'yellow':
            #            signo = 'icon-warning  yellow-text'
            #        else:
            #            signo = 'icon-checkmark good-text'
            #        problemas.append([inspeccion.posicion, observacion, signo])
        for llanta in llantas_actuales:
            for observacion in llanta.observaciones.all():
                color_obs = functions.color_observaciones_one(observacion)
                if color_obs == 'bad':
                    signo = 'icon-cross bad-text'
                elif color_obs == 'yellow':
                    signo = 'icon-warning  yellow-text'
                else:
                    signo = 'icon-checkmark good-text'
                problemas.append([llanta.posicion, observacion, signo])
        context['problemas'] = problemas
        context["llantas"] = llantas
        context["llantas_ordenadas"] = llantas_ordenadas
        return context

class reporteEdicionLlanta(ListView):
    template_name= 'reporteEdicion.html'
    model = Inspeccion
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inspeccion = Inspeccion.objects.get(pk = self.kwargs['pk'])
        cambios = []
        if inspeccion.edicion_manual == True:
            original = []
            cambio = []
            evento_raw =inspeccion.evento
            evento_raw = evento_raw.replace("\'", "\"")
            evento_act = json.loads(evento_raw)
            if evento_act['llanta_mod'] != '':
                original.append(f"Llanta: {evento_act['llanta_inicial']}")
                cambio.append(f"Llanta: {evento_act['llanta_mod']}")
            
            if evento_act['numero_economico_mod'] != '':
                original.append(f"Llanta: {evento_act['numero_economico']}")
                cambio.append(f"Llanta: {evento_act['numero_economico_mod']}")
            
            if evento_act['producto_mod'] != '':
                original.append(f"Producto: {evento_act['producto_inicial']}")
                cambio.append(f"Producto: {evento_act['producto_mod']}")
                
            if evento_act['vida_mod'] != '':
                original.append(f"Vida: {evento_act['vida_inicial']}")
                cambio.append(f"Vida: {evento_act['vida_mod']}")
            
            if evento_act['km_mod'] != '':
                original.append(f"Km: {evento_act['km_inicial']}")
                cambio.append(f"Km: {evento_act['km_mod']}")
                
            if evento_act['presion_mod'] != '':
                original.append(f"Presión: {evento_act['presion_inicial']}")
                cambio.append(f"Presión: {evento_act['presion_mod']}")
                
            if evento_act['profundidad_izquierda_mod'] != '':
                original.append(f"Profundidad izq: {evento_act['profundidad_izquierda_inicial']}")
                cambio.append(f"Profundidad izq: {evento_act['profundidad_izquierda_mod']}")
                
            if evento_act['profundidad_central_mod'] != '':
                original.append(f"Profundidad cen: {evento_act['profundidad_central_inicial']}")
                cambio.append(f"Profundidad cen: {evento_act['profundidad_central_mod']}")
                
            if evento_act['profundidad_derecha_mod'] != '':
                original.append(f"Profundidad der: {evento_act['profundidad_derecha_inicial']}")
                cambio.append(f"Profundidad der: {evento_act['profundidad_derecha_mod']}")
            cambios.append(
                {
                    'posicion': inspeccion.posicion,
                    'original': original,
                    'cambio': cambio
                }
            )
        #for i in cambios:
        #    print(i['posicion'])
        #    print('original')
        #    for a in i['original']:
        #        print(a)
        #    print('cambios')
        #    for a in i['cambio']:
        #        print(a)
        context['cambios'] = cambios        
        return context
    
class ReporteEdicion(ListView):
    template_name= 'reporteEdicion.html'
    model = InspeccionVehiculo
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inspeccion_vehiculo = InspeccionVehiculo.objects.get(pk = self.kwargs['pk'])
        inspecciones_actuales = Inspeccion.objects.filter(inspeccion_vehiculo = inspeccion_vehiculo)
        cambios = []
        for inspeccion in inspecciones_actuales:
            if inspeccion.edicion_manual == True:
                original = []
                cambio = []
                evento_raw =inspeccion.evento
                evento_raw = evento_raw.replace("\'", "\"")
                evento_act = json.loads(evento_raw)
                if evento_act['llanta_mod'] != '':
                    original.append(f"Llanta: {evento_act['llanta_inicial']}")
                    cambio.append(f"Llanta: {evento_act['llanta_mod']}")
                
                if evento_act['numero_economico_mod'] != '':
                    original.append(f"Llanta: {evento_act['numero_economico']}")
                    cambio.append(f"Llanta: {evento_act['numero_economico_mod']}")
                
                if evento_act['producto_mod'] != '':
                    original.append(f"Producto: {evento_act['producto_inicial']}")
                    cambio.append(f"Producto: {evento_act['producto_mod']}")
                    
                if evento_act['vida_mod'] != '':
                    original.append(f"Vida: {evento_act['vida_inicial']}")
                    cambio.append(f"Vida: {evento_act['vida_mod']}")
                
                if evento_act['km_mod'] != '':
                    original.append(f"Km: {evento_act['km_inicial']}")
                    cambio.append(f"Km: {evento_act['km_mod']}")
                    
                if evento_act['presion_mod'] != '':
                    original.append(f"Presión: {evento_act['presion_inicial']}")
                    cambio.append(f"Presión: {evento_act['presion_mod']}")
                    
                if evento_act['profundidad_izquierda_mod'] != '':
                    original.append(f"Profundidad izq: {evento_act['profundidad_izquierda_inicial']}")
                    cambio.append(f"Profundidad izq: {evento_act['profundidad_izquierda_mod']}")
                    
                if evento_act['profundidad_central_mod'] != '':
                    original.append(f"Profundidad cen: {evento_act['profundidad_central_inicial']}")
                    cambio.append(f"Profundidad cen: {evento_act['profundidad_central_mod']}")
                    
                if evento_act['profundidad_derecha_mod'] != '':
                    original.append(f"Profundidad der: {evento_act['profundidad_derecha_inicial']}")
                    cambio.append(f"Profundidad der: {evento_act['profundidad_derecha_mod']}")
                cambios.append(
                    {
                        'posicion': inspeccion.posicion,
                        'original': original,
                        'cambio': cambio
                    }
                )
        #for i in cambios:
        #    print(i['posicion'])
        #    print('original')
        #    for a in i['original']:
        #        print(a)
        #    print('cambios')
        #    for a in i['cambio']:
        #        print(a)
        context['cambios'] = cambios        
        return context
    
class ReporteInspeccion(ListView):
    template_name= 'reporteInspeccion.html'
    model = InspeccionVehiculo
    
    def get_queryset(self):
        """return Bitacora_Edicion.objects.filter(pk = self.kwargs['pk'])"""
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inspeccion_vehiculo = InspeccionVehiculo.objects.get(pk = self.kwargs['pk'])
        vehiculo = inspeccion_vehiculo.vehiculo
        context['inspeccion_vehiculo'] = inspeccion_vehiculo
        context['vehiculo'] = vehiculo
        num_ejes = vehiculo.configuracion.split('.')
        context['cant_ejes'] = len(num_ejes)
        #Generacion de ejes dinamico
        vehiculo_actual = vehiculo
        #llantas_actuales = Llanta.objects.filter(vehiculo = self.kwargs['pk'], tirecheck=False, inventario="Rodante")
        inspecciones_actuales = Inspeccion.objects.filter(inspeccion_vehiculo = inspeccion_vehiculo)

        
        #Obtencion de la data
        num_ejes = vehiculo_actual.configuracion.split('.')
        ejes_no_ordenados = []
        ejes = []
        eje = 1
        color_profundidad = ""
        presiones_establecida=[
            vehiculo_actual.presion_establecida_1,
            vehiculo_actual.presion_establecida_2,
            vehiculo_actual.presion_establecida_3,
            vehiculo_actual.presion_establecida_4,
            vehiculo_actual.presion_establecida_5,
            vehiculo_actual.presion_establecida_6,
            vehiculo_actual.presion_establecida_7,
        ]
        numero = 0
        for num in num_ejes:
            list_temp = []
            for llanta in inspecciones_actuales:
                
                objetivo = llanta.vehiculo.compania.objetivo / 100
                presion_act = llanta.presion
                presion_minima = presiones_establecida[numero] - (presiones_establecida[numero] * objetivo)
                presion_maxima = presiones_establecida[numero] + (presiones_establecida[numero] * objetivo)
                min_produndidad = functions.min_profundidad(llanta)
                punto_retiro = functions.punto_de_retiro(llanta)
                color_profundidad = functions.color_profundidad(min_produndidad, punto_retiro)
                signo = ('icon-checkmark' if color_profundidad=='good' else 'icon-cross')
                color_llanta = functions.color_observaciones(llanta.observaciones.all())
                try:  
                    if presion_act >= presion_minima and presion_act <= presion_maxima:
                        color_presion = 'good'
                    elif presion_act < presion_minima:
                        color_presion = 'bad'
                    elif presion_act > presion_maxima:
                        color_presion = 'yellow'
                except:
                    color_presion = 'good'
                
                if llanta.eje == eje:
                    #print(presion_act > presion_minima and presion_act < presion_maxima)
                    list_temp.append([llanta, color_profundidad, signo, color_presion, min_produndidad, color_llanta,  int(presion_minima), int(presion_maxima)])
            eje += 1
            ejes_no_ordenados.append(list_temp)
            numero += 1

        ejes = functions.acomodo_ejes(ejes_no_ordenados)
        context['ejes'] = ejes
        #Paso de los problemas
        problemas = []
        for observacion in vehiculo.observaciones.all():
            problemas.append(['VH', observacion])
        for inspeccion in inspecciones_actuales:
            for observacion in inspeccion.observaciones.all():
                problemas.append([inspeccion.posicion, observacion])
                
        context['problemas'] = problemas
        
        return context


def download_rendimiento_de_llanta(request):
    # Define Django project base directory
    # content-type of response
    vehiculos = Vehiculo.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania))
    llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=request.user.perfil.compania), vehiculo__in=vehiculos)
    inspecciones = Inspeccion.objects.filter(llanta__in=llantas)

    regresion = functions.km_proyectado(inspecciones, True)
    llantas_limpias = regresion[4]
    llantas_analizadas = llantas.filter(numero_economico__in=llantas_limpias)

    response = HttpResponse(content_type='application/ms-excel')

	#decide file name
    response['Content-Disposition'] = 'attachment; filename="LlantasAnalizadas.xls"'

	#creating workbook
    wb = xlwt.Workbook(encoding='utf-8')

	#adding sheet
    ws = wb.add_sheet("sheet1")

	# Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
	# headers are bold
    font_style.font.bold = True

	#column header names, you can use your own headers here
    columns = ['Llanta', 'Vehiculo', 'Posición', 'Km actual', 'Km proyectado', 'CPK', 'Sucursal', 'Aplicación', 'Clase', 'Nombre de eje', 'Producto', 'Min profundidad', "Eje"]

    #write column headers in sheet
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    #get your data, from database or from a text file...
    for my_row in range(len(llantas_analizadas)):
        ws.write(my_row + 1, 0, str(llantas_analizadas[my_row]), font_style)
        ws.write(my_row + 1, 1, str(llantas_analizadas.values("vehiculo__numero_economico")[my_row]["vehiculo__numero_economico"]), font_style)
        ws.write(my_row + 1, 2, str(llantas_analizadas.values("posicion")[my_row]["posicion"]), font_style)
        ws.write(my_row + 1, 3, str(llantas_analizadas.values("km_actual")[my_row]["km_actual"]), font_style)
        ws.write(my_row + 1, 4, str(regresion[5][my_row]), font_style)
        ws.write(my_row + 1, 5, str(regresion[3][my_row]), font_style)
        ws.write(my_row + 1, 6, str(llantas_analizadas.values("vehiculo__ubicacion__nombre")[my_row]["vehiculo__ubicacion__nombre"]), font_style)
        ws.write(my_row + 1, 7, str(llantas_analizadas.values("vehiculo__aplicacion__nombre")[my_row]["vehiculo__aplicacion__nombre"]), font_style)
        ws.write(my_row + 1, 8, str(llantas_analizadas.values("vehiculo__clase")[my_row]["vehiculo__clase"]), font_style)
        ws.write(my_row + 1, 9, str(llantas_analizadas.values("nombre_de_eje")[my_row]["nombre_de_eje"]), font_style)
        ws.write(my_row + 1, 10, str(llantas_analizadas.values("producto__producto")[my_row]["producto__producto"]), font_style)
        ws.write(my_row + 1, 11, str(llantas_analizadas.annotate(min_profundidad=Least("profundidad_izquierda", "profundidad_central", "profundidad_derecha")).values("min_profundidad")[my_row]["min_profundidad"]), font_style)
        ws.write(my_row + 1, 12, str(llantas_analizadas.values("nombre_de_eje")[my_row]["nombre_de_eje"]), font_style)


    wb.save(response)
    return response

def download_reemplazo_estimado(request):
    # Define Django project base directory
    # content-type of response
    llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=request.user.perfil.compania))
    inspecciones = Inspeccion.objects.filter(llanta__vehiculo__compania=Compania.objects.get(compania=request.user.perfil.compania))
    ubicacion = Ubicacion.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania))[0]

    embudo_vida1 = functions.embudo_vidas(llantas)
    embudo_vida2 = functions.embudo_vidas_con_regresion(inspecciones, ubicacion, 30)
    embudo_vida3 = functions.embudo_vidas_con_regresion(inspecciones, ubicacion, 60)
    embudo_vida4 = functions.embudo_vidas_con_regresion(inspecciones, ubicacion, 90)

    periodo = []
    for llanta in llantas:
        if llanta in embudo_vida1[0]:
            periodo.append("Hoy")
        elif llanta in embudo_vida2[0]:
            periodo.append("30 días")
        elif llanta in embudo_vida3[0]:
            periodo.append("60 días")
        elif llanta in embudo_vida4[0]:
            periodo.append("90 días")
        else:
            periodo.append("-")

    response = HttpResponse(content_type='application/ms-excel')

	#decide file name
    response['Content-Disposition'] = 'attachment; filename="LlantasReemplazoEstimado.xls"'

	#creating workbook
    wb = xlwt.Workbook(encoding='utf-8')

	#adding sheet
    ws = wb.add_sheet("sheet1")

	# Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
	# headers are bold
    font_style.font.bold = True

	#column header names, you can use your own headers here
    columns = ['Llanta', 'Vehiculo', 'Posición', 'Sucursal', 'Aplicación', 'Clase', 'Nombre de eje', 'Producto', 'Min profundidad', 'Periodo']

    #write column headers in sheet
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    #get your data, from database or from a text file...
    for my_row in range(len(llantas)):
        ws.write(my_row + 1, 0, str(llantas[my_row]), font_style)
        ws.write(my_row + 1, 1, str(llantas.values("vehiculo__numero_economico")[my_row]["vehiculo__numero_economico"]), font_style)
        ws.write(my_row + 1, 2, str(llantas.values("posicion")[my_row]["posicion"]), font_style)
        ws.write(my_row + 1, 3, str(llantas.values("vehiculo__ubicacion__nombre")[my_row]["vehiculo__ubicacion__nombre"]), font_style)
        ws.write(my_row + 1, 4, str(llantas.values("vehiculo__aplicacion__nombre")[my_row]["vehiculo__aplicacion__nombre"]), font_style)
        ws.write(my_row + 1, 5, str(llantas.values("vehiculo__clase")[my_row]["vehiculo__clase"]), font_style)
        ws.write(my_row + 1, 6, str(llantas.values("nombre_de_eje")[my_row]["nombre_de_eje"]), font_style)
        ws.write(my_row + 1, 7, str(llantas.values("producto__producto")[my_row]["producto__producto"]), font_style)
        ws.write(my_row + 1, 8, str(llantas.values("ultima_inspeccion__min_profundidad")[my_row]["ultima_inspeccion__min_profundidad"]), font_style)
        ws.write(my_row + 1, 9, str(periodo[my_row]), font_style)

    wb.save(response)
    return response

def informe_de_perdida_y_rendimiento(request):    

    if request.method =="POST":
        flota1 = request.POST.getlist("sucursal")
        aplicacion1 = request.POST.getlist("aplicacion")
        fecha1 = request.POST.get("fechaInicio")
        fecha2 = request.POST.get("fechaFin")
        fecha1 = functions.convertir_rango(fecha1)
        fecha2 = functions.convertir_rango(fecha2)
        primera_fecha = datetime.strptime(fecha1, "%Y/%m/%d").date()
        segunda_fecha = datetime.strptime(fecha2, "%Y/%m/%d").date()

        compania = Compania.objects.get(compania=request.user.perfil.compania)
        inicio = datetime(primera_fecha.year, primera_fecha.month, primera_fecha.day)
        fin    = datetime(segunda_fecha.year, segunda_fecha.month, segunda_fecha.day)
        lista_fechas = [(inicio + timedelta(days=d)).strftime("%Y-%m-%d") for d in range((fin - inicio).days + 1)]

        vehiculos = Vehiculo.objects.filter(compania=Compania.objects.get(compania=compania))

        if flota1:
            vehiculos = vehiculos.filter(functions.reduce(or_, [Q(ubicacion=Ubicacion.objects.get(nombre=f)) for f in flota1]))
        if aplicacion1:
            vehiculos = vehiculos.filter(functions.reduce(or_, [Q(aplicacion=Aplicacion.objects.get(nombre=a)) for a in aplicacion1]))

        llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania=compania), vehiculo__in=vehiculos)
        inspecciones = Inspeccion.objects.filter(llanta__in=llantas)
        productos_llanta = llantas.values("producto").distinct()
        productos = Producto.objects.filter(id__in=productos_llanta)
        flotas_vehiculo = vehiculos.values("ubicacion__nombre").distinct()
        flotas = Ubicacion.objects.filter(compania=compania, nombre__in=flotas_vehiculo)
        aplicaciones_vehiculo = vehiculos.values("aplicacion__nombre").distinct()
        aplicaciones = Aplicacion.objects.filter(compania=Compania.objects.get(compania=compania), nombre__in=aplicaciones_vehiculo)
        ejes = llantas.values("nombre_de_eje").distinct()
        clases = vehiculos.values("clase").distinct()

        regresion = functions.km_proyectado(inspecciones, True)
        llantas_limpias = regresion[4]

        comparativa_de_productos = {}
        for producto in productos:
            valores_producto = []

            llantas_producto_total = llantas.filter(producto=producto)
            llantas_producto = llantas.filter(producto=producto, numero_economico__in=llantas_limpias)

            print("llantas_producto", llantas_producto)

            inspecciones_producto = Inspeccion.objects.filter(llanta__in=llantas_producto)
            regresion_producto = functions.km_proyectado(inspecciones_producto, False)
            km_proyectado_producto = regresion_producto[0]
            cpk_producto = regresion_producto[2]

            valores_producto.append(km_proyectado_producto)
            valores_producto.append(cpk_producto)
            print("valores_producto", valores_producto)

            dibujo = producto.dibujo
            valores_producto.append(dibujo)
            if dibujo and km_proyectado_producto != 0:
                comparativa_de_productos[producto] = valores_producto

        comparativa_de_flotas = {}
        for flota in flotas:
            valores_flota = []

            llantas_flota = llantas.filter(vehiculo__ubicacion=flota, numero_economico__in=llantas_limpias)
            if llantas_flota:
                inspecciones_flota = Inspeccion.objects.filter(llanta__in=llantas_flota)
                regresion_flota = functions.km_proyectado(inspecciones_flota, False)
                km_proyectado_flota = regresion_flota[0]
                cpk_flota = regresion_flota[2]

                valores_flota.append(km_proyectado_flota)
                valores_flota.append(cpk_flota)

                comparativa_de_flotas[flota] = valores_flota

        comparativa_de_aplicaciones = {}
        for aplicacion in aplicaciones:
            valores_aplicacion = []

            llantas_aplicacion = llantas.filter(vehiculo__aplicacion =aplicacion, numero_economico__in=llantas_limpias)
            if llantas_aplicacion:

                inspecciones_aplicacion = Inspeccion.objects.filter(llanta__in=llantas_aplicacion)
                regresion_aplicacion = functions.km_proyectado(inspecciones_aplicacion, False)
                km_proyectado_aplicacion = regresion_aplicacion[0]
                cpk_aplicacion = regresion_aplicacion[2]

                valores_aplicacion.append(km_proyectado_aplicacion)
                valores_aplicacion.append(cpk_aplicacion)

                comparativa_de_aplicaciones[aplicacion] = valores_aplicacion

        comparativa_de_ejes = {}
        for eje in ejes:
            valores_eje = []

            llantas_eje = llantas.filter(nombre_de_eje=eje["nombre_de_eje"], numero_economico__in=llantas_limpias)
            inspecciones_eje = Inspeccion.objects.filter(llanta__in=llantas_eje)
            if inspecciones_eje.exists():
                regresion_eje = functions.km_proyectado(inspecciones_eje, False)
                km_proyectado_eje = regresion_eje[0]
                cpk_eje = regresion_eje[2]

                valores_eje.append(km_proyectado_eje)
                valores_eje.append(cpk_eje)

                comparativa_de_ejes[eje["nombre_de_eje"]] = valores_eje

        comparativa_de_clases = {}
        for clase in clases:
            valores_clase = []

            llantas_clase = llantas.filter(vehiculo__clase=clase["clase"].upper(), numero_economico__in=llantas_limpias)
            inspecciones_clase = Inspeccion.objects.filter(llanta__in=llantas_clase)
            if inspecciones_clase.exists():
                regresion_clase = functions.km_proyectado(inspecciones_clase, False)
                km_proyectado_clase = regresion_clase[0]
                cpk_clase = regresion_clase[2]

                valores_clase.append(km_proyectado_clase)
                valores_clase.append(cpk_clase)

                comparativa_de_clases[clase["clase"]] = valores_clase

        print(comparativa_de_flotas)
        print(comparativa_de_aplicaciones)
        print(comparativa_de_ejes)
        print(comparativa_de_productos)
        print(comparativa_de_clases)


        eje_titulos_flota = []
        for i in range(len(comparativa_de_flotas)):
            com_flota_titulo = list(comparativa_de_flotas.keys())[i]
            eje_titulos_flota.append(com_flota_titulo)

        eje_y1_flota = []
        for i in range(len(comparativa_de_flotas)):
            com_flota_uno_valor = list(comparativa_de_flotas.values())[i][0]
            eje_y1_flota.append(com_flota_uno_valor)

        eje_y2_flota = []
        for i in range(len(comparativa_de_flotas)):
            com_flota_uno_valor = list(comparativa_de_flotas.values())[i][1]
            eje_y2_flota.append(com_flota_uno_valor)

        x_pos = np.arange(len(eje_titulos_flota))

        fig, ax = plt.subplots()
        fontsize = 14

        ax2 = ax.twinx()

        ax.bar(x_pos, eje_y1_flota, align='center', alpha=0.5, ecolor='black',
        capsize=3, width=0.2, color='r', label='data1')

        ax2.bar(x_pos+0.2, eje_y2_flota, align='center', alpha=0.5, ecolor='black',
        capsize=3, width=0.2, color='b', label='data2')

        for p in ax.patches:
            ax.annotate(np.round(p.get_height(),decimals=2), (p.get_x()+p.get_width()/2., p.get_height()), ha='center', va='top', xytext=(0, 10), textcoords='offset points')
        for p in ax2.patches:
            ax2.annotate(np.round(p.get_height(),decimals=3), (p.get_x()+p.get_width()/2., p.get_height()), ha='center', va='top', xytext=(0, 10), textcoords='offset points')

        ax.set_ylabel('Km proyectado')
        ax2.set_ylabel('CPK')

        ax.set_xticks(x_pos)
        ax.set_xticklabels(eje_titulos_flota, fontsize=fontsize)
        ax.set_title('Comparativa flotas')


        plt.grid()
        plt.savefig(os.path.abspath(os.getcwd()) + r"/aetoweb/files/files/ComparacionFlotas.png", dpi = 70, bbox_inches="tight")

        img_flotas = openpyxl.drawing.image.Image(os.path.abspath(os.getcwd()) + r"/aetoweb/files/files/ComparacionFlotas.png")

        eje_titulos_aplicaciones = []
        for i in range(len(comparativa_de_aplicaciones)):
            com_titulo = list(comparativa_de_aplicaciones.keys())[i]
            eje_titulos_aplicaciones.append(com_titulo)

        eje_y1_aplicaciones = []
        for i in range(len(comparativa_de_aplicaciones)):
            com_uno_valor = list(comparativa_de_aplicaciones.values())[i][0]
            eje_y1_aplicaciones.append(com_uno_valor)

        eje_y2_aplicaciones = []
        for i in range(len(comparativa_de_aplicaciones)):
            com_uno_valor = list(comparativa_de_aplicaciones.values())[i][1]
            eje_y2_aplicaciones.append(com_uno_valor)
        x_pos = np.arange(len(eje_titulos_aplicaciones))
        fig, ax = plt.subplots()
        fontsize = 14
        ax2 = ax.twinx()
        ax.bar(x_pos, eje_y1_aplicaciones, align='center', alpha=0.5, ecolor='black',
        capsize=3, width=0.2, color='r', label='data1')
        ax2.bar(x_pos+0.2, eje_y2_aplicaciones, align='center', alpha=0.5, ecolor='black',
        capsize=3, width=0.2, color='b', label='data2')

        for p in ax.patches:
            ax.annotate(np.round(p.get_height(),decimals=2), (p.get_x()+p.get_width()/2., p.get_height()), ha='center', va='top', xytext=(0, 10), textcoords='offset points')
        for p in ax2.patches:
            ax2.annotate(np.round(p.get_height(),decimals=3), (p.get_x()+p.get_width()/2., p.get_height()), ha='center', va='top', xytext=(0, 10), textcoords='offset points')

        ax.set_ylabel('Km proyectado')
        ax2.set_ylabel('CPK')
        ax.set_xticks(x_pos)
        ax.set_xticklabels(eje_titulos_aplicaciones, fontsize=fontsize)
        ax.set_title('Comparación aplicaciones')
        plt.grid()
        plt.savefig(os.path.abspath(os.getcwd()) + r"/aetoweb/files/files/ComparacionAplicaciones.png", dpi = 70, bbox_inches="tight")

        img_aplicaciones = openpyxl.drawing.image.Image(os.path.abspath(os.getcwd()) + r"/aetoweb/files/files/ComparacionAplicaciones.png")

        eje_titulos_ejes = []
        for i in range(len(comparativa_de_ejes)):
            com_ejes_titulo = list(comparativa_de_ejes.keys())[i]
            eje_titulos_ejes.append(com_ejes_titulo)

        eje_y1_ejes = []
        for i in range(len(comparativa_de_ejes)):
            com_ejes_valor1 = list(comparativa_de_ejes.values())[i][0]
            eje_y1_ejes.append(com_ejes_valor1)

        eje_y2_ejes = []
        for i in range(len(comparativa_de_ejes)):
            com_ejes_valor2 = list(comparativa_de_ejes.values())[i][1]
            eje_y2_ejes.append(com_ejes_valor2)

        x_pos = np.arange(len(eje_titulos_ejes))

        fig, ax = plt.subplots()
        fontsize = 14

        ax2 = ax.twinx()

        ax.bar(x_pos, eje_y1_ejes, alpha=0.5, ecolor='black',
        capsize=3, width=0.2, color='r', label='data1')

        ax2.bar(x_pos+0.2, eje_y2_ejes, alpha=0.5, ecolor='black',
        capsize=3, width=0.2, color='b', label='data2')

        for p in ax.patches:
            ax.annotate(np.round(p.get_height(),decimals=2), (p.get_x()+p.get_width()/2., p.get_height()), ha='center', va='top', xytext=(0, 10), textcoords='offset points')
        for p in ax2.patches:
            ax2.annotate(np.round(p.get_height(),decimals=3), (p.get_x()+p.get_width()/2., p.get_height()), ha='center', va='top', xytext=(0, 10), textcoords='offset points')

        ax.set_ylabel('Km proyectado')
        ax2.set_ylabel('CPK')


        ax.set_xlabel(eje_titulos_ejes[int(x_pos)])
        ax.set_xticklabels(eje_titulos_ejes, fontsize=fontsize)
        ax.set_title('Comparativa ejes')

        plt.grid()
        plt.savefig(os.path.abspath(os.getcwd()) + r"/aetoweb/files/files/ComparacionEjes.png", dpi = 70, bbox_inches="tight")

        img_ejes = openpyxl.drawing.image.Image(os.path.abspath(os.getcwd()) + r"/aetoweb/files/files/ComparacionEjes.png")

        eje_titulos_producto = []
        for i in range(len(comparativa_de_productos)):
            com_titulo = list(comparativa_de_productos.values())[i][2]
            eje_titulos_producto.append(com_titulo)

        eje_y1_producto = []
        for i in range(len(comparativa_de_productos)):
            com_uno_valor = list(comparativa_de_productos.values())[i][0]
            eje_y1_producto.append(com_uno_valor)

        eje_y2_producto = []
        for i in range(len(comparativa_de_productos)):
            com_uno_valor = list(comparativa_de_productos.values())[i][1]
            eje_y2_producto.append(com_uno_valor)

        x_pos = np.arange(len(eje_titulos_producto))

        fig, ax = plt.subplots()
        fontsize = 14

        ax2 = ax.twinx()

        ax.bar(x_pos, eje_y1_producto, align='center', alpha=0.5, ecolor='black',
        capsize=3, width=0.2, color='r', label='data1')

        ax2.bar(x_pos+0.2, eje_y2_producto, align='center', alpha=0.5, ecolor='black',
        capsize=3, width=0.2, color='b', label='data2')

        for p in ax.patches:
            ax.annotate(np.round(p.get_height(),decimals=2), (p.get_x()+p.get_width()/2., p.get_height()), ha='center', va='top', xytext=(0, 10), textcoords='offset points')
        for p in ax2.patches:
            ax2.annotate(np.round(p.get_height(),decimals=3), (p.get_x()+p.get_width()/2., p.get_height()), ha='center', va='top', xytext=(0, 10), textcoords='offset points')

        ax.set_ylabel('Km proyectado')
        ax2.set_ylabel('CPK')


        ax.set_xticks(x_pos)
        ax.set_xticklabels(eje_titulos_producto, fontsize=fontsize)
        ax.set_title('Comparativa productos')
        plt.grid()
        plt.savefig(os.path.abspath(os.getcwd()) + r"/aetoweb/files/files/ComparacionProductos.png", dpi = 70, bbox_inches="tight")

        img_producto = openpyxl.drawing.image.Image(os.path.abspath(os.getcwd()) + r"/aetoweb/files/files/ComparacionProductos.png")

        eje_titulos_clase = []
        for i in range(len(comparativa_de_clases)):
            com_clase_titulo = list(comparativa_de_clases.keys())[i]
            eje_titulos_clase.append(com_clase_titulo)

        eje_y1_clase = []
        for i in range(len(comparativa_de_clases)):
            com_clase_uno_valor = list(comparativa_de_clases.values())[i][0]
            eje_y1_clase.append(com_clase_uno_valor)

        eje_y2_clase = []
        for i in range(len(comparativa_de_clases)):
            com_clase_uno_valor = list(comparativa_de_clases.values())[i][1]
            eje_y2_clase.append(com_clase_uno_valor)

        x_pos = np.arange(len(eje_titulos_clase))

        fig, ax = plt.subplots()
        fontsize = 14

        ax2 = ax.twinx()

        ax.bar(x_pos, eje_y1_clase, align='center', alpha=0.5, ecolor='black',
        capsize=3, width=0.2, color='r', label='data1')

        ax2.bar(x_pos+0.2, eje_y2_clase, align='center', alpha=0.5, ecolor='black',
        capsize=3, width=0.2, color='b', label='data2')

        for p in ax.patches:
            ax.annotate(np.round(p.get_height(),decimals=2), (p.get_x()+p.get_width()/2., p.get_height()), ha='center', va='top', xytext=(0, 10), textcoords='offset points')
        for p in ax2.patches:
            ax2.annotate(np.round(p.get_height(),decimals=3), (p.get_x()+p.get_width()/2., p.get_height()), ha='center', va='top', xytext=(0, 10), textcoords='offset points')

        ax.set_ylabel('Km proyectado')
        ax2.set_ylabel('CPK')

        ax.set_xticks(x_pos)
        ax.set_xticklabels(eje_titulos_clase, fontsize=fontsize)
        ax.set_title('Comparativa clases')
        plt.grid()
        plt.savefig(os.path.abspath(os.getcwd()) + r"/aetoweb/files/files/ComparacionClases.png", dpi = 70, bbox_inches="tight")

        img_clases = openpyxl.drawing.image.Image(os.path.abspath(os.getcwd()) + r"/aetoweb/files/files/ComparacionClases.png")

        #creating workbook
        wb = openpyxl.Workbook()
        wb.create_sheet("Perdida")
        reporte = wb.get_sheet_by_name('Perdida')
        wb.remove(wb['Sheet'])

        wb.create_sheet("Rendimiento")
        reporte2 = wb.get_sheet_by_name('Rendimiento')

        reporte2.add_image(img_flotas, 'A1')
        reporte2.add_image(img_aplicaciones, 'L1')
        reporte2.add_image(img_ejes, 'A20')
        reporte2.add_image(img_producto, 'L20')
        reporte2.add_image(img_clases, 'A40')

        e1 = reporte.cell(row=1, column=1, value='VehicleRegistrationNumber')
        e2 = reporte.cell(row=1, column=2, value='CompanyName')
        e3 = reporte.cell(row=1, column=3, value='FleetName')
        e4 = reporte.cell(row=1, column=4, value='LocationName')
        e5 = reporte.cell(row=1, column=5, value='InspectionBeginTime')
        e6 = reporte.cell(row=1, column=6, value='InspectionFullNumber')
        e7 = reporte.cell(row=1, column=7, value='Inspection_Mileage')
        e8 = reporte.cell(row=1, column=8, value='VehicleClassName')
        e9 = reporte.cell(row=1, column=9, value='AxleConfigurationName')
        e10 = reporte.cell(row=1, column=10, value='VehicleMakeName')
        e11 = reporte.cell(row=1, column=11, value='VehicleModelName')
        e12 = reporte.cell(row=1, column=12, value='VehicleStatusId')
        e12 = reporte.cell(row=1, column=13, value='TyreSerialNumber')
        e12 = reporte.cell(row=1, column=14, value='Wheel_Position')
        e12 = reporte.cell(row=1, column=15, value='TyreMakeName')
        e12 = reporte.cell(row=1, column=16, value='TyrePatternName')
        e12 = reporte.cell(row=1, column=17, value='TyreSizeName')
        e12 = reporte.cell(row=1, column=18, value='TyreLifeName')
        e12 = reporte.cell(row=1, column=19, value='TD1')
        e12 = reporte.cell(row=1, column=20, value='TD2')
        e12 = reporte.cell(row=1, column=21, value='TD3')
        e12 = reporte.cell(row=1, column=22, value='Profundidad inicial')
        e12 = reporte.cell(row=1, column=23, value='Precio')
        e12 = reporte.cell(row=1, column=24, value='Min profundidad')
        e12 = reporte.cell(row=1, column=25, value='Prom')
        e12 = reporte.cell(row=1, column=26, value='%')
        e12 = reporte.cell(row=1, column=27, value='Dinero perdido 1')
        e12 = reporte.cell(row=1, column=28, value='Punto de retiro')
        e12 = reporte.cell(row=1, column=29, value='Pérdida total')

        FILE_PATH = os.path.abspath(os.getcwd()) + r"/aetoweb/files/files/Inspections_Bulk.csv"
        file = open(FILE_PATH, "r", encoding="latin-1", newline='')
        next(file, None)
        reader = csv.reader(file, delimiter=",")
        iteracion = 0

        for row in reader:
            if row[4][:10] in lista_fechas:
                iteracion += 1
                llanta = row[12]
                FILE_PATH = os.path.abspath(os.getcwd()) + r"/aetoweb/files/files/RollingStock2022_03_25_040126.csv"
                file2 = open(FILE_PATH, "r", encoding="latin-1", newline='')
                next(file2, None)
                reader2 = csv.reader(file2, delimiter=",")
                for row2 in reader2:
                    try:
                        llanta2 = row2[9]
                        if llanta == llanta2:

                            producto = row2[10]
                            FILE_PATH = os.path.abspath(os.getcwd()) + r"/aetoweb/files/files/Products2022_03_25_043513.csv"
                            file3 = open(FILE_PATH, "r", encoding="latin-1", newline='')
                            next(file3, None)
                            reader3 = csv.reader(file3, delimiter=",")
                            for row3 in reader3:
                                producto2 = row3[4]
                                if producto == producto2:
                                    profundidad_inicial = float(row3[10])
                                    precio = float(row3[12])


                            numero_de_eje = int(row2[17]) - 1
                            vehiculo = row2[6]
                            FILE_PATH = os.path.abspath(os.getcwd()) + "/aetoweb/files/files/Vehicles2022_03_25_043019.csv"
                            file4 = open(FILE_PATH, "r", encoding="latin-1", newline='')
                            next(file4, None)
                            reader4 = csv.reader(file4, delimiter=",")
                            for row4 in reader4:
                                vehiculo2 = row4[9]
                                if vehiculo == vehiculo2:
                                    configuracion = row4[14].split(".")
                                    if configuracion[numero_de_eje][0] == "S":
                                        punto_de_retiro = compania.punto_retiro_eje_direccion
                                    elif configuracion[numero_de_eje][0] == "D":
                                        punto_de_retiro = compania.punto_retiro_eje_traccion
                                    elif configuracion[numero_de_eje][0] == "T":
                                        punto_de_retiro = compania.punto_retiro_eje_arrastre
                                    elif configuracion[numero_de_eje][0] == "C":
                                        punto_de_retiro = compania.punto_retiro_eje_loco
                                    elif configuracion[numero_de_eje][0] == "L":
                                        punto_de_retiro = compania.punto_retiro_eje_retractil

                    except:
                        pass

                tds = [float(row[18]), float(row[19]), float(row[20])]
                min_profundidad = min(tds)
                prom = round(statistics.mean(tds), 2)
                try:
                    if profundidad_inicial == 0:
                        porcentaje = 0
                    else:
                        porcentaje = round((prom - min_profundidad) / profundidad_inicial, 2)
                except:
                    porcentaje = None
                    profundidad_inicial = None

                try:
                    if min_profundidad >= 1000:
                        dinero_perdido = 0
                    else:
                        dinero_perdido = round((precio * porcentaje), 2)
                except:
                    dinero_perdido = None
                    precio = None

                try:
                    perdida_total = round((min_profundidad - punto_de_retiro) * dinero_perdido)
                except:
                    perdida_total = None
                    punto_de_retiro = None

                column1 = reporte.cell(row=iteracion + 1, column=1)
                column1.value = row[0]
                column2 = reporte.cell(row=iteracion + 1, column=2)
                column2.value = row[1]
                column3 = reporte.cell(row=iteracion + 1, column=3)
                column3.value = row[2]
                column4 = reporte.cell(row=iteracion + 1, column=4)
                column4.value = row[3]
                column5 = reporte.cell(row=iteracion + 1, column=5)
                column5.value = row[4]
                column6 = reporte.cell(row=iteracion + 1, column=6)
                column6.value = row[5]
                column7 = reporte.cell(row=iteracion + 1, column=7)
                column7.value = row[6]
                column8 = reporte.cell(row=iteracion + 1, column=8)
                column8.value = row[7]
                column9 = reporte.cell(row=iteracion + 1, column=9)
                column9.value = row[8]
                column10 = reporte.cell(row=iteracion + 1, column=10)
                column10.value = row[9]
                column11 = reporte.cell(row=iteracion + 1, column=11)
                column11.value = row[10]
                column12 = reporte.cell(row=iteracion + 1, column=12)
                column12.value = row[11]
                column13 = reporte.cell(row=iteracion + 1, column=13)
                column13.value = row[12]
                column14 = reporte.cell(row=iteracion + 1, column=14)
                column14.value = row[13]
                column15 = reporte.cell(row=iteracion + 1, column=15)
                column15.value = row[14]
                column16 = reporte.cell(row=iteracion + 1, column=16)
                column16.value = row[15]
                column17 = reporte.cell(row=iteracion + 1, column=17)
                column17.value = row[16]
                column18 = reporte.cell(row=iteracion + 1, column=18)
                column18.value = row[17]
                column19 = reporte.cell(row=iteracion + 1, column=19)
                column19.value = row[18]
                column20 = reporte.cell(row=iteracion + 1, column=20)
                column20.value = row[19]
                column21 = reporte.cell(row=iteracion + 1, column=21)
                column21.value = row[20]
                column22 = reporte.cell(row=iteracion + 1, column=22)
                column22.value = profundidad_inicial
                column23 = reporte.cell(row=iteracion + 1, column=23)
                column23.value = precio
                column24 = reporte.cell(row=iteracion + 1, column=24)
                column24.value = min_profundidad
                column25 = reporte.cell(row=iteracion + 1, column=25)
                column25.value = prom
                column26 = reporte.cell(row=iteracion + 1, column=26)
                column26.value = porcentaje
                column27 = reporte.cell(row=iteracion + 1, column=27)
                column27.value = dinero_perdido
                column28 = reporte.cell(row=iteracion + 1, column=28)
                column28.value = punto_de_retiro
                column29 = reporte.cell(row=iteracion + 1, column=29)
                column29.value = perdida_total

        file.close()

        response = HttpResponse(content_type='application/ms-excel')

        #decide file name
        response['Content-Disposition'] = 'attachment; filename="InformePerdidaYRendimiento.xlsx"'

        wb.save(response)
        return response

def download_todo(request):
    # Define Django project base directory
    # content-type of response
    vehiculos = Vehiculo.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania))
    llantas = Llanta.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania))
    productos = Producto.objects.filter(compania=Compania.objects.get(compania=request.user.perfil.compania))
    inspecciones = Inspeccion.objects.filter(llanta__compania=Compania.objects.get(compania=request.user.perfil.compania))

    regresion = functions.km_proyectado(inspecciones, True)
    llantas_limpias = regresion[4]
    llantas_analizadas = llantas.filter(numero_economico__in=llantas_limpias)

    response = HttpResponse(content_type='application/ms-excel')

	#decide file name
    response['Content-Disposition'] = 'attachment; filename="Info.xls"'

	#creating workbook
    wb = xlwt.Workbook(encoding='utf-8')

	#adding sheet
    ws = wb.add_sheet("Vehiculos")

	# Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
	# headers are bold
    font_style.font.bold = True

	#column header names, you can use your own headers here
    columns = ['Número Económico', 'Modelo', 'Marca', 'Compania', 'Sucursal', 'Aplicacion', 'Clase', 'Configuracion', 'Presion establecida 1', 'Presion establecida 2', 'Presion establecida 3', 'Presion establecida 4', 'Presion establecida 5', 'Presion establecida 6', 'Presion establecida 7', "Km", "Km diario máximo"]

    #write column headers in sheet
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    #get your data, from database or from a text file...
    for my_row in range(len(vehiculos)):
        ws.write(my_row + 1, 0, str(vehiculos[my_row]), font_style)
        ws.write(my_row + 1, 1, str(vehiculos.values("modelo")[my_row]["modelo"]), font_style)
        ws.write(my_row + 1, 2, str(vehiculos.values("marca")[my_row]["marca"]), font_style)
        ws.write(my_row + 1, 3, str(vehiculos.values("compania__compania")[my_row]["compania__compania"]), font_style)
        ws.write(my_row + 1, 4, str(vehiculos.values("ubicacion__nombre")[my_row]["ubicacion__nombre"]), font_style)
        ws.write(my_row + 1, 5, str(vehiculos.values("aplicacion__nombre")[my_row]["aplicacion__nombre"]), font_style)
        ws.write(my_row + 1, 6, str(vehiculos.values("clase")[my_row]["clase"]), font_style)
        ws.write(my_row + 1, 7, str(vehiculos.values("configuracion")[my_row]["configuracion"]), font_style)
        ws.write(my_row + 1, 8, str(vehiculos.values("presion_establecida_1")[my_row]["presion_establecida_1"]), font_style)
        ws.write(my_row + 1, 9, str(vehiculos.values("presion_establecida_2")[my_row]["presion_establecida_2"]), font_style)
        ws.write(my_row + 1, 10, str(vehiculos.values("presion_establecida_3")[my_row]["presion_establecida_3"]), font_style)
        ws.write(my_row + 1, 11, str(vehiculos.values("presion_establecida_4")[my_row]["presion_establecida_4"]), font_style)
        ws.write(my_row + 1, 12, str(vehiculos.values("presion_establecida_5")[my_row]["presion_establecida_5"]), font_style)
        ws.write(my_row + 1, 13, str(vehiculos.values("presion_establecida_6")[my_row]["presion_establecida_6"]), font_style)
        ws.write(my_row + 1, 14, str(vehiculos.values("presion_establecida_7")[my_row]["presion_establecida_7"]), font_style)
        ws.write(my_row + 1, 15, str(vehiculos.values("km")[my_row]["km"]), font_style)
        ws.write(my_row + 1, 16, str(vehiculos.values("km_diario_maximo")[my_row]["km_diario_maximo"]), font_style)


	#adding sheet
    ws2 = wb.add_sheet("Llantas")

	# Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
	# headers are bold
    font_style.font.bold = True

	#column header names, you can use your own headers here
    columns = ['Número Económico', 'Compania', 'Vehiculo', 'Ubicacion', 'Aplicacion', 'Taller', 'Vida', 'Tipo de Eje', 'Nombre de Eje', 'Posición', 'Presion actual', 'Profundidad izquierda', 'Profundidad central', "Profundidad derecha", "Km actual", "Km montado", "Producto", "Inventario", "Fecha de entrada inventario", "Condición de desecho", "Razón de desecho"]

    #write column headers in sheet
    for col_num in range(len(columns)):
        ws2.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    #get your data, from database or from a text file...
    for my_row in range(len(llantas)):
        ws2.write(my_row + 1, 0, str(llantas[my_row]), font_style)
        ws2.write(my_row + 1, 1, str(llantas.values("vehiculo__compania__compania")[my_row]["vehiculo__compania__compania"]), font_style)
        ws2.write(my_row + 1, 2, str(llantas.values("vehiculo__numero_economico")[my_row]["vehiculo__numero_economico"]), font_style)
        ws2.write(my_row + 1, 3, str(llantas.values("ubicacion__nombre")[my_row]["ubicacion__nombre"]), font_style)
        ws2.write(my_row + 1, 4, str(llantas.values("aplicacion__nombre")[my_row]["aplicacion__nombre"]), font_style)
        ws2.write(my_row + 1, 5, str(llantas.values("taller")[my_row]["taller"]), font_style)
        ws2.write(my_row + 1, 6, str(llantas.values("vida")[my_row]["vida"]), font_style)
        ws2.write(my_row + 1, 7, str(llantas.values("tipo_de_eje")[my_row]["tipo_de_eje"]), font_style)
        ws2.write(my_row + 1, 8, str(llantas.values("nombre_de_eje")[my_row]["nombre_de_eje"]), font_style)
        ws2.write(my_row + 1, 9, str(llantas.values("posicion")[my_row]["posicion"]), font_style)
        ws2.write(my_row + 1, 10, str(llantas.values("presion_actual")[my_row]["presion_actual"]), font_style)
        ws2.write(my_row + 1, 11, str(llantas.values("profundidad_izquierda")[my_row]["profundidad_izquierda"]), font_style)
        ws2.write(my_row + 1, 12, str(llantas.values("profundidad_central")[my_row]["profundidad_central"]), font_style)
        ws2.write(my_row + 1, 13, str(llantas.values("profundidad_derecha")[my_row]["profundidad_derecha"]), font_style)
        ws2.write(my_row + 1, 14, str(llantas.values("km_actual")[my_row]["km_actual"]), font_style)
        ws2.write(my_row + 1, 15, str(llantas.values("km_montado")[my_row]["km_montado"]), font_style)
        ws2.write(my_row + 1, 16, str(llantas.values("producto__producto")[my_row]["producto__producto"]), font_style)
        ws2.write(my_row + 1, 17, str(llantas.values("inventario")[my_row]["inventario"]), font_style)
        ws2.write(my_row + 1, 18, str(llantas.values("fecha_de_entrada_inventario")[my_row]["fecha_de_entrada_inventario"]), font_style)
        ws2.write(my_row + 1, 19, str(llantas.values("desecho__condicion")[my_row]["desecho__condicion"]), font_style)
        ws2.write(my_row + 1, 20, str(llantas.values("desecho__razon")[my_row]["desecho__razon"]), font_style)

	#adding sheet
    ws3 = wb.add_sheet("Productos")

	# Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
	# headers are bold
    font_style.font.bold = True

	#column header names, you can use your own headers here
    columns = ['Producto', 'Marca', 'Dibujo', 'Rango', 'Dimensión', 'Profundidad inicial', 'Vida', 'Precio']

    #write column headers in sheet
    for col_num in range(len(columns)):
        ws3.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    #get your data, from database or from a text file...
    for my_row in range(len(productos)):
        ws3.write(my_row + 1, 0, str(productos[my_row]), font_style)
        ws3.write(my_row + 1, 1, str(productos.values("marca")[my_row]["marca"]), font_style)
        ws3.write(my_row + 1, 2, str(productos.values("dibujo")[my_row]["dibujo"]), font_style)
        ws3.write(my_row + 1, 3, str(productos.values("rango")[my_row]["rango"]), font_style)
        ws3.write(my_row + 1, 4, str(productos.values("dimension")[my_row]["dimension"]), font_style)
        ws3.write(my_row + 1, 5, str(productos.values("profundidad_inicial")[my_row]["profundidad_inicial"]), font_style)
        ws3.write(my_row + 1, 6, str(productos.values("vida")[my_row]["vida"]), font_style)
        ws3.write(my_row + 1, 7, str(productos.values("precio")[my_row]["precio"]), font_style)


	#adding sheet
    ws4 = wb.add_sheet("Inspecciones")

	# Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
	# headers are bold
    font_style.font.bold = True

	#column header names, you can use your own headers here
    columns = ['Llanta', 'Usuario', 'Vehículo', 'Fecha', 'Vida', 'Km', 'Profundidad izquierda', 'Profundidad central', 'Profundidad derecha', 'Presión']

    #write column headers in sheet
    for col_num in range(len(columns)):
        ws4.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    #get your data, from database or from a text file...
    for my_row in range(len(inspecciones)):
        ws4.write(my_row + 1, 0, str(inspecciones.values("llanta__numero_economico")[my_row]["llanta__numero_economico"]), font_style)
        ws4.write(my_row + 1, 1, str(inspecciones.values("usuario__user__username")[my_row]["usuario__user__username"]), font_style)
        ws4.write(my_row + 1, 2, str(inspecciones.values("vehiculo__numero_economico")[my_row]["vehiculo__numero_economico"]), font_style)
        ws4.write(my_row + 1, 3, str(inspecciones.values("fecha_hora")[my_row]["fecha_hora"]), font_style)
        ws4.write(my_row + 1, 4, str(inspecciones.values("vida")[my_row]["vida"]), font_style)
        ws4.write(my_row + 1, 5, str(inspecciones.values("km_vehiculo")[my_row]["km_vehiculo"]), font_style)
        ws4.write(my_row + 1, 6, str(inspecciones.values("profundidad_izquierda")[my_row]["profundidad_izquierda"]), font_style)
        ws4.write(my_row + 1, 7, str(inspecciones.values("profundidad_central")[my_row]["profundidad_central"]), font_style)
        ws4.write(my_row + 1, 8, str(inspecciones.values("profundidad_derecha")[my_row]["profundidad_derecha"]), font_style)
        ws4.write(my_row + 1, 9, str(inspecciones.values("presion")[my_row]["presion"]), font_style)

	#adding sheet
    ws5 = wb.add_sheet("Llantas Analizadas")

	# Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
	# headers are bold
    font_style.font.bold = True

	#column header names, you can use your own headers here
    columns = ['Número Económico', 'Compania', 'Vehiculo', 'Ubicacion', 'Aplicacion', 'Taller', 'Vida', 'Tipo de Eje', 'Nombre de Eje', 'Posición', 'Presion actual', 'Profundidad izquierda', 'Profundidad central', "Profundidad derecha", "Km actual", "Km montado", "Producto", "Inventario", "Fecha de entrada inventario", "CPK", "Km proyectado", "Condición de desecho", "Razón de desecho"]

    #write column headers in sheet
    for col_num in range(len(columns)):
        ws5.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    #get your data, from database or from a text file...
    for my_row in range(len(llantas_analizadas)):
        ws5.write(my_row + 1, 0, str(llantas_analizadas[my_row]), font_style)
        ws5.write(my_row + 1, 1, str(llantas_analizadas.values("vehiculo__compania__compania")[my_row]["vehiculo__compania__compania"]), font_style)
        ws5.write(my_row + 1, 2, str(llantas_analizadas.values("vehiculo__numero_economico")[my_row]["vehiculo__numero_economico"]), font_style)
        ws5.write(my_row + 1, 3, str(llantas_analizadas.values("ubicacion__nombre")[my_row]["ubicacion__nombre"]), font_style)
        ws5.write(my_row + 1, 4, str(llantas_analizadas.values("aplicacion__nombre")[my_row]["aplicacion__nombre"]), font_style)
        ws5.write(my_row + 1, 5, str(llantas_analizadas.values("taller")[my_row]["taller"]), font_style)
        ws5.write(my_row + 1, 6, str(llantas_analizadas.values("vida")[my_row]["vida"]), font_style)
        ws5.write(my_row + 1, 7, str(llantas_analizadas.values("tipo_de_eje")[my_row]["tipo_de_eje"]), font_style)
        ws5.write(my_row + 1, 8, str(llantas_analizadas.values("nombre_de_eje")[my_row]["nombre_de_eje"]), font_style)
        ws5.write(my_row + 1, 9, str(llantas_analizadas.values("posicion")[my_row]["posicion"]), font_style)
        ws5.write(my_row + 1, 10, str(llantas_analizadas.values("presion_actual")[my_row]["presion_actual"]), font_style)
        ws5.write(my_row + 1, 11, str(llantas_analizadas.values("profundidad_izquierda")[my_row]["profundidad_izquierda"]), font_style)
        ws5.write(my_row + 1, 12, str(llantas_analizadas.values("profundidad_central")[my_row]["profundidad_central"]), font_style)
        ws5.write(my_row + 1, 13, str(llantas_analizadas.values("profundidad_derecha")[my_row]["profundidad_derecha"]), font_style)
        ws5.write(my_row + 1, 14, str(llantas_analizadas.values("km_actual")[my_row]["km_actual"]), font_style)
        ws5.write(my_row + 1, 15, str(llantas_analizadas.values("km_montado")[my_row]["km_montado"]), font_style)
        ws5.write(my_row + 1, 16, str(llantas_analizadas.values("producto__producto")[my_row]["producto__producto"]), font_style)
        ws5.write(my_row + 1, 17, str(llantas_analizadas.values("inventario")[my_row]["inventario"]), font_style)
        ws5.write(my_row + 1, 18, str(llantas_analizadas.values("fecha_de_entrada_inventario")[my_row]["fecha_de_entrada_inventario"]), font_style)
        ws5.write(my_row + 1, 19, str(regresion[3][my_row]), font_style)
        ws5.write(my_row + 1, 20, str(regresion[5][my_row]), font_style)
        ws5.write(my_row + 1, 21, str(llantas_analizadas.values("desecho__condicion")[my_row]["desecho__condicion"]), font_style)
        ws5.write(my_row + 1, 22, str(llantas_analizadas.values("desecho__razon")[my_row]["desecho__razon"]), font_style)


    wb.save(response)
    return response

def ftp_newpick(request):

    functions_ftp.ftp_descarga()

def download_llantas_rodantes(request):
    # Define Django project base directory
    # content-type of response
    #compania=Compania.objects.get(compania=request.user.perfil.compania)
    #print(compania)
    #llantas_patito = Llanta.objects.filter(compania = compania, patito=True)
#
    #response = HttpResponse(content_type='application/ms-excel')
#
	##decide file name
    #response['Content-Disposition'] = 'attachment; filename="Llantas_info.xls"'
#
	##creating workbook
    #wb = xlwt.Workbook(encoding='utf-8')
#
	##adding sheet
    #ws = wb.add_sheet("Llantas")
#
	## Sheet header, first row
    #row_num = 0
#
    #font_style = xlwt.XFStyle()
	## headers are bold
    #font_style.font.bold = True
#
	##column header names, you can use your own headers here
    #columns = ['Número Económico', 'Número económico nuevo', 'Vehículo', 'Sucursal', 'Aplicación', 'Vida', 'Posición', 'Presión actual', 'Profundidad izquierda', "Profundidad central", "Profundidad derecha", "Km actual", "Producto", "Inventario", "Km de montado"]
#
    ##write column headers in sheet
    #for col_num in range(len(columns)):
    #    ws.write(row_num, col_num, columns[col_num], font_style)
#
    ## Sheet body, remaining rows
    #font_style = xlwt.XFStyle()
#
    ##get your data, from database or from a text file...
    #for my_row in range(len(llantas_patito)):
    #    ws.write(my_row + 1, 0, str(llantas_patito.values("numero_economico")[my_row]["numero_economico"]), font_style)
    #    ws.write(my_row + 1, 2, str(llantas_patito.values("vehiculo__numero_economico")[my_row]["vehiculo__numero_economico"]), font_style)
    #    ws.write(my_row + 1, 3, str(llantas_patito.values("ubicacion__nombre")[my_row]["ubicacion__nombre"]), font_style)
    #    ws.write(my_row + 1, 4, str(llantas_patito.values("aplicacion__nombre")[my_row]["aplicacion__nombre"]), font_style)
    #    ws.write(my_row + 1, 5, str(llantas_patito.values("vida")[my_row]["vida"]), font_style)
    #    ws.write(my_row + 1, 6, str(llantas_patito.values("posicion")[my_row]["posicion"]), font_style)
    #    ws.write(my_row + 1, 7, str(llantas_patito.values("presion_actual")[my_row]["presion_actual"]), font_style)
    #    ws.write(my_row + 1, 13, str(llantas_patito.values("inventario")[my_row]["inventario"]), font_style)
#
    #wb.save(response)
    
    
    usuario = request.user
    perfil = Perfil.objects.get(user = usuario)
    compania = perfil.compania
    llantas_patito = Llanta.objects.filter(compania = compania, patito=True)
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append( (
            'Número Económico',
            'Número económico nuevo',
            'Vehículo',
            'Sucursal',
            'Aplicación',
            'Vida',
            'Posición',
            'Presión actual',
            'Profundidad izquierda',
            "Profundidad central",
            "Profundidad derecha",
            "Km actual",
            "Producto",
            "Inventario",
            "Km de montado"
        ) )
    fuente = Font(bold=True)
    celdas_en_negrita = sheet["A1":"O1"]
    for row in celdas_en_negrita:
        for cell in row:
            cell.font = fuente

    
    for llanta in llantas_patito:
        sheet.append( (
                llanta.numero_economico,
                None,
                llanta.vehiculo.numero_economico,
                llanta.ubicacion.nombre,
                llanta.aplicacion.nombre,
                llanta.vida,
                llanta.posicion,
                llanta.presion_actual,
                llanta.profundidad_izquierda,
                llanta.profundidad_central,
                llanta.profundidad_derecha,
                llanta.km_actual,
                llanta.producto.producto,
                llanta.inventario,
                llanta.km_montado,
        ) )
    sheet.column_dimensions [ 'A' ]. ancho = 500
    sheet.column_dimensions [ 'B' ]. ancho = 500
    sheet.column_dimensions [ 'C' ]. ancho = 500
    sheet.column_dimensions [ 'D' ]. ancho = 500
    sheet.column_dimensions [ 'E' ]. ancho = 500
    sheet.column_dimensions [ 'F' ]. ancho = 500
    sheet.column_dimensions [ 'G' ]. ancho = 500
    sheet.column_dimensions [ 'H' ]. ancho = 500
    sheet.column_dimensions [ 'I' ]. ancho = 500
    sheet.column_dimensions [ 'J' ]. ancho = 500
    sheet.column_dimensions [ 'K' ]. ancho = 500
    sheet.column_dimensions [ 'L' ]. ancho = 500
    sheet.column_dimensions [ 'M' ]. ancho = 500
    sheet.column_dimensions [ 'N' ]. ancho = 500
    sheet.column_dimensions [ 'O' ]. ancho = 500
    response = HttpResponse(content_type='application/msexcel')
    response['Content-Disposition'] = f'attachment; filename=Catalogo de Productos {compania}.xlsx'
    book.save(response)
    return response
    return response

# Consumo de llantas = Cantidad llantas rodantes * Km recorridos / Rendimiento de la llanta
def redireccionOrden(request, id):
    orden = Orden.objects.get(pk=id)
    usuario = request.user
    perfil = Perfil.objects.get(user = usuario)
    
    fecha = orden.fecha
    folio = orden.folio
    status = orden.status
    if status == 'PreOrden' or status == 'Orden':
        
        datos = json.loads(orden.datos)
        data_list = datos['productos'][0]
        
        taller = orden.taller
        
        productos = data_list['producto']
        cantidad = data_list['cantidad']
        num_ecos = data_list['num_economicos']
    
        print(productos)
        print(cantidad)
        print(num_ecos)
        cont = 0
        main_cont = 0
        templeta_data = []
        for num in cantidad:
            produc = productos[main_cont]
            nums_ecos = []
            for i in range(0, int(num)):
                #print(i)
                #print(produc)
                #print(num_ecos[cont])
                nums_ecos.append(num_ecos[cont])
                cont += 1
            templeta_data.append(
            {
                'producto': produc,
                'nums_ecos': nums_ecos
            }
            )
            main_cont += 1
        
        return render(request, 'ordenLlanta.html', 
                        { 
                            "fecha": fecha,
                            "folio": folio,
                            "id": orden.id,
                            "taller": taller,
                            "total": cont,
                            "status": status,
                            "templeta_data": templeta_data,
                            'perfil': perfil
                          }
                      )
    elif status == 'AntesRenovar':
        
        #Formateo de la data
        
        #ids = functions.int_list_element(ids)
        #llantas = Llanta.objects.filter(pk__in = ids)
        
        fecha = date.today()
        renovador = orden.renovador
        data = json.loads((orden.datos).replace("\'", "\""))
        ids = data['productos'][0]['ids']
        llantas = Llanta.objects.filter(pk__in = ids)
        print(llantas)
        print(renovador)
        print(ids)
        
        
        return render(request, 'ordenSalidaRen.html', 
                        { 
                            'llantas': llantas,
                            'fecha': fecha,
                            'perfil': perfil,
                            'renovador': renovador,
                            'ids': ids,
                            'historico': True
                          }
                      )
    elif status == 'CambioTaller':
        return redirect('dashboards:ordenEntradaTaller', orden.id)
    
    elif status == 'CambioStock':
        return redirect('dashboards:ordenEntradaStock', orden.id)
    
    elif status == 'ConRenovador':
        url = f"%s?id={orden.id}" % reverse('dashboards:ordenEntrada')
        return redirect(url)
    
def descargarCatalogoProducto(request):
    usuario = request.user
    perfil = Perfil.objects.get(user = usuario)
    compania = perfil.compania
    productos = Producto.objects.filter(compania = compania)
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append( (
        "Marca",
        "Dibujo",
        "Rango",
        "Dimencion",
        "Producto",
        "Profundidad Inicial",
        "Aplicacion",
        "Vida",
        "Precio actual",
        "KM esperado"
        ) )
    fuente = Font(bold=True)
    celdas_en_negrita = sheet["A1":"J1"]
    for row in celdas_en_negrita:
        for cell in row:
            cell.font = fuente

    
    for producto in productos:
        sheet.append( (
            producto.marca,
            producto.dibujo,
            producto.rango,
            producto.dimension,
            producto.producto,
            producto.profundidad_inicial,
            producto.aplicacion,
            producto.vida,
            producto.precio,
            producto.km_esperado,
        ) )
    sheet.column_dimensions [ 'A' ]. ancho = 500
    sheet.column_dimensions [ 'B' ]. ancho = 500
    sheet.column_dimensions [ 'C' ]. ancho = 500
    sheet.column_dimensions [ 'D' ]. ancho = 500
    sheet.column_dimensions [ 'E' ]. ancho = 500
    sheet.column_dimensions [ 'F' ]. ancho = 500
    sheet.column_dimensions [ 'G' ]. ancho = 500
    sheet.column_dimensions [ 'H' ]. ancho = 500
    sheet.column_dimensions [ 'I' ]. ancho = 500
    sheet.column_dimensions [ 'J' ]. ancho = 500
    response = HttpResponse(content_type='application/msexcel')
    response['Content-Disposition'] = f'attachment; filename=Catalogo de Productos {compania}.xlsx'
    book.save(response)
    return response


def descargarCatalogoVehiculos(request):
    usuario = request.user
    perfil = Perfil.objects.get(user = usuario)
    compania = perfil.compania
    vehiculos = Vehiculo.objects.filter(compania = compania)
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append( (
        "Numero Economico",
        "Modelo",
        "Marca",
        "Aplicacion",
        "clase",
        "configuracion",
        "Presion Establecida 1",
        "Presion Establecida 2",
        "Presion Establecida 3",
        "Presion Establecida 4",
        'Presion Establecida 5',
        'Presion Establecida 6',
        'Presion Establecida 7',
        'km diario maximo',
        'Estado',
        'Nuevo',
        ) )
    fuente = Font(bold=True)
    celdas_en_negrita = sheet["A1":"J1"]
    for row in celdas_en_negrita:
        for cell in row:
            cell.font = fuente

    
    for vehiculo in vehiculos:
        if vehiculo.estatus_activo == True:
            status = "Activo"
        else:
            status = "Inactivo"
    
        if vehiculo.nuevo == True:
            new = "Nuevo"
        else: 
            new = "Usado"
            
        sheet.append( (
            vehiculo.numero_economico,
            vehiculo.modelo,
            vehiculo.marca,
            vehiculo.aplicacion.nombre,
            vehiculo.clase,
            vehiculo.configuracion,
            vehiculo.presion_establecida_1,
            vehiculo.presion_establecida_2,
            vehiculo.presion_establecida_3,
            vehiculo.presion_establecida_4,
            vehiculo.presion_establecida_5,
            vehiculo.presion_establecida_6,
            vehiculo.presion_establecida_7,
            vehiculo.km_diario_maximo,
            status,
            new,
        ) ) 
    sheet.column_dimensions [ 'A' ]. ancho = 500
    sheet.column_dimensions [ 'B' ]. ancho = 500
    sheet.column_dimensions [ 'C' ]. ancho = 500
    sheet.column_dimensions [ 'D' ]. ancho = 500
    sheet.column_dimensions [ 'E' ]. ancho = 500
    sheet.column_dimensions [ 'F' ]. ancho = 500
    sheet.column_dimensions [ 'G' ]. ancho = 500
    sheet.column_dimensions [ 'H' ]. ancho = 500
    sheet.column_dimensions [ 'I' ]. ancho = 500
    sheet.column_dimensions [ 'J' ]. ancho = 500
    response = HttpResponse(content_type='application/msexcel')
    response['Content-Disposition'] = f'attachment; filename=Catalogo de vehiculos {compania}.xlsx'
    book.save(response)
    return response


def descargarCatalogoPresupuesto(request):
    usuario = request.user
    perfil = Perfil.objects.get(user = usuario)
    compania = perfil.compania
    Presupuestos = Presupuesto.objects.filter(compania = compania)
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append( (
        "Mes y año",
        "Compania",
        "Sucursal",
        "Presupuesto",
        "Gasto Real",
        "Kilometros Recorridos",
        ) )
    fuente = Font(bold=True)
    celdas_en_negrita = sheet["A1":"J1"]
    for row in celdas_en_negrita:
        for cell in row:
            cell.font = fuente

    
    for presupuesto in Presupuestos:
        sheet.append( (
            presupuesto.mes_ano ,
            presupuesto.compania.compania,
            presupuesto.ubicacion.nombre,
            presupuesto.presupuesto,
            presupuesto.gasto_real,
            presupuesto.km_recorridos,
        ) ) 
    sheet.column_dimensions [ 'A' ]. ancho = 500
    sheet.column_dimensions [ 'B' ]. ancho = 500
    sheet.column_dimensions [ 'C' ]. ancho = 500
    sheet.column_dimensions [ 'D' ]. ancho = 500
    sheet.column_dimensions [ 'E' ]. ancho = 500
    sheet.column_dimensions [ 'F' ]. ancho = 500
    
    response = HttpResponse(content_type='application/msexcel')
    response['Content-Disposition'] = f'attachment; filename=Catalogo de vehiculos {compania}.xlsx'
    book.save(response)
    return response


def descargarCatalogoProductoVacio(request):
    usuario = request.user
    perfil = Perfil.objects.get(user = usuario)
    compania = perfil.compania
    productos = Producto.objects.filter(compania = compania)
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append( (
        "Marca",
        "Dibujo",
        "Rango",
        "Dimencion",
        "Profundidad Inicial",
        "Aplicacion",
        "Vida",
        "Precio actual",
        "KM esperado"
        ) )
    fuente = Font(bold=True)
    celdas_en_negrita = sheet["A1":"J1"]
    for row in celdas_en_negrita:
        for cell in row:
            cell.font = fuente
            
    sheet.column_dimensions [ 'A' ]. ancho = 500
    sheet.column_dimensions [ 'B' ]. ancho = 500
    sheet.column_dimensions [ 'C' ]. ancho = 500
    sheet.column_dimensions [ 'D' ]. ancho = 500
    sheet.column_dimensions [ 'E' ]. ancho = 500
    sheet.column_dimensions [ 'F' ]. ancho = 500
    sheet.column_dimensions [ 'G' ]. ancho = 500
    sheet.column_dimensions [ 'H' ]. ancho = 500
    sheet.column_dimensions [ 'I' ]. ancho = 500
    sheet.column_dimensions [ 'J' ]. ancho = 500
    response = HttpResponse(content_type='application/msexcel')
    response['Content-Disposition'] = f'attachment; filename=Catalogo de Productos.xlsx'
    book.save(response)
    return response