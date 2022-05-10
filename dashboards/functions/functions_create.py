# Utilities
import csv
from dashboards.models import Aplicacion, Bitacora, Compania, Inspeccion, Llanta, Perfil, Producto, Ubicacion, Vehiculo
from datetime import date, datetime
from openpyxl import load_workbook
from itertools import count
import json
from random import sample, randint, uniform, random
import os
import openpyxl
import pandas as pd

def borrar_ubicaciones():
    ubicaciones = Ubicacion.objects.filter(compania=Compania.objects.get(compania="Tramo"))
    while ubicaciones.count():
        ids = ubicaciones.values_list('pk', flat=True)[:100]
        Ubicacion.objects.filter(pk__in = ids, compania=Compania.objects.get(compania="Tramo")).delete()

def borrar_llantas():
    llantas = Llanta.objects.filter(compania=Compania.objects.get(compania="Tramo"))
    while llantas.count():
        ids = llantas.values_list('pk', flat=True)[:100]
        Llanta.objects.filter(pk__in = ids, compania=Compania.objects.get(compania="Tramo")).delete()

def borrar_km_actuales():
    llantas = Llanta.objects.filter(compania=Compania.objects.get(compania="Tramo"))
    for llanta in llantas:
        llanta.km_actual = None
        llanta.save()

def convertir_vehiculos():
    vehiculos = Vehiculo.objects.filter(compania=Compania.objects.get(compania="Tramo"))
    for vehiculo in vehiculos:
        vehiculo.tirecheck = False
        vehiculo.save()



def borrar_productos():
    productos = Producto.objects.filter(compania=None)
    while productos.count():
        ids = productos.values_list('pk', flat=True)[:100]
        Producto.objects.filter(pk__in = ids, compania=None).delete()


def borrar_inspecciones():
    inspecciones = Inspeccion.objects.filter(llanta__compania=Compania.objects.get(compania="Tramo"))
    while inspecciones.count():
        ids = inspecciones.values_list('pk', flat=True)[:100]
        Inspeccion.objects.filter(pk__in = ids, llanta__compania=Compania.objects.get(compania="Tramo")).delete()

def borrar_vehiculos():
    vehiculos = Vehiculo.objects.filter(compania=Compania.objects.get(compania="Tramo"))
    while vehiculos.count():
        ids = vehiculos.values_list('pk', flat=True)[:100]
        Vehiculo.objects.filter(pk__in = ids, compania=Compania.objects.get(compania="Tramo")).delete()

def configurar_producto():
    productos = Producto.objects.all()
    for producto in productos:
        palabras = str(producto).split()
        marca = ""
        dibujo = ""
        rango = ""
        dimension = ""
        try:
            if (".5" in palabras[4]) or ("R17" in palabras[4]):
                marca = palabras[1]
                dibujo = palabras[2]
                rango = palabras[3]
                dimension = palabras[4]                
            else:
                marca = palabras[-4]
                dibujo = palabras[-3]
                rango = palabras[-2]
                dimension = palabras[-1]
        except:
            try:
                if (".5" in palabras[3]) or ("R17" in palabras[3]):
                    marca = palabras[0]
                    dibujo = palabras[1]
                    rango = palabras[2]
                    dimension = palabras[3]
                elif (".5" in palabras[2]) or ("R17" in palabras[2]):
                    marca = palabras[0]
                    dibujo = palabras[1]
                    dimension = palabras[2]
                    rango = palabras[3]
                elif (".5" in palabras[1]) or ("R17" in palabras[1]):
                    marca = palabras[0]
                    dimension = palabras[1]
                    dibujo = palabras[2]
                    rango = palabras[3]
            except:
                try:
                    if (".5" in palabras[2]) or ("R17" in palabras[2]):
                        marca = palabras[0]
                        dibujo = palabras[1]
                        rango = ""
                        dimension = palabras[2]
                except:
                    try:
                        marca = palabras[0]
                        dibujo = palabras[1]
                    except:
                        pass
        producto.marca = marca
        producto.dibujo = dibujo
        producto.rango = rango
        producto.dimension = dimension
        producto.save()

def crear_clase(clase):
    clases = Vehiculo.opciones_clase
    for c in clases:
        if c[0] == clase.upper():
            return None
    clase_mayuscula = clase.upper()
    path = os.path.abspath(os.getcwd()) + "\dashboards\models.py"
    with open(path) as file:
        lines = file.readlines()
        lines.insert(135, f'                    ("{clase_mayuscula}", "{clase}"),\n')
        file.close()

    archivo_nuevo = open(path, "w")
    archivo_nuevo.writelines(lines)
    archivo_nuevo.close()

def crear_km():
    llantas = Llanta.objects.filter(compania=Compania.objects.get(compania="pruebacal"))
    for llanta in llantas:
        llanta.km_actual = 9000
        llanta.save()

def crear_km_montado():
    llantas = Llanta.objects.filter(compania=Compania.objects.get(compania="pruebacarlos"))
    for llanta in llantas:
        llanta.km_montado = None
        llanta.save()        
    llantas = Llanta.objects.filter(compania=Compania.objects.get(compania="pruebacarlos"), vehiculo__numero_economico="C1")
    i = 0
    for llanta in llantas:
        if i < 2:
            llanta.km_montado = 50000
            llanta.save()
        else:
            llanta.km_montado = 0
            llanta.save()
        i += 1

def crear_inspecciones(user):
    llantas = Llanta.objects.filter(compania=Compania.objects.get(compania="pruebacarlos"), vehiculo__numero_economico="C3")
    loop = 0
    for llanta in llantas:
        """if llanta.profundidad_izquierda:
            profundidad_izquierda= llanta.profundidad_izquierda - 0.5
            profundidad_central= llanta.profundidad_central - 0.5
            profundidad_derecha= llanta.profundidad_derecha - 0.5
        else:
            profundidad_izquierda=round(uniform(4,6), 1)
            profundidad_central=round(uniform(4,6), 1)
            profundidad_derecha=round(uniform(4,6), 1)"""
        loop += 1
        if loop <= 1:
            profundidad_izquierda=4
        elif loop <= 2:
            profundidad_izquierda=4.8
        elif loop <= 3:
            profundidad_izquierda=4.9
        elif loop <= 4:
            profundidad_izquierda=4.6
        elif loop <= 5:
            profundidad_izquierda=5.5
        elif loop <= 6:
            profundidad_izquierda=5.4
        elif loop <= 7:
            profundidad_izquierda=5.6
        elif loop <= 8:
            profundidad_izquierda=6

        #profundidad_central=round(uniform(1,3), 1)
        #profundidad_derecha=round(uniform(1,3), 1)
        km_vehiculo = 120000
        if llanta == "1" or llanta == "2":
            inspeccion_creada = Inspeccion.objects.create(tipo_de_evento="Inspección",
                                    llanta=llanta,
                                    usuario=user,
                                    vehiculo=llanta.vehiculo,
                                    fecha_hora=date.today(),
                                    vida=llanta.vida,
                                    km_vehiculo=km_vehiculo,
                                    profundidad_izquierda=profundidad_izquierda,
                                    #profundidad_central=profundidad_central,
                                    #profundidad_derecha=profundidad_derecha,
                                    evento=str({\
                        "llanta_inicial" : llanta.id, "llanta_mod" : "",\
                        "producto_inicial" : str(llanta.producto), "producto_mod" : "",\
                        "vida_inicial" : llanta.vida, "vida_mod" : "",\
                        "km_inicial" : km_vehiculo, "km_mod" : "",\
                        "presion_inicial" : "", "presion_mod" : "",\
                        "profundidad_izquierda_inicial" : profundidad_izquierda, "profundidad_izquierda_mod" : "",\
                        "profundidad_central_inicial" : "", "profundidad_central_mod" : "",\
                        "profundidad_derecha_inicial" : "", "profundidad_derecha_mod" : ""\
                        })
            )
        else:
            inspeccion_creada = Inspeccion.objects.create(tipo_de_evento="Inspección",
                                    llanta=llanta,
                                    usuario=user,
                                    vehiculo=llanta.vehiculo,
                                    fecha_hora=date.today(),
                                    vida=llanta.vida,
                                    km_vehiculo=km_vehiculo,
                                    profundidad_izquierda=profundidad_izquierda,
                                    #profundidad_central=profundidad_central,
                                    #profundidad_derecha=profundidad_derecha,
                                    evento=str({\
                        "llanta_inicial" : llanta.id, "llanta_mod" : "",\
                        "producto_inicial" : str(llanta.producto), "producto_mod" : "",\
                        "vida_inicial" : llanta.vida, "vida_mod" : "",\
                        "km_inicial" : km_vehiculo, "km_mod" : "",\
                        "presion_inicial" : "", "presion_mod" : "",\
                        "profundidad_izquierda_inicial" : profundidad_izquierda, "profundidad_izquierda_mod" : "",\
                        "profundidad_central_inicial" : "", "profundidad_central_mod" : "",\
                        "profundidad_derecha_inicial" : "", "profundidad_derecha_mod" : ""\
                        })
            )
        llanta.profundidad_izquierda = profundidad_izquierda
        #llanta.profundidad_central = profundidad_central
        #llanta.profundidad_derecha = profundidad_derecha
        #llanta.km_actual = km_vehiculo - llanta.km_montado
        llanta.save()
        vehiculo = Vehiculo.objects.get(numero_economico=llanta.vehiculo.numero_economico)
        vehiculo.ultima_inspeccion = inspeccion_creada
        vehiculo.save()

def crear_profundidad(user):
    llantas = Llanta.objects.filter(compania=Compania.objects.get(compania="pruebacal"))
    for llanta in llantas:
        p1 = llanta.profundidad_izquierda
        p2 = llanta.profundidad_central
        p3 = llanta.profundidad_derecha
        random1 = round(uniform(0,2), 1)
        llanta.profundidad_izquierda = round(p1 - random1, 1)
        random2 = round(uniform(0,2), 1)
        llanta.profundidad_central = round(p2 - random2, 1)
        random3 = round(uniform(0,2), 1)
        llanta.profundidad_derecha = round(p3 - random3, 1)
        llanta.save()
        inspeccion = Inspeccion.objects.create(tipo_de_evento="Inspección",
                                llanta=llanta,
                                usuario=user,
                                vehiculo=llanta.vehiculo,
                                fecha_hora=date(2022, 4, 18),
                                vida="Nueva",
                                km=9000,
                                profundidad_izquierda=round(p1 - random1, 1),
                                profundidad_central=round(p2 - random2, 1),
                                profundidad_derecha=round(p3 - random3, 1),
                                evento=str({\
                                    "llanta_inicial" : llanta, "llanta_mod" : "",\
                                    "producto_inicial" : llanta.producto, "producto_mod" : "",\
                                    "vida_inicial" : llanta.vida, "vida_mod" : "",\
                                    "km_inicial" : 9000, "km_mod" : "",\
                                    "presion_inicial" : "", "presion_mod" : "",\
                                    "profundidad_izquierda_inicial" : round(p1 - random1, 1), "profundidad_izquierda_mod" : "",\
                                    "profundidad_central_inicial" : round(p2 - random2, 1), "profundidad_central_mod" : "",\
                                    "profundidad_derecha_inicial" : round(p3 - random3, 1), "profundidad_derecha_mod" : ""\
                                    })
        )
        llanta.ultima_inspeccion = inspeccion
        llanta.save()

def tirecheck_llanta():
    llantas = Llanta.objects.filter(compania=Compania.objects.get(compania="New Pick"))
    for llanta in llantas:
        llanta.tirecheck = True
        llanta.save()

def crear_ultima_inspeccion():
    llantas = Llanta.objects.filter(compania=Compania.objects.get(compania="AGA"))
    for llanta in llantas:
        if llanta.ultima_inspeccion:
            llanta.ultima_inspeccion = None
            llanta.save()

def borrar_ultima_inspeccion_llanta():
    llantas = Llanta.objects.filter(compania=Compania.objects.get(compania="New Pick"))
    for llanta in llantas:
        if llanta.ultima_inspeccion:
            llanta.ultima_inspeccion = None
            llanta.save()

def borrar_ultima_inspeccion_vehiculo():
    vehiculos = Vehiculo.objects.filter(compania=Compania.objects.get(compania="Tramo"))
    for vehiculo in vehiculos:
        if vehiculo.ultima_inspeccion:
            vehiculo.ultima_inspeccion = None
            vehiculo.save()

def crear_configuracion():
    vehiculos = Vehiculo.objects.filter(compania=Compania.objects.get(compania="Bisonte"))
    for vehiculo in vehiculos:
        if vehiculo.clase == "UTILITARIO TALLER" or vehiculo.clase == "UTILITARIO ADMINISTRATIVO" or vehiculo.clase == "CAMIONETA" or vehiculo.clase == "MONTACARGAS":
            vehiculo.configuracion = "S2.D2"
            vehiculo.numero_de_llantas = 4
            vehiculo.save()
        if vehiculo.clase == "AUTOTANQUE QUIMICOS" or vehiculo.clase == "AUTOTANQUE ALIMENTICIO" or vehiculo.clase == "AUTOTANQUE COMBUSTIBLE" or vehiculo.clase == "TORTHON SECO" or vehiculo.clase == "TORTHON REFRIGERADO" or vehiculo.clase == "TRACTOCAMION" or vehiculo.clase == "YARD TRUCK":
            vehiculo.configuracion = "S2.D4.D4"
            vehiculo.numero_de_llantas = 10
            vehiculo.save()
        if vehiculo.clase == "CAJA SECA 40" or vehiculo.clase == "CAJA SECA 48" or vehiculo.clase == "CAJA SECA 53" or vehiculo.clase == "PLATAFORMA 53" or vehiculo.clase == "CAJA SECA 53 (3 EJES)" or vehiculo.clase == "CAJA REFRIGERADO 48" or vehiculo.clase == "CAJA REFRIGERADO 53" or vehiculo.clase == "CAJA REFRIGERADO 40" or vehiculo.clase == "THERMOKING THORTON CAJA 25" or vehiculo.clase == "DOLLY" or vehiculo.clase == "PLATAFORMA 40" or vehiculo.clase == "CORTINA" or vehiculo.clase == "PLATAFORMA 35" or vehiculo.clase == "CORTINA 38" or vehiculo.clase == "TOLVA" or vehiculo.clase == "PORTACONTENEDOR":
            vehiculo.configuracion = "T4.T4"
            vehiculo.numero_de_llantas = 8
            vehiculo.save()
        if vehiculo.clase == "PLATAFORMA 53 (3 EJES)":
            vehiculo.configuracion = "T4.T4.T4"
            vehiculo.numero_de_llantas = 12
            vehiculo.save()
        if vehiculo.clase == "RABON":
            vehiculo.configuracion = "S2.D4"
            vehiculo.numero_de_llantas = 6
            vehiculo.save()


def crear_configuracion2():
    inspecciones = Inspeccion.objects.filter(id__in=range(1030, 1115))
    for inspeccion in inspecciones:
        inspeccion.fecha_hora = date(2022, 2, 2)
        inspeccion.save()
    

def crear_llantas():
    vehiculos = Vehiculo.objects.filter(compania=Compania.objects.get(compania="Tramo"))
    for vehiculo in vehiculos:
        posiciones = []
        ejes = vehiculo.configuracion.split(".")
        if vehiculo.configuracion == "T4.T4.SP1":
            posiciones.append("1LO")
            posiciones.append("1LI")
            posiciones.append("1RI")
            posiciones.append("1RO")
            posiciones.append("2LO")
            posiciones.append("2LI")
            posiciones.append("2RI")
            posiciones.append("2RO")
            posiciones.append("3LI")
        if vehiculo.configuracion == "T4.T4.SP2":
            posiciones.append("1LO")
            posiciones.append("1LI")
            posiciones.append("1RI")
            posiciones.append("1RO")
            posiciones.append("2LO")
            posiciones.append("2LI")
            posiciones.append("2RI")
            posiciones.append("2RO")
            posiciones.append("3LI")
            posiciones.append("3RI")
        if vehiculo.configuracion == "S2.C4.D4":
            posiciones.append("1LI")
            posiciones.append("1RI")
            posiciones.append("2LO")
            posiciones.append("2LI")
            posiciones.append("2RI")
            posiciones.append("2RO")
            posiciones.append("3LO")
            posiciones.append("3LI")
            posiciones.append("3RI")
            posiciones.append("3RO")
        if vehiculo.configuracion == "S2.D4.D4.SP1":
            posiciones.append("1LI")
            posiciones.append("1RI")
            posiciones.append("2LO")
            posiciones.append("2LI")
            posiciones.append("2RI")
            posiciones.append("2RO")
            posiciones.append("3LO")
            posiciones.append("3LI")
            posiciones.append("3RI")
            posiciones.append("3RO")
            posiciones.append("4LI")
        if vehiculo.configuracion == "T4.T4.T4.SP2":
            posiciones.append("1LO")
            posiciones.append("1LI")
            posiciones.append("1RI")
            posiciones.append("1RO")
            posiciones.append("2LO")
            posiciones.append("2LI")
            posiciones.append("2RI")
            posiciones.append("2RO")
            posiciones.append("3LO")
            posiciones.append("3LI")
            posiciones.append("3RI")
            posiciones.append("3RO")
            posiciones.append("4LI")
            posiciones.append("4RI")
        if vehiculo.configuracion == "T4.T4":
            posiciones.append("1LO")
            posiciones.append("1LI")
            posiciones.append("1RI")
            posiciones.append("1RO")
            posiciones.append("2LO")
            posiciones.append("2LI")
            posiciones.append("2RI")
            posiciones.append("2RO")
        if vehiculo.configuracion == "T4.T4.T4":
            posiciones.append("1LO")
            posiciones.append("1LI")
            posiciones.append("1RI")
            posiciones.append("1RO")
            posiciones.append("2LO")
            posiciones.append("2LI")
            posiciones.append("2RI")
            posiciones.append("2RO")
            posiciones.append("3LO")
            posiciones.append("3LI")
            posiciones.append("3RI")
            posiciones.append("3RO")
        for i in range(vehiculo.numero_de_llantas):
            posicion = posiciones[i]

            if ejes[int(posicion[0]) - 1][0] == "S":
                nombre_de_eje = "Dirección"
            if ejes[int(posicion[0]) - 1][0] == "D":
                nombre_de_eje = "Tracción"
            if ejes[int(posicion[0]) - 1][0] == "T":
                nombre_de_eje = "Arrastre"
            if ejes[int(posicion[0]) - 1][0] == "C":
                nombre_de_eje = "Loco"
            if ejes[int(posicion[0]) - 1][0] == "L":
                nombre_de_eje = "Retractil"

            Llanta.objects.create(numero_economico=f"{vehiculo}-{i}",
                                compania=vehiculo.compania,
                                vehiculo=vehiculo,
                                tipo_de_eje=ejes[int(posicion[0]) - 1],
                                eje=posicion[0],
                                posicion=posicion,
                                nombre_de_eje=nombre_de_eje,
                                fecha_de_entrada_inventario=date.today()
            )


def crear_nombre_de_eje():
    """llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania="Compania Prueba"))
    for llanta in llantas:
        num_aleatorio = sorted(sample(range(2, 5), 2))
        i = Inspeccion.objects.create(
            llanta=llanta,
            fecha_hora=datetime.today(),
            tiempo_de_inspeccion=2,
            min_profundidad=num_aleatorio[0],
            max_profundidad=num_aleatorio[1],
        )
        llanta.ultima_inspeccion = i
        llanta.save()"""

def crear_de_bitacora_el_vehiculo():
    bitacoras = Bitacora.objects.filter(compania=Compania.objects.get(compania="New Pick"))
    for b in bitacoras:
        vehiculo = Vehiculo.objects.get(numero_economico=b.numero_economico)
        vehiculo.fecha_de_inflado = b.fecha_de_inflado
        vehiculo.tiempo_de_inflado = b.tiempo_de_inflado
        vehiculo.presion_de_entrada = b.presion_de_entrada
        vehiculo.presion_de_salida = b.presion_de_salida
        vehiculo.presion_establecida = b.presion_establecida
        vehiculo.save()

def crear_1(numero_economico, compania, ubicacion, aplicacion, clase, configuracion, tiempo_de_inflado, presion_de_entrada, presion_de_salida, presion_establecida):
    """Vehiculo.objects.create(numero_economico=numero_economico,
                                modelo="Kenworth",
                                marca="International",
                                compania=Compania.objects.get(compania=compania),
                                ubicacion=Ubicacion.objects.get(nombre=ubicacion),
                                aplicacion=Aplicacion.objects.get(nombre=aplicacion),
                                clase=clase,
                                configuracion=configuracion,
                                fecha_de_inflado=date.today(),
                                tiempo_de_inflado=tiempo_de_inflado,
                                presion_de_entrada=presion_de_entrada,
                                presion_de_salida=presion_de_salida,
                                presion_establecida=presion_establecida
                            )
    Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=numero_economico),
        compania=Compania.objects.get(compania=Compania.objects.get(compania=compania)),
        fecha_de_inflado=date.today(),
        tiempo_de_inflado=tiempo_de_inflado,
        presion_de_entrada=presion_de_entrada,
        presion_de_salida=presion_de_salida,
        presion_establecida=presion_establecida
        )"""

def crear_2(vehiculo):
    """num_llantas = 0
    for i in vehiculo.configuracion:
        try:
            i = int(i)
            num_llantas += i
        except:
            pass
    iteracion_ejes = {}
    ejes = vehiculo.configuracion.split(".")
    iteracion_ejes[1] = [ejes[0], int(ejes[0][1])]
    iteracion_ejes[2] = [ejes[1], int(ejes[1][1])]
    try:
        iteracion_ejes[ejes[2]] = [ejes[2], int(ejes[2][1])]
    except:
        pass
    familia_ejes = 0
    iteracion_familia_ejes = 0
    posiciones = ["LI", "RI", "LO", "RO"]
    for i in range(1, num_llantas + 1):            
        Llanta.objects.create(numero_economico=f"{vehiculo}-{i}",
            usuario=Perfil.objects.get(user__username="pruebacal"),
            vehiculo=vehiculo,
            producto="NUEVA Michelin XDA5+ LRH 11R22.5",
            marca="Michelin",
            tipo_de_eje=iteracion_ejes[familia_ejes + 1][0],
            eje=familia_ejes + 1,
            posicion=f"{familia_ejes + 1}{posiciones[iteracion_familia_ejes]}",
            fecha_de_inflado=date.today()
        )
        iteracion_familia_ejes += 1
        if iteracion_familia_ejes == iteracion_ejes[familia_ejes + 1][1]:
            familia_ejes += 1
            iteracion_familia_ejes = 0"""

def crear_3(vehiculo):
    """llantas = Llanta.objects.filter(vehiculo=vehiculo)
    for llanta in llantas:
        num_aleatorio = sorted(sample(range(2, 10), 2))

        i = Inspeccion.objects.create(
            llanta=llanta,
            fecha_hora=datetime(2021, 11, 20, 12, 0, 1),
            tiempo_de_inspeccion=2,
            min_profundidad=num_aleatorio[0],
            max_profundidad=num_aleatorio[1],
        )
        llanta.ultima_inspeccion = i
        llanta.save()"""

def crear_llanta(num_aleatorio, iteracion_llantas_totales, vehiculo, eje, iteracion_ejes, posicion, user):
    """presion_encontrada = randint(80, 100)
    tiempo_de_inspeccion = round(uniform(2, 3), 1)
    profundidad = [randint(2, 8), randint(2, 8)]
    primera_profundidad = [randint(2, 8), randint(2, 8)]
    profundidad.sort()
    primera_profundidad.sort()
    inicio = datetime(2021, 8, 1)
    final =  datetime(2021, 11, 3)
    fecha_hora_de_inspeccion = inicio + (final - inicio) * random()

    vehiculo.fecha_hora_de_inspeccion = fecha_hora_de_inspeccion
    vehiculo.save()

    Llanta.objects.create(numero_economico=f"AETO-LLANTA-{num_aleatorio[iteracion_llantas_totales]}",
                            usuario=Perfil.objects.get(user=user),
                            vehiculo=vehiculo,
                            producto="Bridgstone A1",
                            marca="Freightliner",
                            tipo_de_eje=eje,
                            eje=iteracion_ejes,
                            posicion=posicion,
                            presion_encontrada=presion_encontrada,
                            presion_actual=vehiculo.presion_de_entrada,
                            presion_establecida=vehiculo.presion_establecida,
                            fecha_de_inflado=vehiculo.fecha_de_inflado,
                            fecha_hora_de_inspeccion=fecha_hora_de_inspeccion,
                            tiempo_de_inspeccion=tiempo_de_inspeccion,
                            primera_min_profundidad=primera_profundidad[0],
                            primera_max_profundidad=primera_profundidad[1],
                            min_profundidad=profundidad[0],
                            max_profundidad=profundidad[1],
                            )"""

def crear_llanta_sin_vehiculo(user):

    """for i in range(1, 51):
        tiempo_de_inspeccion = round(uniform(2, 3), 1)
        profundidad = [randint(2, 8), randint(2, 8)]
        primera_profundidad = [randint(2, 8), randint(2, 8)]
        profundidad.sort()
        primera_profundidad.sort()
        inicio = datetime(2021, 8, 1)
        final =  datetime(2021, 11, 3)
        fecha_hora_de_inspeccion = inicio + (final - inicio) * random()
        producto = randint(1, 3)
        vida = randint(1, 3)
        inventario = randint(1, 6)

        if vida == 1:
            vida = "Nueva"
        elif vida == 2:
            vida = "1R"
        elif vida == 3:
            vida = "2R"

        if vida == "Nueva":
            inventario = "Nueva"
        else:
            if inventario == 1:
                inventario = "Antes de Renovar"
            if inventario == 2:
                inventario = "Antes de Desechar"
            if inventario == 3:
                inventario = "Renovada"
            if inventario == 4:
                inventario = "Con renovador"
            if inventario == 5:
                inventario = "Desecho Final"
            if inventario == 6:
                inventario = "Servicio"


        Llanta.objects.create(numero_economico=f"PruebaCal-{i}",
                                usuario=Perfil.objects.get(user=user),
                                producto=Producto.objects.get(id=producto),
                                vida=vida,
                                inventario=inventario
                                )"""

def crear_producto():
    """llantas = Llanta.objects.filter(vehiculo__compania=Compania.objects.get(compania="pruebacal"))
    for llanta in llantas:

        if llanta.numero_economico[0:2] == "P1":
            llanta.nombre_de_eje = "Arrastre"
            llanta.save()
        if llanta.numero_economico[0:2] == "P2":
            llanta.nombre_de_eje = "Arrastre"
            llanta.save()
        if llanta.numero_economico[0:2] == "P3":
            llanta.nombre_de_eje = "Arrastre"
            llanta.save()
        if llanta.numero_economico[0:2] == "P4":
            llanta.nombre_de_eje = "Arrastre"
            llanta.save()"""

def crear_mm():
    """inspecciones = Inspeccion.objects.filter(llanta__vehiculo__compania=Compania.objects.get(compania="pruebacal"))
    for inspeccion in inspecciones:
        num_aleatorio = sorted(sample(range(4, 10), 2))
        
        inspeccion.min_profundidad = num_aleatorio[0]
        inspeccion.max_profundidad = num_aleatorio[1]
        inspeccion.save()"""

def crear_fecha():
    """inspecciones = Inspeccion.objects.filter(id__in=[115, 116, 117, 118, 119, 120, 121, 122])
    for inspeccion in inspecciones:
        inspeccion.fecha_hora = date(2022, 1, 30)
        inspeccion.save()"""

def excel(user):
    
    """num_aleatorio = sample(range(1, 100000), 36000)
    vehiculos = Vehiculo.objects.filter(compania=Compania.objects.get(compania="Compania Prueba"))
    iteracion_vehiculo = 0
    iteracion_llantas_totales = 0
    iteracion_configuraciones = ["S2.D2", "S2.D2.D2", "S2.D4", "S2.D2.D2.T4.T4"]
    for vehiculo in vehiculos:
        iteracion_ejes = {}
        if iteracion_vehiculo < 400:
            vehiculo.configuracion = iteracion_configuraciones[0]
            vehiculo.save()
            ejes = iteracion_configuraciones[0].split(".")
            iteracion_ejes[ejes[0]] = [int(ejes[0][1]), 1]
            iteracion_ejes[ejes[1]] = [int(ejes[1][1]), 2]
        elif iteracion_vehiculo < 1800:
            vehiculo.configuracion = iteracion_configuraciones[1]
            vehiculo.save()
            ejes = iteracion_configuraciones[1].split(".")
            iteracion_ejes[ejes[0]] = [int(ejes[0][1]), 1]
            iteracion_ejes[ejes[1]] = [int(ejes[1][1]), 2]
            iteracion_ejes[ejes[2]] = [int(ejes[2][1]), 3]
        elif iteracion_vehiculo < 2600:
            vehiculo.configuracion = iteracion_configuraciones[2]
            vehiculo.save()
            ejes = iteracion_configuraciones[2].split(".")
            iteracion_ejes[ejes[0]] = [int(ejes[0][1]), 1]
            iteracion_ejes[ejes[1]] = [int(ejes[1][1]), 2]
        else:
            vehiculo.configuracion = iteracion_configuraciones[3]
            vehiculo.save()
            ejes = iteracion_configuraciones[3].split(".")
            iteracion_ejes[ejes[0]] = [int(ejes[0][1]), 1]
            iteracion_ejes[ejes[1]] = [int(ejes[1][1]), 2]
            iteracion_ejes[ejes[2]] = [int(ejes[2][1]), 3]
            iteracion_ejes[ejes[3]] = [int(ejes[3][1]), 4]

        for eje in iteracion_ejes:

            if iteracion_ejes[eje][0] == 2:
                posicion = f"{iteracion_ejes[eje][1]}RI"
                crear_llanta(num_aleatorio, iteracion_llantas_totales, vehiculo, eje, iteracion_ejes[eje][1], posicion, user)
                iteracion_llantas_totales += 1
                posicion = f"{iteracion_ejes[eje][1]}LI"
                crear_llanta(num_aleatorio, iteracion_llantas_totales, vehiculo, eje, iteracion_ejes[eje][1], posicion, user)
                iteracion_llantas_totales += 1
            elif iteracion_ejes[eje][0] == 4:
                posicion = f"{iteracion_ejes[eje][1]}RI"
                crear_llanta(num_aleatorio, iteracion_llantas_totales, vehiculo, eje, iteracion_ejes[eje][1], posicion, user)
                iteracion_llantas_totales += 1
                posicion = f"{iteracion_ejes[eje][1]}RO"
                crear_llanta(num_aleatorio, iteracion_llantas_totales, vehiculo, eje, iteracion_ejes[eje][1], posicion, user)
                iteracion_llantas_totales += 1
                posicion = f"{iteracion_ejes[eje][1]}LI"
                crear_llanta(num_aleatorio, iteracion_llantas_totales, vehiculo, eje, iteracion_ejes[eje][1], posicion, user)
                iteracion_llantas_totales += 1
                posicion = f"{iteracion_ejes[eje][1]}LO"
                crear_llanta(num_aleatorio, iteracion_llantas_totales, vehiculo, eje, iteracion_ejes[eje][1], posicion, user)
                iteracion_llantas_totales += 1
        iteracion_vehiculo += 1"""

    """num_aleatorio = sample(range(1, 100000), 3000)
    iteracion = 0
    for num in num_aleatorio:
        iteracion += 1
        if iteracion <= 400:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Kenworth",
                                        marca="International",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="CAMIONETA",
                                        ubicacion=Ubicacion.objects.get(nombre="Torreón Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Camionera Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.4,
                                        presion_de_entrada=100,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.4,
                presion_de_entrada=100,
                presion_de_salida=100,
                )
        elif iteracion <= 500:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Tesla",
                                        marca="International",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="AUTOTANQUE ALIMENTICIO",
                                        ubicacion=Ubicacion.objects.get(nombre="Torreón Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Camionera Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.3,
                                        presion_de_entrada=100,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.4,
                presion_de_entrada=100,
                presion_de_salida=100,
            )
        elif iteracion <= 640:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Tesla",
                                        marca="International",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="AUTOTANQUE ALIMENTICIO",
                                        ubicacion=Ubicacion.objects.get(nombre="Torreón Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Camionera Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.0,
                                        presion_de_entrada=75,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.0,
                presion_de_entrada=75,
                presion_de_salida=100,
            )
        elif iteracion <= 700:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Cascadia",
                                        marca="Freightliner",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="AUTOTANQUE ALIMENTICIO",
                                        ubicacion=Ubicacion.objects.get(nombre="Torreón Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Camionera Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.0,
                                        presion_de_entrada=100,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.0,
                presion_de_entrada=100,
                presion_de_salida=100,
            )
        elif iteracion <= 720:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Cascadia",
                                        marca="Freightliner",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="AUTOTANQUE ALIMENTICIO",
                                        ubicacion=Ubicacion.objects.get(nombre="Torreón Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Camionera Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.8,
                                        presion_de_entrada=80,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.8,
                presion_de_entrada=80,
                presion_de_salida=100,
            )
        elif iteracion <= 720:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Cascadia",
                                        marca="Freightliner",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="AUTOTANQUE ALIMENTICIO",
                                        ubicacion=Ubicacion.objects.get(nombre="Torreón Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Transportes Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.7,
                                        presion_de_entrada=82,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.7,
                presion_de_entrada=82,
                presion_de_salida=100,
            )
        elif iteracion <= 790:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Cascadia",
                                        marca="Freightliner",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="AUTOTANQUE ALIMENTICIO",
                                        ubicacion=Ubicacion.objects.get(nombre="Torreón Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Transportes Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.5,
                                        presion_de_entrada=100,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.5,
                presion_de_entrada=100,
                presion_de_salida=100,
            )
        elif iteracion <= 1200:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Cascadia",
                                        marca="Freightliner",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="REMOLQUE",
                                        ubicacion=Ubicacion.objects.get(nombre="Monterrey Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Transportes Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.5,
                                        presion_de_entrada=100,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.5,
                presion_de_entrada=100,
                presion_de_salida=100,
            )
        elif iteracion <= 2200:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Cascadia",
                                        marca="Freightliner",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="TOLVA",
                                        ubicacion=Ubicacion.objects.get(nombre="Monterrey Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Transportes Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.2,
                                        presion_de_entrada=100,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.2,
                presion_de_entrada=100,
                presion_de_salida=100,
            )
        elif iteracion <= 3000:
            Vehiculo.objects.create(numero_economico=f"AETO-{num}",
                                        modelo="Cascadia",
                                        marca="Freightliner",
                                        compania=Compania.objects.get(compania="Compania Prueba"),
                                        clase="TOLVA",
                                        ubicacion=Ubicacion.objects.get(nombre="Monterrey Prueba"),
                                        aplicacion=Aplicacion.objects.get(nombre="Camionera Prueba"),
                                        fecha_de_inflado=date.today(),
                                        tiempo_de_inflado=2.6,
                                        presion_de_entrada=72,
                                        presion_de_salida=100,
                                        presion_establecida=100
                                    )
            Bitacora.objects.create(numero_economico=Vehiculo.objects.get(numero_economico=f"AETO-{num}"),
                compania=Compania.objects.get(compania=Compania.objects.get(compania="Compania Prueba")),
                fecha_de_inflado=date.today(),
                tiempo_de_inflado=2.6,
                presion_de_entrada=72,
                presion_de_salida=100,
            )"""
    
    """FILE_PATH = "D:/Descargas/Tablas AETO Tire.xlsx"
    SHEET = "Inspeccion"
    workbook = load_workbook(FILE_PATH, read_only=True)
    sheet = workbook[SHEET]

    for row in sheet.iter_rows(min_row=2):
        numero_economico_de_llanta = row[1].value
        min_profundidad = int(row[4].value)

        Inspeccion.objects.create(llanta=Llanta.objects.get(numero_economico=numero_economico_de_llanta),
                            fecha_hora=date.today(),
                            tiempo_de_inspeccion=1,
                            min_profundidad=min_profundidad,
                            max_profundidad=min_profundidad,
                            )"""

    """while vehiculos_chidos.count():
        ids = Vehiculo.objects.values_list('pk', flat=True)[:100]
        Vehiculo.objects.filter(pk__in = ids, compania=Compania.objects.get(compania="Idealease")).delete()
    """

