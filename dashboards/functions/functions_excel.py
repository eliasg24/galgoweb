# Utilities
from dashboards.models import Aplicacion, Bitacora, Compania, Inspeccion, Llanta, Perfil, Producto, Ubicacion, Vehiculo, Observacion
from datetime import date, datetime
from openpyxl import load_workbook
import csv
import os
import openpyxl
from random import randint

# Functions
from dashboards.functions import functions, functions_create

def excel_productos():
    FILE_PATH = "D:/Aetoweb/aeto/files/files/Products2022_03_25_043513.csv"
    file = open(FILE_PATH, "r", encoding="utf-8-sig", newline='')
    next(file, None)
    reader = csv.reader(file, delimiter=",")

    for row in reader:
        producto = row[4]
        try:
            producto_hecho = Producto.objects.get(producto=producto)
        except:
            producto_hecho = Producto.objects.create(producto=producto)
        if producto_hecho:
            marca = row[6]
            profundidad_inicial = int(float(row[10]))
            vida = row[9]
            if vida == "New":
                vida = "Nueva"
            elif vida == "Retread":
                vida = "Renovada"

            precio = row[12]
            if precio == "":
                precio = 2000
            else:
                precio = int(float(row[12]))
            
            producto_hecho.marca=marca
            producto_hecho.profundidad_inicial=profundidad_inicial
            producto_hecho.vida=vida
            producto_hecho.precio=precio
            producto_hecho.save()

def excel_vehiculos():
    FILE_PATH = "D:/Aetoweb/aeto/files/files/Vehicles2022_03_25_043019.csv"
    file = open(FILE_PATH, "r", encoding="utf-8-sig", newline='')
    next(file, None)
    reader = csv.reader(file, delimiter=",")

    for row in reader:
        compania = row[3].capitalize()
        if compania == "Corcelip":
            numero_economico = row[9]
            flota = row[5]
            aplicacion = row[7]
            clase = row[12]
            configuracion = row[14]
            marca = row[16]
            modelo = row[18]
            fecha_de_creacion = row[21]

            functions_create.crear_clase(clase)
            fecha_de_creacion = functions.convertir_fecha2(fecha_de_creacion)

            try:
                ubicacion = Ubicacion.objects.get(nombre=flota, compania=Compania.objects.get(compania=compania))
            except:
                ubicacion = Ubicacion.objects.create(nombre=flota, compania=Compania.objects.get(compania=compania))

            try:
                aplicacion = Aplicacion.objects.get(nombre=aplicacion, compania=Compania.objects.get(compania=compania))
            except:
                aplicacion = Aplicacion.objects.create(nombre=aplicacion, compania=Compania.objects.get(compania=compania))

            numero_de_llantas = functions.cantidad_llantas(configuracion)

            Vehiculo.objects.create(numero_economico=numero_economico,
                                modelo=modelo,
                                marca=marca,
                                compania=Compania.objects.get(compania=compania),
                                ubicacion=ubicacion,
                                aplicacion=aplicacion,
                                numero_de_llantas=numero_de_llantas,
                                clase=clase.upper(),
                                configuracion=configuracion,
                                fecha_de_creacion=fecha_de_creacion
                                )
                                
def excel_llantas_rodantes(user):
    FILE_PATH = "D:/Aetoweb/aeto/files/files/Stock2022_03_25_041155.csv"
    file = open(FILE_PATH, "r", encoding="utf-8-sig", newline='')
    next(file, None)
    reader = csv.reader(file, delimiter=",")


    for row in reader:
        compania = row[1].capitalize()
        if compania == "Corcelip":
            numero_economico = row[9]
            usuario = Perfil.objects.get(user=user)
            vehiculo = Vehiculo.objects.get(numero_economico=row[6])
            vida = row[15]
            if vida == "New":
                vida = "Nueva"
            elif vida == "1st Retread":
                vida = "1R"
            elif vida == "2st Retread":
                vida = "2R"
            elif vida == "3st Retread":
                vida = "3R"
            elif vida == "4st Retread":
                vida = "4R"
            elif vida == "Retread":
                vida = "1R"

            posicion = row[7]
            tipo_de_eje = functions.sacar_eje(int(posicion[0]), vehiculo)
            eje = int(posicion[0])
            presion_de_entrada = row[11]
            if presion_de_entrada == "":
                presion_de_entrada = None
                presion_de_salida = None
            else:
                presion_de_entrada = int(float(row[11]))
                presion_de_salida = int(float(row[11]))
                fecha_de_inflado = date.today()
            
            if tipo_de_eje[0] == "S":
                nombre_de_eje = "Dirección"
            elif tipo_de_eje[0] == "D":
                nombre_de_eje = "Tracción"
            elif tipo_de_eje[0] == "T":
                nombre_de_eje = "Arrastre"
            elif tipo_de_eje[0] == "C":
                nombre_de_eje = "Loco"
            elif tipo_de_eje[0] == "L":
                nombre_de_eje = "Retractil"

            producto = row[10]
            try:
                producto = Producto.objects.get(producto=producto)
            except:
                producto = Producto.objects.create(producto=producto)

            inventario = "Rodante"
            km_montado = row[13]
            if km_montado == "":
                km_montado = None
            else:
                km_montado = int(float(row[13]))

            Llanta.objects.create(numero_economico=numero_economico,
                                usuario=usuario,
                                compania=Compania.objects.get(compania=compania),
                                vehiculo=vehiculo,
                                ubicacion=vehiculo.ubicacion,
                                vida=vida,
                                tipo_de_eje=tipo_de_eje,
                                eje=eje,
                                posicion=posicion,
                                presion_de_entrada=presion_de_entrada,
                                presion_de_salida=presion_de_salida,
                                fecha_de_inflado=fecha_de_inflado,
                                nombre_de_eje=nombre_de_eje,
                                producto=producto,
                                inventario=inventario,
                                km_montado=km_montado,
                                )

def excel_llantas(user):
    FILE_PATH = "D:/Aetoweb/aeto/files/files/Stock2022_03_25_041155.csv"
    file = open(FILE_PATH, "r", encoding="utf-8-sig", newline='')
    next(file, None)
    reader = csv.reader(file, delimiter=",")


    i = []
    for row in reader:
        compania = row[1].capitalize()
        if compania == "Corcelip":
            numero_economico = row[7]
            try:
                llanta = Llanta.objects.filter(numero_economico=numero_economico, compania=Compania.objects.get(compania=compania))
                i.append(llanta)

            except:
                print("no", numero_economico)
                usuario = Perfil.objects.get(user=user)
                vida = row[14]
                if vida == "New":
                    vida = "Nueva"
                elif vida == "1st Retread":
                    vida = "1R"
                elif vida == "2st Retread":
                    vida = "2R"
                elif vida == "3st Retread":
                    vida = "3R"
                elif vida == "4st Retread":
                    vida = "4R"
                elif vida == "Retread":
                    vida = "1R"

                presion_de_entrada = row[9]
                if presion_de_entrada == "":
                    presion_de_entrada = None
                    presion_de_salida = None
                    fecha_de_inflado = None
                else:
                    presion_de_entrada = int(float(row[9]))
                    presion_de_salida = int(float(row[9]))
                    fecha_de_inflado = date.today()
                
                producto = row[8]
                try:
                    producto = Producto.objects.get(producto=producto)
                except:
                    producto = Producto.objects.create(producto=producto)

                inventario = row[5]
                if inventario == "RollingStock":
                    inventario = "Rodante"
                elif inventario == "ForScrapStock":
                    inventario = "Antes de Desechar"
                elif inventario == "ForServiceStock":
                    inventario = "Servicio"
                else:
                    inventario = None
                try:
                    km_montado = int(row[12])
                except:
                    km_montado = None
                Llanta.objects.create(numero_economico=numero_economico,
                                    usuario=usuario,
                                    compania=Compania.objects.get(compania=compania),
                                    vida=vida,
                                    presion_de_entrada=presion_de_entrada,
                                    presion_de_salida=presion_de_salida,
                                    fecha_de_inflado=fecha_de_inflado,
                                    producto=producto,
                                    inventario=inventario,
                                    km_montado=km_montado,
                                    )
    print("i", len(i))
    my_list = list(set(i))
    print("my_list", len(my_list))




def excel_inspecciones():
    FILE_PATH = "D:/Aetoweb/aeto/files/files/Inspections_Bulk.csv"
    file = open(FILE_PATH, "r", encoding="latin-1", newline='')
    next(file, None)
    reader = csv.reader(file, delimiter=",")

    for row in reader:
        llanta = row[12]

        try:
            llanta_hecha = Llanta.objects.get(numero_economico=llanta)
        except:
            llanta_hecha = None
        print(llanta_hecha)
        if llanta_hecha:
            fecha_hora = row[4]
            fecha_hora = functions.convertir_fecha2(fecha_hora)
            km = row[6]
            if km == "":
                km = 2000
            else:
                km = int(float(row[6]))

            profundidades = [float(row[18]), float(row[19]), float(row[20])]
            min_profundidad = min(profundidades)
            max_profundidad = max(profundidades)

            inspeccion_creada = Inspeccion.objects.create(llanta=llanta_hecha,
                                    fecha_hora=fecha_hora,
                                    km=km,
                                    min_profundidad=min_profundidad,
                                    max_profundidad=max_profundidad,
            )
            llanta_hecha.ultima_inspeccion = inspeccion_creada
            llanta_hecha.save()
            try:
                vehiculo = Vehiculo.objects.get(numero_economico=llanta_hecha.vehiculo.numero_economico)
                vehiculo.ultima_inspeccion = inspeccion_creada
                vehiculo.save()
            except:
                pass



def ExcelAeto(llanta, vehiculo, posicion, km_actual, km_proyectado, cpk, sucursal, aplicacion, clase, nomeje, producto, min_profundidad):
    wb = openpyxl.Workbook()
    reporte = wb.active
    e1 = reporte.cell(row=1, column=1, value='Llanta')
    e2 = reporte.cell(row=1, column=2, value='Vehiculo')
    e3 = reporte.cell(row=1, column=3, value='Posición')
    e4 = reporte.cell(row=1, column=4, value='Km actual')
    e5 = reporte.cell(row=1, column=5, value='Km proyectado')
    e6 = reporte.cell(row=1, column=6, value='CPK')
    e7 = reporte.cell(row=1, column=7, value='Sucursal')
    e8 = reporte.cell(row=1, column=8, value='Aplicación')
    e9 = reporte.cell(row=1, column=9, value='Clase')
    e10 = reporte.cell(row=1, column=10, value='Nombre de eje')
    e11 = reporte.cell(row=1, column=11, value='Producto')
    e12 = reporte.cell(row=1, column=12, value='Min profundidad')
    for llt in range(len(llanta)):
        llantacld = reporte.cell(row=llt + 2, column=1)
        llantacld.value = str(llanta[llt])
    for vhc in range(len(vehiculo)):
        vehiculocld = reporte.cell(row=vhc + 2, column=2)
        vehiculocld.value = str(vehiculo[vhc]["vehiculo__numero_economico"])
    for pos in range(len(posicion)):
        posicioncld = reporte.cell(row=pos + 2, column=3)
        posicioncld.value = str(posicion[pos]["posicion"])
    for kma in range(len(km_actual)):
        nomejecld = reporte.cell(row=kma + 2, column=4)
        nomejecld.value = str(km_actual[kma]["ultima_inspeccion__km"])
    for kmp in range(len(km_proyectado)):
        nomejecld = reporte.cell(row=kmp + 2, column=5)
        nomejecld.value = str(km_proyectado[kmp])
    for cp in range(len(cpk)):
        nomejecld = reporte.cell(row=cp + 2, column=6)
        nomejecld.value = str(cpk[cp])
    for suc in range(len(sucursal)):
        nomejecld = reporte.cell(row=suc + 2, column=7)
        nomejecld.value = str(sucursal[suc]["vehiculo__ubicacion__nombre"])
    for app in range(len(aplicacion)):
        nomejecld = reporte.cell(row=app + 2, column=8)
        nomejecld.value = str(aplicacion[app]["vehiculo__aplicacion__nombre"])
    for cls in range(len(clase)):
        nomejecld = reporte.cell(row=cls + 2, column=9)
        nomejecld.value = str(clase[cls]["vehiculo__clase"])
    for nde in range(len(nomeje)):
        nomejecld = reporte.cell(row=nde + 2, column=10)
        nomejecld.value = str(nomeje[nde]["nombre_de_eje"])
    for pro in range(len(producto)):
        productocld = reporte.cell(row=pro + 2, column=11)
        productocld.value = str(producto[pro]["producto__producto"])
    for min in range(len(min_profundidad)):
        productocld = reporte.cell(row=min + 2, column=12)
        productocld.value = str(min_profundidad[min]["ultima_inspeccion__min_profundidad"])
    wb.save('reporteLlantas.xlsx')

def agregarExcel():
    path = "D://Aetoweb/aeto/inspectionss.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row = 1, column=22)
    #print(cell_obj.value)
    datos_prof = [5,9,20,'Juan', 'Daniel']
    c_prec = ['El', 'Precio', 'de', 'la', 'Historia']
    min_cad = ['**[--@', 'gmail', '.com', 6777, 9999]
    max_cad = [3.1415, 6.7839, 4+5.9, 'suma'+'resta', 'multi']
    porcen = [5,4, 'setvalues', 'dictionary', 'lista']
    dinero = [50000,40000, 1000000,-5000, -100000]
    eje = ['5.67E, 48.91S','Nuevo León, México', 'QWErty asdv',33.44 * 22, 50/10]
    perdida = [3.1415-9*6.7, 6/3+7, 4+5*9-12, 'operaciones básicas', '3-20-2022']
    i = 0
    for i in range(3683):
        profundidad_inicial = randint(15, 25)
        min_profundidad = profundidad_inicial + randint(8, 13)
        tipo_prom = randint(0, 3)
        prom = min_profundidad + tipo_prom
        if tipo_prom == 0:
            porc = 0
        elif tipo_prom == 1:
            porc = 2
        elif tipo_prom == 2:
            porc = 4
        elif tipo_prom == 3:
            porc = 5
        
        precio = randint(1, 15)
        precio_list = [0, 4200, 8500, 3800, 7500, 7800, 1635, 2800, 3900, 9000, 3400, 7900, 4000, 7668, 2850, 6700, 3938]
        precio = precio_list[precio]
        dinero_perdido = precio * porc
        punto_de_retiro_list = [2, 3, 4]
        punto_de_retiro = randint(0, 2)
        punto_de_retiro = punto_de_retiro_list[punto_de_retiro]
        perdida_total = (min_profundidad - punto_de_retiro) * dinero_perdido

        prof_cld = sheet_obj.cell(row = i + 2, column=22)
        c_cld = sheet_obj.cell(row = i + 2, column=23)
        min_cld = sheet_obj.cell(row = i + 2, column=24)
        max_cld = sheet_obj.cell(row = i + 2, column=25)
        porct_cld = sheet_obj.cell(row = i + 2, column=26)
        din_cld = sheet_obj.cell(row = i + 2, column=27)
        eje_cld = sheet_obj.cell(row = i + 2, column=28)
        perd_cld = sheet_obj.cell(row = i + 2, column=29)
        prof_cld.value = profundidad_inicial
        c_cld.value = precio
        min_cld.value = min_profundidad
        max_cld.value = prom
        porct_cld.value = porc
        din_cld.value = dinero_perdido
        eje_cld.value = punto_de_retiro
        perd_cld.value = perdida_total
        #print(c_cld.value)
        i += 1
    wb_obj.save(filename = 'informedeperdidayrendimiento.xlsx')

def excel_vehiculos2():
    FILE_PATH = "C:/Users/elias/Downloads/Compañia AGA pulpo.csv"
    file = open(FILE_PATH, "r", encoding="latin-1", newline='')
    next(file, None)
    reader = csv.reader(file, delimiter=",")


    for row in reader:
        compania = row[0]
        if compania == "AGA":
            numero_economico = row[3]
            flota = row[1]
            aplicacion = row[2]
            marca = row[4]
            clase = row[6]
            presion_establecida = row[7]

            try:
                ubicacion = Ubicacion.objects.get(nombre=flota, compania=Compania.objects.get(compania=compania))
            except:
                ubicacion = Ubicacion.objects.create(nombre=flota, compania=Compania.objects.get(compania=compania))

            try:
                aplicacion = Aplicacion.objects.get(nombre=aplicacion, compania=Compania.objects.get(compania=compania))
            except:
                aplicacion = Aplicacion.objects.create(nombre=aplicacion, compania=Compania.objects.get(compania=compania))

            Vehiculo.objects.create(numero_economico=numero_economico,
                                marca=marca,
                                compania=Compania.objects.get(compania=compania),
                                ubicacion=ubicacion,
                                aplicacion=aplicacion,
                                clase=clase.upper(),
                                presion_establecida=presion_establecida
                                )

def excel_observaciones():
    FILE_PATH = "D:/aeto/Tablas AETO Tire - Catalogo de observacion.csv"
    file = open(FILE_PATH, "r", encoding="utf-8-sig", newline='')
    next(file, None)
    reader = csv.reader(file, delimiter=",")

    for row in reader:
    
        observacion = row[1]
        color = row[2]
        nivel = row[3]
        activacion = row[4]

        if color == "roja":
            color = "Rojo"
        elif color == "amarilla":
            color = "Amarillo"
        elif color == "sin color":
            color = "NA"

        if nivel == "llanta":
            nivel = "Llanta"
        if nivel == "vehiculo":
            nivel = "Vehiculo"

        if activacion == "Automatica":
            automatico = True
        else:
            automatico = False

        Observacion.objects.create(observacion=observacion,
                            color=color,
                            nivel=nivel,
                            automatico=automatico
                            )

def excel_productos():
    FILE_PATH = "C:/Users/elias/Downloads/Productos_de_neumáticos_20220426_135532.xlsx"
    wb_obj = openpyxl.load_workbook(FILE_PATH)
    sheet_obj = wb_obj.active

    for i in range(sheet_obj.max_row):
    
        producto = str(sheet_obj.cell(row=i + 2, column=1).value)
        marca = str(sheet_obj.cell(row=i + 2, column=4).value)
        dibujo = str(sheet_obj.cell(row=i + 2, column=5).value)
        dimension = str(sheet_obj.cell(row=i + 2, column=2).value)
        vida = str(sheet_obj.cell(row=i + 2, column=6).value)
        precio = str(sheet_obj.cell(row=i + 2, column=8).value)
        if vida == "Nuevo":
            vida = "Nueva"
        elif vida == "Renovado":
            vida = "Renovada"
        
        try:
            profundidad_inicial = float(str(sheet_obj.cell(row=i + 2, column=23).value))
        except:
            profundidad_inicial = None

        try:
            producto = Producto.objects.get(producto=producto)
            producto.delete()
        except:
            pass

        try:
            Producto.objects.create(producto=producto,
                                compania=Compania.objects.get(compania="Tramo"),
                                marca=marca,
                                dibujo=dibujo,
                                dimension=dimension,
                                vida=vida,
                                precio=int(precio),
                                profundidad_inicial=profundidad_inicial
                                )
        except:
            pass